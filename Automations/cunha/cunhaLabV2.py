import requests
import json
import pandas as pd
from typing import List, Dict, Tuple
import time
from tqdm import tqdm
import os
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
import seaborn as sns
from collections import Counter
import re
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
import sys
import schedule
import logging
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
from dotenv import load_dotenv

# Carregar variáveis de ambiente
load_dotenv()

# Configurar logging
logging.basicConfig(filename='D:\\OneDrive - Synvia Group\\Data Analysis\\Churn PCLs\\Automations\\cunha\\logs.txt',
                    level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')

# Caminho base para salvamento
BASE_DIR = r'D:\OneDrive - Synvia Group\Data Analysis\Churn PCLs\Automations\cunha'

MUNICIPIOS_POR_ESTADO = {
    "AC": [
        "Acrelândia", "Assis Brasil", "Brasiléia", "Bujari", "Capixaba", "Cruzeiro do Sul",
        "Epitaciolândia", "Feijó", "Jordão", "Manoel Urbano", "Marechal Thaumaturgo",
        "Mâncio Lima", "Plácido de Castro", "Porto Acre", "Porto Walter", "Rio Branco",
        "Rodrigues Alves", "Santa Rosa do Purus", "Sena Madureira", "Senador Guiomard",
        "Tarauacá", "Xapuri"
    ],
    "AL": [
        "Água Branca", "Anadia", "Arapiraca", "Atalaia", "Barra de Santo Antônio",
        "Barra de São Miguel", "Batalha", "Belém", "Belo Monte", "Boca da Mata",
        "Branquinha", "Cacimbinhas", "Cajueiro", "Campestre", "Campo Alegre",
        "Campo Grande", "Canapi", "Capela", "Carneiros", "Chã Preta", "Coité do Noia",
        "Colônia Leopoldina", "Coqueiro Seco", "Coruripe", "Craíbas", "Delmiro Gouveia",
        "Dois Riachos", "Estrela de Alagoas", "Feira Grande", "Feliz Deserto", "Flexeiras",
        "Girau do Ponciano", "Ibateguara", "Igaci", "Igreja Nova", "Inhapi",
        "Jacaré dos Homens", "Jacuípe", "Japaratinga", "Jaramataia", "Jequiá da Praia",
        "Joaquim Gomes", "Jundiá", "Junqueiro", "Lagoa da Canoa", "Limoeiro de Anadia",
        "Maceió", "Major Izidoro", "Mar Vermelho", "Maragogi", "Maravilha",
        "Marechal Deodoro", "Maribondo", "Mata Grande", "Matriz de Camaragibe", "Messias",
        "Minador do Negrão", "Monteirópolis", "Murici", "Novo Lino", "Olho d'Água das Flores",
        "Olho d'Água do Casado", "Olho d'Água Grande", "Olivença", "Ouro Branco", "Palestina",
        "Palmeira dos Índios", "Pão de Açúcar", "Pariconha", "Paripueira", "Passo de Camaragibe",
        "Paulo Jacinto", "Penedo", "Piaçabuçu", "Pilar", "Pindoba", "Piranhas",
        "Poço das Trincheiras", "Porto Calvo", "Porto de Pedras", "Porto Real do Colégio",
        "Quebrangulo", "Rio Largo", "Roteiro", "Santa Luzia do Norte", "Santana do Ipanema",
        "Santana do Mundaú", "São Brás", "São José da Laje", "São José da Tapera",
        "São Luís do Quitunde", "São Miguel dos Campos", "São Miguel dos Milagres",
        "São Sebastião", "Satuba", "Senador Rui Palmeira", "Tanque d'Arca", "Taquarana",
        "Teotônio Vilela", "Traipu", "União dos Palmares", "Viçosa"
    ],
    "AP": [
        "Amapá", "Calçoene", "Cutias", "Ferreira Gomes", "Itaubal", "Laranjal do Jari",
        "Macapá", "Mazagão", "Oiapoque", "Pedra Branca do Amapari", "Porto Grande",
        "Pracuúba", "Santana", "Serra do Navio", "Tartarugalzinho", "Vitória do Jari"
    ],
    "AM": [
        "Alvarães", "Amaturá", "Anamã", "Anori", "Apuí", "Atalaia do Norte", "Autazes",
        "Barcelos", "Barreirinha", "Benjamin Constant", "Beruri", "Boa Vista do Ramos",
        "Boca do Acre", "Borba", "Caapiranga", "Canutama", "Carauari", "Careiro",
        "Careiro da Várzea", "Coari", "Codajás", "Eirunepé", "Envira", "Fonte Boa",
        "Guajará", "Humaitá", "Ipixuna", "Iranduba", "Itacoatiara", "Itamarati",
        "Itapiranga", "Japurá", "Juruá", "Jutaí", "Lábrea", "Manacapuru", "Manaquiri",
        "Manaus", "Manicoré", "Maraã", "Maués", "Nhamundá", "Nova Olinda do Norte",
        "Novo Airão", "Novo Aripuanã", "Parintins", "Pauini", "Presidente Figueiredo",
        "Rio Preto da Eva", "Santa Isabel do Rio Negro", "Santo Antônio do Içá",
        "São Gabriel da Cachoeira", "São Paulo de Olivença", "São Sebastião do Uatumã",
        "Silves", "Tabatinga", "Tapauá", "Tefé", "Tonantins", "Uarini", "Urucará",
        "Urucurituba"
    ],
    "BA": [
        "Abaíra", "Abaré", "Acajutiba", "Adustina", "Água Fria", "Aiquara", "Alagoinhas",
        "Alcobaça", "Almadina", "Amargosa", "Amélia Rodrigues", "América Dourada", "Anagé",
        "Andaraí", "Andorinha", "Angical", "Anguera", "Antas", "Antônio Cardoso",
        "Antônio Gonçalves", "Aporá", "Apuarema", "Araçás", "Aracatu", "Araci", "Aramari",
        "Arataca", "Aratuípe", "Aurelino Leal", "Baianópolis", "Baixa Grande", "Banzaê",
        "Barra", "Barra da Estiva", "Barra do Choça", "Barra do Mendes", "Barra do Rocha",
        "Barreiras", "Barro Alto", "Barro Preto", "Barrocas", "Belmonte", "Belo Campo",
        "Biritinga", "Boa Nova", "Boa Vista do Tupim", "Bom Jesus da Lapa", "Bom Jesus da Serra",
        "Boninal", "Bonito", "Boquira", "Botuporã", "Brejões", "Brejolândia", "Brotas de Macaúbas",
        "Brumado", "Buerarema", "Buritirama", "Caatiba", "Cabaceiras do Paraguaçu", "Cachoeira",
        "Caculé", "Caém", "Caetanos", "Caetité", "Cafarnaum", "Cairu", "Caldeirão Grande",
        "Camacan", "Camaçari", "Camamu", "Campo Alegre de Lourdes", "Campo Formoso", "Canápolis",
        "Canarana", "Canavieiras", "Candeal", "Candeias", "Candiba", "Cândido Sales", "Cansanção",
        "Canudos", "Capela do Alto Alegre", "Capim Grosso", "Caraíbas", "Caravelas",
        "Cardeal da Silva", "Carinhanha", "Casa Nova", "Castro Alves", "Catolândia", "Catu",
        "Caturama", "Central", "Chorrochó", "Cícero Dantas", "Cipó", "Coaraci", "Cocos",
        "Conceição da Feira", "Conceição do Almeida", "Conceição do Coité", "Conceição do Jacuípe",
        "Conde", "Condeúba", "Contendas do Sincorá", "Coração de Maria", "Cordeiros", "Coribe",
        "Coronel João Sá", "Correntina", "Cotegipe", "Cravolândia", "Crisópolis", "Cristópolis",
        "Cruz das Almas", "Curaçá", "Dário Meira", "Dias d'Ávila", "Dom Basílio", "Dom Macedo Costa",
        "Elísio Medrado", "Encruzilhada", "Entre Rios", "Érico Cardoso", "Esplanada",
        "Euclides da Cunha", "Eunápolis", "Fátima", "Feira da Mata", "Feira de Santana",
        "Filadélfia", "Firmino Alves", "Floresta Azul", "Formosa do Rio Preto", "Gandu", "Gavião",
        "Gentio do Ouro", "Glória", "Gongogi", "Governador Mangabeira", "Guajeru", "Guanambi",
        "Guaratinga", "Heliópolis", "Iaçu", "Ibiassucê", "Ibicaraí", "Ibicoara", "Ibicuí",
        "Ibipeba", "Ibipitanga", "Ibiquera", "Ibirapitanga", "Ibirapuã", "Ibirataia", "Ibitiara",
        "Ibititá", "Ibotirama", "Ichu", "Igaporã", "Igrapiúna", "Iguaí", "Ilhéus", "Inhambupe",
        "Ipecaetá", "Ipiaú", "Ipirá", "Ipupiara", "Irajuba", "Iramaia", "Iraquara", "Irará",
        "Irecê", "Itabela", "Itaberaba", "Itabuna", "Itacaré", "Itaeté", "Itagi", "Itagibá",
        "Itagimirim", "Itaguaçu da Bahia", "Itaju do Colônia", "Itajuípe", "Itamaraju",
        "Itamari", "Itambé", "Itanagra", "Itaparica", "Itapé", "Itapebi", "Itapetinga",
        "Itapicuru", "Itapitanga", "Itaquara", "Itarantim", "Itatim", "Itiruçu", "Itiúba",
        "Itororó", "Ituaçu", "Ituberá", "Iuiú", "Jaborandi", "Jacaraci", "Jacobina", "Jaguaquara",
        "Jaguarari", "Jaguaripe", "Jandaíra", "Jequié", "Jeremoabo", "Jiquiriçá", "Jitaúna",
        "João Dourado", "Juazeiro", "Jucuruçu", "Jussara", "Jussari", "Jussiape", "Lafaiete Coutinho",
        "Lagoa Real", "Laje", "Lajedão", "Lajedinho", "Lajedo do Tabocal", "Lamarão", "Lapão",
        "Lauro de Freitas", "Lençóis", "Licínio de Almeida", "Livramento de Nossa Senhora",
        "Luís Eduardo Magalhães", "Macajuba", "Macarani", "Macaúbas", "Macururé", "Madre de Deus",
        "Maetinga", "Maiquinique", "Mairi", "Malhada", "Malhada de Pedras", "Manoel Vitorino",
        "Mansidão", "Maracás", "Maragogipe", "Maraú", "Marcionílio Souza", "Mascote", "Mata de São João",
        "Matina", "Medeiros Neto", "Miguel Calmon", "Milagres", "Mirangaba", "Mirante", "Monte Santo",
        "Morpará", "Morro do Chapéu", "Mortugaba", "Mucugê", "Mucuri", "Mulungu do Morro", "Mundo Novo",
        "Muniz Ferreira", "Muquém do São Francisco", "Muritiba", "Mutuípe", "Nazaré", "Nilo Peçanha",
        "Nordestina", "Nova Canaã", "Nova Fátima", "Nova Ibiá", "Nova Itarana", "Nova Redenção",
        "Nova Soure", "Nova Viçosa", "Novo Horizonte", "Novo Triunfo", "Olindina", "Oliveira dos Brejinhos",
        "Ouriçangas", "Ourolândia", "Palmas de Monte Alto", "Palmeiras", "Paramirim", "Paratinga",
        "Paripiranga", "Pau Brasil", "Paulo Afonso", "Pé de Serra", "Pedrão", "Pedro Alexandre",
        "Piatã", "Pilão Arcado", "Pindaí", "Pindobaçu", "Pintadas", "Piraí do Norte", "Piripá",
        "Piritiba", "Planaltino", "Planalto", "Poções", "Pojuca", "Ponto Novo", "Porto Seguro",
        "Potiraguá", "Prado", "Presidente Dutra", "Presidente Jânio Quadros", "Presidente Tancredo Neves",
        "Queimadas", "Quijingue", "Quixabeira", "Rafael Jambeiro", "Remanso", "Retirolândia", "Riachão das Neves",
        "Riachão do Jacuípe", "Riacho de Santana", "Ribeira do Amparo", "Ribeira do Pombal", "Ribeirão do Largo",
        "Rio de Contas", "Rio do Antônio", "Rio do Pires", "Rio Real", "Rodelas", "Ruy Barbosa",
        "Salinas da Margarida", "Salvador", "Santa Bárbara", "Santa Brígida", "Santa Cruz Cabrália",
        "Santa Cruz da Vitória", "Santa Inês", "Santa Luzia", "Santa Maria da Vitória", "Santa Rita de Cássia",
        "Santa Terezinha", "Santaluz", "Santana", "Santanópolis", "Santo Amaro", "Santo Antônio de Jesus",
        "Santo Estêvão", "São Desidério", "São Domingos", "São Felipe", "São Félix", "São Félix do Coribe",
        "São Francisco do Conde", "São Gabriel", "São Gonçalo dos Campos", "São José da Vitória",
        "São José do Jacuípe", "São Miguel das Matas", "São Sebastião do Passé", "Sapeaçu", "Sátiro Dias",
        "Saubara", "Saúde", "Seabra", "Sebastião Laranjeiras", "Senhor do Bonfim", "Sento Sé", "Serra do Ramalho",
        "Serra Dourada", "Serra Preta", "Serrinha", "Serrolândia", "Simões Filho", "Sítio do Mato",
        "Sítio do Quinto", "Sobradinho", "Souto Soares", "Tabocas do Brejo Velho", "Tanque Novo", "Tanquinho",
        "Taperoá", "Tapiramutá", "Teixeira de Freitas", "Teodoro Sampaio", "Teofilândia", "Teolândia",
        "Terra Nova", "Tremedal", "Tucano", "Uauá", "Ubaíra", "Ubaitaba", "Ubatã", "Uibaí", "Umburanas",
        "Una", "Urandi", "Uruçuca", "Utinga", "Valença", "Valente", "Várzea da Roça", "Várzea do Poço",
        "Várzea Nova", "Varzedo", "Vera Cruz", "Vereda", "Vitória da Conquista", "Wagner", "Wanderley",
        "Wenceslau Guimarães", "Xique-Xique"
    ],
    "CE": [
        "Abaiara", "Acarapé", "Acaraú", "Acopiara", "Aiuaba", "Alcântaras", "Altaneira", "Alto Santo",
        "Amontada", "Antonina do Norte", "Apuiarés", "Aquiraz", "Aracati", "Aracoiaba", "Ararendá", "Araripe",
        "Aratuba", "Arneiroz", "Assaré", "Aurora", "Baixio", "Banabuiú", "Barbalha", "Barreira", "Barro",
        "Barroquinha", "Baturité", "Beberibe", "Bela Cruz", "Boa Viagem", "Brejo Santo", "Camocim", "Campos Sales",
        "Canindé", "Capistrano", "Caridade", "Cariré", "Caririaçu", "Cariús", "Carnaubal", "Cascavel",
        "Catarina", "Catunda", "Caucaia", "Cedro", "Chaval", "Choró", "Chorozinho", "Coreau", "Crateús",
        "Crato", "Croatá", "Cruz", "Deputado Irapuan Pinheiro", "Ererê", "Eusébio", "Farias Brito", "Forquilha",
        "Fortaleza", "Fortim", "Frecheirinha", "General Sampaio", "Graça", "Granja", "Granjeiro", "Groaíras",
        "Guaiúba", "Guaraciaba do Norte", "Guaramiranga", "Hidrolândia", "Horizonte", "Ibaretama", "Ibiapina",
        "Ibicuitinga", "Icapuí", "Icó", "Iguatu", "Independência", "Ipaporanga", "Ipaumirim", "Ipu", "Ipueiras",
        "Iracema", "Irauçuba", "Itaiçaba", "Itaitinga", "Itapagé", "Itapipoca", "Itapiúna", "Itarema", "Itatira",
        "Jaguaretama", "Jaguaribara", "Jaguaribe", "Jaguaruana", "Jardim", "Jati", "Jijoca de Jericoacoara",
        "Juazeiro do Norte", "Jucás", "Lavras da Mangabeira", "Limoeiro do Norte", "Madalena", "Maracanaú",
        "Maranguape", "Marco", "Martinópole", "Massapê", "Mauriti", "Meruoca", "Milagres", "Milhã", "Miraíma",
        "Missão Velha", "Mombaça", "Monsenhor Tabosa", "Morada Nova", "Moraújo", "Morrinhos", "Mucambo",
        "Mulungu", "Nova Olinda", "Nova Russas", "Novo Orientes", "Ocara", "Orós", "Pacajus", "Pacatuba",
        "Pacoti", "Pacujá", "Palhano", "Palmácia", "Paracuru", "Paraipaba", "Parambu", "Paramoti", "Penaforte",
        "Pentecoste", "Pereiro", "Pindoretama", "Piquet Carneiro", "Pires Ferreira", "Poranga", "Porteiras",
        "Potengi", "Potiretama", "Quiterianópolis", "Quixadá", "Quixelô", "Quixeramobim", "Quixeré", "Redenção",
        "Reriutaba", "Russas", "Saboeiro", "Salitre", "Santa Quitéria", "Santana do Acaraú", "Santana do Cariri",
        "São Benedito", "São Gonçalo do Amarante", "São João do Jaguaribe", "São Luís do Curu", "Senador Pompeu",
        "Senador Sá", "Sobral", "Solonópole", "Tabuleiro do Norte", "Tamboril", "Tarrafas", "Tauá", "Tejuçuoca",
        "Tianguá", "Trairi", "Tururu", "Ubajara", "Umari", "Umirim", "Uruburetama", "Uruoca", "Varjota",
        "Várzea Alegre", "Viçosa do Ceará"
    ],
    # Adicione listas completas para os outros estados de forma similar. Para o exemplo, adicionei CE. Em produção, use um JSON completo de fontes como IBGE ou GitHub.
    # Para os restantes (DF, ES, GO, MA, MT, MS, MG, PA, PB, PR, PE, PI, RJ, RN, RS, RO, RR, SC, SP, SE, TO), adicione listas semelhantes.
    "DF": ["Brasília"],
    "ES": ["Vitória", "Vila Velha", "Serra", "Cariacica", "Cachoeiro de Itapemirim" , "Linhares", "Colatina", "Guarapari", "Aracruz", "Viana", "Nova Venécia", "Barra de São Francisco", "Santa Maria de Jetibá", "Castelo", "Marataízes", "São Mateus", "Itapemirim", "Guaçuí", "Alegre", "Domingos Martins", "Iúna", "Jaguaré", "Baixo Guandu", "Conceição da Barra", "Pedro Canário", "Pinheiros", "Ibatiba", "Presidente Kennedy", "Rio Novo do Sul", "Mimoso do Sul", "Apiacá", "Bom Jesus do Norte", "São José do Calçado", "Vargem Alta", "Laranja da Terra", "Itaguaçu", "Muniz Freire", "Irupi", "Muqui", "Atílio Vivácqua", "Jerônimo Monteiro", "Cachoeiro de Itapemirim", "Anchieta", "Iconha", "Piúma", "Rio Bananal", "Sooretama", "Montanha", "Mucurici", "Ponto Belo", "Água Doce do Norte", "Ecoporanga", "Mantenópolis", "Alto Rio Novo", "Pancas", "Águia Branca", "São Gabriel da Palha", "Vila Pavão", "João Neiva", "Ibiraçu", "Fundão", "Santa Teresa", "Santa Leopoldina", "Itarana", "São Roque do Canaã", "São Domingos do Norte", "Vila Valério", "Governador Lindenberg", "Marilândia", "Jaguaré", "Conceição do Castelo", "Brejetuba", "Afonso Cláudio", "Divino de São Lourenço", "Dores do Rio Preto", "São José do Calçado"],
    # Continue adicionando para os outros estados. Para obter listas completas, use fontes como IBGE[](https://www.ibge.gov.br/cidades-e-estados.html).
    "GO": ["Goiânia", "Aparecida de Goiânia", "Anápolis", "Rio Verde", "Luziânia", "Águas Lindas de Goiás", "Valparaíso de Goiás", "Trindade", "Formosa", "Novo Gama", "Senador Canedo", "Catalão", "Itumbiara", "Jataí", "Planaltina", "Caldas Novas", "Goianésia", "Santo Antônio do Descoberto", "Cidade Ocidental", "Mineiros", "Cristalina", "Inhumas", "Jaraguá", "Niquelândia", "Morrinhos", "Goianira", "Porangatu", "Itaberaí", "Uruaçu", "Bela Vista de Goiás", "Goianápolis", "Hidrolândia", "Goiatuba", "Itapuranga", "São Miguel do Araguaia", "Ipameri", "São Simão", "Pires do Rio", "Padre Bernardo", "Campos Belos", "Quirinópolis", "Posse", "Abadiânia", "Alexânia", "Piracanjuba", "Silvânia", "Vianópolis", "Palmeiras de Goiás", "Caiapônia", "Indiara", "Paraúna", "Ceres", "Rialma", "Anicuns", "Buriti Alegre", "Itapirapuã", "Orizona", "Sanclerlândia", "São Luís de Montes Belos", "Firminópolis", "Turvânia", "Uruana", "Carmópolis de Minas", "Nazário", "Aragarças", "Arenópolis", "Heitoraí", "Santa Fé de Goiás", "Britânia", "Jussara", "Matrinchã", "Santa Rita do Novo Destino", "Nova Crixás", "Mundo Novo", "Morro Agudo de Goiás", "Amaralina", "Bonópolis", "Novo Mundo", "Trombas", "Montividiu do Norte", "Mutunópolis", "Porangatu", "Estrela do Norte", "Mara Rosa", "Campinorte", "Campinaçu", "Minaçu", "Colinas do Sul", "Niquelândia", "Alto Horizonte", "Nova Iguaçu de Goiás", "Campos Verdes", "Santa Terezinha de Goiás", "Crixás", "Uirapuru", "Nova América", "Rialma", "Rianápolis", "Santa Isabel", "Rubiataba", "Nova Glória", "Guarinos", "Pilar de Goiás", "Amorinópolis", "Ipiranga de Goiás", "Córrego do Ouro", "Moiporá", "Israelândia", "Jaupaci", "Taquaral de Goiás", "Jesúpolis", "Nova Veneza", "Ouro Verde de Goiás", "Fazenda Nova", "Goianésia", "Vilaboim", "Barro Alto", "Santa Rita do Trivelato", "Itaguari", "Itaguaru", "Jaraguá", "Itaberaí", "Itaguaí", "Santa Rosa de Goiás", "Petrolina de Goiás", "Ouro Verde de Goiás", "Sanclerlândia", "Adelândia", "Americano do Brasil", "Anicuns", "Buriti de Goiás", "Mossâmedes", "Nazário", "Turvânia", "Avelinópolis", "Santa Bárbara de Goiás", "Aurilândia", "Cachoeira de Goiás", "Paraúna", "Vicentinópolis", "Palminópolis", "São João da Paraúna", "Cezarina", "Varjão", "Santa Cruz de Goiás", "Palmelo", "Caldazinha", "Leopoldo de Bulhões", "Silvânia", "Vianópolis", "Orizona", "Pires do Rio", "Urutaí", "Ipameri", "Catalão", "Ouvidor", "Três Ranchos", "Anhanguera", "Cumari", "Nova Aurora", "Corumbaíba", "Água Limpa", "Marzagão", "Rio Quente", "Caldas Novas", "Buriti Alegre", "Joviânia", "Aloândia", "Pontalina", "Mairipotaba", "Piracanjuba", "Professor Jamil", "Cromínia", "Vicentinópolis", "Porteirão", "Edealina", "Edéia", "Indiara", "Acreúna", "Turvelândia", "Mauriti", "Santo Antônio do Descoberto", "Alexânia", "Abadiânia", "Corumbá de Goiás", "Pir часópolis", "Damianópolis", "Sítio d'Abadia", "Cabeceiras", "Formosa", "Vila Boa", "Buritinópolis", "Planaltina", "Água Fria de Goiás", "Mimoso de Goiás", "Padre Bernardo", "Cocalzinho de Goiás", "Vila Propício", "Niquelândia", "Colinas do Sul", "Cavalcante", "Campos Alegres de Goiás", "Monte Alegre de Goiás", "Teresina de Goiás", "Nova Roma", "São João d'Aliança", "Flores de Goiás", "Alto Paraíso de Goiás", "São Gabriel de Goiás", "Cristalina", "Luziânia", "Novo Gama", "Valparaíso de Goiás", "Cidade Ocidental", "Santo Antônio do Descoberto", "Cocalzinho de Goiás", "Padre Bernardo", "Águas Lindas de Goiás"],
    # Complete as listas para os outros estados usando fontes confiáveis como o IBGE. Esta é uma versão parcial; expanda conforme necessário.
    "MA": [
        "Açailândia", "Afonso Cunha", "Água Doce do Maranhão", "Alcântara", "Aldeias Altas",
        "Altamira do Maranhão", "Alto Alegre do Maranhão", "Alto Alegre do Pindaré",
        "Alto Parnaíba", "Amapá do Maranhão", "Amarante do Maranhão", "Anajatuba", "Anapurus",
        "Apicum-Açu", "Araguanã", "Araioses", "Arame", "Arari", "Axixá", "Bacabal", "Bacabeira",
        "Bacuri", "Bacurituba", "Balsas", "Barão de Grajaú", "Barra do Corda", "Barreirinhas",
        "Bela Vista do Maranhão", "Belágua", "Benedito Leite", "Bequimão", "Bernardo do Mearim",
        "Boa Vista do Gurupi", "Bom Jardim", "Bom Jesus das Selvas", "Bom Lugar", "Brejo",
        "Brejo de Areia", "Buriti", "Buriti Bravo", "Buriticupu", "Buritirana", "Cachoeira Grande",
        "Cajapió", "Cajari", "Campestre do Maranhão", "Cândido Mendes", "Cantanhede",
        "Capinzal do Norte", "Carolina", "Carutapera", "Caxias", "Cedral", "Central do Maranhão",
        "Centro do Guilherme", "Centro Novo do Maranhão", "Chapadinha", "Cidelândia", "Codó",
        "Coelho Neto", "Colinas", "Conceição do Lago Açu", "Coroatá", "Cururupu", "Davinópolis",
        "Dom Pedro", "Duque Bacelar", "Esperantinópolis", "Estreito", "Feira Nova do Maranhão",
        "Fernando Falcão", "Formosa da Serra Negra", "Fortaleza dos Nogueiras", "Fortuna",
        "Godofredo Viana", "Gonçalves Dias", "Governador Archer", "Governador Edison Lobão",
        "Governador Eugênio Barros", "Governador Luiz Rocha", "Governador Newton Bello",
        "Governador Nunes Freire", "Graça Aranha", "Grajaú", "Guimarães", "Humberto de Campos",
        "Icatu", "Igarapé do Meio", "Igarapé Grande", "Imperatriz", "Itaipava do Grajaú",
        "Itapecuru-Mirim", "Itinga do Maranhão", "Jatobá", "Jenipapo dos Vieiras", "João Lisboa",
        "Joselândia", "Junco do Maranhão", "Lago da Pedra", "Lago do Junco", "Lago dos Rodrigues",
        "Lago Verde", "Lagoa do Mato", "Lagoa Grande do Maranhão", "Lajeado Novo", "Lima Campos",
        "Loreto", "Luís Domingues", "Magalhães de Almeida", "Maracaçumé", "Marajá do Sena",
        "Maranhãozinho", "Mata Roma", "Matinha", "Matões", "Matões do Norte", "Milagres do Maranhão",
        "Mirador", "Miranda do Norte", "Mirinzal", "Monção", "Montes Altos", "Morros",
        "Nina Rodrigues", "Nova Colinas", "Nova Iorque", "Nova Olinda do Maranhão",
        "Olho d'Água das Cunhãs", "Olinda Nova do Maranhão", "Paço do Lumiar", "Palmeirândia",
        "Paraibano", "Parnarama", "Passagem Franca", "Pastos Bons", "Paulino Neves", "Paulo Ramos",
        "Pedreiras", "Pedro do Rosário", "Penalva", "Peri Mirim", "Peritoró", "Pindaré-Mirim",
        "Pinheiro", "Pio XII", "Pirapemas", "Poção de Pedras", "Porto Franco", "Porto Rico do Maranhão",
        "Presidente Dutra", "Presidente Juscelino", "Presidente Médici", "Presidente Sarney",
        "Presidente Vargas", "Primeira Cruz", "Raposa", "Riachão", "Ribamar Fiquene", "Rosário",
        "Sambaíba", "Santa Filomena do Maranhão", "Santa Helena", "Santa Inês", "Santa Luzia",
        "Santa Luzia do Paruá", "Santa Quitéria do Maranhão", "Santa Rita", "Santana do Maranhão",
        "Santo Amaro do Maranhão", "Santo Antônio dos Lopes", "São Benedito do Rio Preto",
        "São Bento", "São Bernardo", "São Domingos do Azeitão", "São Domingos do Maranhão",
        "São Félix de Balsas", "São Francisco do Brejão", "São Francisco do Maranhão",
        "São João Batista", "São João do Caru", "São João do Paraíso", "São João do Soter",
        "São João dos Patos", "São José de Ribamar", "São José dos Basílios", "São Luís",
        "São Luís Gonzaga do Maranhão", "São Mateus do Maranhão", "São Pedro da Água Branca",
        "São Pedro dos Crentes", "São Raimundo das Mangabeiras", "São Raimundo do Doca Bezerra",
        "São Roberto", "São Vicente Ferrer", "Satubinha", "Senador Alexandre Costa",
        "Senador La Rocque", "Serrano do Maranhão", "Sítio Novo", "Sucupira do Norte",
        "Sucupira do Riachão", "Tasso Fragoso", "Timbiras", "Timon", "Trizidela do Vale",
        "Tufilândia", "Tuntum", "Turiaçu", "Turilândia", "Tutóia", "Urbano Santos",
        "Vargem Grande", "Viana", "Vila Nova dos Martírios", "Vitória do Mearim", "Vitorino Freire",
        "Zé Doca"
    ],
    "MT": [
        "Acorizal", "Água Boa", "Alta Floresta", "Alto Araguaia", "Alto Boa Vista", "Alto Garças",
        "Alto Paraguai", "Alto Taquari", "Apiacás", "Araguaiana", "Araguainha", "Araputanga",
        "Arenápolis", "Aripuanã", "Barão de Melgaço", "Barra do Bugres", "Barra do Garças",
        "Boa Esperança do Norte", "Bom Jesus do Araguaia", "Brasnorte", "Cáceres", "Campinápolis",
        "Campo Novo do Parecis", "Campo Verde", "Campos de Júlio", "Canabrava do Norte", "Canarana",
        "Carlinda", "Castanheira", "Chapada dos Guimarães", "Cláudia", "Cocalinho", "Colíder",
        "Colniza", "Comodoro", "Confresa", "Conquista d'Oeste", "Cotriguaçu", "Cuiabá",
        "Curvelândia", "Denise", "Diamantino", "Dom Aquino", "Feliz Natal", "Figueirópolis d'Oeste",
        "Gaúcha do Norte", "General Carneiro", "Glória d'Oeste", "Guarantã do Norte", "Guiratinga",
        "Indiavaí", "Ipiranga do Norte", "Itanhangá", "Itaubá", "Itiquira", "Jaciara", "Jangada",
        "Jauru", "Juara", "Juína", "Juruena", "Juscimeira", "Lambari d'Oeste", "Lucas do Rio Verde",
        "Luciara", "Marcelândia", "Matupá", "Mirassol d'Oeste", "Nobres", "Nortelândia",
        "Nossa Senhora do Livramento", "Nova Bandeirantes", "Nova Brasilândia", "Nova Canaã do Norte",
        "Nova Guarita", "Nova Lacerda", "Nova Marilândia", "Nova Maringá", "Nova Monte Verde",
        "Nova Mutum", "Nova Nazaré", "Nova Olímpia", "Nova Santa Helena", "Nova Ubiratã",
        "Nova Xavantina", "Novo Horizonte do Norte", "Novo Mundo", "Novo Santo Antônio",
        "Novo São Joaquim", "Paranaíta", "Paranatinga", "Pedra Preta", "Peixoto de Azevedo",
        "Planalto da Serra", "Poconé", "Pontal do Araguaia", "Ponte Branca", "Pontes e Lacerda",
        "Porto Alegre do Norte", "Porto dos Gaúchos", "Porto Esperidião", "Porto Estrela", "Poxoréu",
        "Primavera do Leste", "Querência", "Reserva do Cabaçal", "Ribeirão Cascalheira",
        "Ribeirãozinho", "Rio Branco", "Rondolândia", "Rondonópolis", "Rosário Oeste",
        "Salto do Céu", "Santa Carmem", "Santa Cruz do Xingu", "Santa Rita do Trivelato",
        "Santa Terezinha", "Santo Afonso", "Santo Antônio do Leste", "Santo Antônio de Leverger",
        "São Félix do Araguaia", "São José do Povo", "São José do Rio Claro", "São José do Xingu",
        "São José dos Quatro Marcos", "São Pedro da Cipa", "Sapezal", "Serra Nova Dourada",
        "Sinop", "Sorriso", "Tabaporã", "Tangará da Serra", "Tapurah", "Terra Nova do Norte",
        "Tesouro", "Torixoréu", "União do Sul", "Vale de São Domingos", "Várzea Grande", "Vera",
        "Vila Bela da Santíssima Trindade", "Vila Rica"
    ],
    "MS": [
        "Água Clara", "Alcinópolis", "Amambai", "Anastácio", "Anaurilândia", "Angélica",
        "Antônio João", "Aparecida do Taboado", "Aquidauana", "Aral Moreira", "Bandeirantes",
        "Bataguassu", "Batayporã", "Bela Vista", "Bodoquena", "Bonito", "Brasilândia", "Caarapó",
        "Camapuã", "Campo Grande", "Caracol", "Cassilândia", "Chapadão do Sul", "Corguinho",
        "Coronel Sapucaia", "Corumbá", "Costa Rica", "Coxim", "Deodápolis", "Dois Irmãos do Buriti",
        "Douradina", "Dourados", "Eldorado", "Fátima do Sul", "Figueirão", "Glória de Dourados",
        "Guia Lopes da Laguna", "Iguatemi", "Inocência", "Itaporã", "Itaquiraí", "Ivinhema",
        "Japorã", "Jaraguari", "Jardim", "Jateí", "Juti", "Ladário", "Laguna Carapã", "Maracaju",
        "Miranda", "Mundo Novo", "Naviraí", "Nioaque", "Nova Alvorada do Sul", "Nova Andradina",
        "Novo Horizonte do Sul", "Paraíso das Águas", "Paranaíba", "Paranhos", "Pedro Gomes",
        "Ponta Porã", "Porto Murtinho", "Ribas do Rio Pardo", "Rio Brilhante", "Rio Negro",
        "Rio Verde de Mato Grosso", "Rochedo", "Santa Rita do Pardo", "São Gabriel do Oeste",
        "Selvíria", "Sete Quedas", "Sidrolândia", "Sonora", "Tacuru", "Taquarussu", "Terenos",
        "Três Lagoas", "Vicentina"
    ],
    "MG": [
        "Abadia dos Dourados", "Abaeté", "Abre Campo", "Acaiaca", "Açucena", "Água Boa",
        "Água Comprida", "Aguanil", "Águas Formosas", "Águas Vermelhas", "Aimorés", "Aiuruoca",
        "Alagoa", "Albertina", "Além Paraíba", "Alfenas", "Alfredo Vasconcelos", "Almenara",
        "Alpercata", "Alpinópolis", "Alterosa", "Alto Caparaó", "Alto Jequitibá", "Alto Rio Doce",
        "Alvarenga", "Alvinópolis", "Alvorada de Minas", "Amparo do Serra", "Andradas", "Andrelândia",
        "Angelândia", "Antônio Carlos", "Antônio Dias", "Antônio Prado de Minas", "Araçaí", "Aracitaba",
        "Araçuaí", "Araguari", "Arantina", "Araponga", "Araporã", "Arapuá", "Araújos", "Araxá",
        "Arceburgo", "Arcos", "Areado", "Argirita", "Aricanduva", "Arinos", "Astolfo Dutra",
        "Ataléia", "Augusto de Lima", "Baependi", "Baldim", "Bambuí", "Bandeira", "Bandeira do Sul",
        "Barão de Cocais", "Barão de Monte Alto", "Barbacena", "Barra Longa", "Barroso",
        "Bela Vista de Minas", "Belmiro Braga", "Belo Horizonte", "Belo Oriente", "Belo Vale",
        "Berilo", "Berizal", "Bertópolis", "Betim", "Bias Fortes", "Bicas", "Biquinhas",
        "Boa Esperança", "Bocaina de Minas", "Bocaiuva", "Bom Despacho", "Bom Jardim de Minas",
        "Bom Jesus da Penha", "Bom Jesus do Amparo", "Bom Jesus do Galho", "Bom Repouso",
        "Bom Sucesso", "Bonfim", "Bonfinópolis de Minas", "Bonito de Minas", "Borda da Mata",
        "Botelhos", "Botumirim", "Brás Pires", "Brasilândia de Minas", "Brasília de Minas",
        "Braúnas", "Brazópolis", "Brumadinho", "Bueno Brandão", "Buenópolis", "Bugre", "Buritis",
        "Buritizeiro", "Cabeceira Grande", "Cabo Verde", "Cachoeira da Prata", "Cachoeira de Minas",
        "Cachoeira de Pajeú", "Cachoeira Dourada", "Caetanópolis", "Caeté", "Caiana", "Cajuri",
        "Caldas", "Camacho", "Camanducaia", "Cambuí", "Cambuquira", "Campanário", "Campanha",
        "Campestre", "Campina Verde", "Campo Azul", "Campo Belo", "Campo do Meio", "Campo Florido",
        "Campos Altos", "Campos Gerais", "Cana Verde", "Canaã", "Canápolis", "Candeias", "Cantagalo",
        "Caparaó", "Capela Nova", "Capelinha", "Capetinga", "Capim Branco", "Capinópolis",
        "Capitão Andrade", "Capitão Enéas", "Capitólio", "Caputira", "Caraí", "Caranaíba", "Carandaí",
        "Carangola", "Caratinga", "Carbonita", "Careaçu", "Carlos Chagas", "Carmésia",
        "Carmo da Cachoeira", "Carmo da Mata", "Carmo de Minas", "Carmo do Cajuru",
        "Carmo do Paranaíba", "Carmo do Rio Claro", "Carmópolis de Minas", "Carneirinho", "Carrancas",
        "Carvalhópolis", "Carvalhos", "Casa Grande", "Cascalho Rico", "Cássia", "Cataguases",
        "Catas Altas", "Catas Altas da Noruega", "Catuji", "Catuti", "Caxambu", "Cedro do Abaeté",
        "Central de Minas", "Centralina", "Chácara", "Chalé", "Chapada do Norte", "Chapada Gaúcha",
        "Chiador", "Cipotânea", "Claraval", "Claro dos Poções", "Cláudio", "Coimbra", "Coluna",
        "Comendador Gomes", "Comercinho", "Conceição da Aparecida", "Conceição da Barra de Minas",
        "Conceição das Alagoas", "Conceição das Pedras", "Conceição de Ipanema",
        "Conceição do Mato Dentro", "Conceição do Pará", "Conceição do Rio Verde",
        "Conceição dos Ouros", "Cônego Marinho", "Confins", "Congonhal", "Congonhas",
        "Congonhas do Norte", "Conquista", "Conselheiro Lafaiete", "Conselheiro Pena", "Consolação",
        "Contagem", "Coqueiral", "Coração de Jesus", "Cordisburgo", "Cordislândia", "Corinto",
        "Coroaci", "Coromandel", "Coronel Fabriciano", "Coronel Murta", "Coronel Pacheco",
        "Coronel Xavier Chaves", "Córrego Danta", "Córrego do Bom Jesus", "Córrego Fundo",
        "Córrego Novo", "Couto de Magalhães de Minas", "Crisólita", "Cristais", "Cristália",
        "Cristiano Otoni", "Cristina", "Crucilândia", "Cruzeiro da Fortaleza", "Cruzília", "Cuparaque",
        "Curral de Dentro", "Curvelo", "Datas", "Delfim Moreira", "Delfinópolis", "Delta", "Descoberto",
        "Desterro de Entre Rios", "Desterro do Melo", "Diamantina", "Diogo de Vasconcelos", "Dionísio",
        "Divinésia", "Divino", "Divino das Laranjeiras", "Divinolândia de Minas", "Divinópolis",
        "Divisa Alegre", "Divisa Nova", "Divisópolis", "Dom Bosco", "Dom Cavati", "Dom Joaquim",
        "Dom Silvério", "Dom Viçoso", "Dona Euzébia", "Dores de Campos", "Dores de Guanhães",
        "Dores do Indaiá", "Dores do Turvo", "Doresópolis", "Douradoquara", "Durandé", "Elói Mendes",
        "Engenheiro Caldas", "Engenheiro Navarro", "Entre Folhas", "Entre Rios de Minas", "Ervália",
        "Esmeraldas", "Espera Feliz", "Espinosa", "Espírito Santo do Dourado", "Estiva",
        "Estrela Dalva", "Estrela do Indaiá", "Estrela do Sul", "Eugenópolis", "Ewbank da Câmara",
        "Extrema", "Fama", "Faria Lemos", "Felício dos Santos", "Felisburgo", "Felixlândia",
        "Fernandes Tourinho", "Ferros", "Fervedouro", "Florestal", "Formiga", "Formoso",
        "Fortaleza de Minas", "Fortuna de Minas", "Francisco Badaró", "Francisco Dumont",
        "Francisco Sá", "Franciscópolis", "Frei Gaspar", "Frei Inocêncio", "Frei Lagonegro",
        "Fronteira", "Fronteira dos Vales", "Fruta de Leite", "Frutal", "Funilândia", "Galiléia",
        "Gameleiras", "Glaucilândia", "Goiabeira", "Goianá", "Gonçalves", "Gonzaga", "Gouveia",
        "Governador Valadares", "Grão Mogol", "Grupiara", "Guanhães", "Guapé", "Guaraciaba",
        "Guaraciama", "Guaranésia", "Guarani", "Guarará", "Guarda-Mor", "Guaxupé", "Guidoval",
        "Guimarânia", "Guiricema", "Gurinhatã", "Heliodora", "Iapu", "Ibertioga", "Ibiá", "Ibiaí",
        "Ibiracatu", "Ibiraci", "Ibirité", "Ibitiúra de Minas", "Ibituruna", "Icaraí de Minas",
        "Igarapé", "Igaratinga", "Iguatama", "Ijaci", "Ilicínea", "Imbé de Minas", "Inconfidentes",
        "Indaiabira", "Indianópolis", "Ingaí", "Inhapim", "Inhaúma", "Inimutaba", "Ipaba", "Ipanema",
        "Ipatinga", "Ipiaçu", "Ipuiuna", "Iraí de Minas", "Itabira", "Itabirinha", "Itabirito",
        "Itacambira", "Itacarambi", "Itaguara", "Itaipé", "Itajubá", "Itamarandiba",
        "Itamarati de Minas", "Itambacuri", "Itambé do Mato Dentro", "Itamogi", "Itamonte",
        "Itanhandu", "Itanhomi", "Itaobim", "Itapagipe", "Itapecerica", "Itapeva", "Itatiaiuçu",
        "Itau de Minas", "Itaúna", "Itaverava", "Itinga", "Itueta", "Ituiutaba", "Itumirim",
        "Iturama", "Itutinga", "Jaboticatubas", "Jacinto", "Jacuí", "Jacutinga", "Jaguaraçu",
        "Jaíba", "Jampruca", "Janaúba", "Januária", "Japaraíba", "Japonvar", "Jeceaba",
        "Jenipapo de Minas", "Jequeri", "Jequitaí", "Jequitibá", "Jequitinhonha", "Jesuânia",
        "Joaíma", "Joanésia", "João Monlevade", "João Pinheiro", "Joaquim Felício", "Jordânia",
        "José Gonçalves de Minas", "José Raydan", "Josenópolis", "Juatuba", "Juiz de Fora",
        "Juramento", "Juruaia", "Juvenília", "Ladainha", "Lagamar", "Lagoa da Prata", "Lagoa dos Patos",
        "Lagoa Dourada", "Lagoa Formosa", "Lagoa Grande", "Lagoa Santa", "Lajinha", "Lambari", "Lamim",
        "Laranjal", "Lassance", "Lavras", "Leandro Ferreira", "Leme do Prado", "Leopoldina",
        "Liberdade", "Lima Duarte", "Limeira do Oeste", "Lontra", "Luisburgo", "Luislândia",
        "Luminárias", "Luz", "Machacalis", "Machado", "Madre de Deus de Minas", "Malacacheta",
        "Mamonas", "Manga", "Manhuaçu", "Manhumirim", "Mantena", "Mar de Espanha", "Maravilhas",
        "Maria da Fé", "Mariana", "Marilac", "Mário Campos", "Maripá de Minas", "Marliéria",
        "Marmelópolis", "Martinho Campos", "Martins Soares", "Mata Verde", "Materlândia",
        "Mateus Leme", "Mathias Lobato", "Matias Barbosa", "Matias Cardoso", "Matipó", "Mato Verde",
        "Matozinhos", "Matutina", "Medeiros", "Medina", "Mendes Pimentel", "Mercês", "Mesquita",
        "Minas Novas", "Minduri", "Mirabela", "Miradouro", "Miraí", "Miravânia", "Moeda", "Moema",
        "Monjolos", "Monsenhor Paulo", "Montalvânia", "Monte Alegre de Minas", "Monte Azul",
        "Monte Belo", "Monte Carmelo", "Monte Formoso", "Monte Santo de Minas", "Monte Sião",
        "Montes Claros", "Montezuma", "Morada Nova de Minas", "Morro da Garça", "Morro do Pilar",
        "Munhoz", "Muriaé", "Mutum", "Muzambinho", "Nacip Raydan", "Nanuque", "Naque", "Natalândia",
        "Natércia", "Nazareno", "Nepomuceno", "Ninheira", "Nova Belém", "Nova Era", "Nova Lima",
        "Nova Módica", "Nova Ponte", "Nova Porteirinha", "Nova Resende", "Nova Serrana",
        "Nova União", "Novo Cruzeiro", "Novo Oriente de Minas", "Novorizonte", "Olaria",
        "Olhos-d'Água", "Olímpio Noronha", "Oliveira", "Oliveira Fortes", "Onça de Pitangui",
        "Oratórios", "Orizânia", "Ouro Branco", "Ouro Fino", "Ouro Preto", "Ouro Verde de Minas",
        "Padre Carvalho", "Padre Paraíso", "Pai Pedro", "Paineiras", "Pains", "Paiva", "Palma",
        "Palmópolis", "Papagaios", "Pará de Minas", "Paracatu", "Paraguaçu", "Paraisópolis",
        "Paraopeba", "Passa Quatro", "Passa Tempo", "Passa Vinte", "Passabém", "Passos", "Patis",
        "Patos de Minas", "Patrocínio", "Patrocínio do Muriaé", "Paula Cândido", "Paulistas", "Pavão",
        "Peçanha", "Pedra Azul", "Pedra Bonita", "Pedra do Anta", "Pedra do Indaiá", "Pedra Dourada",
        "Pedralva", "Pedras de Maria da Cruz", "Pedrinópolis", "Pedro Leopoldo", "Pedro Teixeira",
        "Pequeri", "Pequi", "Perdigão", "Perdizes", "Perdões", "Periquito", "Pescador", "Piau",
        "Piedade de Caratinga", "Piedade de Ponte Nova", "Piedade do Rio Grande",
        "Piedade dos Gerais", "Pimenta", "Pingo-d'Água", "Pintópolis", "Piracema", "Pirajuba",
        "Piranga", "Piranguçu", "Piranguinho", "Pirapetinga", "Pirapora", "Piraúba", "Pitangui",
        "Piumhi", "Planura", "Poço Fundo", "Poços de Caldas", "Pocrane", "Pompéu", "Ponte Nova",
        "Ponto Chique", "Ponto dos Volantes", "Porteirinha", "Porto Firme", "Poté", "Pouso Alegre",
        "Pouso Alto", "Prados", "Prata", "Pratápolis", "Pratinha", "Presidente Bernardes",
        "Presidente Juscelino", "Presidente Kubitschek", "Presidente Olegário", "Prudente de Morais",
        "Quartel Geral", "Queluzito", "Raposos", "Raul Soares", "Recreio", "Reduto", "Resende Costa",
        "Resplendor", "Ressaquinha", "Riachinho", "Riacho dos Machados", "Ribeirão das Neves",
        "Ribeirão Vermelho", "Rio Acima", "Rio Casca", "Rio do Prado", "Rio Doce", "Rio Espera",
        "Rio Manso", "Rio Novo", "Rio Paranaíba", "Rio Pardo de Minas", "Rio Piracicaba", "Rio Pomba",
        "Rio Preto", "Rio Vermelho", "Ritápolis", "Rochedo de Minas", "Rodeiro", "Romaria",
        "Rosário da Limeira", "Rubelita", "Rubim", "Sabará", "Sabinópolis", "Sacramento", "Salinas",
        "Salto da Divisa", "Santa Bárbara", "Santa Bárbara do Leste", "Santa Bárbara do Monte Verde",
        "Santa Bárbara do Tugúrio", "Santa Cruz de Minas", "Santa Cruz de Salinas",
        "Santa Cruz do Escalvado", "Santa Efigênia de Minas", "Santa Fé de Minas",
        "Santa Helena de Minas", "Santa Juliana", "Santa Luzia", "Santa Margarida",
        "Santa Maria de Itabira", "Santa Maria do Salto", "Santa Maria do Suaçuí",
        "Santa Rita de Caldas", "Santa Rita de Ibitipoca", "Santa Rita de Jacutinga",
        "Santa Rita de Minas", "Santa Rita do Itueto", "Santa Rita do Sapucaí", "Santa Rosa da Serra",
        "Santa Vitória", "Santana da Vargem", "Santana de Cataguases", "Santana de Pirapama",
        "Santana do Deserto", "Santana do Garambéu", "Santana do Jacaré", "Santana do Manhuaçu",
        "Santana do Paraíso", "Santana do Riacho", "Santana dos Montes", "Santo Antônio do Amparo",
        "Santo Antônio do Aventureiro", "Santo Antônio do Grama", "Santo Antônio do Itambé",
        "Santo Antônio do Jacinto", "Santo Antônio do Monte", "Santo Antônio do Retiro",
        "Santo Antônio do Rio Abaixo", "Santo Antônio do Salto", "Santo Hipólito", "Santos Dumont",
        "São Bento Abade", "São Brás do Suaçuí", "São Domingos das Dores", "São Domingos do Prata",
        "São Félix de Minas", "São Francisco", "São Francisco de Paula", "São Francisco de Sales",
        "São Francisco do Glória", "São Geraldo", "São Geraldo da Piedade", "São Geraldo do Baixio",
        "São Gonçalo do Abaeté", "São Gonçalo do Pará", "São Gonçalo do Rio Abaixo",
        "São Gonçalo do Sapucaí", "São Gotardo", "São João Batista do Glória", "São João da Lagoa",
        "São João da Mata", "São João da Ponte", "São João das Missões", "São João del-Rei",
        "São João do Manhuaçu", "São João do Manteninha", "São João do Oriente", "São João do Pacuí",
        "São João Evangelista", "São João Nepomuceno", "São Joaquim de Bicas", "São José da Barra",
        "São José da Lapa", "São José da Safira", "São José da Varginha", "São José do Alegre",
        "São José do Divino", "São José do Goiabal", "São José do Jacuri", "São José do Mantimento",
        "São Lourenço", "São Miguel do Anta", "São Pedro da União", "São Pedro do Suaçuí",
        "São Pedro dos Ferros", "São Romão", "São Roque de Minas", "São Sebastião da Bela Vista",
        "São Sebastião da Vargem Alegre", "São Sebastião do Anta", "São Sebastião do Maranhão",
        "São Sebastião do Oeste", "São Sebastião do Paraíso", "São Sebastião do Rio Preto",
        "São Sebastião do Rio Verde", "São Thomé das Letras", "São Tiago", "São Tomás de Aquino",
        "São Vicente de Minas", "Sapucaí-Mirim", "Sardoá", "Sarzedo", "Sem-Peixe", "Senador Amaral",
        "Senador Cortes", "Senador Firmino", "Senador José Bento", "Senador Modestino Gonçalves",
        "Senhora de Oliveira", "Senhora do Porto", "Senhora dos Remédios", "Sereno",
        "Serra Azul de Minas", "Serra da Saudade", "Serra do Salitre", "Serra dos Aimorés",
        "Serrania", "Serranópolis de Minas", "Serranos", "Serro", "Sete Lagoas", "Setubinha",
        "Silveirânia", "Silvianópolis", "Simão Pereira", "Simonésia", "Sobrália", "Soledade de Minas",
        "Tabuleiro", "Taiobeiras", "Taparuba", "Tapira", "Tapiraí", "Taquaraçu de Minas", "Tarumirim",
        "Teixeiras", "Teófilo Otoni", "Timóteo", "Tiradentes", "Tiros", "Tocantins", "Tocos do Moji",
        "Toledo", "Tombos", "Três Corações", "Três Marias", "Três Pontas", "Tumiritinga",
        "Tupaciguara", "Turmalina", "Turvolândia", "Ubá", "Ubaí", "Ubaporanga", "Uberaba",
        "Uberlândia", "Umburatiba", "Unaí", "União de Minas", "Uruana de Minas", "Urucânia",
        "Urucuia", "Vargem Alegre", "Vargem Bonita", "Vargem Grande do Rio Pardo", "Varginha",
        "Varjão de Minas", "Várzea da Palma", "Varzelândia", "Vazante", "Verdelândia", "Veredinha",
        "Veríssimo", "Vermelho Novo", "Vespasiano", "Viçosa", "Vieiras", "Virgem da Lapa", "Virgínia",
        "Virginópolis", "Virgolândia", "Visconde do Rio Branco", "Volta Grande", "Wenceslau Braz"
    ],
    "PA": [
        "Abaetetuba", "Abel Figueiredo", "Acará", "Afuá", "Água Azul do Norte", "Alenquer", "Almeirim", "Altamira", "Anajás", "Ananindeua", "Anapu", "Augusto Corrêa", "Aurora do Pará", "Aveiro", "Bagre", "Baião", "Bannach", "Barcarena", "Belém", "Belterra", "Benevides", "Bom Jesus do Tocantins", "Bonito", "Bragança", "Brasil Novo", "Brejo Grande do Araguaia", "Breu Branco", "Breves", "Bujaru", "Cachoeira do Arari", "Cachoeira do Piriá", "Cametá", "Canaã dos Carajás", "Capanema", "Capitão Poço", "Castanhal", "Chaves", "Colares", "Conceição do Araguaia", "Concórdia do Pará", "Cumaru do Norte", "Curionópolis", "Curralinho", "Curuá", "Curuçá", "Dom Eliseu", "Eldorado do Carajás", "Faro", "Floresta do Araguaia", "Garrafão do Norte", "Goianésia do Pará", "Gurupá", "Igarapé-Açu", "Igarapé-Miri", "Inhangapi", "Ipixuna do Pará", "Irituia", "Itaituba", "Itupiranga", "Jacareacanga", "Jacundá", "Juruti", "Limoeiro do Ajuru", "Mãe do Rio", "Magalhães Barata", "Marabá", "Maracanã", "Marapanim", "Marituba", "Medicilândia", "Melgaço", "Mocajuba", "Moju", "Mojuí dos Campos", "Monte Alegre", "Muaná", "Nova Esperança do Piriá", "Nova Ipixuna", "Nova Timboteua", "Novo Progresso", "Novo Repartimento", "Óbidos", "Oeiras do Pará", "Oriximiná", "Ourém", "Ourilândia do Norte",
        "Pacajá", "Palestina do Pará", "Paragominas", "Parauapebas", "Pau-d'Arco", "Peixe-Boi",
        "Piçarra", "Placas", "Ponta de Pedras", "Portel", "Porto de Moz", "Prainha", "Primavera",
        "Quatipuru", "Redenção", "Rio Maria", "Rondon do Pará", "Rurópolis", "Salinópolis",
        "Salvaterra", "Santa Bárbara do Pará", "Santa Cruz do Arari", "Santa Izabel do Pará",
        "Santa Luzia do Pará", "Santa Maria das Barreiras", "Santa Maria do Pará",
        "Santana do Araguaia", "Santarém", "Santarém Novo", "Santo Antônio do Tauá",
        "São Caetano de Odivelas", "São Domingos do Araguaia", "São Domingos do Capim",
        "São Félix do Xingu", "São Francisco do Pará", "São Geraldo do Araguaia", "São João da Ponta",
        "São João de Pirabas", "São João do Araguaia", "São Miguel do Guamá",
        "São Sebastião da Boa Vista", "Sapucaia", "Senador José Porfírio", "Soure", "Tailândia",
        "Terra Alta", "Terra Santa", "Tomé-Açu", "Tracuateua", "Trairão", "Tucumã", "Tucuruí",
        "Ulianópolis", "Uruará", "Vigia", "Viseu", "Vitória do Xingu", "Xinguara"
    ],
    "PB": [
        "Água Branca", "Aguiar", "Alagoa Grande", "Alagoa Nova", "Alagoinha", "Alcantil",
        "Algodão de Jandaíra", "Alhandra", "Amparo", "Aparecida", "Araçagi", "Arara", "Araruna",
        "Areia", "Areia de Baraúnas", "Areial", "Aroeiras", "Assunção", "Baía da Traição",
        "Bananeiras", "Baraúna", "Barra de Santa Rosa", "Barra de Santana", "Barra de São Miguel",
        "Bayeux", "Belém", "Belém do Brejo do Cruz", "Bernardino Batista", "Boa Ventura", "Boa Vista",
        "Bom Jesus", "Bom Sucesso", "Bonito de Santa Fé", "Boqueirão", "Borborema", "Brejo do Cruz",
        "Brejo dos Santos", "Caaporã", "Cabaceiras", "Cabedelo", "Cachoeira dos Índios",
        "Cacimba de Areia", "Cacimba de Dentro", "Cacimbas", "Caiçara", "Cajazeiras", "Cajazeirinhas",
        "Caldas Brandão", "Camalaú", "Campina Grande", "Capim", "Caraúbas", "Carrapateira",
        "Casserengue", "Catingueira", "Catolé do Rocha", "Caturité", "Conceição", "Condado", "Conde",
        "Congo", "Coremas", "Coxixola", "Cruz do Espírito Santo", "Cubati", "Cuité",
        "Cuité de Mamanguape", "Cuitegi", "Curral de Cima", "Curral Velho", "Damião", "Desterro",
        "Diamante", "Dona Inês", "Duas Estradas", "Emas", "Esperança", "Fagundes", "Frei Martinho",
        "Gado Bravo", "Guarabira", "Gurinhém", "Gurjão", "Ibiara", "Igaracy", "Imaculada", "Ingá",
        "Itabaiana", "Itaporanga", "Itapororoca", "Itatuba", "Jacaraú", "Jericó", "João Pessoa",
        "Joca Claudino", "Juarez Távora", "Juazeirinho", "Junco do Seridó", "Juripiranga", "Juru",
        "Lagoa", "Lagoa de Dentro", "Lagoa Seca", "Lastro", "Livramento", "Logradouro", "Lucena",
        "Mãe d'Água", "Malta", "Mamanguape", "Manaíra", "Marcação", "Mari", "Marizópolis",
        "Massaranduba", "Mataraca", "Matinhas", "Mato Grosso", "Maturéia", "Mogeiro", "Montadas",
        "Monte Horebe", "Monteiro", "Mulungu", "Natuba", "Nazarezinho", "Nova Floresta",
        "Nova Olinda", "Nova Palmeira", "Olho d'Água", "Olivedos", "Ouro Velho", "Parari", "Passagem",
        "Patos", "Paulista", "Pedra Branca", "Pedra Lavrada", "Pedras de Fogo", "Pedro Régis",
        "Piancó", "Picuí", "Pilar", "Pilões", "Pilõezinhos", "Pirpirituba", "Pitimbu", "Pocinhos",
        "Poço Dantas", "Poço de José de Moura", "Pombal", "Prata", "Princesa Isabel", "Puxinanã",
        "Queimadas", "Quixaba", "Remígio", "Riachão", "Riachão do Bacamarte", "Riachão do Poço",
        "Riacho de Santo Antônio", "Riacho dos Cavalos", "Rio Tinto", "Salgadinho",
        "Salgado de São Félix", "Santa Cecília", "Santa Cruz", "Santa Helena", "Santa Inês",
        "Santa Luzia", "Santa Rita", "Santa Terezinha", "Santana de Mangueira", "Santana dos Garrotes",
        "Santo André", "São Bentinho", "São Bento", "São Domingos", "São Domingos do Cariri",
        "São Francisco", "São João do Cariri", "São João do Rio do Peixe", "São João do Tigre",
        "São José da Lagoa Tapada", "São José de Caiana", "São José de Espinharas",
        "São José de Piranhas", "São José de Princesa", "São José do Bonfim",
        "São José do Brejo do Cruz", "São José do Sabugi", "São José dos Cordeiros",
        "São José dos Ramos", "São Mamede", "São Miguel de Taipu", "São Sebastião de Lagoa de Roça",
        "São Sebastião do Umbuzeiro", "São Vicente do Seridó", "Sapé", "Serra Branca",
        "Serra da Raiz", "Serra Grande", "Serra Redonda", "Serraria", "Sertãozinho", "Sobrado",
        "Solânea", "Soledade", "Sossêgo", "Sousa", "Sumé", "Tacima", "Taperoá", "Tavares",
        "Teixeira", "Tenório", "Triunfo", "Uiraúna", "Umbuzeiro", "Várzea", "Vieirópolis",
        "Vista Serrana", "Zabelê"
    ],
    # Para os estados restantes, adicione listas similares. Exemplo para PR:
    "PR": [
        "Abatiá", "Adrianópolis", "Agudos do Sul", "Almirante Tamandaré", "Altamira do Paraná", "Alto Paraíso", "Alto Paraná", "Alto Piquiri", "Altônia", "Alvorada do Sul", "Amaporã", "Ampére", "Anahy", "Andirá", "Ângulo", "Antonina", "Antônio Olinto", "Apucarana", "Arapongas", "Arapoti", "Arapuã", "Araruna", "Araucária", "Ariranha do Ivaí", "Assaí", "Assis Chateaubriand", "Astorga", "Atalaia", "Balsa Nova", "Bandeirantes", "Barbosa Ferraz", "Barra do Jacaré", "Barracão", "Bela Vista da Caroba", "Bela Vista do Paraíso", "Bituruna", "Boa Esperança", "Boa Esperança do Iguaçu", "Boa Ventura de São Roque", "Boa Vista da Aparecida", "Bocaiúva do Sul", "Bom Jesus do Sul", "Bom Sucesso", "Bom Sucesso do Sul", "Borrazópolis", "Braganey", "Brasilândia do Sul", "Cafeara", "Cafelândia", "Cafezal do Sul", "Califórnia", "Cambará", "Cambé", "Cambira", "Campina da Lagoa", "Campina do Simão", "Campina Grande do Sul", "Campo Bonito", "Campo do Tenente", "Campo Largo", "Campo Magro", "Campo Mourão", "Candói", "Cantagalo", "Capanema", "Capitão Leônidas Marques", "Carambeí", "Carlópolis", "Cascavel", "Castro", "Catanduvas", "Centenário do Sul", "Cerro Azul", "Céu Azul", "Chopinzinho", "Cianorte", "Cidade Gaúcha", "Clevelândia", "Colombo", "Colorado", "Congonhinhas", "Conselheiro Mairinck", "Contenda", "Corbélia", "Cornélio Procópio", "Coronel Domingos Soares", "Coronel Vivida", "Corumbataí do Sul", "Cruzeiro do Iguaçu", "Cruzeiro do Oeste", "Cruzeiro do Sul", "Cruz Machado", "Cruzmaltina", "Curitiba", "Curiúva", "Diamante d'Oeste", "Diamante do Norte", "Diamante do Sul", "Dois Vizinhos", "Douradina", "Doutor Camargo", "Enéas Marques", "Engenheiro Beltrão", "Esperança Nova", "Entre Rios do Oeste", "Espigão Alto do Iguaçu", "Farol", "Faxinal", "Fazenda Rio Grande", "Fênix", "Fernandes Pinheiro", "Figueira", "Floraí", "Flor da Serra do Sul", "Floresta", "Florestópolis", "Flórida", "Formosa do Oeste", "Foz do Iguaçu", "Francisco Alves", "Francisco Beltrão", "Foz do Jordão", "General Carneiro", "Godoy Moreira", "Goioerê", "Goioxim", "Grandes Rios", "Guaíra", "Guairaçá", "Guamiranga", "Guapirama", "Guaporema", "Guaraci", "Guaraniaçu", "Guarapuava", "Guaraqueçaba", "Guaratuba", "Honório Serpa", "Ibaiti", "Ibema", "Ibiporã", "Icaraíma", "Iguaraçu", "Iguatu", "Imbaú", "Imbituva", "Inácio Martins", "Inajá", "Indianópolis", "Ipiranga", "Iporã", "Iracema do Oeste", "Irati", "Iretama", "Itaguajé", "Itaipulândia", "Itambaracá", "Itambé", "Itapejara d'Oeste", "Itaperuçu", "Itaúna do Sul", "Ivaí", "Ivaiporã", "Ivaté", "Ivatuba", "Jaboti", "Jacarezinho", "Jaguapitã", "Jaguariaíva", "Jandaia do Sul", "Janiópolis", "Japira", "Japurá", "Jardim Alegre", "Jardim Olinda", "Jataizinho", "Jesuítas", "Joaquim Távora", "Jundiaí do Sul", "Juranda", "Jussara", "Kaloré", "Lapa", "Laranjal", "Laranjeiras do Sul", "Leópolis", "Lidianópolis", "Lindoeste", "Loanda", "Lobato", "Londrina", "Luiziana", "Lunardelli", "Lupionópolis", "Mallet", "Mamborê", "Mandaguaçu", "Mandaguari", "Mandirituba", "Manfrinópolis", "Mangueirinha", "Manoel Ribas", "Marechal Cândido Rondon", "Maria Helena", "Marialva", "Marilândia do Sul", "Marilena", "Mariluz", "Maringá", "Mariópolis", "Maripá", "Marmeleiro", "Marquinho", "Marumbi", "Matelândia", "Matinhos", "Mato Rico", "Mauá da Serra", "Medianeira", "Mercedes", "Mirador", "Miraselva", "Missal", "Moreira Sales", "Morretes", "Munhoz de Melo", "Nossa Senhora das Graças", "Nova Aliança do Ivaí", "Nova América da Colina", "Nova Aurora", "Nova Cantu", "Nova Esperança", "Nova Esperança do Sudoeste", "Nova Fátima", "Nova Laranjeiras", "Nova Londrina", "Nova Olímpia", "Nova Santa Bárbara", "Nova Santa Rosa", "Nova Prata do Iguaçu", "Nova Tebas", "Novo Itacolomi", "Ortigueira", "Ourizona", "Ouro Verde do Oeste", "Paiçandu", "Palmas", "Palmeira", "Palmital", "Palotina", "Paraíso do Norte", "Paranacity", "Paranaguá", "Paranapoema", "Paranavaí", "Pato Bragado", "Pato Branco", "Paula Freitas", "Paulo Frontin", "Peabiru", "Perobal", "Pérola", "Pérola d'Oeste", "Piên", "Pinhais", "Pinhalão", "Pinhal de São Bento", "Pinhão", "Piraí do Sul", "Piraquara", "Pitanga", "Pitangueiras", "Planaltina do Paraná", "Planalto", "Ponta Grossa", "Pontal do Paraná", "Porecatu", "Porto Amazonas", "Porto Barreiro", "Porto Rico", "Porto Vitória", "Prado Ferreira", "Pranchita", "Presidente Castelo Branco", "Primeiro de Maio", "Prudentópolis", "Quarto Centenário", "Quatiguá", "Quatro Barras", "Quatro Pontes", "Quedas do Iguaçu", "Querência do Norte", "Quinta do Sol", "Quitandinha", "Ramilândia", "Rancho Alegre", "Rancho Alegre D'Oeste", "Realeza", "Rebouças", "Renascença", "Reserva", "Reserva do Iguaçu", "Ribeirão Claro", "Ribeirão do Pinhal", "Rio Azul", "Rio Bom", "Rio Bonito do Iguaçu", "Rio Branco do Ivaí", "Rio Branco do Sul", "Rio Negro", "Rolândia", "Roncador", "Rondon", "Rosário do Ivaí", "Sabáudia", "Salgado Filho", "Salto do Itararé", "Salto do Lontra", "Santa Amélia", "Santa Cecília do Pavão", "Santa Cruz de Monte Castelo", "Santa Fé", "Santa Helena", "Santa Izabel do Oeste", "Santa Lúcia", "Santa Maria do Oeste", "Santa Mariana", "Santa Mônica", "Santana do Itararé", "Santa Tereza do Oeste", "Santa Terezinha de Itaipu", "Santo Antônio da Platina", "Santo Antônio do Caiuá", "Santo Antônio do Paraíso", "Santo Antônio do Sudoeste", "Santo Inácio", "São Carlos do Ivaí", "São Jerônimo da Serra", "São João", "São João do Caiuá", "São João do Ivaí", "São João do Triunfo", "São Jorge d'Oeste", "São Jorge do Ivaí", "São Jorge do Patrocínio", "São José da Boa Vista", "São José das Palmeiras", "São José do Itavo", "São Manoel do Paraná", "São Mateus do Sul", "São Miguel do Iguaçu", "São Pedro do Iguaçu", "São Pedro do Ivaí", "São Pedro do Paraná", "São Sebastião da Amoreira", "São Tomé", "Sapopema", "Sarandi", "Saudade do Iguaçu", "Sengés", "Serranópolis do Iguaçu", "Sertaneja", "Sertanópolis", "Siqueira Campos", "Sulina", "Tamarana", "Tamboara", "Tapejara", "Teixeira Soares", "Telêmaco Borba", "Terra Boa", "Terra Rica", "Terra Roxa", "Tibagi", "Tijucas do Sul", "Toledo", "Tomazina", "Três Barras do Paraná", "Tunas do Paraná", "Tuneiras do Oeste", "Tupãssi", "Turvo", "Ubiratã", "Umuarama", "União da Vitória", "Uniflor", "Uraí", "Wenceslau Braz", "Ventania", "Vera Cruz do Oeste", "Verê", "Alto Paraíso", "Virmond", "Vitorino", "Xambrê"
    ],
    # Adicione os outros estados de forma similar. Para completude, busque listas atualizadas do IBGE.
    # Exemplo para SP:
    "SP": [
        "Adamantina", "Adolfo", "Aguaí", "Águas da Prata", "Águas de Lindóia", "Águas de Santa Bárbara", "Águas de São Pedro", "Agudos", "Alambari", "Alfredo Marcondes", "Altair", "Altinópolis", "Alto Alegre", "Alumínio", "Álvares Florence", "Álvares Machado", "Álvaro de Carvalho", "Alvinlândia", "Americana", "Américo Brasiliense", "Américo de Campos", "Amparo", "Analândia", "Andradina", "Angatuba", "Anhembi", "Anhumas", "Aparecida", "Aparecida d'Oeste", "Apiaí", "Araçariguama", "Araçatuba", "Araçoiaba da Serra", "Aramina", "Arandu", "Arapeí", "Araraquara", "Araras", "Arco-Íris", "Arealva", "Areias", "Areiópolis", "Ariranha", "Artur Nogueira", "Arujá", "Aspásia", "Assis", "Atibaia", "Auriflama", "Avaí", "Avanhandava", "Avaré", "Bady Bassitt", "Balbinos", "Bálsamo", "Bananal", "Barão de Antonina", "Barbosa", "Bariri", "Barra Bonita", "Barra do Chapéu", "Barra do Turvo", "Barretos", "Barrinha", "Barueri", "Bastos", "Batatais", "Bauru", "Bebedouro", "Bento de Abreu", "Bernardino de Campos", "Bertioga", "Bilac", "Birigui", "Biritiba-Mirim", "Boa Esperança do Sul", "Bocaina", "Bofete", "Boituva", "Bom Jesus dos Perdões", "Bom Sucesso de Itararé", "Borá", "Boracéia", "Borborema", "Borebi", "Botucatu", "Bragança Paulista", "Braúna", "Brejo Alegre", "Brodowski", "Brotas", "Buri", "Buritama", "Buritizal", "Cabrália Paulista", "Cabreúva", "Caçapava", "Cachoeira Paulista", "Caconde", "Cafelândia", "Caiabu", "Caieiras", "Caiuá", "Cajamar", "Cajati", "Cajobi", "Cajuru", "Campina do Monte Alegre", "Campinas", "Campo Limpo Paulista", "Campos do Jordão", "Campos Novos Paulista", "Cananéia", "Canas", "Cândido Mota", "Cândido Rodrigues", "Canitar", "Capão Bonito", "Capela do Alto", "Capivari", "Caraguatatuba", "Carapicuíba", "Cardoso", "Casa Branca", "Cássia dos Coqueiros", "Castilho", "Catanduva", "Catiguá", "Cedral", "Cerqueira César", "Cerquilho", "Cesário Lange", "Charqueada", "Chavantes", "Clementina", "Colina", "Colômbia", "Conchal", "Conchas", "Cordeirópolis", "Coroados", "Coronel Macedo", "Corumbataí", "Cosmópolis", "Cosmorama", "Cotia", "Cravinhos", "Cristais Paulista", "Cruzália", "Cruzeiro", "Cubatão", "Cunha", "Descalvado", "Diadema", "Dirce Reis", "Divinolândia", "Dobrada", "Dois Córregos", "Dolcinópolis", "Dourado", "Dracena", "Duartina", "Dumont", "Echaporã", "Eldorado", "Elias Fausto", "Elisiário", "Embaúba", "Embu das Artes", "Embu-Guaçu", "Emilianópolis", "Engenheiro Coelho", "Espírito Santo do Pinhal", "Espírito Santo do Turvo", "Estiva Gerbi", "Estrela d'Oeste", "Estrela do Norte", "Euclides da Cunha Paulista", "Fartura", "Fernando Prestes", "Fernandópolis", "Fernão", "Ferraz de Vasconcelos", "Flora Rica", "Floreal", "Flórida Paulista", "Florínea", "Francisco Morato", "Franco da Rocha", "Gabriel Monteiro", "Gália", "Garça", "Gastão Vidigal", "Gavião Peixoto", "General Salgado", "Getulina", "Glicério", "Guaiçara", "Guaimbê", "Guaíra", "Guapiaçu", "Guapiara", "Guará", "Guaraçaí", "Guaraci", "Guarani d'Oeste", "Guarantã", "Guararapes", "Guararema", "Guaratinguetá", "Guareí", "Guariba", "Guarujá", "Guarulhos", "Guatapará", "Guzolândia", "Herculândia", "Holambra", "Hortolândia", "Iacanga", "Iacri", "Iaras", "Ibaté", "Ibirá", "Ibirarema", "Ibitinga", "Ibiúna", "Icém", "Iepê", "Igaraçu do Tietê", "Igarapava", "Igaratá", "Iguape", "Ilha Comprida", "Ilha Solteira", "Ilhabela", "Indaiatuba", "Indiana", "Indiaporã", "Inúbia Paulista", "Ipaussu", "Iperó", "Ipeúna", "Ipiguá", "Iporanga", "Ipuã", "Iracemápolis", "Irapuã", "Irapuru", "Itaberá", "Itaí", "Itajobi", "Itaju", "Itanhaém", "Itaóca", "Itapecerica da Serra", "Itapetininga", "Itapeva", "Itapevi", "Itapira", "Itapirapuã Paulista", "Itaporanga", "Itapuí", "Itapura", "Itaquaquecetuba", "Itararé", "Itariri", "Itatiba", "Itatinga", "Itirapina", "Itirapuã", "Itobi", "Itu", "Itupeva", "Ituverava", "Jaborandi", "Jaboticabal", "Jacareí", "Jaci", "Jacupiranga", "Jaguariúna", "Jales", "Jambeiro", "Jandira", "Jardinópolis", "Jarinu", "Jaú", "Jeriquara", "Joanópolis", "João Ramalho", "José Bonifácio", "Júlio Mesquita", "Jumirim", "Jundiaí", "Junqueirópolis", "Juquiá", "Juquitiba", "Lagoinha", "Laranjal Paulista", "Lavínia", "Lavrinhas", "Leme", "Lençóis Paulista", "Limeira", "Lindóia", "Lins", "Lorena", "Lourdes", "Louveira", "Lucélia", "Lucianópolis", "Luís Antônio", "Luiziânia", "Lupércio", "Lutécia", "Macatuba", "Macaubal", "Macedônia", "Magda", "Mairinque", "Mairiporã", "Manduri", "Marabá Paulista", "Maracaí", "Marapoama", "Mariápolis", "Marília", "Marinópolis", "Martinópolis", "Matão", "Mauá", "Mendonça", "Meridiano", "Mesópolis", "Miguelópolis", "Mineiros do Tietê", "Mira Estrela", "Miracatu", "Mirandópolis", "Mirante do Paranapanema", "Mirassol", "Mirassolândia", "Mococa", "Mogi das Cruzes", "Mogi Guaçu", "Mogi Mirim", "Mombuca", "Monções", "Mongaguá", "Monte Alegre do Sul", "Monte Alto", "Monte Aprazível", "Monte Azul Paulista", "Monte Castelo", "Monte Mor", "Monteiro Lobato", "Morro Agudo", "Morungaba", "Motuca", "Murutinga do Sul", "Nantes", "Narandiba", "Natividade da Serra", "Nazaré Paulista", "Neves Paulista", "Nhandeara", "Nipoã", "Nova Aliança", "Nova Campina", "Nova Canaã Paulista", "Nova Castilho", "Nova Europa", "Nova Granada", "Nova Guataporanga", "Nova Independência", "Nova Luzitânia", "Nova Odessa", "Novais", "Novo Horizonte", "Nuporanga", "Ocauçu", "Óleo", "Olímpia", "Onda Verde", "Oriente", "Orindiúva", "Orlândia", "Osasco", "Oscar Bressane", "Oswaldo Cruz", "Ourinhos", "Ouro Verde", "Ouroeste", "Pacaembu", "Palestina", "Palmares Paulista", "Palmeira d'Oeste", "Palmital", "Panorama", "Paraguaçu Paulista", "Paraibuna", "Paraíso", "Paranapanema", "Paranapuã", "Parapuã", "Pardinho", "Pariquera-Açu", "Parisi", "Patrocínio Paulista", "Paulicéia", "Paulistânia", "Paulo de Faria", "Pederneiras", "Pedra Bela", "Pedranópolis", "Pedregulho", "Pedreira", "Pedrinhas Paulista", "Pedro de Toledo", "Penápolis", "Pereira Barreto", "Pereiras", "Peruíbe", "Piacatu", "Piedade", "Pilar do Sul", "Pindamonhangaba", "Pindorama", "Pinhalzinho", "Piquerobi", "Piquete", "Piracaia", "Piracicaba", "Piraju", "Pirajuí", "Pirangi", "Pirapora do Bom Jesus", "Pirapozinho", "Pirassununga", "Piratininga", "Pitangueiras", "Planalto", "Platina", "Poá", "Poloni", "Pompéia", "Pongaí", "Pontal", "Pontalinda", "Pontes Gestal", "Populina", "Porangaba", "Porto Feliz", "Porto Ferreira", "Potim", "Potirendaba", "Pracinha", "Pradópolis", "Praia Grande", "Pratânia", "Presidente Bernardes", "Presidente Epitácio", "Presidente Prudente", "Presidente Venceslau", "Promissão", "Quadra", "Quatá", "Queiroz", "Queluz", "Quintana", "Rafard", "Rancharia", "Redenção da Serra", "Regente Feijó", "Reginópolis", "Registro", "Restinga", "Ribeira", "Ribeirão Bonito", "Ribeirão Branco", "Ribeirão Corrente", "Ribeirão do Sul", "Ribeirão dos Índios", "Ribeirão Grande", "Ribeirão Pires", "Ribeirão Preto", "Rifaina", "Rincão", "Rinópolis", "Rio Claro", "Rio das Pedras", "Rio Grande da Serra", "Riolândia", "Riversul", "Rosana", "Roseira", "Rubiácea", "Rubinéia", "Sabino", "Sagres", "Sales", "Sales Oliveira", "Salesópolis", "Salmourão", "Saltinho", "Salto", "Salto de Pirapora", "Salto Grande", "Sandovalina", "Santa Adélia", "Santa Albertina", "Santa Bárbara d'Oeste", "Santa Branca", "Santa Clara d'Oeste", "Santa Cruz da Conceição", "Santa Cruz da Esperança", "Santa Cruz das Palmeiras", "Santa Cruz do Rio Pardo", "Santa Ernestina", "Santa Fé do Sul", "Santa Gertrudes", "Santa Isabel", "Santa Lúcia", "Santa Maria da Serra", "Santa Mercedes", "Santa Rita d'Oeste", "Santa Rita do Passa Quatro", "Santa Rosa de Viterbo", "Santa Salete", "Santana da Ponte Pensa", "Santana de Parnaíba", "Santo Anastácio", "Santo André", "Santo Antônio da Alegria", "Santo Antônio de Posse", "Santo Antônio do Aracanguá", "Santo Antônio do Jardim", "Santo Antônio do Pinhal", "Santo Expedito", "Santópolis do Aguapeí", "Santos", "São Bento do Sapucaí", "São Bernardo do Campo", "São Caetano do Sul", "São Carlos", "São Francisco", "São João da Boa Vista", "São João das Duas Pontes", "São João de Iracema", "São João do Pau d'Alho", "São Joaquim da Barra", "São José da Bela Vista", "São José do Barreiro", "São José do Rio Pardo", "São José do Rio Preto", "São José dos Campos", "São Lourenço da Serra", "São Luís do Paraitinga", "São Manuel", "São Miguel Arcanjo", "São Paulo", "São Pedro", "São Pedro do Turvo", "São Roque", "São Sebastião", "São Sebastião da Grama", "São Simão", "São Vicente", "Sarapuí", "Sarutaiá", "Sebastianópolis do Sul", "Serra Azul", "Serra Negra", "Serrana", "Sertãozinho", "Sete Barras", "Severínia", "Silveiras", "Socorro", "Sorocaba", "Sud Mennucci", "Sumaré", "Suzanápolis", "Suzano", "Tabapuã", "Tabatinga", "Taboão da Serra", "Taciba", "Taguaí", "Taiaçu", "Taiúva", "Tambaú", "Tanabi", "Tapiraí", "Tapiratiba", "Taquaral", "Taquaritinga", "Taquarituba", "Taquarivaí", "Tarabai", "Tarumã", "Tatuí", "Taubaté", "Tejupá", "Teodoro Sampaio", "Terra Roxa", "Tietê", "Timburi", "Torre de Pedra", "Torrinha", "Trabiju", "Tremembé", "Três Fronteiras", "Tuiuti", "Tupã", "Tupi Paulista", "Turiúba", "Turmalina", "Ubarana", "Ubatuba", "Ubirajara", "Uchoa", "União Paulista", "Urânia", "Uru", "Urupês", "Valentim Gentil", "Valinhos", "Valparaíso", "Vargem", "Vargem Grande do Sul", "Vargem Grande Paulista", "Várzea Paulista", "Vera Cruz", "Vinhedo", "Viradouro", "Vista Alegre do Alto", "Vitória Brasil", "Votorantim", "Votuporanga", "Zacarias"
    ],
    # Adicione os restantes (PE, PI, RJ, RN, RS, RO, RR, SC, SE, TO) de forma similar.
    "PE": [
        "Abreu e Lima", "Afogados da Ingazeira", "Afrânio", "Agrestina", "Água Preta", "Águas Belas", "Alagoinha", "Aliança", "Altinho", "Amaraji", "Angelim", "Araçoiaba", "Araripina", "Arcoverde", "Barra de Guabiraba", "Barreiros", "Belém de Maria", "Belém do São Francisco", "Belo Jardim", "Betânia", "Bezerros", "Bodocó", "Bom Conselho", "Bom Jardim", "Bonito", "Brejão", "Brejinho", "Brejo da Madre de Deus", "Buenos Aires", "Buíque", "Cabo de Santo Agostinho", "Cabrobó", "Cachoeirinha", "Caetés", "Calçado", "Calumbi", "Camaragibe", "Camocim de São Félix", "Camutanga", "Canhotinho", "Capoeiras", "Carnaíba", "Carnaubeira da Penha", "Carpina", "Caruaru", "Casinhas", "Catende", "Cedro", "Chã de Alegria", "Chã Grande", "Condado", "Correntes", "Cortês", "Cumaru", "Cupira", "Custódia", "Dormentes", "Escada", "Exu", "Feira Nova", "Fernando de Noronha", "Ferreiros", "Flores", "Floresta", "Frei Miguelinho", "Gameleira", "Garanhuns", "Glória do Goitá", "Goiana", "Granito", "Gravatá", "Iati", "Ibimirim", "Ibirajuba", "Igarassu", "Iguaracy", "Inajá", "Ingazeira", "Ipojuca", "Ipubi", "Itacuruba", "Itaíba", "Ilha de Itamaracá", "Itambé", "Itapetim", "Itapissuma", "Itaquitinga", "Jaboatão dos Guararapes", "Jaqueira", "Jataúba", "Jatobá", "João Alfredo", "Joaquim Nabuco", "Jucati", "Jupi", "Jurema", "Lagoa de Itaenga", "Lagoa do Carro", "Lagoa do Ouro", "Lagoa dos Gatos", "Lagoa Grande", "Lajedo", "Limoeiro", "Macaparana", "Machados", "Manari", "Maraial", "Mirandiba", "Moreno", "Nazaré da Mata", "Olinda", "Orobó", "Orocó", "Ouricuri", "Palmares", "Palmeirina", "Panelas", "Paranatama", "Parnamirim", "Passira", "Paudalho", "Paulista", "Pedra", "Pesqueira", "Petrolândia", "Petrolina", "Poção", "Pombos", "Primavera", "Quipapá", "Quixaba", "Recife", "Riacho das Almas", "Ribeirão", "Rio Formoso", "Sairé", "Salgadinho", "Salgueiro", "Saloá", "Sanharó", "Santa Cruz", "Santa Cruz da Baixa Verde", "Santa Cruz do Capibaribe", "Santa Filomena", "Santa Maria da Boa Vista", "Santa Maria do Cambucá", "Santa Terezinha", "São Benedito do Sul", "São Bento do Una", "São Caitano", "São João", "São Joaquim do Monte", "São José da Coroa Grande", "São José do Belmonte", "São José do Egito", "São Lourenço da Mata", "São Vicente Ferrer", "Serra Talhada", "Serrita", "Sertânia", "Sirinhaém", "Moreilândia", "Solidão", "Surubim", "Tabira", "Tacaimbó", "Tacaratu", "Tamandaré", "Taquaritinga do Norte", "Terezinha", "Terra Nova", "Timbaúba", "Toritama", "Tracunhaém", "Trindade", "Triunfo", "Tupanatinga", "Tuparetama", "Venturosa", "Verdejante", "Vertente do Lério", "Vertentes", "Vicência", "Vitória de Santo Antão", "Xexéu"
    ],
    "PI": [
        "Acauã", "Agricolândia", "Água Branca", "Alagoinha do Piauí", "Alegrete do Piauí", "Alto Longá", "Altos", "Alvorada do Gurguéia", "Amarante", "Angical do Piauí", "Anísio de Abreu", "Antônio Almeida", "Aroazes", "Aroeiras do Itaim", "Arraial", "Assunção do Piauí", "Avelino Lopes", "Baixa Grande do Ribeiro", "Barra d'Alcântara", "Barras", "Barreiras do Piauí", "Barro Duro", "Batalha", "Bela Vista do Piauí", "Belém do Piauí", "Beneditinos", "Bertolínia", "Betânia do Piauí", "Boa Hora", "Bocaina", "Bom Jesus", "Bom Princípio do Piauí", "Bonfim do Piauí", "Boqueirão do Piauí", "Brasileira", "Brejo do Piauí", "Buriti dos Lopes", "Buriti dos Montes", "Cabeceiras do Piauí", "Cajazeiras do Piauí", "Cajueiro da Praia", "Caldeirão Grande do Piauí", "Campinas do Piauí", "Campo Alegre do Fidalgo", "Campo Grande do Piauí", "Campo Largo do Piauí", "Campo Maior", "Canavieira", "Canto do Buriti", "Capitão de Campos", "Capitão Gervásio Oliveira", "Caracol", "Caraúbas do Piauí", "Caridade do Piauí", "Castelo do Piauí", "Caxingó", "Cocal", "Cocal de Telha", "Cocal dos Alves", "Coivaras", "Colônia do Gurguéia", "Colônia do Piauí", "Conceição do Canindé", "Coronel José Dias", "Corrente", "Cristalândia do Piauí", "Cristino Castro", "Curimatá", "Currais", "Curralinhos", "Curral Novo do Piauí", "Demerval Lobão", "Dirceu Arcoverde", "Dom Expedito Lopes", "Domingos Mourão", "Dom Inocêncio", "Elesbão Veloso", "Eliseu Martins", "Esperantina", "Fartura do Piauí", "Flores do Piauí", "Floresta do Piauí", "Floriano", "Francinópolis", "Francisco Ayres", "Francisco Macedo", "Francisco Santos", "Fronteiras", "Geminiano", "Gilbués", "Guadalupe", "Guaribas", "Hugo Napoleão", "Ilha Grande", "Inhuma", "Ipiranga do Piauí", "Isaías Coelho", "Itainópolis", "Itaueira", "Jacobina do Piauí", "Jaicós", "Jardim do Mulato", "Jatobá do Piauí", "Jerumenha", "João Costa", "Joaquim Pires", "Joca Marques", "José de Freitas", "Juazeiro do Piauí", "Júlio Borges", "Jurema", "Lagoinha do Piauí", "Lagoa Alegre", "Lagoa do Barro do Piauí", "Lagoa de São Francisco", "Lagoa do Piauí", "Lagoa do Sítio", "Landri Sales", "Luís Correia", "Luzilândia", "Madeiro", "Manoel Emídio", "Marcolândia", "Marcos Parente", "Massapê do Piauí", "Matias Olímpio", "Miguel Alves", "Miguel Leão", "Milton Brandão", "Monsenhor Gil", "Monsenhor Hipólito", "Monte Alegre do Piauí", "Morro Cabeça no Tempo", "Morro do Chapéu do Piauí", "Murici dos Portelas", "Nazaré do Piauí", "Nazária", "Nossa Senhora de Nazaré", "Nossa Senhora dos Remédios", "Novo Oriente do Piauí", "Novo Santo Antônio", "Oeiras", "Padre Marcos", "Paes Landim", "Pajeú do Piauí", "Palmeira do Piauí", "Palmeirais", "Paquetá", "Parnaguá", "Parnaíba", "Passagem Franca do Piauí", "Patos do Piauí", "Pau d'Arco do Piauí", "Paulistana", "Pavussu", "Pedro II", "Pedro Laurentino", "Nova Santa Rita", "Picos", "Pimenteiras", "Pio IX", "Piracuruca", "Piripiri", "Porto", "Porto Alegre do Piauí", "Prata do Piauí", "Queimada Nova", "Redenção do Gurguéia", "Regeneração", "Riacho Frio", "Ribeira do Piauí", "Rio Grande do Piauí", "Santa Cruz do Piauí", "Santa Cruz dos Milagres", "Santa Filomena", "Santa Luz", "Santana do Piauí", "Santa Rosa do Piauí", "Santo Antônio de Lisboa", "Santo Antônio dos Milagres", "Santo Inácio do Piauí", "São Braz do Piauí", "São Félix do Piauí", "São Francisco de Assis do Piauí", "São Francisco do Piauí", "São Gonçalo do Gurguéia", "São Gonçalo do Piauí", "São João da Canabrava", "São João da Fronteira", "São João da Serra", "São João da Varjota", "São João do Arraial", "São João do Piauí", "São José do Divino", "São José do Peixe", "São José do Piauí", "São Julião", "São Lourenço do Piauí", "São Luis do Piauí", "São Miguel da Baixa Grande", "São Miguel do Fidalgo", "São Miguel do Tapuio", "São Pedro do Piauí", "São Raimundo Nonato", "Sebastião Barros", "Sebastião Leal", "Sigefredo Pacheco", "Simões", "Simplício Mendes", "Socorro do Piauí", "Sussuapara", "Tamboril do Piauí", "Tanque do Piauí", "Teresina", "União", "Uruçuí", "Valença do Piauí", "Várzea Branca", "Várzea Grande", "Vera Mendes", "Vila Nova do Piauí", "Wall Ferraz"
    ],
    "RJ": [
        "Angra dos Reis", "Aperibé", "Araruama", "Areal", "Armação dos Búzios", "Arraial do Cabo",
        "Barra Mansa", "Barra do Piraí", "Belford Roxo", "Bom Jardim", "Bom Jesus do Itabapoana", "Cabo Frio",
        "Cachoeiras de Macacu", "Cambuci", "Campos dos Goytacazes", "Cantagalo", "Carapebus", "Cardoso Moreira",
        "Carmo", "Casimiro de Abreu", "Comendador Levy Gasparian", "Conceição de Macabu", "Cordeiro", "Duas Barras",
        "Duque de Caxias", "Engenheiro Paulo de Frontin", "Guapimirim", "Iguaba Grande", "Itaboraí", "Itaguaí",
        "Italva", "Itaocara", "Itaperuna", "Itatiaia", "Japeri", "Laje do Muriaé",
        "Macaé", "Macuco", "Magé", "Mangaratiba", "Maricá", "Mendes",
        "Mesquita", "Miguel Pereira", "Miracema", "Natividade", "Nilópolis", "Niterói",
        "Nova Friburgo", "Nova Iguaçu", "Paracambi", "Paraty", "Paraíba do Sul", "Paty do Alferes",
        "Petrópolis", "Pinheiral", "Piraí", "Porciúncula", "Porto Real", "Quatis",
        "Queimados", "Quissamã", "Resende", "Rio Bonito", "Rio Claro", "Rio das Flores",
        "Rio das Ostras", "Rio de Janeiro", "Santa Maria Madalena", "Santo Antônio de Pádua", "Sapucaia", "Saquarema",
        "Seropédica", "Silva Jardim", "Sumidouro", "São Fidélis", "São Francisco de Itabapoana", "São Gonçalo",
        "São José de Ubá", "São José do Vale do Rio Preto", "São João da Barra", "São João de Meriti", "São Pedro da Aldeia", "São Sebastião do Alto",
        "Tanguá", "Teresópolis", "Trajano de Moraes", "Três Rios", "Valença", "Varre-Sai",
        "Vassouras", "Volta Redonda"
    ],
    "RN": [
        "Acari", "Afonso Bezerra", "Alexandria", "Almino Afonso", "Alto do Rodrigues", "Angicos",
        "Antônio Martins", "Apodi", "Areia Branca", "Arês", "Açu", "Baraúna",
        "Barcelona", "Baía Formosa", "Bento Fernandes", "Bodó", "Bom Jesus", "Brejinho",
        "Caicó", "Caiçara do Norte", "Caiçara do Rio do Vento", "Campo Grande", "Campo Redondo", "Canguaretama",
        "Caraúbas", "Carnaubais", "Carnaúba dos Dantas", "Ceará-Mirim", "Cerro Corá", "Coronel Ezequiel",
        "Coronel João Pessoa", "Cruzeta", "Currais Novos", "Doutor Severiano", "Encanto", "Equador",
        "Espírito Santo", "Extremoz", "Felipe Guerra", "Fernando Pedroza", "Florânia", "Francisco Dantas",
        "Frutuoso Gomes", "Galinhos", "Goianinha", "Governador Dix-Sept Rosado", "Grossos", "Guamaré",
        "Ielmo Marinho", "Ipanguaçu", "Ipueira", "Itajá", "Itaú", "Jandaíra",
        "Janduís", "Januário Cicco", "Japi", "Jardim de Angicos", "Jardim de Piranhas", "Jardim do Seridó",
        "Jaçanã", "José da Penha", "João Câmara", "João Dias", "Jucurutu", "Jundiá",
        "Lagoa Nova", "Lagoa Salgada", "Lagoa d'Anta", "Lagoa de Pedras", "Lagoa de Velhos", "Lajes",
        "Lajes Pintadas", "Lucrécia", "Luís Gomes", "Macau", "Macaíba", "Major Sales",
        "Marcelino Vieira", "Martins", "Maxaranguape", "Messias Targino", "Montanhas", "Monte Alegre",
        "Monte das Gameleiras", "Mossoró", "Natal", "Nova Cruz", "Nísia Floresta", "Olho d'Água do Borges",
        "Ouro Branco", "Paraná", "Parazinho", "Paraú", "Parelhas", "Parnamirim",
        "Passa e Fica", "Passagem", "Patu", "Pau dos Ferros", "Pedra Grande", "Pedra Preta",
        "Pedro Avelino", "Pedro Velho", "Pendências", "Pilões", "Portalegre", "Porto do Mangue",
        "Poço Branco", "Pureza", "Rafael Fernandes", "Rafael Godeiro", "Riacho da Cruz", "Riacho de Santana",
        "Riachuelo", "Rio do Fogo", "Rodolfo Fernandes", "Ruy Barbosa", "Santa Cruz", "Santa Maria",
        "Santana do Matos", "Santana do Seridó", "Santo Antônio", "Senador Elói de Souza", "Senador Georgino Avelino", "Serra Caiada",
        "Serra Negra do Norte", "Serra de São Bento", "Serra do Mel", "Serrinha", "Serrinha dos Pintos", "Severiano Melo",
        "São Bento do Norte", "São Bento do Trairí", "São Fernando", "São Francisco do Oeste", "São Gonçalo do Amarante", "São José de Mipibu",
        "São José do Campestre", "São José do Seridó", "São João do Sabugi", "São Miguel", "São Miguel do Gostoso", "São Paulo do Potengi",
        "São Pedro", "São Rafael", "São Tomé", "São Vicente", "Sítio Novo", "Taboleiro Grande",
        "Taipu", "Tangará", "Tenente Ananias", "Tenente Laurentino Cruz", "Tibau", "Tibau do Sul",
        "Timbaúba dos Batistas", "Touros", "Triunfo Potiguar", "Umarizal", "Upanema", "Venha-Ver",
        "Vera Cruz", "Vila Flor", "Viçosa", "Várzea", "Água Nova"
    ],
    "RO": [
        "Alta Floresta D'Oeste", "Alto Alegre dos Parecis", "Alto Paraíso", "Alvorada D'Oeste", "Ariquemes", "Buritis",
        "Cabixi", "Cacaulândia", "Cacoal", "Campo Novo de Rondônia", "Candeias do Jamari", "Castanheiras",
        "Cerejeiras", "Chupinguaia", "Colorado do Oeste", "Corumbiara", "Costa Marques", "Cujubim",
        "Espigão D'Oeste", "Governador Jorge Teixeira", "Guajará-Mirim", "Itapuã do Oeste", "Jaru", "Ji-Paraná",
        "Machadinho D'Oeste", "Ministro Andreazza", "Mirante da Serra", "Monte Negro", "Nova Brasilândia D'Oeste", "Nova Mamoré",
        "Nova União", "Novo Horizonte do Oeste", "Ouro Preto do Oeste", "Parecis", "Pimenta Bueno", "Pimenteiras do Oeste",
        "Porto Velho", "Presidente Médici", "Primavera de Rondônia", "Rio Crespo", "Rolim de Moura", "Santa Luzia D'Oeste",
        "Seringueiras", "São Felipe D'Oeste", "São Francisco do Guaporé", "São Miguel do Guaporé", "Teixeirópolis", "Theobroma",
        "Urupá", "Vale do Anari", "Vale do Paraíso", "Vilhena"
    ],
    "RR": [
        "Alto Alegre", "Amajari", "Boa Vista", "Bonfim", "Cantá", "Caracaraí",
        "Caroebe", "Iracema", "Mucajaí", "Normandia", "Pacaraima", "Rorainópolis",
        "São João da Baliza", "São Luiz", "Uiramutã"
    ],
    "RS": [
        "Aceguá", "Agudo", "Ajuricaba", "Alecrim", "Alegrete", "Alegria",
        "Almirante Tamandaré do Sul", "Alpestre", "Alto Alegre", "Alto Feliz", "Alvorada", "Amaral Ferrador",
        "Ametista do Sul", "André da Rocha", "Anta Gorda", "Antônio Prado", "Arambaré", "Araricá",
        "Aratiba", "Arroio Grande", "Arroio do Meio", "Arroio do Padre", "Arroio do Sal", "Arroio do Tigre",
        "Arroio dos Ratos", "Arvorezinha", "Augusto Pestana", "Bagé", "Balneário Pinhal", "Barra Funda",
        "Barra do Guarita", "Barra do Quaraí", "Barra do Ribeiro", "Barra do Rio Azul", "Barracão", "Barros Cassal",
        "Barão", "Barão de Cotegipe", "Barão do Triunfo", "Benjamin Constant do Sul", "Bento Gonçalves", "Boa Vista das Missões",
        "Boa Vista do Buricá", "Boa Vista do Cadeado", "Boa Vista do Incra", "Boa Vista do Sul", "Bom Jesus", "Bom Princípio",
        "Bom Progresso", "Bom Retiro do Sul", "Boqueirão do Leão", "Bossoroca", "Bozano", "Braga",
        "Brochier", "Butiá", "Cacequi", "Cachoeira do Sul", "Cachoeirinha", "Cacique Doble",
        "Caibaté", "Caiçara", "Camaquã", "Camargo", "Cambará do Sul", "Campestre da Serra",
        "Campina das Missões", "Campinas do Sul", "Campo Bom", "Campo Novo", "Campos Borges", "Candelária",
        "Candiota", "Canela", "Canguçu", "Canoas", "Canudos do Vale", "Capela de Santana",
        "Capitão", "Capivari do Sul", "Capão Bonito do Sul", "Capão da Canoa", "Capão do Cipó", "Capão do Leão",
        "Carazinho", "Caraá", "Carlos Barbosa", "Carlos Gomes", "Casca", "Caseiros",
        "Catuípe", "Caxias do Sul", "Caçapava do Sul", "Centenário", "Cerrito", "Cerro Branco",
        "Cerro Grande", "Cerro Grande do Sul", "Cerro Largo", "Chapada", "Charqueadas", "Charrua",
        "Chiapetta", "Chuvisca", "Chuí", "Cidreira", "Ciríaco", "Colinas",
        "Colorado", "Condor", "Constantina", "Coqueiro Baixo", "Coqueiros do Sul", "Coronel Barros",
        "Coronel Bicaco", "Coronel Pilar", "Cotiporã", "Coxilha", "Crissiumal", "Cristal",
        "Cristal do Sul", "Cruz Alta", "Cruzaltense", "Cruzeiro do Sul", "Cândido Godói", "David Canabarro",
        "Derrubadas", "Dezesseis de Novembro", "Dilermando de Aguiar", "Dois Irmãos", "Dois Irmãos das Missões", "Dois Lajeados",
        "Dom Feliciano", "Dom Pedrito", "Dom Pedro de Alcântara", "Dona Francisca", "Doutor Maurício Cardoso", "Doutor Ricardo",
        "Eldorado do Sul", "Encantado", "Encruzilhada do Sul", "Engenho Velho", "Entre Rios do Sul", "Entre-Ijuís",
        "Erebango", "Erechim", "Ernestina", "Erval Grande", "Erval Seco", "Esmeralda",
        "Esperança do Sul", "Espumoso", "Estação", "Esteio", "Estrela", "Estrela Velha",
        "Estância Velha", "Eugênio de Castro", "Fagundes Varela", "Farroupilha", "Faxinal do Soturno", "Faxinalzinho",
        "Fazenda Vilanova", "Feliz", "Flores da Cunha", "Floriano Peixoto", "Fontoura Xavier", "Formigueiro",
        "Forquetinha", "Fortaleza dos Valos", "Frederico Westphalen", "Garibaldi", "Garruchos", "Gaurama",
        "General Câmara", "Gentil", "Getúlio Vargas", "Giruá", "Glorinha", "Gramado",
        "Gramado Xavier", "Gramado dos Loureiros", "Gravataí", "Guabiju", "Guaporé", "Guarani das Missões",
        "Guaíba", "Harmonia", "Herval", "Herveiras", "Horizontina", "Hulha Negra",
        "Humaitá", "Ibarama", "Ibiaçá", "Ibiraiaras", "Ibirapuitã", "Ibirubá",
        "Igrejinha", "Ijuí", "Ilópolis", "Imbé", "Imigrante", "Independência",
        "Inhacorá", "Ipiranga do Sul", "Ipê", "Iraí", "Itaara", "Itacurubi",
        "Itapuca", "Itaqui", "Itati", "Itatiba do Sul", "Ivorá", "Ivoti",
        "Jaboticaba", "Jacuizinho", "Jacutinga", "Jaguari", "Jaguarão", "Jaquirana",
        "Jari", "Jóia", "Júlio de Castilhos", "Lagoa Bonita do Sul", "Lagoa Vermelha", "Lagoa dos Três Cantos",
        "Lagoão", "Lajeado", "Lajeado do Bugre", "Lavras do Sul", "Liberato Salzano", "Lindolfo Collor",
        "Linha Nova", "Machadinho", "Mampituba", "Manoel Viana", "Maquiné", "Maratá",
        "Marau", "Marcelino Ramos", "Mariana Pimentel", "Mariano Moro", "Marques de Souza", "Mata",
        "Mato Castelhano", "Mato Leitão", "Mato Queimado", "Maximiliano de Almeida", "Maçambará", "Minas do Leão",
        "Miraguaí", "Montauri", "Monte Alegre dos Campos", "Monte Belo do Sul", "Montenegro", "Mormaço",
        "Morrinhos do Sul", "Morro Redondo", "Morro Reuter", "Mostardas", "Muitos Capões", "Muliterno",
        "Muçum", "Nicolau Vergueiro", "Nonoai", "Nova Alvorada", "Nova Araçá", "Nova Bassano",
        "Nova Boa Vista", "Nova Bréscia", "Nova Candelária", "Nova Esperança do Sul", "Nova Hartz", "Nova Palma",
        "Nova Petrópolis", "Nova Prata", "Nova Pádua", "Nova Ramada", "Nova Roma do Sul", "Nova Santa Rita",
        "Novo Barreiro", "Novo Cabrais", "Novo Hamburgo", "Novo Machado", "Novo Tiradentes", "Novo Xingu",
        "Não-Me-Toque", "Osório", "Paim Filho", "Palmares do Sul", "Palmeira das Missões", "Palmitinho",
        "Panambi", "Pantano Grande", "Paraí", "Paraíso do Sul", "Pareci Novo", "Parobé",
        "Passa Sete", "Passo Fundo", "Passo do Sobrado", "Paulo Bento", "Paverama", "Pedras Altas",
        "Pedro Osório", "Pejuçara", "Pelotas", "Picada Café", "Pinhal", "Pinhal Grande",
        "Pinhal da Serra", "Pinheirinho do Vale", "Pinheiro Machado", "Pinto Bandeira", "Pirapó", "Piratini",
        "Planalto", "Ponte Preta", "Pontão", "Porto Alegre", "Porto Lucena", "Porto Mauá",
        "Porto Vera Cruz", "Porto Xavier", "Portão", "Pouso Novo", "Poço das Antas", "Presidente Lucena",
        "Progresso", "Protásio Alves", "Putinga", "Quaraí", "Quatro Irmãos", "Quevedos",
        "Quinze de Novembro", "Redentora", "Relvado", "Restinga Sêca", "Rio Grande", "Rio Pardo",
        "Rio dos Índios", "Riozinho", "Roca Sales", "Rodeio Bonito", "Rolador", "Rolante",
        "Ronda Alta", "Rondinha", "Roque Gonzales", "Rosário do Sul", "Sagrada Família", "Saldanha Marinho",
        "Salto do Jacuí", "Salvador das Missões", "Salvador do Sul", "Sananduva", "Sant'Ana do Livramento", "Santa Bárbara do Sul",
        "Santa Cecília do Sul", "Santa Clara do Sul", "Santa Cruz do Sul", "Santa Margarida do Sul", "Santa Maria", "Santa Maria do Herval",
        "Santa Rosa", "Santa Tereza", "Santa Vitória do Palmar", "Santana da Boa Vista", "Santiago", "Santo Antônio da Patrulha",
        "Santo Antônio das Missões", "Santo Antônio do Palma", "Santo Antônio do Planalto", "Santo Augusto", "Santo Cristo", "Santo Expedito do Sul",
        "Santo Ângelo", "Sapiranga", "Sapucaia do Sul", "Sarandi", "Seberi", "Sede Nova",
        "Segredo", "Selbach", "Senador Salgado Filho", "Sentinela do Sul", "Serafina Corrêa", "Sertão",
        "Sertão Santana", "Sete de Setembro", "Severiano de Almeida", "Silveira Martins", "Sinimbu", "Sobradinho",
        "Soledade", "São Borja", "São Domingos do Sul", "São Francisco de Assis", "São Francisco de Paula", "São Gabriel",
        "São Jerônimo", "São Jorge", "São José das Missões", "São José do Herval", "São José do Hortêncio", "São José do Inhacorá",
        "São José do Norte", "São José do Ouro", "São José do Sul", "São José dos Ausentes", "São João da Urtiga", "São João do Polêsine",
        "São Leopoldo", "São Lourenço do Sul", "São Luiz Gonzaga", "São Marcos", "São Martinho", "São Martinho da Serra",
        "São Miguel das Missões", "São Nicolau", "São Paulo das Missões", "São Pedro da Serra", "São Pedro das Missões", "São Pedro do Butiá",
        "São Pedro do Sul", "São Sebastião do Caí", "São Sepé", "São Valentim", "São Valentim do Sul", "São Valério do Sul",
        "São Vendelino", "São Vicente do Sul", "Sério", "Tabaí", "Tapejara", "Tapera",
        "Tapes", "Taquara", "Taquari", "Taquaruçu do Sul", "Tavares", "Tenente Portela",
        "Terra de Areia", "Teutônia", "Tio Hugo", "Tiradentes do Sul", "Toropi", "Torres",
        "Tramandaí", "Travesseiro", "Trindade do Sul", "Triunfo", "Três Arroios", "Três Cachoeiras",
        "Três Coroas", "Três Forquilhas", "Três Palmeiras", "Três Passos", "Três de Maio", "Tucunduva",
        "Tunas", "Tupanci do Sul", "Tupanciretã", "Tupandi", "Tuparendi", "Turuçu",
        "Ubiretama", "Unistalda", "União da Serra", "Uruguaiana", "Vacaria", "Vale Real",
        "Vale Verde", "Vale do Sol", "Vanini", "Venâncio Aires", "Vera Cruz", "Veranópolis",
        "Vespasiano Corrêa", "Viadutos", "Viamão", "Vicente Dutra", "Victor Graeff", "Vila Flores",
        "Vila Lângaro", "Vila Maria", "Vila Nova do Sul", "Vista Alegre", "Vista Alegre do Prata", "Vista Gaúcha",
        "Vitória das Missões", "Westfália", "Xangri-lá", "Água Santa", "Áurea"
    ],
    "SC": [
        "Abdon Batista", "Abelardo Luz", "Agrolândia", "Agronômica", "Alfredo Wagner", "Alto Bela Vista",
        "Anchieta", "Angelina", "Anita Garibaldi", "Anitápolis", "Antônio Carlos", "Apiúna",
        "Arabutã", "Araquari", "Araranguá", "Armazém", "Arroio Trinta", "Arvoredo",
        "Ascurra", "Atalanta", "Aurora", "Balneário Arroio do Silva", "Balneário Barra do Sul", "Balneário Camboriú",
        "Balneário Gaivota", "Balneário Piçarras", "Balneário Rincão", "Bandeirante", "Barra Bonita", "Barra Velha",
        "Bela Vista do Toldo", "Belmonte", "Benedito Novo", "Biguaçu", "Blumenau", "Bocaina do Sul",
        "Bom Jardim da Serra", "Bom Jesus", "Bom Jesus do Oeste", "Bom Retiro", "Bombinhas", "Botuverá",
        "Braço do Norte", "Braço do Trombudo", "Brunópolis", "Brusque", "Caibi", "Calmon",
        "Camboriú", "Campo Alegre", "Campo Belo do Sul", "Campo Erê", "Campos Novos", "Canelinha",
        "Canoinhas", "Capinzal", "Capivari de Baixo", "Capão Alto", "Catanduvas", "Caxambu do Sul",
        "Caçador", "Celso Ramos", "Cerro Negro", "Chapadão do Lageado", "Chapecó", "Cocal do Sul",
        "Concórdia", "Cordilheira Alta", "Coronel Freitas", "Coronel Martins", "Correia Pinto", "Corupá",
        "Criciúma", "Cunha Porã", "Cunhataí", "Curitibanos", "Descanso", "Dionísio Cerqueira",
        "Dona Emma", "Doutor Pedrinho", "Entre Rios", "Ermo", "Erval Velho", "Faxinal dos Guedes",
        "Flor do Sertão", "Florianópolis", "Formosa do Sul", "Forquilhinha", "Fraiburgo", "Frei Rogério",
        "Galvão", "Garopaba", "Garuva", "Gaspar", "Governador Celso Ramos", "Gravatal",
        "Grão-Pará", "Guabiruba", "Guaraciaba", "Guaramirim", "Guarujá do Sul", "Guatambú",
        "Herval d'Oeste", "Ibiam", "Ibicaré", "Ibirama", "Ilhota", "Imaruí",
        "Imbituba", "Imbuia", "Indaial", "Iomerê", "Ipira", "Iporã do Oeste",
        "Ipuaçu", "Ipumirim", "Iraceminha", "Irani", "Irati", "Irineópolis",
        "Itaiópolis", "Itajaí", "Itapema", "Itapiranga", "Itapoá", "Ituporanga",
        "Itá", "Içara", "Jaborá", "Jacinto Machado", "Jaguaruna", "Jaraguá do Sul",
        "Jardinópolis", "Joaçaba", "Joinville", "José Boiteux", "Jupiá", "Lacerdópolis",
        "Lages", "Laguna", "Lajeado Grande", "Laurentino", "Lauro Müller", "Lebon Régis",
        "Leoberto Leal", "Lindóia do Sul", "Lontras", "Luiz Alves", "Luzerna", "Macieira",
        "Mafra", "Major Gercino", "Major Vieira", "Maracajá", "Maravilha", "Marema",
        "Massaranduba", "Matos Costa", "Meleiro", "Mirim Doce", "Modelo", "Mondaí",
        "Monte Carlo", "Monte Castelo", "Morro Grande", "Morro da Fumaça", "Navegantes", "Nova Erechim",
        "Nova Itaberaba", "Nova Trento", "Nova Veneza", "Novo Horizonte", "Orleans", "Otacílio Costa",
        "Ouro", "Ouro Verde", "Paial", "Painel", "Palhoça", "Palma Sola",
        "Palmeira", "Palmitos", "Papanduva", "Paraíso", "Passo de Torres", "Passos Maia",
        "Paulo Lopes", "Pedras Grandes", "Penha", "Peritiba", "Pescaria Brava", "Petrolândia",
        "Pinhalzinho", "Pinheiro Preto", "Piratuba", "Planalto Alegre", "Pomerode", "Ponte Alta",
        "Ponte Alta do Norte", "Ponte Serrada", "Porto Belo", "Porto União", "Pouso Redondo", "Praia Grande",
        "Presidente Castello Branco", "Presidente Getúlio", "Presidente Nereu", "Princesa", "Quilombo", "Rancho Queimado",
        "Rio Fortuna", "Rio Negrinho", "Rio Rufino", "Rio das Antas", "Rio do Campo", "Rio do Oeste",
        "Rio do Sul", "Rio dos Cedros", "Riqueza", "Rodeio", "Romelândia", "Salete",
        "Saltinho", "Salto Veloso", "Sangão", "Santa Cecília", "Santa Helena", "Santa Rosa de Lima",
        "Santa Rosa do Sul", "Santa Terezinha", "Santa Terezinha do Progresso", "Santiago do Sul", "Santo Amaro da Imperatriz", "Saudades",
        "Schroeder", "Seara", "Serra Alta", "Siderópolis", "Sombrio", "Sul Brasil",
        "São Bento do Sul", "São Bernardino", "São Bonifácio", "São Carlos", "São Cristóvão do Sul", "São Domingos",
        "São Francisco do Sul", "São Joaquim", "São José", "São José do Cedro", "São José do Cerrito", "São João Batista",
        "São João do Itaperiú", "São João do Oeste", "São João do Sul", "São Lourenço do Oeste", "São Ludgero", "São Martinho",
        "São Miguel da Boa Vista", "São Miguel do Oeste", "São Pedro de Alcântara", "Taió", "Tangará", "Tigrinhos",
        "Tijucas", "Timbé do Sul", "Timbó", "Timbó Grande", "Treviso", "Treze Tílias",
        "Treze de Maio", "Trombudo Central", "Três Barras", "Tubarão", "Tunápolis", "Turvo",
        "União do Oeste", "Urubici", "Urupema", "Urussanga", "Vargem", "Vargem Bonita",
        "Vargeão", "Vidal Ramos", "Videira", "Vitor Meireles", "Witmarsum", "Xanxerê",
        "Xavantina", "Xaxim", "Zortéa", "Água Doce", "Águas Frias", "Águas Mornas",
        "Águas de Chapecó"
    ],
    "SE": [
        "Amparo do São Francisco", "Aquidabã", "Aracaju", "Arauá", "Areia Branca", "Barra dos Coqueiros",
        "Boquim", "Brejo Grande", "Campo do Brito", "Canhoba", "Canindé de São Francisco", "Capela",
        "Carira", "Carmópolis", "Cedro de São João", "Cristinápolis", "Cumbe", "Divina Pastora",
        "Estância", "Feira Nova", "Frei Paulo", "Gararu", "General Maynard", "Graccho Cardoso",
        "Ilha das Flores", "Indiaroba", "Itabaiana", "Itabaianinha", "Itabi", "Itaporanga d'Ajuda",
        "Japaratuba", "Japoatã", "Lagarto", "Laranjeiras", "Macambira", "Malhada dos Bois",
        "Malhador", "Maruim", "Moita Bonita", "Monte Alegre de Sergipe", "Muribeca", "Neópolis",
        "Nossa Senhora Aparecida", "Nossa Senhora da Glória", "Nossa Senhora das Dores", "Nossa Senhora de Lourdes", "Nossa Senhora do Socorro", "Pacatuba",
        "Pedra Mole", "Pedrinhas", "Pinhão", "Pirambu", "Porto da Folha", "Poço Redondo",
        "Poço Verde", "Propriá", "Riachuelo", "Riachão do Dantas", "Ribeirópolis", "Rosário do Catete",
        "Salgado", "Santa Luzia do Itanhy", "Santa Rosa de Lima", "Santana do São Francisco", "Santo Amaro das Brotas", "Simão Dias",
        "Siriri", "São Cristóvão", "São Domingos", "São Francisco", "São Miguel do Aleixo", "Telha",
        "Tobias Barreto", "Tomar do Geru", "Umbaúba"
    ],
    "TO": [
        "Abreulândia", "Aguiarnópolis", "Aliança do Tocantins", "Almas", "Alvorada", "Ananás",
        "Angico", "Aparecida do Rio Negro", "Aragominas", "Araguacema", "Araguanã", "Araguatins",
        "Araguaçu", "Araguaína", "Arapoema", "Arraias", "Augustinópolis", "Aurora do Tocantins",
        "Axixá do Tocantins", "Babaçulândia", "Bandeirantes do Tocantins", "Barra do Ouro", "Barrolândia", "Bernardo Sayão",
        "Bom Jesus do Tocantins", "Brasilândia do Tocantins", "Brejinho de Nazaré", "Buriti do Tocantins", "Cachoeirinha", "Campos Lindos",
        "Cariri do Tocantins", "Carmolândia", "Carrasco Bonito", "Caseara", "Centenário", "Chapada da Natividade",
        "Chapada de Areia", "Colinas do Tocantins", "Colméia", "Combinado", "Conceição do Tocantins", "Couto Magalhães",
        "Cristalândia", "Crixás do Tocantins", "Darcinópolis", "Dianópolis", "Divinópolis do Tocantins", "Dois Irmãos do Tocantins",
        "Dueré", "Esperantina", "Figueirópolis", "Filadélfia", "Formoso do Araguaia", "Fátima",
        "Goianorte", "Goiatins", "Guaraí", "Gurupi", "Ipueiras", "Itacajá",
        "Itaguatins", "Itapiratins", "Itaporã do Tocantins", "Jaú do Tocantins", "Juarina", "Lagoa da Confusão",
        "Lagoa do Tocantins", "Lajeado", "Lavandeira", "Lizarda", "Luzinópolis", "Marianópolis do Tocantins",
        "Mateiros", "Maurilândia do Tocantins", "Miracema do Tocantins", "Miranorte", "Monte Santo do Tocantins", "Monte do Carmo",
        "Muricilândia", "Natividade", "Nazaré", "Nova Olinda", "Nova Rosalândia", "Novo Acordo",
        "Novo Alegre", "Novo Jardim", "Oliveira de Fátima", "Palmas", "Palmeirante", "Palmeiras do Tocantins",
        "Palmeirópolis", "Paranã", "Paraíso do Tocantins", "Pau D'Arco", "Pedro Afonso", "Peixe",
        "Pequizeiro", "Pindorama do Tocantins", "Piraquê", "Pium", "Ponte Alta do Bom Jesus", "Ponte Alta do Tocantins",
        "Porto Alegre do Tocantins", "Porto Nacional", "Praia Norte", "Presidente Kennedy", "Pugmil", "Recursolândia",
        "Riachinho", "Rio Sono", "Rio da Conceição", "Rio dos Bois", "Sampaio", "Sandolândia",
        "Santa Fé do Araguaia", "Santa Maria do Tocantins", "Santa Rita do Tocantins", "Santa Rosa do Tocantins", "Santa Tereza do Tocantins", "Santa Terezinha do Tocantins",
        "Silvanópolis", "Sucupira", "São Bento do Tocantins", "São Félix do Tocantins", "São Miguel do Tocantins", "São Salvador do Tocantins",
        "São Sebastião do Tocantins", "São Valério", "Sítio Novo do Tocantins", "Tabocão", "Taguatinga", "Taipas do Tocantins",
        "Talismã", "Tocantinópolis", "Tocantínia", "Tupirama", "Tupiratins", "Wanderlândia",
        "Xambioá"
    ]
}

class LabScraperV2:
    """Classe para fazer scraping dos dados de laboratórios do Gralab"""
    
    def __init__(self):
        self.base_url = "https://clab-af-ecommerce.azurewebsites.net/eccommerce"
        # IMPORTANTE: Keys do concorrente - manter em .env e NUNCA commitar
        self.postos_code = os.getenv("AZURE_POSTOS_CODE")
        self.cidades_code = os.getenv("AZURE_CIDADES_CODE")
        self.empresa_id = 3
        
        # Validar que as variáveis de ambiente foram carregadas
        if not self.postos_code or not self.cidades_code:
            raise ValueError(
                "❌ ERRO: Azure Function Keys não encontradas!\n"
                "Configure as variáveis AZURE_POSTOS_CODE e AZURE_CIDADES_CODE no arquivo .env\n"
                "O arquivo .env deve estar em: Automations/cunha/.env"
            )
        self.checkpoint_file = os.path.join(BASE_DIR, "checkpoint_scraper.json")
        self.global_report_file = os.path.join(BASE_DIR, "relatorio_completo_laboratorios_gralab.xlsx")
        self.municipios = self.carregar_municipios_brasil()
        
    def carregar_municipios_brasil(self) -> List[Tuple[str, str]]:
        """
        Carrega todos os municípios do Brasil do dict MUNICIPIOS_POR_ESTADO
        
        Returns:
            Lista de tuplas (município, UF)
        """
        municipios_lista = []
        for uf, municipios in MUNICIPIOS_POR_ESTADO.items():
            for municipio in municipios:
                municipios_lista.append((municipio, uf))
            
        logging.info(f"[INFO] Carregados {len(municipios_lista)} municípios de {len(MUNICIPIOS_POR_ESTADO)} estados")
        return municipios_lista
    
    def buscar_postos_multiplos_formatos(self, municipio: str, uf: str) -> List[Dict]:
        """
        Busca postos testando múltiplos formatos de nome de cidade
        
        Args:
            municipio: Nome do município
            uf: Sigla do estado (UF)
            
        Returns:
            Lista única de postos encontrados (sem duplicatas)
        """
        formatos = [
            municipio.upper(),  # Ex: "CAMPINAS"
            f"{municipio.upper()} - {uf}",  # Ex: "CAMPINAS - SP"
            f"{municipio.upper()}-{uf}",  # Ex: "CAMPINAS-SP"
        ]
        
        # Mapeamento de UF para nome completo do estado
        estados_nomes = {
            "AC": "ACRE", "AL": "ALAGOAS", "AP": "AMAPÁ", "AM": "AMAZONAS",
            "BA": "BAHIA", "CE": "CEARÁ", "DF": "DISTRITO FEDERAL", "ES": "ESPÍRITO SANTO",
            "GO": "GOIÁS", "MA": "MARANHÃO", "MT": "MATO GROSSO", "MS": "MATO GROSSO DO SUL",
            "MG": "MINAS GERAIS", "PA": "PARÁ", "PB": "PARAÍBA", "PR": "PARANÁ",
            "PE": "PERNAMBUCO", "PI": "PIAUÍ", "RJ": "RIO DE JANEIRO", "RN": "RIO GRANDE DO NORTE",
            "RS": "RIO GRANDE DO SUL", "RO": "RONDÔNIA", "RR": "RORAIMA",
            "SC": "SANTA CATARINA", "SP": "SÃO PAULO", "SE": "SERGIPE", "TO": "TOCANTINS"
        }
        
        # Adicionar formato com nome completo do estado se disponível
        if uf in estados_nomes:
            formatos.append(f"{municipio.upper()} - {estados_nomes[uf]}")
        
        postos_unicos = {}
        
        for formato in formatos:
            try:
                postos = self.buscar_postos(formato)
                if postos:
                    for posto in postos:
                        entidade_id = posto.get('EntidadeId')
                        if entidade_id and entidade_id not in postos_unicos:
                            postos_unicos[entidade_id] = posto
                time.sleep(0.05)  # Delay reduzido para otimização com threading
            except Exception as e:
                continue
        
        return list(postos_unicos.values())
    
    def salvar_checkpoint(self, municipios_processados: int, total_labs: int):
        """Salva checkpoint do progresso"""
        checkpoint = {
            'municipios_processados': municipios_processados,
            'total_labs_coletados': total_labs,
            'timestamp': time.time()
        }
        with open(self.checkpoint_file, 'w') as f:
            json.dump(checkpoint, f)
    
    def carregar_checkpoint(self) -> Dict:
        """Carrega checkpoint se existir"""
        if os.path.exists(self.checkpoint_file):
            try:
                with open(self.checkpoint_file, 'r') as f:
                    return json.load(f)
            except:
                return {}
        return {}
        
    def buscar_cidades(self, busca: str = "") -> List[Dict]:
        """
        Busca cidades disponíveis
        
        Args:
            busca: Termo de busca para filtrar cidades
            
        Returns:
            Lista de dicionários com informações das cidades
        """
        url = f"{self.base_url}/BuscarCidades"
        params = {
            "code": self.cidades_code,
            "busca": busca
        }
        
        try:
            response = requests.get(url, params=params)
            response.raise_for_status()
            return response.json()
        except Exception as e:
            logging.error(f"Erro ao buscar cidades com termo '{busca}': {e}")
            return []
    
    def buscar_postos(self, cidade: str) -> List[Dict]:
        """
        Busca postos/laboratórios em uma cidade específica
        
        Args:
            cidade: Nome da cidade (ex: "CAMPINAS - SP")
            
        Returns:
            Lista de dicionários com informações dos postos
        """
        url = f"{self.base_url}/BuscarPostos"
        params = {
            "code": self.postos_code,
            "empresaId": self.empresa_id,
            "cidade": cidade
        }
        
        try:
            response = requests.get(url, params=params)
            response.raise_for_status()
            return response.json()
        except Exception as e:
            logging.error(f"Erro ao buscar postos na cidade '{cidade}': {e}")
            return []
    
    @staticmethod
    def converter_para_excel(valor):
        """Converte valores complexos para strings compatíveis com Excel"""
        # Verificar tipos complexos primeiro, antes de pd.isna()
        if isinstance(valor, (list, tuple)):
            return '; '.join(str(v) for v in valor)
        elif isinstance(valor, dict):
            return str(valor)
        # Verificar NaN (após tratar tipos complexos)
        try:
            if pd.isna(valor):
                return ''
        except (ValueError, TypeError):
            # Se pd.isna() falhar (arrays), tratar como string
            pass
        # Tipos numéricos
        if isinstance(valor, (int, float)):
            return valor
        # Tudo mais vira string
        return str(valor)
    
    @staticmethod
    def extrair_precos_servicos(servicos_str):
        """
        Extrai os preços de CNH, Concurso e CLT da coluna Servicos
        
        Args:
            servicos_str: String representando lista de serviços (ex: "['Concurso - R$ 180.00', ...]")
            
        Returns:
            Dict com {'preco_cnh': '139.00', 'preco_concurso': '180.00', 'preco_clt': '139.00'}
        """
        import ast
        
        precos = {'preco_cnh': '', 'preco_concurso': '', 'preco_clt': ''}
        
        try:
            if pd.isna(servicos_str) or servicos_str == '' or servicos_str is None:
                return precos
            
            # Converter string para lista
            if isinstance(servicos_str, str):
                servicos = ast.literal_eval(servicos_str)
            elif isinstance(servicos_str, list):
                servicos = servicos_str
            else:
                return precos
            
            # Extrair preços de cada serviço
            for servico in servicos:
                servico_upper = servico.upper()
                
                if 'CONCURSO' in servico_upper and 'R$' in servico:
                    try:
                        preco = servico.split('R$')[1].strip().split()[0]
                        precos['preco_concurso'] = preco
                    except:
                        pass
                        
                elif 'C.N.H' in servico_upper or 'CNH' in servico_upper:
                    if 'R$' in servico:
                        try:
                            preco = servico.split('R$')[1].strip().split()[0]
                            precos['preco_cnh'] = preco
                        except:
                            pass
                            
                elif 'CLT' in servico_upper or 'EMPREGADO' in servico_upper:
                    if 'R$' in servico:
                        try:
                            preco = servico.split('R$')[1].strip().split()[0]
                            precos['preco_clt'] = preco
                        except:
                            pass
            
        except Exception as e:
            logging.warning(f"Erro ao extrair preços de serviços: {e}")
        
        return precos
    
    def descobrir_todas_cidades(self) -> List[str]:
        """
        Descobre todas as cidades disponíveis na API usando busca recursiva
        A API retorna no máximo 50 resultados por busca, então precisamos
        aprofundar a busca quando atingimos esse limite
        
        Returns:
            Lista com nomes de todas as cidades disponíveis
        """
        logging.info("Descobrindo TODAS as cidades disponíveis (busca recursiva)...")
        
        cidades_set = set()  # Usar set para evitar duplicatas
        
        def buscar_recursivo(prefixo: str, nivel: int = 0):
            indent = "  " * (nivel + 1)
            
            resultados = self.buscar_cidades(prefixo)
            
            if not resultados:
                return
            
            # Extrai os nomes das cidades
            cidades_encontradas = []
            for cidade in resultados:
                if isinstance(cidade, str):
                    cidades_encontradas.append(cidade)
                elif isinstance(cidade, dict) and 'nome' in cidade:
                    cidades_encontradas.append(cidade['nome'])
                elif isinstance(cidade, dict) and 'cidade' in cidade:
                    cidades_encontradas.append(cidade['cidade'])
            
            # Adiciona ao set
            cidades_set.update(cidades_encontradas)
            
            # Se retornou 50 resultados (limite), precisamos aprofundar
            if len(resultados) >= 50:
                logging.info(f"{indent}[!] '{prefixo}' retornou {len(resultados)} (LIMITE!) - Aprofundando busca...")
                
                # Busca com mais uma letra
                letras = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
                for letra in letras:
                    novo_prefixo = prefixo + letra
                    buscar_recursivo(novo_prefixo, nivel + 1)
                    time.sleep(0.1)
            else:
                logging.info(f"{indent}[OK] '{prefixo}' retornou {len(resultados)} cidades")
            
            time.sleep(0.2)
        
        # Começa buscando com todas as letras do alfabeto
        letras = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        
        for letra in letras:
            logging.info(f"\n  Buscando cidades com '{letra}'...")
            buscar_recursivo(letra)
        
        cidades_lista = sorted(list(cidades_set))
        logging.info(f"\n  [SUCESSO!] Total de {len(cidades_lista)} cidades UNICAS encontradas")
        return cidades_lista
    
    def coletar_todos_postos(self, usar_lista_completa: bool = True, retomar: bool = True, max_workers: int = 16) -> pd.DataFrame:
        """
        Coleta informações de todos os postos em TODOS os municípios do Brasil usando threading
        
        Args:
            usar_lista_completa: Se True, usa lista completa de municípios do Brasil
            retomar: Se True, tenta retomar de checkpoint anterior
            max_workers: Número máximo de threads simultâneas (padrão: 16, otimizado para Ryzen 7 7500X)
                        Para operações I/O bound pode usar mais threads (ex: 24-32)
            
        Returns:
            DataFrame do pandas com todos os postos encontrados
        """
        # Carregar municípios do Brasil
        if usar_lista_completa:
            municipios = self.carregar_municipios_brasil()
            if not municipios:
                logging.error("[ERRO] Não foi possível carregar a lista de municípios!")
                return pd.DataFrame()
        else:
            # Fallback: usa lista de cidades principais
            logging.warning("\n[AVISO] Usando lista manual de cidades principais...")
            municipios = [
                ("São Paulo", "SP"),
                ("Campinas", "SP"),
                ("Rio de Janeiro", "RJ"),
                ("Belo Horizonte", "MG"),
                ("Brasília", "DF"),
                ("Curitiba", "PR"),
                ("Porto Alegre", "RS"),
                ("Salvador", "BA"),
                ("Recife", "PE"),
                ("Fortaleza", "CE")
            ]
        
        # Verificar checkpoint
        checkpoint = self.carregar_checkpoint() if retomar else {}
        inicio_idx = checkpoint.get('municipios_processados', 0)
        
        if inicio_idx > 0:
            logging.info(f"\n[INFO] Retomando coleta do município {inicio_idx + 1}/{len(municipios)}")
            logging.info(f"[INFO] Laboratórios já coletados: {checkpoint.get('total_labs_coletados', 0)}")
        
        logging.info(f"\nIniciando coleta em {len(municipios)} municípios do Brasil usando {max_workers} threads...")
        print(f"\n{'='*60}")
        print(f"Iniciando coleta em {len(municipios)} municípios")
        print(f"Usando {max_workers} threads simultâneas")
        print(f"{'='*60}\n")
        sys.stdout.flush()
        
        # Estruturas compartilhadas thread-safe
        postos_unicos = {}  # Dicionário para controlar duplicatas por EntidadeId
        todos_postos = []
        municipios_com_labs = 0
        municipios_sem_labs = 0
        municipios_processados = inicio_idx
        resultados_processados = 0  # Contador para atualização da barra de progresso
        
        # Locks para proteger estruturas compartilhadas
        lock_postos = Lock()
        lock_contadores = Lock()
        lock_checkpoint = Lock()
        lock_progresso = Lock()
        
        def processar_municipio(args):
            """Função worker para processar um município"""
            idx, municipio, uf = args
            
            try:
                # Buscar com múltiplos formatos
                postos = self.buscar_postos_multiplos_formatos(municipio, uf)
                
                novos_labs = 0
                
                if postos:
                    for posto in postos:
                        posto['cidade_busca'] = f"{municipio} - {uf}"
                        posto['municipio'] = municipio
                        posto['uf_busca'] = uf
                        entidade_id = posto.get('EntidadeId')
                        
                        # Usar lock para proteger operações no dicionário compartilhado
                        with lock_postos:
                            # Evita duplicatas usando EntidadeId como chave única
                            if entidade_id and entidade_id not in postos_unicos:
                                postos_unicos[entidade_id] = posto
                                todos_postos.append(posto)
                                novos_labs += 1
                
                # Atualizar contadores com lock
                with lock_contadores:
                    if novos_labs > 0:
                        municipios_com_labs += 1
                    else:
                        municipios_sem_labs += 1
                    municipios_processados += 1
                
                return {
                    'idx': idx,
                    'municipio': municipio,
                    'uf': uf,
                    'novos_labs': novos_labs,
                    'sucesso': True
                }
            except Exception as e:
                logging.error(f"Erro ao processar {municipio}-{uf}: {e}")
                with lock_contadores:
                    municipios_sem_labs += 1
                    municipios_processados += 1
                return {
                    'idx': idx,
                    'municipio': municipio,
                    'uf': uf,
                    'novos_labs': 0,
                    'sucesso': False
                }
        
        # Preparar lista de tarefas (apenas municípios não processados)
        tarefas = [(idx, municipio, uf) for idx, (municipio, uf) in enumerate(municipios[inicio_idx:], start=inicio_idx)]
        
        # Barra de progresso com tqdm (thread-safe)
        pbar = tqdm(total=len(municipios), initial=inicio_idx, desc="Coletando laboratórios", 
                   unit=" municípios", ncols=120, mininterval=0.1, maxinterval=1.0,
                   file=sys.stdout, dynamic_ncols=False)
        
        try:
            # Usar ThreadPoolExecutor para processar em paralelo
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                # Submeter todas as tarefas
                print(f"Submetendo {len(tarefas)} tarefas para processamento...")
                sys.stdout.flush()
                future_to_municipio = {executor.submit(processar_municipio, tarefa): tarefa for tarefa in tarefas}
                print(f"✓ Todas as tarefas submetidas. Processando em paralelo...\n")
                sys.stdout.flush()
                
                # Processar resultados conforme completam
                primeiro_resultado = True
                for future in as_completed(future_to_municipio):
                    try:
                        resultado = future.result()
                        
                        if primeiro_resultado:
                            print(f"✓ Primeiro resultado recebido: {resultado['municipio']}-{resultado['uf']}")
                            sys.stdout.flush()
                            primeiro_resultado = False
                        
                        # Atualizar contador de progresso
                        with lock_progresso:
                            resultados_processados += 1
                            total_processados = resultados_processados
                        
                        # Atualizar barra de progresso (thread-safe)
                        pbar.update(1)
                        pbar.set_postfix({
                            'Labs': len(todos_postos),
                            'Último': f"{resultado['municipio'][:15]}-{resultado['uf']}",
                            'Threads': max_workers
                        })
                        pbar.refresh()  # Força atualização visual
                        
                        # Log periódico para debug
                        if total_processados % 50 == 0:
                            logging.info(f"Progresso: {total_processados}/{len(tarefas)} municípios processados | {len(todos_postos)} laboratórios encontrados")
                        
                        # Salvar checkpoint a cada 100 municípios processados
                        with lock_checkpoint:
                            if total_processados % 100 == 0:
                                self.salvar_checkpoint(inicio_idx + total_processados, len(todos_postos))
                    except Exception as e:
                        logging.error(f"Erro ao processar resultado: {e}")
                        with lock_progresso:
                            resultados_processados += 1
                        pbar.update(1)
        finally:
            pbar.close()
        
        # Salvar checkpoint final
        self.salvar_checkpoint(len(municipios), len(todos_postos))
        
        logging.info(f"\nRESUMO DA COLETA:")
        logging.info(f"  - Municípios pesquisados: {len(municipios)}")
        logging.info(f"  - Municípios com laboratórios: {municipios_com_labs}")
        logging.info(f"  - Municípios sem laboratórios: {municipios_sem_labs}")
        logging.info(f"  - Total de laboratórios únicos: {len(todos_postos)}")
        logging.info(f"  - Threads utilizadas: {max_workers}")
        
        # Limpar checkpoint após conclusão
        if os.path.exists(self.checkpoint_file):
            os.remove(self.checkpoint_file)
        
        if todos_postos:
            df = pd.DataFrame(todos_postos)
            today = datetime.now().strftime('%Y-%m-%d')
            daily_file = os.path.join(BASE_DIR, f"dados_{today}.csv")
            df.to_csv(daily_file, index=False, encoding='utf-8-sig')
            logging.info(f"Dados diários salvos: {daily_file}")
            return df
        else:
            logging.error("\n[ERRO] Nenhum posto foi encontrado.")
            return pd.DataFrame()
    
    def detectar_churn(self, df_atual: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """Detecta credenciamentos e descredenciamentos comparando com dia anterior via CNPJ"""
        yesterday = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
        yesterday_file = os.path.join(BASE_DIR, f"dados_{yesterday}.csv")
        
        if not os.path.exists(yesterday_file):
            logging.warning("Não há dados do dia anterior para comparação de churn.")
            return pd.DataFrame(), pd.DataFrame()
        
        df_anterior = pd.read_csv(yesterday_file, encoding='utf-8-sig')
        
        cnpjs_anterior = set(df_anterior['Cnpj'].dropna().unique())
        cnpjs_atual = set(df_atual['Cnpj'].dropna().unique())
        
        novos_cnpjs = cnpjs_atual - cnpjs_anterior
        removidos_cnpjs = cnpjs_anterior - cnpjs_atual
        
        df_novos = df_atual[df_atual['Cnpj'].isin(novos_cnpjs)]
        df_removidos = df_anterior[df_anterior['Cnpj'].isin(removidos_cnpjs)]
        
        # Converter DataFrames para formato compatível com Excel
        df_novos_excel = df_novos.copy()
        df_removidos_excel = df_removidos.copy()
        
        for col in df_novos_excel.columns:
            df_novos_excel[col] = df_novos_excel[col].apply(self.converter_para_excel)
        for col in df_removidos_excel.columns:
            df_removidos_excel[col] = df_removidos_excel[col].apply(self.converter_para_excel)
        
        today = datetime.now().strftime('%Y-%m-%d')
        churn_file = os.path.join(BASE_DIR, f"churn_{today}.xlsx")
        with pd.ExcelWriter(churn_file) as writer:
            df_novos_excel.to_excel(writer, sheet_name='Credenciamentos', index=False)
            df_removidos_excel.to_excel(writer, sheet_name='Descredenciamentos', index=False)
        logging.info(f"Churn diário salvo: {churn_file}")
        
        return df_novos, df_removidos
    
    def atualizar_entrada_saida(self):
        """
        Gera/atualiza aba EntradaSaida com histórico completo de credenciamentos/descredenciamentos
        Analisa todos os arquivos dados_*.csv e mantém um registro único por CNPJ sem duplicatas
        """
        # Buscar todos os arquivos dados_*.csv
        csv_files = []
        for file in os.listdir(BASE_DIR):
            if file.startswith('dados_') and file.endswith('.csv'):
                csv_files.append(file)
        
        if not csv_files:
            logging.warning("Nenhum arquivo dados_*.csv encontrado para gerar EntradaSaida.")
            return
        
        # Extrair datas e ordenar
        datas_arquivos = []
        for file in csv_files:
            try:
                data_str = file.replace('dados_', '').replace('.csv', '')
                data_obj = datetime.strptime(data_str, '%Y-%m-%d')
                datas_arquivos.append((data_obj, file))
            except ValueError:
                continue
        
        datas_arquivos.sort(key=lambda x: x[0])
        
        if not datas_arquivos:
            logging.warning("Nenhuma data válida encontrada nos arquivos CSV.")
            return
        
        # Dicionário para rastrear cada CNPJ (chave: CNPJ, valor: dict com informações)
        cnpj_tracker = {}
        
        # Processar cada arquivo na ordem cronológica
        for idx_arquivo, (data_obj, file) in enumerate(datas_arquivos):
            data_str = data_obj.strftime('%Y-%m-%d')
            file_path = os.path.join(BASE_DIR, file)
            eh_primeiro_arquivo = (idx_arquivo == 0)
            
            try:
                df = pd.read_csv(file_path, encoding='utf-8-sig')
                cnpjs_presentes = set()
                
                # Processar cada laboratório do arquivo
                for _, row in df.iterrows():
                    cnpj = str(row.get('Cnpj', '')).strip()
                    if not cnpj or cnpj == '' or pd.isna(cnpj):
                        continue
                    
                    cnpjs_presentes.add(cnpj)
                    
                    # Se CNPJ não existe no tracker, criar entrada
                    if cnpj not in cnpj_tracker:
                        # Extrair preços dos serviços
                        precos = self.extrair_precos_servicos(row.get('Servicos', ''))
                        
                        # Se é o primeiro arquivo, marcar como base inicial, não como credenciamento
                        # Pois não sabemos quando ele realmente entrou
                        if eh_primeiro_arquivo:
                            tipo_movimentacao = ''  # Vazio = já existia na base inicial
                        else:
                            tipo_movimentacao = 'Credenciamento'  # Apareceu depois = novo credenciamento
                        
                        cnpj_tracker[cnpj] = {
                            'CNPJ': cnpj,
                            'Nome': str(row.get('Nome', '')),
                            'Cidade': str(row.get('Cidade', '')),
                            'UF': str(row.get('UF', '')),
                            'Endereco': str(row.get('Endereco', '')),
                            'Telefone': str(row.get('Telefone', '')),
                            'Preco_CNH': precos['preco_cnh'],
                            'Preco_Concurso': precos['preco_concurso'],
                            'Preco_CLT': precos['preco_clt'],
                            'Data_Entrada': data_str,
                            'Data_Saida': None,
                            'Status': 'Ativo',
                            'Tipo_Movimentacao': tipo_movimentacao,
                            'Ultima_Verificacao': data_str,
                            'historico_datas': [data_str]
                        }
                    else:
                        # CNPJ já existe - atualizar informações
                        info = cnpj_tracker[cnpj]
                        
                        # Extrair preços dos serviços
                        precos = self.extrair_precos_servicos(row.get('Servicos', ''))
                        
                        # Se estava inativo, reativar
                        if info['Status'] == 'Inativo':
                            info['Status'] = 'Ativo'
                            info['Data_Saida'] = None
                            info['Data_Entrada'] = data_str  # Nova data de entrada
                            info['Tipo_Movimentacao'] = 'Recredenciamento'
                        
                        # Atualizar última verificação
                        info['Ultima_Verificacao'] = data_str
                        
                        # Adicionar data ao histórico se não existe
                        if data_str not in info['historico_datas']:
                            info['historico_datas'].append(data_str)
                        
                        # Atualizar informações (podem ter mudado)
                        info['Nome'] = str(row.get('Nome', ''))
                        info['Cidade'] = str(row.get('Cidade', ''))
                        info['UF'] = str(row.get('UF', ''))
                        info['Endereco'] = str(row.get('Endereco', ''))
                        info['Telefone'] = str(row.get('Telefone', ''))
                        info['Preco_CNH'] = precos['preco_cnh']
                        info['Preco_Concurso'] = precos['preco_concurso']
                        info['Preco_CLT'] = precos['preco_clt']
                
                # Verificar CNPJs que não estão mais presentes (descredenciamentos)
                for cnpj, info in cnpj_tracker.items():
                    if info['Status'] == 'Ativo' and cnpj not in cnpjs_presentes:
                        # Só marcar como inativo se a última verificação foi na data anterior
                        # (para não marcar como inativo em datas futuras que ainda não processamos)
                        if info['Ultima_Verificacao'] < data_str:
                            info['Status'] = 'Inativo'
                            info['Data_Saida'] = data_str
                            info['Tipo_Movimentacao'] = 'Descredenciamento'
                
            except Exception as e:
                logging.error(f"Erro ao processar arquivo {file} para EntradaSaida: {e}")
                continue
        
        # Criar/atualizar planilha
        if os.path.exists(self.global_report_file):
            wb = openpyxl.load_workbook(self.global_report_file)
        else:
            wb = Workbook()
            wb.remove(wb.active)
        
        # Remover aba antiga se existir
        if 'Historico Churn' in wb.sheetnames:
            wb.remove(wb['Historico Churn'])
        if 'EntradaSaida' in wb.sheetnames:
            wb.remove(wb['EntradaSaida'])
        
        # Criar nova aba EntradaSaida (inserir como primeira aba)
        ws = wb.create_sheet('EntradaSaida', 0)
        
        # Cabeçalho
        headers = ['CNPJ', 'Nome', 'Cidade', 'UF', 'Endereco', 'Telefone', 
                   'Preço CNH', 'Preço Concurso', 'Preço CLT',
                   'Data Entrada', 'Data Saída', 'Status', 'Tipo Movimentação',
                   'Última Verificação', 'Histórico Datas']
        ws.append(headers)
        
        # Formatação do cabeçalho
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # FILTRAR: Apenas laboratórios com movimentação real (não incluir base inicial)
        # Base inicial = labs sem tipo de movimentação
        registros_com_movimentacao = [
            info for info in cnpj_tracker.values()
            if info.get('Tipo_Movimentacao', '') != ''  # Tem tipo de movimentação
        ]
        
        # Ordenar por data de entrada (mais recente primeiro), depois por status
        registros_ordenados = sorted(
            registros_com_movimentacao,
            key=lambda x: (x['Data_Entrada'], x['Status'] == 'Inativo'),
            reverse=True
        )
        
        # Adicionar dados (apenas labs com movimentação)
        for info in registros_ordenados:
            historico = ', '.join(sorted(info['historico_datas']))
            
            row = [
                self.converter_para_excel(info['CNPJ']),
                self.converter_para_excel(info['Nome']),
                self.converter_para_excel(info['Cidade']),
                self.converter_para_excel(info['UF']),
                self.converter_para_excel(info['Endereco']),
                self.converter_para_excel(info['Telefone']),
                info.get('Preco_CNH', ''),
                info.get('Preco_Concurso', ''),
                info.get('Preco_CLT', ''),
                info['Data_Entrada'],
                info['Data_Saida'] if info['Data_Saida'] else '',
                info['Status'],
                info.get('Tipo_Movimentacao', 'Ativo'),
                info['Ultima_Verificacao'],
                historico
            ]
            ws.append(row)
            
            # Aplicar bordas e cores
            for cell in ws[ws.max_row]:
                cell.border = border
                # Colorir linha baseado no status
                if info['Status'] == 'Ativo':
                    cell.fill = PatternFill(start_color="E7F4E4", end_color="E7F4E4", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 18  # CNPJ
        ws.column_dimensions['B'].width = 35  # Nome
        ws.column_dimensions['C'].width = 20  # Cidade
        ws.column_dimensions['D'].width = 5   # UF
        ws.column_dimensions['E'].width = 40  # Endereco
        ws.column_dimensions['F'].width = 18  # Telefone
        ws.column_dimensions['G'].width = 12  # Preço CNH
        ws.column_dimensions['H'].width = 14  # Preço Concurso
        ws.column_dimensions['I'].width = 12  # Preço CLT
        ws.column_dimensions['J'].width = 13  # Data Entrada
        ws.column_dimensions['K'].width = 13  # Data Saída
        ws.column_dimensions['L'].width = 10  # Status
        ws.column_dimensions['M'].width = 18  # Tipo Movimentação
        ws.column_dimensions['N'].width = 18  # Última Verificação
        ws.column_dimensions['O'].width = 30  # Histórico Datas
        
        # Salvar
        wb.save(self.global_report_file)
        
        # Estatísticas (apenas labs com movimentação)
        total_movimentacoes = len(registros_com_movimentacao)
        credenciamentos = sum(1 for info in registros_com_movimentacao if info.get('Tipo_Movimentacao') == 'Credenciamento')
        descredenciamentos = sum(1 for info in registros_com_movimentacao if info.get('Tipo_Movimentacao') == 'Descredenciamento')
        recredenciamentos = sum(1 for info in registros_com_movimentacao if info.get('Tipo_Movimentacao') == 'Recredenciamento')
        
        logging.info(f"Aba 'EntradaSaida' criada/atualizada: {total_movimentacoes} movimentações ({credenciamentos} credenciamentos, {descredenciamentos} descredenciamentos, {recredenciamentos} recredenciamentos)")
        
        return total_movimentacoes, credenciamentos, descredenciamentos
    
    def gerar_resumo_credenciamentos(self):
        """
        Gera resumo diário e acumulado de credenciamentos/descredenciamentos
        analisando todos os arquivos dados_*.csv da pasta
        """
        # Buscar todos os arquivos dados_*.csv
        csv_files = []
        for file in os.listdir(BASE_DIR):
            if file.startswith('dados_') and file.endswith('.csv'):
                csv_files.append(file)
        
        if not csv_files:
            logging.warning("Nenhum arquivo dados_*.csv encontrado para gerar resumo.")
            return
        
        # Extrair datas e ordenar
        datas = []
        for file in csv_files:
            try:
                # Formato: dados_YYYY-MM-DD.csv
                data_str = file.replace('dados_', '').replace('.csv', '')
                data_obj = datetime.strptime(data_str, '%Y-%m-%d')
                datas.append((data_obj, file))
            except ValueError:
                continue
        
        datas.sort(key=lambda x: x[0])  # Ordenar por data
        
        if not datas:
            logging.warning("Nenhuma data válida encontrada nos arquivos CSV.")
            return
        
        # Preparar lista de resumos
        resumos = []
        total_acumulado_credenciamentos = 0
        total_acumulado_descredenciamentos = 0
        
        # Processar cada dia comparando com o anterior
        for idx, (data_obj, file) in enumerate(datas):
            data_str = data_obj.strftime('%Y-%m-%d')
            file_path = os.path.join(BASE_DIR, file)
            
            try:
                df_atual = pd.read_csv(file_path, encoding='utf-8-sig')
                cnpjs_atual = set(df_atual['Cnpj'].dropna().unique())
                
                credenciamentos = 0
                descredenciamentos = 0
                
                # Se não é o primeiro dia, comparar com o anterior
                if idx > 0:
                    data_anterior_obj, file_anterior = datas[idx - 1]
                    file_anterior_path = os.path.join(BASE_DIR, file_anterior)
                    
                    try:
                        df_anterior = pd.read_csv(file_anterior_path, encoding='utf-8-sig')
                        cnpjs_anterior = set(df_anterior['Cnpj'].dropna().unique())
                        
                        novos_cnpjs = cnpjs_atual - cnpjs_anterior
                        removidos_cnpjs = cnpjs_anterior - cnpjs_atual
                        
                        credenciamentos = len(novos_cnpjs)
                        descredenciamentos = len(removidos_cnpjs)
                    except Exception as e:
                        logging.warning(f"Erro ao ler arquivo anterior {file_anterior}: {e}")
                
                # Atualizar totais acumulados
                total_acumulado_credenciamentos += credenciamentos
                total_acumulado_descredenciamentos += descredenciamentos
                
                saldo = credenciamentos - descredenciamentos
                
                resumos.append({
                    'Data': data_str,
                    'Credenciamentos': credenciamentos,
                    'Descredenciamentos': descredenciamentos,
                    'Saldo': saldo,
                    'Total Acumulado Credenciamentos': total_acumulado_credenciamentos,
                    'Total Acumulado Descredenciamentos': total_acumulado_descredenciamentos
                })
                
            except Exception as e:
                logging.error(f"Erro ao processar arquivo {file}: {e}")
                continue
        
        if not resumos:
            logging.warning("Nenhum resumo foi gerado.")
            return
        
        # Carregar ou criar arquivo Excel
        if os.path.exists(self.global_report_file):
            wb = openpyxl.load_workbook(self.global_report_file)
        else:
            wb = Workbook()
            wb.remove(wb.active)  # Remove aba padrão
        
        # Criar ou atualizar aba de resumo
        ws_name = 'Resumo Credenciamentos'
        if ws_name in wb.sheetnames:
            wb.remove(wb[ws_name])
        
        # Inserir aba após "Resumo Geografico" se existir, senão criar no final
        if 'Resumo Geografico' in wb.sheetnames:
            idx = wb.sheetnames.index('Resumo Geografico') + 1
            ws = wb.create_sheet(ws_name, idx)
        else:
            ws = wb.create_sheet(ws_name)
        
        # Cabeçalho
        headers = ['Data', 'Credenciamentos', 'Descredenciamentos', 'Saldo', 
                   'Total Acumulado Credenciamentos', 'Total Acumulado Descredenciamentos']
        ws.append(headers)
        
        # Formatação do cabeçalho
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Adicionar dados
        for resumo in resumos:
            row = [
                resumo['Data'],
                resumo['Credenciamentos'],
                resumo['Descredenciamentos'],
                resumo['Saldo'],
                resumo['Total Acumulado Credenciamentos'],
                resumo['Total Acumulado Descredenciamentos']
            ]
            ws.append(row)
            
            # Aplicar bordas nas células de dados
            for cell in ws[ws.max_row]:
                cell.border = border
                if cell.column == 1:  # Coluna Data
                    cell.alignment = Alignment(horizontal='center')
                else:  # Colunas numéricas
                    cell.alignment = Alignment(horizontal='right')
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 15  # Data
        ws.column_dimensions['B'].width = 18  # Credenciamentos
        ws.column_dimensions['C'].width = 20  # Descredenciamentos
        ws.column_dimensions['D'].width = 12  # Saldo
        ws.column_dimensions['E'].width = 35  # Total Acumulado Credenciamentos
        ws.column_dimensions['F'].width = 35  # Total Acumulado Descredenciamentos
        
        # Salvar
        wb.save(self.global_report_file)
        logging.info(f"Aba '{ws_name}' criada/atualizada no relatório global.")
    
    def gerar_relatorio_global(self, df: pd.DataFrame):
        """
        Gera o relatório completo global consolidado com preços e estilização
        Atualiza abas existentes sem sobrescrever o arquivo completo
        """
        # Carregar arquivo existente ou criar novo
        if os.path.exists(self.global_report_file):
            wb = openpyxl.load_workbook(self.global_report_file)
        else:
            wb = Workbook()
            wb.remove(wb.active)  # Remove aba padrão
        
        # Extrair preços da coluna Servicos e adicionar ao DataFrame
        df_com_precos = df.copy()
        precos_list = df_com_precos['Servicos'].apply(self.extrair_precos_servicos)
        df_com_precos['Preço CNH'] = precos_list.apply(lambda x: x['preco_cnh'])
        df_com_precos['Preço Concurso'] = precos_list.apply(lambda x: x['preco_concurso'])
        df_com_precos['Preço CLT'] = precos_list.apply(lambda x: x['preco_clt'])
        
        # Reordenar colunas para colocar preços após Telefone
        colunas_base = ['Nome', 'Cnpj', 'Telefone', 'Preço CNH', 'Preço Concurso', 'Preço CLT',
                        'Endereco', 'Bairro', 'Complemento', 'Cidade', 'UF', 'Latitude', 'Longitude',
                        'HorarioFuncionamento', 'UrlProduto', 'EntidadeId', 'Servicos',
                        'cidade_busca', 'municipio', 'uf_busca']
        
        # Usar apenas colunas que existem
        colunas_ordenadas = [col for col in colunas_base if col in df_com_precos.columns]
        df_com_precos = df_com_precos[colunas_ordenadas]
        
        # Converter DataFrame para formato compatível com Excel
        df_excel = df_com_precos.copy()
        for col in df_excel.columns:
            df_excel[col] = df_excel[col].apply(self.converter_para_excel)
        
        # ==== ABA DADOS COMPLETOS ====
        ws_name = "Dados Completos"
        if ws_name in wb.sheetnames:
            wb.remove(wb[ws_name])
        ws = wb.create_sheet(ws_name)
        
        # Adicionar dados
        for r in dataframe_to_rows(df_excel, index=False, header=True):
            ws.append(r)
        
        # Estilização
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Formatar cabeçalho
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Aplicar bordas e zebrar linhas
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            for cell in row:
                cell.border = border
                # Zebrar linhas (cinza claro em linhas pares)
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 35  # Nome
        ws.column_dimensions['B'].width = 18  # CNPJ
        ws.column_dimensions['C'].width = 18  # Telefone
        ws.column_dimensions['D'].width = 12  # Preço CNH
        ws.column_dimensions['E'].width = 14  # Preço Concurso
        ws.column_dimensions['F'].width = 12  # Preço CLT
        ws.column_dimensions['G'].width = 40  # Endereço
        
        # Congelar primeira linha (cabeçalho)
        ws.freeze_panes = 'A2'
        
        # ==== ABA RESUMO GEOGRÁFICO ====
        ws_name = "Resumo Geografico"
        if ws_name in wb.sheetnames:
            wb.remove(wb[ws_name])
        ws = wb.create_sheet(ws_name)
        
        # Preparar dados
        dist_uf = df['UF'].value_counts().sort_index()
        ws.append(['UF', 'Quantidade de Laboratórios', 'Percentual (%)'])
        
        total = len(df)
        for uf, count in dist_uf.items():
            percentual = (count / total * 100) if total > 0 else 0
            ws.append([uf, count, f'{percentual:.2f}%'])
        
        # Adicionar linha de total
        ws.append(['TOTAL', total, '100%'])
        
        # Estilização
        # Formatar cabeçalho
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Formatar dados
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row-1), start=2):
            for cell in row:
                cell.border = border
                if cell.column == 1:  # Coluna UF
                    cell.alignment = Alignment(horizontal='center')
                    cell.font = Font(bold=True)
                elif cell.column == 2:  # Coluna Quantidade
                    cell.alignment = Alignment(horizontal='right')
                else:  # Coluna Percentual
                    cell.alignment = Alignment(horizontal='center')
                
                # Zebrar linhas
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        # Formatar linha de total
        total_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        total_font = Font(bold=True, color="FFFFFF", size=12)
        for cell in ws[ws.max_row]:
            cell.fill = total_fill
            cell.font = total_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if cell.column != 2 else 'right', vertical='center')
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 8   # UF
        ws.column_dimensions['B'].width = 30  # Quantidade
        ws.column_dimensions['C'].width = 18  # Percentual
        
        # Congelar primeira linha
        ws.freeze_panes = 'A2'
        
        # Salvar
        wb.save(self.global_report_file)
        logging.info(f"Relatório global atualizado: {self.global_report_file}")
    
    def run_pipeline(self, max_workers: int = 16):
        """
        Executa o pipeline completo
        
        Args:
            max_workers: Número máximo de threads simultâneas para scraping (padrão: 16, otimizado para Ryzen 7 7500X)
        """
        today = datetime.now().strftime('%Y-%m-%d')
        daily_file = os.path.join(BASE_DIR, f"dados_{today}.csv")
        
        # Verificar se já executou hoje
        if os.path.exists(daily_file):
            print(f"\n{'='*70}")
            print(f"✓ Arquivo do dia atual já existe!")
            print(f"{'='*70}")
            print(f"📁 Arquivo: {os.path.basename(daily_file)}")
            print(f"⏭️  Pulando coleta de dados...")
            print(f"📊 Carregando dados existentes e atualizando relatórios...\n")
            
            logging.info(f"Arquivo do dia atual já existe: {daily_file}")
            logging.info("Pulando coleta de dados. Carregando dados existentes e atualizando relatórios...")
            
            try:
                df = pd.read_csv(daily_file, encoding='utf-8-sig')
                print(f"✓ Dados carregados: {len(df)} laboratórios encontrados\n")
                logging.info(f"Dados carregados: {len(df)} laboratórios encontrados")
            except Exception as e:
                print(f"✗ Erro ao carregar arquivo: {e}")
                print(f"⚠️  Executando coleta completa...\n")
                logging.error(f"Erro ao carregar arquivo {daily_file}: {e}")
                logging.info("Executando coleta completa devido ao erro...")
                df = self.coletar_todos_postos(max_workers=max_workers)
        else:
            print(f"\n{'='*70}")
            print(f"🔍 Arquivo do dia não encontrado - Iniciando coleta completa")
            print(f"{'='*70}")
            print(f"📅 Data: {today}")
            print(f"🧵 Threads: {max_workers}\n")
            
            logging.info(f"Iniciando pipeline com {max_workers} threads...")
            df = self.coletar_todos_postos(max_workers=max_workers)
        
        # Processar churn (sempre compara com dia anterior)
        print("🔄 Processando churn (comparando com dia anterior)...")
        df_novos, df_removidos = self.detectar_churn(df)
        print(f"   ➕ Credenciamentos: {len(df_novos)}")
        print(f"   ➖ Descredenciamentos: {len(df_removidos)}\n")
        
        # Atualizar aba EntradaSaida (análise completa de todos os arquivos)
        print("📝 Atualizando aba EntradaSaida (apenas movimentações)...")
        try:
            total_mov, credenc, descredenc = self.atualizar_entrada_saida()
            print(f"   ✓ {total_mov} movimentações (+{credenc} credenciamentos, -{descredenc} descredenciamentos)\n")
        except Exception as e:
            print(f"   ✗ Erro ao atualizar EntradaSaida: {e}\n")
            logging.error(f"Erro ao atualizar EntradaSaida: {e}")
        
        # Gerar/atualizar relatório global
        print("📊 Gerando/atualizando relatório global...")
        self.gerar_relatorio_global(df)
        print("   ✓ Relatório atualizado\n")
        
        # Gerar resumo de credenciamentos (analisa todos os arquivos históricos)
        print("📈 Gerando resumo de credenciamentos...")
        self.gerar_resumo_credenciamentos()
        print("   ✓ Resumo gerado\n")
        
        print(f"{'='*70}")
        print("✅ Pipeline concluído com sucesso!")
        print(f"{'='*70}\n")
        
        logging.info("Pipeline concluído.")

def main():
    scraper = LabScraperV2()
    
    # Verificar argumentos de linha de comando
    max_workers = 16  # Padrão: 16 threads (otimizado para Ryzen 7 7500X - 8 cores / 16 threads)
                       # Para operações I/O bound (requisições HTTP), pode usar mais threads (ex: 24-32)
    daemon_mode = False
    gerar_relatorio_apenas = False
    
    for arg in sys.argv[1:]:
        if arg == '--daemon':
            daemon_mode = True
        elif arg == '--gerar-relatorio':
            gerar_relatorio_apenas = True
        elif arg.startswith('--threads='):
            try:
                max_workers = int(arg.split('=')[1])
                logging.info(f"Usando {max_workers} threads conforme especificado.")
            except ValueError:
                logging.warning(f"Valor inválido para threads: {arg}. Usando padrão: 16")
                max_workers = 16
    
    if gerar_relatorio_apenas:
        # Gerar relatório a partir do CSV do dia atual
        today = datetime.now().strftime('%Y-%m-%d')
        daily_file = os.path.join(BASE_DIR, f"dados_{today}.csv")
        
        if os.path.exists(daily_file):
            print(f"Carregando dados de {daily_file}...")
            df = pd.read_csv(daily_file, encoding='utf-8-sig')
            print(f"Dados carregados: {len(df)} laboratórios")
            
            # Gerar relatório global
            scraper.gerar_relatorio_global(df)
            print("Relatório global gerado com sucesso!")
            
            # Tentar gerar churn se houver dados do dia anterior
            try:
                df_novos, df_removidos = scraper.detectar_churn(df)
                print(f"Churn detectado: {len(df_novos)} credenciamentos, {len(df_removidos)} descredenciamentos")
            except Exception as e:
                logging.warning(f"Não foi possível processar churn: {e}")
            
            # Atualizar aba EntradaSaida
            try:
                total_mov, credenc, descredenc = scraper.atualizar_entrada_saida()
                print(f"EntradaSaida atualizado: {total_mov} movimentações (+{credenc} credenciamentos, -{descredenc} descredenciamentos)")
            except Exception as e:
                logging.warning(f"Não foi possível atualizar EntradaSaida: {e}")
            
            # Gerar resumo de credenciamentos
            scraper.gerar_resumo_credenciamentos()
            print("Resumo de credenciamentos gerado com sucesso!")
        else:
            print(f"Arquivo não encontrado: {daily_file}")
            print("Execute o script normalmente para coletar os dados primeiro.")
    elif daemon_mode:
        schedule.every().day.at("02:00").do(scraper.run_pipeline, max_workers=max_workers)
        logging.info(f"Daemon iniciado. Agendado para 02:00 diariamente com {max_workers} threads.")
        while True:
            schedule.run_pending()
            time.sleep(60)
    else:
        scraper.run_pipeline(max_workers=max_workers)

if __name__ == "__main__":
    main()