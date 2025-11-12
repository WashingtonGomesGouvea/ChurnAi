"""
Sistema Syntox Churn
Dashboard moderno e profissional para an√°lise de reten√ß√£o de laborat√≥rios
"""
import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, timedelta
import os
from typing import Optional, List, Dict, Any, Tuple
from io import BytesIO
from dataclasses import dataclass
import warnings
warnings.filterwarnings('ignore')
# Importar configura√ß√µes
from config_churn import *
from pandas.tseries.offsets import BDay
# Importar sistema de autentica√ß√£o Microsoft
from auth_microsoft import MicrosoftAuth, AuthManager, create_login_page, create_user_header
# ============================================
# FUN√á√ïES DE INTEGRA√á√ÉO SHAREPOINT/ONEDRIVE
# ============================================
def _get_graph_config() -> Optional[Dict[str, Any]]:
    """Extrai configura√ß√µes do Graph API dos secrets do Streamlit."""
    try:
        graph = st.secrets.get("graph", {})
        files = st.secrets.get("files", {})
        onedrive = st.secrets.get("onedrive", {})
        if not graph:
            return None
        return {
            "tenant_id": graph.get("tenant_id", ""),
            "client_id": graph.get("client_id", ""),
            "client_secret": graph.get("client_secret", ""),
            "hostname": graph.get("hostname", ""),
            "site_path": graph.get("site_path", ""),
            "library_name": graph.get("library_name", "Documents"),
            "user_upn": onedrive.get("user_upn", ""),
            "arquivo": files.get("arquivo", ""),
        }
    except Exception:
        return None
def _is_valid_csv(path: str) -> bool:
    """Verifica se arquivo CSV √© v√°lido."""
    try:
        if not os.path.exists(path):
            return False
        df = pd.read_csv(path, nrows=5)
        return len(df.columns) > 0
    except:
        return False
def _is_valid_parquet(path: str) -> bool:
    """Verifica se arquivo Parquet √© v√°lido."""
    try:
        if not os.path.exists(path):
            return False
        df = pd.read_parquet(path)
        return len(df.columns) > 0
    except:
        return False
def should_download_sharepoint(arquivo_remoto: str = None, force: bool = False) -> bool:
    """Verifica se deve baixar arquivo do SharePoint."""
    if force:
        return True
    # Determinar qual arquivo verificar (baseado no arquivo remoto solicitado)
    if arquivo_remoto:
        base_name = os.path.basename(arquivo_remoto)
        if base_name:
            arquivo_local = os.path.join(OUTPUT_DIR, base_name)
        else:
            arquivo_local = os.path.join(OUTPUT_DIR, "churn_analysis_latest.csv")
    else:
        arquivo_local = os.path.join(OUTPUT_DIR, "churn_analysis_latest.csv")
    # Verificar se existe arquivo local recente (< 5 minutos)
    if os.path.exists(arquivo_local):
        import time
        idade_arquivo = time.time() - os.path.getmtime(arquivo_local)
        if idade_arquivo < CACHE_TTL: # CACHE_TTL definido em config_churn.py
            return False
    return True
def baixar_sharepoint(arquivo_remoto: str = None, force: bool = False) -> Optional[str]:
    """
    Baixa arquivo do OneDrive/SharePoint via Microsoft Graph.
 
    Args:
        arquivo_remoto: Caminho do arquivo no OneDrive (usa config padr√£o se None)
        force: For√ßa download mesmo se cache v√°lido
 
    Returns:
        Caminho local do arquivo baixado ou None se falhar
    """
    cfg = _get_graph_config()
 
    # Sem configura√ß√£o Graph, retornar arquivo local se existir
    if not cfg or not (cfg.get("tenant_id") and cfg.get("client_id") and cfg.get("client_secret")):
        arquivo_local = os.path.join(OUTPUT_DIR, "churn_analysis_latest.csv")
        if os.path.exists(arquivo_local):
            return arquivo_local
        return None
 
    # Verificar se precisa baixar
    if not should_download_sharepoint(arquivo_remoto=arquivo_remoto, force=force):
        # Retornar o arquivo local correspondente ao solicitado
        if arquivo_remoto:
            base_name = os.path.basename(arquivo_remoto)
            if base_name:
                arquivo_local = os.path.join(OUTPUT_DIR, base_name)
            else:
                arquivo_local = os.path.join(OUTPUT_DIR, "churn_analysis_latest.csv")
        else:
            arquivo_local = os.path.join(OUTPUT_DIR, "churn_analysis_latest.csv")
        if os.path.exists(arquivo_local):
            return arquivo_local
 
    try:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
     
        # Usar ChurnSPConnector
        from churn_sp_connector import ChurnSPConnector
     
        connector = ChurnSPConnector(config=st.secrets)
     
        # Determinar arquivo remoto
        if arquivo_remoto is None:
            arquivo_remoto = cfg.get("arquivo", "Data Analysis/Churn PCLs/churn_analysis_latest.csv")
     
        # Baixar arquivo
        content = connector.download(arquivo_remoto)
     
        # Salvar localmente
        base_name = os.path.basename(arquivo_remoto)
        if not base_name:
            base_name = "churn_analysis_latest.csv"
     
        local_path = os.path.join(OUTPUT_DIR, base_name)
     
        with open(local_path, "wb") as f:
            f.write(content)
     
        # Validar arquivo baixado
        if _is_valid_csv(local_path) or _is_valid_parquet(local_path):
            return local_path
     
        return None
     
    except Exception as e:
        st.warning(f"‚ö†Ô∏è N√£o foi poss√≠vel baixar do SharePoint: {e}")
        # Tentar usar arquivo local se existir
        arquivo_local = os.path.join(OUTPUT_DIR, "churn_analysis_latest.csv")
        if os.path.exists(arquivo_local):
            return arquivo_local
        return None

def baixar_excel_gralab(force: bool = False) -> Optional[str]:
    """
    Baixa arquivo Excel do Gralab do SharePoint.
    
    Args:
        force: For√ßa download mesmo se cache v√°lido
    
    Returns:
        Caminho local do arquivo baixado ou None se falhar
    """
    arquivo_remoto = "/personal/washington_gouvea_synvia_com_/Documents/Data Analysis/Churn PCLs/Automations/cunha/relatorio_completo_laboratorios_gralab.xlsx"
    base_name = "relatorio_completo_laboratorios_gralab.xlsx"
    arquivo_local = os.path.join(OUTPUT_DIR, base_name)
    
    # Verificar cache (4 horas = 14400 segundos)
    if not force and os.path.exists(arquivo_local):
        import time
        idade_arquivo = time.time() - os.path.getmtime(arquivo_local)
        if idade_arquivo < 14400:  # 4 horas
            return arquivo_local
    
    cfg = _get_graph_config()
    
    # Sem configura√ß√£o Graph, retornar arquivo local se existir
    if not cfg or not (cfg.get("tenant_id") and cfg.get("client_id") and cfg.get("client_secret")):
        if os.path.exists(arquivo_local):
            return arquivo_local
        return None
    
    try:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        
        # Usar ChurnSPConnector
        from churn_sp_connector import ChurnSPConnector
        
        connector = ChurnSPConnector(config=st.secrets)
        
        # Baixar arquivo
        content = connector.download(arquivo_remoto)
        
        # Salvar localmente
        with open(arquivo_local, "wb") as f:
            f.write(content)
        
        # Validar se √© Excel v√°lido
        try:
            import openpyxl
            openpyxl.load_workbook(arquivo_local, read_only=True)
            return arquivo_local
        except Exception:
            return None
        
    except Exception as e:
        # Tentar usar arquivo local se existir
        if os.path.exists(arquivo_local):
            return arquivo_local
        return None

# Configura√ß√£o da p√°gina
st.set_page_config(
    page_title="üìä Syntox Churn",
    page_icon="üìä",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        'About': "Syntox Churn - Sistema profissional para monitoramento de reten√ß√£o de PCLs"
    }
)
# CSS moderno e profissional - Atualizado com melhorias de layout
CSS_STYLES = """
<style>
    /* Tema profissional atualizado - Synvia */
    :root {
        --primary-color: #6BBF47;
        --secondary-color: #52B54B;
        --success-color: #6BBF47;
        --warning-color: #F59E0B;
        --danger-color: #DC2626;
        --info-color: #3B82F6;
        --light-bg: #F5F7FA;
        --dark-bg: #262730;
        --border-radius: 12px; /* Aumentado para visual mais moderno */
        --shadow: 0 4px 8px rgba(0,0,0,0.1); /* Sombra mais suave */
        --transition: all 0.3s ease;
    }
    /* Reset e base */
    * { box-sizing: border-box; }
    /* Header profissional */
    .main-header {
        background: linear-gradient(135deg, var(--primary-color), var(--secondary-color));
        color: white;
        padding: 2.5rem 1.5rem; /* Aumentado padding */
        border-radius: var(--border-radius);
        margin-bottom: 2.5rem;
        text-align: center;
        box-shadow: var(--shadow);
    }
    .main-header h1 {
        margin: 0;
        font-size: 2.8rem; /* Aumentado tamanho */
        font-weight: 400;
        text-shadow: 0 2px 4px rgba(0,0,0,0.3);
    }
    .main-header p {
        margin: 0.5rem 0 0 0;
        opacity: 0.9;
        font-size: 1.2rem;
    }
    /* Cards de m√©tricas modernas - Melhorados */
    .metric-card {
        background: white;
        border-radius: var(--border-radius);
        padding: 1.5rem;
        box-shadow: var(--shadow);
        border: 1px solid #e9ecef;
        transition: var(--transition);
        text-align: center;
        margin-bottom: 1.5rem; /* Aumentado espa√ßamento */
        display: flex;               /* Estabilidade de altura */
        flex-direction: column;      /* Empilha valor, label, delta */
        justify-content: center;     /* Centraliza verticalmente */
        min-height: 140px;           /* Altura m√≠nima consistente */
    }
    .metric-card:hover {
        transform: translateY(-4px); /* Mais eleva√ß√£o */
        box-shadow: 0 6px 12px rgba(0,0,0,0.15);
    }
    .metric-value {
        font-size: 2.2rem; /* Aumentado */
        font-weight: 700;
        margin: 0.5rem 0;
        color: var(--primary-color);
    }
    .metric-label {
        font-size: 1rem; /* Ajustado */
        color: #6c757d;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        margin: 0;
    }
    .metric-delta {
        font-size: 0.9rem;
        margin-top: 0.5rem;
        min-height: 1rem;            /* Reserva espa√ßo mesmo vazia */
    }
    .metric-delta.positive { color: var(--success-color); }
    .metric-delta.negative { color: var(--danger-color); }
    /* Status badges - Ajustados */
    .status-badge {
        display: inline-block;
        padding: 0.35rem 0.85rem; /* Ajustado espa√ßamento */
        border-radius: 20px;
        font-size: 0.9rem;
        font-weight: 600;
        text-transform: uppercase;
    }
    .status-alto { background-color: #fee; color: var(--danger-color); border: 1px solid #fcc; }
    .status-medio { background-color: #ffeaa7; color: var(--warning-color); border: 1px solid #ffeaa7; }
    .status-baixo { background-color: #d4edda; color: var(--success-color); border: 1px solid #c3e6cb; }
    .status-inativo { background-color: #f8f9fa; color: #6c757d; border: 1px solid #dee2e6; }
    /* Bot√µes modernos */
    .stButton > button {
        background: linear-gradient(135deg, var(--primary-color), var(--secondary-color));
        color: white;
        border: none;
        border-radius: var(--border-radius);
        padding: 0.85rem 1.75rem; /* Ajustado */
        font-weight: 600;
        transition: var(--transition);
        box-shadow: var(--shadow);
    }
    .stButton > button:hover {
        transform: translateY(-2px); /* Mais eleva√ß√£o */
        box-shadow: 0 6px 12px rgba(0,0,0,0.2);
    }
    /* Sidebar moderna */
    .sidebar-header {
        background: var(--light-bg);
        padding: 1.2rem;
        border-radius: var(--border-radius);
        margin-bottom: 1.2rem;
        border-left: 5px solid var(--primary-color);
    }
    .sidebar-header h3 {
        margin: 0;
        color: var(--primary-color);
        font-size: 1.2rem;
        font-weight: 600;
    }
    /* Tabelas modernas */
    .dataframe-container {
        background: white;
        border-radius: var(--border-radius);
        padding: 1.2rem;
        box-shadow: var(--shadow);
        overflow: hidden;
    }
    /* Expander styling */
    .streamlit-expanderHeader {
        background: var(--light-bg);
        border-radius: var(--border-radius);
        font-weight: 600;
        color: var(--primary-color);
    }
    /* Loading states */
    .loading-container {
        text-align: center;
        padding: 3rem;
        color: #6c757d;
    }
    .loading-spinner {
        border: 4px solid #f3f3f3;
        border-top: 4px solid var(--primary-color);
        border-radius: 50%;
        width: 40px;
        height: 40px;
        animation: spin 1s linear infinite;
        margin: 0 auto 1rem;
    }
    .overlay-loader {
        position: fixed;
        inset: 0;
        background: rgba(15, 23, 42, 0.92);
        backdrop-filter: blur(4px);
        z-index: 10000;
        display: flex;
        align-items: center;
        justify-content: center;
    }
    .overlay-loader__content {
        text-align: center;
        color: #f8fafc;
        max-width: 480px;
        padding: 2.5rem;
        border-radius: 20px;
        background: rgba(17, 24, 39, 0.55);
        box-shadow: 0 20px 45px rgba(15, 23, 42, 0.45);
        border: 1px solid rgba(148, 163, 184, 0.2);
    }
    .overlay-loader__spinner {
        border: 6px solid rgba(148, 163, 184, 0.3);
        border-top: 6px solid var(--secondary-color);
        border-radius: 50%;
        width: 90px;
        height: 90px;
        animation: spin 1s linear infinite;
        margin: 0 auto 1.5rem;
    }
    .overlay-loader__title {
        font-size: 1.6rem;
        font-weight: 600;
        margin-bottom: 0.5rem;
    }
    .overlay-loader__subtitle {
        font-size: 1rem;
        opacity: 0.85;
        line-height: 1.5;
    }
    @keyframes spin {
        0% { transform: rotate(0deg); }
        100% { transform: rotate(360deg); }
    }
    /* Responsividade */
    @media (max-width: 768px) {
        .metric-card {
            margin-bottom: 1.5rem;
        }
        .main-header h1 {
            font-size: 2.2rem;
        }
        .metric-value {
            font-size: 1.8rem;
        }
    }
    /* Custom scrollbar */
    ::-webkit-scrollbar {
        width: 8px;
        height: 8px;
    }
    ::-webkit-scrollbar-track {
        background: #f1f1f1;
        border-radius: 4px;
    }
    ::-webkit-scrollbar-thumb {
        background: var(--primary-color);
        border-radius: 4px;
    }
    ::-webkit-scrollbar-thumb:hover {
        background: var(--secondary-color);
    }
    /* Footer */
    .footer {
        text-align: center;
        padding: 2rem 0;
        color: #6c757d;
        border-top: 1px solid #e9ecef;
        margin-top: 3rem;
    }
    /* Dark mode support */
    @media (prefers-color-scheme: dark) {
        :root {
            --light-bg: #2d3748;
            --dark-bg: #1a202c;
        }
        .metric-card {
            background: var(--dark-bg);
            border-color: #4a5568;
            color: white;
        }
        .metric-label {
            color: #a0aec0;
        }
    }
    /* Melhorias de espa√ßamento e layout */
    section[data-testid="stExpander"] > div {
        margin-bottom: 1rem;
    }
    .stTabs [data-testid="stMarkdownContainer"] {
        font-size: 1.1rem;
        font-weight: 600;
    }
    /* Ajuste para gr√°ficos */
    .plotly-chart {
        margin: 1rem 0;
        border-radius: var(--border-radius);
        box-shadow: var(--shadow);
        padding: 1rem;
        background: white;
    }
</style>
"""
# Injetar CSS
st.markdown(CSS_STYLES, unsafe_allow_html=True)
# ========================================
# CLASSES DO SISTEMA v2.0 - Atualizado com corre√ß√µes de bugs
# ========================================
@dataclass
class KPIMetrics:
    """Classe para armazenar m√©tricas calculadas."""
    total_labs: int = 0
    churn_rate: float = 0.0
    total_coletas: int = 0
    labs_em_risco: int = 0
    ativos_7d: float = 0.0
    ativos_30d: float = 0.0
    labs_alto_risco: int = 0
    labs_medio_risco: int = 0
    labs_baixo_risco: int = 0
    labs_inativos: int = 0
    labs_critico: int = 0
    labs_recuperando: int = 0
    labs_sem_coleta_48h: int = 0
    vol_hoje_total: int = 0
    vol_d1_total: int = 0
    ativos_7d_count: int = 0
    ativos_30d_count: int = 0
    labs_normal_count: int = 0
    labs_atencao_count: int = 0
    labs_moderado_count: int = 0
    labs_alto_count: int = 0
    labs_critico_count: int = 0
    labs_abaixo_mm7_br: int = 0
    labs_abaixo_mm7_br_pct: float = 0.0
    labs_abaixo_mm7_uf: int = 0
    labs_abaixo_mm7_uf_pct: float = 0.0
class DataManager:
    """Gerenciador de dados com cache inteligente."""
    @staticmethod
    def normalizar_cnpj(cnpj: str) -> str:
        """Remove formata√ß√£o do CNPJ (pontos, tra√ßos, barras)"""
        if pd.isna(cnpj) or cnpj == '':
            return ''
        # Converter num√©ricos para string sem decimais (evita sufixo '.0')
        if isinstance(cnpj, (int, float)):
            try:
                cnpj = str(int(cnpj))
            except Exception:
                cnpj = str(cnpj)
        # Remove tudo exceto d√≠gitos
        cnpj_limpo = ''.join(filter(str.isdigit, str(cnpj)))
        # Garantir 14 d√≠gitos
        if len(cnpj_limpo) < 14:
            cnpj_limpo = cnpj_limpo.zfill(14)
        elif len(cnpj_limpo) > 14:
            cnpj_limpo = cnpj_limpo[-14:]
        return cnpj_limpo
    @staticmethod
    @st.cache_data(ttl=CACHE_TTL)
    def carregar_dados_churn() -> Optional[pd.DataFrame]:
        """Carrega dados de an√°lise de churn com cache inteligente."""
        try:
            # PRIMEIRO: Tentar baixar do SharePoint/OneDrive
            arquivo_sharepoint = baixar_sharepoint()
         
            if arquivo_sharepoint and os.path.exists(arquivo_sharepoint):
                # Tentar ler como CSV primeiro
                try:
                    df = pd.read_csv(arquivo_sharepoint, encoding=ENCODING, low_memory=False)
                    return df
                except Exception:
                    # Tentar como Parquet
                    try:
                        df = pd.read_parquet(arquivo_sharepoint, engine='pyarrow')
                        return df
                    except Exception:
                        pass
         
            # FALLBACK: Tentar arquivos locais
            # Primeiro tenta CSV (mais comum)
            arquivo_csv = os.path.join(OUTPUT_DIR, "churn_analysis_latest.csv")
            if os.path.exists(arquivo_csv):
                df = pd.read_csv(arquivo_csv, encoding=ENCODING, low_memory=False)
                return df
         
            # Fallback para parquet
            arquivo_path = os.path.join(OUTPUT_DIR, CHURN_ANALYSIS_FILE)
            if os.path.exists(arquivo_path):
                df = pd.read_parquet(arquivo_path, engine='pyarrow')
                return df
         
            return None
         
        except Exception as e:
            st.error(f"‚ùå Erro ao carregar dados: {e}")
            return None
    @staticmethod
    def preparar_dados(df: pd.DataFrame) -> pd.DataFrame:
        """Prepara e limpa os dados carregados - Atualizado para coer√™ncia entre telas."""
        if df is None or df.empty:
            return pd.DataFrame()
        # Removido bloco de debug da sidebar para manter interface limpa
        # Garantir tipos de dados corretos
        if 'Data_Analise' in df.columns:
            df['Data_Analise'] = pd.to_datetime(df['Data_Analise'], errors='coerce')
        # Calcular volume total se n√£o existir (at√© o m√™s atual)
        try:
            # Fun√ß√£o inline para evitar depend√™ncia circular
            meses_ordem = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
            ano_atual = pd.Timestamp.today().year
            limite_mes = pd.Timestamp.today().month if 2025 == ano_atual else 12
            meses_limite = meses_ordem[:limite_mes]
            sufixo = str(2025)[-2:]
            meses_2025_dyn = [m for m in meses_limite if f'N_Coletas_{m}_{sufixo}' in df.columns]
        except Exception:
            meses_2025_dyn = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out']
        colunas_meses = [f'N_Coletas_{mes}_25' for mes in meses_2025_dyn]
        if 'Volume_Total_2025' not in df.columns:
            df['Volume_Total_2025'] = df[colunas_meses].sum(axis=1, skipna=True) if colunas_meses else 0
        # Adicionar coluna CNPJ normalizado para match com dados VIP
        if 'CNPJ_PCL' in df.columns:
            df['CNPJ_Normalizado'] = df['CNPJ_PCL'].apply(DataManager.normalizar_cnpj)
        # Filtro Active == True para coer√™ncia
        if 'Active' in df.columns:
            df = df[df['Active'] == True]
        # === Nova r√©gua de risco di√°rio ===
        colunas_novas = [
            "Vol_Hoje", "Vol_D1", "MM7", "MM30", "MM90", "DOW_Media",
            "Delta_D1", "Delta_MM7", "Delta_MM30", "Delta_MM90",
            "Risco_Diario", "Recuperacao"
        ]
        try:
            registros = []
            for _, r in df.iterrows():
                res = RiskEngine.classificar(r)
                registros.append(res if res else {c: None for c in colunas_novas})
            df_risk = pd.DataFrame(registros, index=df.index)
            for c in colunas_novas:
                df[c] = df_risk.get(c)
        except Exception:
            for c in colunas_novas:
                if c not in df.columns:
                    df[c] = None
        # Opcional: preservar a coluna antiga para auditoria
        if 'Status_Risco' in df.columns and 'Risco_Diario' in df.columns:
            df.rename(columns={'Status_Risco': 'Status_Risco_Legado'}, inplace=True)

        # Normalizar colunas de recoletas
        recoleta_cols = [c for c in df.columns if c.startswith('Recoletas_')]
        for col in recoleta_cols:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
        for col in ['Total_Recoletas_2024', 'Total_Recoletas_2025']:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)

        # Normalizar colunas de pre√ßo
        price_cols = []
        for cfg in PRICE_CATEGORIES.values():
            prefix = cfg['prefix']
            price_cols.extend([
                f'Preco_{prefix}_Total',
                f'Preco_{prefix}_Coleta',
                f'Preco_{prefix}_Exame'
            ])
        for col in price_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')

        if 'Voucher_Commission' in df.columns:
            df['Voucher_Commission'] = pd.to_numeric(df['Voucher_Commission'], errors='coerce')

        if 'Data_Preco_Atualizacao' in df.columns:
            df['Data_Preco_Atualizacao'] = pd.to_datetime(df['Data_Preco_Atualizacao'], errors='coerce', utc=True)
            try:
                df['Data_Preco_Atualizacao'] = df['Data_Preco_Atualizacao'].dt.tz_convert(TIMEZONE)
            except Exception:
                pass

        return df
    @staticmethod
    @st.cache_data(ttl=CACHE_TTL)
    def carregar_matriz_cs_normalizada() -> Optional[pd.DataFrame]:
        """Carrega dados da matriz CS normalizada com cache inteligente."""
        try:
            # PRIMEIRO: Tentar baixar do SharePoint/OneDrive
            arquivo_vip_remoto = "Data Analysis/Churn PCLs/matriz_cs_normalizada.csv"
            arquivo_sharepoint = baixar_sharepoint(arquivo_remoto=arquivo_vip_remoto)
            if arquivo_sharepoint and os.path.exists(arquivo_sharepoint):
                # Tentar ler como CSV
                try:
                    df = pd.read_csv(arquivo_sharepoint, encoding='utf-8-sig', low_memory=False)
                    # Verificar se tem coluna CNPJ ou CNPJ_PCL
                    if 'CNPJ' in df.columns:
                        coluna_cnpj = 'CNPJ'
                    elif 'CNPJ_PCL' in df.columns:
                        coluna_cnpj = 'CNPJ_PCL'
                        # Renomear para CNPJ para compatibilidade
                        df['CNPJ'] = df['CNPJ_PCL']
                    else:
                    # Warning removido - ser√° tratado onde a fun√ß√£o √© chamada
                        return None
                    # Ler CNPJ como string para preservar zeros √† esquerda
                    df['CNPJ'] = df['CNPJ'].astype(str)
                    df['CNPJ_Normalizado'] = df['CNPJ'].apply(DataManager.normalizar_cnpj)
                    # Toast removido - ser√° exibido onde a fun√ß√£o √© chamada
                    return df
                except Exception as e:
                    # Warning removido - ser√° tratado onde a fun√ß√£o √© chamada
                    pass
            # FALLBACK: Tentar arquivos locais
            caminhos_possiveis = [
                VIP_CSV_FILE,
                os.path.join(OUTPUT_DIR, VIP_CSV_FILE),
                os.path.join(os.path.dirname(OUTPUT_DIR), VIP_CSV_FILE),
            ]
            arquivo_csv = None
            for caminho in caminhos_possiveis:
                if os.path.exists(caminho):
                    arquivo_csv = caminho
                    break
            if arquivo_csv:
                # Ler CNPJ como string para preservar zeros √† esquerda
                df = pd.read_csv(
                    arquivo_csv,
                    encoding='utf-8-sig',
                    dtype={'CNPJ': 'string'},
                    low_memory=False
                )
                # Garantir que CNPJ seja string e normalizar
                df['CNPJ'] = df['CNPJ'].astype(str)
                df['CNPJ_Normalizado'] = df['CNPJ'].apply(DataManager.normalizar_cnpj)
                # Toast removido - ser√° exibido onde a fun√ß√£o √© chamada
                return df
            return None
        except Exception as e:
            # Error removido - ser√° tratado onde a fun√ß√£o √© chamada
            return None
    @staticmethod
    @st.cache_data(ttl=VIP_CACHE_TTL)
    def carregar_dados_vip() -> Optional[pd.DataFrame]:
        """Carrega dados VIP do CSV normalizado com cache."""
        try:
            # Tentar baixar matriz CS do SharePoint
            arquivo_vip_remoto = "Data Analysis/Churn PCLs/matriz_cs_normalizada.csv"
            arquivo_sharepoint = baixar_sharepoint(arquivo_remoto=arquivo_vip_remoto, force=False)
            if arquivo_sharepoint and os.path.exists(arquivo_sharepoint):
                # Ler arquivo VIP
                df_vip = pd.read_csv(
                    arquivo_sharepoint,
                    encoding='utf-8-sig'
                )
                # Verificar se tem coluna CNPJ ou CNPJ_PCL
                if 'CNPJ' in df_vip.columns:
                    coluna_cnpj = 'CNPJ'
                elif 'CNPJ_PCL' in df_vip.columns:
                    coluna_cnpj = 'CNPJ_PCL'
                    # Renomear para CNPJ para compatibilidade
                    df_vip['CNPJ'] = df_vip['CNPJ_PCL']
                else:
                    # Warning removido - ser√° tratado onde a fun√ß√£o √© chamada
                    return None
                # Ler CNPJ como string para preservar zeros √† esquerda
                df_vip['CNPJ'] = df_vip['CNPJ'].astype(str)
                df_vip['CNPJ_Normalizado'] = df_vip['CNPJ'].apply(DataManager.normalizar_cnpj)
                # Toast removido - ser√° exibido onde a fun√ß√£o √© chamada
                return df_vip
         
            # FALLBACK: Tentar m√∫ltiplos caminhos locais
            caminhos_possiveis = [
                VIP_CSV_FILE,
                os.path.join(OUTPUT_DIR, VIP_CSV_FILE),
                os.path.join(os.path.dirname(OUTPUT_DIR), VIP_CSV_FILE),
            ]
            arquivo_csv = None
            for caminho in caminhos_possiveis:
                if os.path.exists(caminho):
                    arquivo_csv = caminho
                    break
            if arquivo_csv:
                # Ler CNPJ como string para preservar zeros √† esquerda
                df_vip = pd.read_csv(
                    arquivo_csv,
                    encoding='utf-8-sig',
                    dtype={'CNPJ': 'string'}
                )
                # Garantir que CNPJ seja string e normalizar
                df_vip['CNPJ'] = df_vip['CNPJ'].astype(str)
                df_vip['CNPJ_Normalizado'] = df_vip['CNPJ'].apply(DataManager.normalizar_cnpj)
                # Toast removido - ser√° exibido onde a fun√ß√£o √© chamada
                return df_vip
            else:
                # Warning removido - ser√° tratado onde a fun√ß√£o √© chamada
                return None
        except Exception as e:
            st.warning(f"Erro ao carregar arquivo VIP: {e}")
            return None
    @staticmethod
    @st.cache_data(ttl=CACHE_TTL)
    def carregar_laboratories() -> Optional[pd.DataFrame]:
        """Carrega dados de laboratories.csv com cache inteligente."""
        try:
            # PRIMEIRO: Tentar baixar do SharePoint/OneDrive
            arquivo_labs_remoto = "Data Analysis/Churn PCLs/laboratories.csv"
            arquivo_sharepoint = baixar_sharepoint(arquivo_remoto=arquivo_labs_remoto)
            
            if arquivo_sharepoint and os.path.exists(arquivo_sharepoint):
                try:
                    df_labs = pd.read_csv(arquivo_sharepoint, encoding=ENCODING, low_memory=False)
                    
                    # Normalizar CNPJ para permitir matching
                    if 'cnpj' in df_labs.columns:
                        df_labs['cnpj'] = df_labs['cnpj'].astype(str)
                        df_labs['CNPJ_Normalizado'] = df_labs['cnpj'].apply(DataManager.normalizar_cnpj)
                    elif 'CNPJ' in df_labs.columns:
                        df_labs['CNPJ'] = df_labs['CNPJ'].astype(str)
                        df_labs['CNPJ_Normalizado'] = df_labs['CNPJ'].apply(DataManager.normalizar_cnpj)
                    
                    return df_labs
                except Exception as e:
                    # Erro silencioso - ser√° tratado onde a fun√ß√£o √© chamada
                    pass
            
            # FALLBACK: Tentar arquivo local
            arquivo_labs = os.path.join(OUTPUT_DIR, LABORATORIES_FILE)
            if os.path.exists(arquivo_labs):
                df_labs = pd.read_csv(arquivo_labs, encoding=ENCODING, low_memory=False)
                
                # Normalizar CNPJ para permitir matching
                if 'cnpj' in df_labs.columns:
                    df_labs['cnpj'] = df_labs['cnpj'].astype(str)
                    df_labs['CNPJ_Normalizado'] = df_labs['cnpj'].apply(DataManager.normalizar_cnpj)
                elif 'CNPJ' in df_labs.columns:
                    df_labs['CNPJ'] = df_labs['CNPJ'].astype(str)
                    df_labs['CNPJ_Normalizado'] = df_labs['CNPJ'].apply(DataManager.normalizar_cnpj)
                
                return df_labs
            
            return None
        except Exception as e:
            # Erro silencioso - ser√° tratado onde a fun√ß√£o √© chamada
            return None
    
    @staticmethod
    @st.cache_data(ttl=14400)  # Cache de 4 horas
    def carregar_dados_gralab() -> Optional[Dict[str, pd.DataFrame]]:
        """
        Carrega dados do Excel do concorrente Gralab com todas as abas.
        
        Returns:
            Dicion√°rio com DataFrames das abas ou None se falhar
        """
        try:
            # Baixar arquivo Excel do SharePoint
            arquivo_excel = baixar_excel_gralab()
            
            if not arquivo_excel or not os.path.exists(arquivo_excel):
                return None
            
            # Ler todas as abas do Excel
            todas_abas = pd.read_excel(arquivo_excel, sheet_name=None, engine='openpyxl')
            
            # Normalizar CNPJ em todas as abas que tenham coluna CNPJ ou Cnpj
            for nome_aba, df in todas_abas.items():
                # Procurar coluna de CNPJ (case insensitive)
                coluna_cnpj = None
                for col in df.columns:
                    if col.upper() == 'CNPJ':
                        coluna_cnpj = col
                        break
                
                if coluna_cnpj:
                    df[coluna_cnpj] = df[coluna_cnpj].astype(str)
                    df['CNPJ_Normalizado'] = df[coluna_cnpj].apply(DataManager.normalizar_cnpj)
            
            return todas_abas
            
        except Exception as e:
            # Erro silencioso - ser√° tratado onde a fun√ß√£o √© chamada
            return None


class RiskEngine:
    """Calcula MM7/MM30/MM90, D-1, DOW e classifica o risco di√°rio (nova r√©gua)."""

    @staticmethod
    def _serie_diaria_from_json(json_str: str) -> pd.Series:
        """Converte 'Dados_Diarios_2025' (dict 'YYYY-MM' -> {dia:coletas}) em s√©rie di√°ria."""
        if pd.isna(json_str) or str(json_str).strip() in ("", "{}", "null"):
            return pd.Series(dtype="float")
        import json
        try:
            j = json.loads(json_str)
        except Exception:
            return pd.Series(dtype="float")
        rows = []
        for ym, dias in j.items():
            try:
                y, m = ym.split("-")
            except Exception:
                continue
            for d_str, v in dias.items():
                try:
                    d = int(d_str)
                    rows.append((pd.Timestamp(int(y), int(m), d), int(v)))
                except Exception:
                    continue
        if not rows:
            return pd.Series(dtype="float")
        s = pd.Series({d: v for d, v in rows}).sort_index()
        return s

    @staticmethod
    def _last_business_day(reference: Optional[pd.Timestamp] = None) -> pd.Timestamp:
        """Retorna a √∫ltima data √∫til (considerando TIMEZONE)."""
        if reference is None:
            reference = pd.Timestamp.now(tz=TIMEZONE)
        else:
            if reference.tzinfo is None:
                reference = reference.tz_localize(TIMEZONE)
            else:
                reference = reference.tz_convert(TIMEZONE)
        reference = reference.normalize()
        reference_naive = reference.tz_localize(None)
        while reference_naive.weekday() >= 5:
            reference_naive = (reference_naive - BDay(1))
        return reference_naive

    @staticmethod
    def _serie_business_day(s: pd.Series, ref_date: pd.Timestamp) -> pd.Series:
        """Reindexa s√©rie para frequ√™ncia de dias √∫teis at√© ref_date."""
        if s.empty:
            return s
        s = s.sort_index()
        start = s.index.min()
        if ref_date < start:
            ref_date = start
        idx = pd.bdate_range(start, ref_date)
        if len(idx) == 0:
            idx = pd.DatetimeIndex([ref_date])
        return s.reindex(idx, fill_value=0)

    @staticmethod
    def _rolling_means(s: pd.Series, ref_date: pd.Timestamp) -> dict:
        """MM7/MM30/MM90, D-1, m√©dia por DOW e contadores auxiliares."""
        if s.empty:
            return dict(MM7=0, MM30=0, MM90=0, D1=0, DOW=0, HOJE=0, zeros_consec=0, quedas50_consec=0)
        if ref_date not in s.index:
            return dict(MM7=0, MM30=0, MM90=0, D1=0, DOW=0, HOJE=0, zeros_consec=0, quedas50_consec=0)

        hoje = float(s.loc[ref_date])
        serie_ate_ref = s.loc[:ref_date]
        if len(serie_ate_ref) > 1:
            d1 = float(serie_ate_ref.iloc[-2])
        else:
            d1 = 0.0
        mm7 = float(serie_ate_ref.tail(7).mean())
        mm30 = float(serie_ate_ref.tail(30).mean())
        mm90 = float(serie_ate_ref.tail(90).mean())
        dow = int(ref_date.weekday())
        dow_vals = serie_ate_ref[serie_ate_ref.index.weekday == dow]
        dow_mean = float(dow_vals.tail(90).mean()) if len(dow_vals) else 0.0
        zeros_consec = 0
        for valor in serie_ate_ref[::-1]:
            if valor == 0:
                zeros_consec += 1
            else:
                break

        def _is_queda50(idx):
            mm7_local = s.loc[:idx].tail(7).mean()
            return s.loc[idx] < 0.5 * mm7_local if mm7_local > 0 else False

        ultimos = s.loc[:ref_date].tail(3)
        quedas50_consec = sum([_is_queda50(idx) for idx in ultimos.index])
        return dict(MM7=mm7, MM30=mm30, MM90=mm90, D1=d1, DOW=dow_mean, HOJE=hoje,
                    zeros_consec=zeros_consec, quedas50_consec=quedas50_consec)

    @staticmethod
    def _to_float(value: Any) -> Optional[float]:
        """Converte valor para float com tratamento de NaN."""
        try:
            if pd.isna(value):
                return None
        except TypeError:
            pass
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def classificar(row: pd.Series) -> dict:
        """Aplica as regras do anexo e retorna m√©tricas + 'Risco_Diario' e 'Recuperacao'."""
        s = RiskEngine._serie_diaria_from_json(row.get("Dados_Diarios_2025", "{}"))
        if s.empty:
            return {}
        ref_date = RiskEngine._last_business_day()
        s = RiskEngine._serie_business_day(s, ref_date)
        if s.empty:
            return {}
        m = RiskEngine._rolling_means(s, ref_date)
        hoje, d1 = m["HOJE"], m["D1"]
        mm7, mm30, mm90, dow = m["MM7"], m["MM30"], m["MM90"], m["DOW"]

        def pct(a, b):
            return (a - b) / b * 100 if b and b != 0 else 0.0

        # L√≥gica h√≠brida para Delta D-1: usar MM7 como fallback quando D-1 = 0
        if d1 > 0:
            d_vs_d1 = pct(hoje, d1)
        elif d1 == 0 and hoje == 0:
            d_vs_d1 = 0.0
        elif d1 == 0 and hoje > 0:
            # Fallback: usar MM7 como refer√™ncia quando D-1 est√° zerado
            d_vs_d1 = pct(hoje, mm7) if mm7 > 0 else 0.0
        else:
            d_vs_d1 = 0.0
        
        d_vs_mm7 = pct(hoje, mm7)
        d_vs_mm30 = pct(hoje, mm30)
        d_vs_mm90 = pct(hoje, mm90)

        mm7_br = RiskEngine._to_float(row.get("MM7_BR"))
        mm7_uf = RiskEngine._to_float(row.get("MM7_UF"))
        mm7_cidade = RiskEngine._to_float(row.get("MM7_CIDADE"))
        contexto_mm = [mm for mm in [mm7_br, mm7_uf, mm7_cidade] if mm is not None and mm > 0]

        reducoes = []
        for mm_ctx in contexto_mm:
            reducao = 1 - (hoje / mm_ctx) if mm_ctx > 0 else 0
            reducoes.append(max(0.0, reducao))
        maior_reducao = max(reducoes) if reducoes else 0.0
        reducao_zero_absoluto = any(mm_ctx > 0 and hoje == 0 for mm_ctx in contexto_mm)

        risco = "üü¢ Normal"
        limiar_medio = REDUCAO_MEDIO_RISCO
        limiar_alto = REDUCAO_ALTO_RISCO

        if reducao_zero_absoluto or maior_reducao >= 1.0 or m["zeros_consec"] >= 7 or m["quedas50_consec"] >= 3:
            risco = "‚ö´ Cr√≠tico"
        elif maior_reducao >= limiar_alto:
            risco = "üî¥ Alto"
        elif maior_reducao >= limiar_medio:
            risco = "üü† Moderado"
        elif maior_reducao > 0:
            risco = "üü° Aten√ß√£o"
        else:
            risco = "üü¢ Normal"

        recuperacao = False
        ultimos_4 = s.loc[:ref_date].tail(4)
        if len(ultimos_4) == 4 and hoje >= mm7 and (ultimos_4.iloc[:3].mean() < 0.9 * mm7):
            recuperacao = True
        return {
            "Vol_Hoje": int(hoje), "Vol_D1": int(d1),
            "MM7": round(mm7, 3), "MM30": round(mm30, 3), "MM90": round(mm90, 3), "DOW_Media": round(dow, 1),
            "Delta_D1": round(d_vs_d1, 1), "Delta_MM7": round(d_vs_mm7, 1),
            "Delta_MM30": round(d_vs_mm30, 1), "Delta_MM90": round(d_vs_mm90, 1),
            "Risco_Diario": risco, "Recuperacao": recuperacao
        }


class VIPManager:
    """Gerenciador de dados VIP."""
    @staticmethod
    def buscar_info_vip(cnpj: str, df_vip: pd.DataFrame) -> Optional[dict]:
        """Busca informa√ß√µes VIP para um CNPJ (apenas ranking e rede, sem contato)."""
        if df_vip is None or df_vip.empty or not cnpj:
            return None
     
        cnpj_normalizado = DataManager.normalizar_cnpj(cnpj)
        if not cnpj_normalizado:
            return None
     
        # Buscar match no DataFrame VIP
        match = df_vip[df_vip['CNPJ_Normalizado'] == cnpj_normalizado]
        if not match.empty:
            row = match.iloc[0]
            return {
                'ranking': row.get('Ranking', ''),
                'ranking_rede': row.get('Ranking Rede', ''),
                'rede': row.get('Rede', '')
                # Campos de contato removidos - agora v√™m de laboratories.csv
            }
        return None
    
    @staticmethod
    def buscar_info_laboratory(cnpj: str, df_labs: pd.DataFrame) -> Optional[dict]:
        """Busca informa√ß√µes de contato do laboratories.csv por CNPJ."""
        if df_labs is None or df_labs.empty or not cnpj:
            return None
        
        cnpj_normalizado = DataManager.normalizar_cnpj(cnpj)
        if not cnpj_normalizado:
            return None
        
        # Buscar match no DataFrame de laboratories
        match = df_labs[df_labs['CNPJ_Normalizado'] == cnpj_normalizado]
        if match.empty:
            return None
        
        row = match.iloc[0]
        
        # Fun√ß√£o auxiliar para extrair valores de campos aninhados ou flattenados
        def extrair_de_dict(valor_dict, chave, subchave=None):
            """Extrai valor de dict aninhado ou string JSON parseada."""
            if pd.isna(valor_dict) or valor_dict == '':
                return ''
            
            # Se √© string, tentar parsear como JSON
            if isinstance(valor_dict, str):
                try:
                    import json
                    import re
                    # Limpar ObjectId e outras strings n√£o-JSON
                    valor_limpo = re.sub(r'ObjectId\([^)]+\)', 'null', valor_dict)
                    valor_limpo = valor_limpo.replace("'", '"')
                    valor_dict = json.loads(valor_limpo)
                except:
                    return ''
            
            # Se √© dict, extrair valor
            if isinstance(valor_dict, dict):
                if subchave:
                    # Acessar chave.subchave (ex: contact.telephone)
                    if chave in valor_dict:
                        sub_dict = valor_dict.get(chave, {})
                        if isinstance(sub_dict, dict) and subchave in sub_dict:
                            valor = sub_dict.get(subchave, '')
                            return str(valor).strip() if valor else ''
                else:
                    # Acessar chave diretamente
                    if chave in valor_dict:
                        valor = valor_dict.get(chave, '')
                        return str(valor).strip() if valor else ''
            
            return ''
        
        # Extrair nome do contato (prefer√™ncia: director.name, fallback: manager.name)
        contato = ''
        
        # Tentar director.name
        if 'director' in df_labs.columns:
            director_data = row.get('director', '')
            contato = extrair_de_dict(director_data, 'name')
        
        # Tentar manager.name como fallback
        if not contato and 'manager' in df_labs.columns:
            manager_data = row.get('manager', '')
            contato = extrair_de_dict(manager_data, 'name')
        
        # Tentar colunas flattenadas (caso o CSV tenha sido flattenado)
        if not contato and 'director.name' in df_labs.columns:
            contato = str(row.get('director.name', '')).strip()
        if not contato and 'manager.name' in df_labs.columns:
            contato = str(row.get('manager.name', '')).strip()
        
        # Extrair telefone de contact.telephone
        telefone = ''
        if 'contact' in df_labs.columns:
            contact_data = row.get('contact', '')
            telefone = extrair_de_dict(contact_data, 'telephone')
        
        # Tentar coluna flattenada
        if not telefone and 'contact.telephone' in df_labs.columns:
            telefone = str(row.get('contact.telephone', '')).strip()
        
        # Extrair email de contact.email
        email = ''
        if 'contact' in df_labs.columns:
            contact_data = row.get('contact', '')
            email = extrair_de_dict(contact_data, 'email')
        
        # Tentar coluna flattenada
        if not email and 'contact.email' in df_labs.columns:
            email = str(row.get('contact.email', '')).strip()
        
        # Fun√ß√£o auxiliar para extrair array de strings
        def extrair_array(campo_dict, chave_array):
            """Extrai array de strings de um dict aninhado ou string JSON."""
            if pd.isna(campo_dict) or campo_dict == '':
                return []
            
            # Se √© string, tentar parsear como JSON
            if isinstance(campo_dict, str):
                try:
                    import json
                    import re
                    valor_limpo = re.sub(r'ObjectId\([^)]+\)', 'null', campo_dict)
                    valor_limpo = valor_limpo.replace("'", '"')
                    campo_dict = json.loads(valor_limpo)
                except:
                    return []
            
            # Se √© dict, extrair array
            if isinstance(campo_dict, dict):
                if chave_array in campo_dict:
                    array_data = campo_dict.get(chave_array, [])
                    if isinstance(array_data, list):
                        return [str(item).strip() for item in array_data if item]
                    elif isinstance(array_data, str):
                        try:
                            import json
                            return json.loads(array_data)
                        except:
                            return [array_data] if array_data else []
            
            return []
        
        # Fun√ß√£o auxiliar para extrair campos booleanos True
        def extrair_booleanos(campo_dict, prefixo):
            """Extrai lista de chaves onde valor √© True."""
            lista = []
            if pd.isna(campo_dict) or campo_dict == '':
                return lista
            
            # Se √© string, tentar parsear como JSON
            if isinstance(campo_dict, str):
                try:
                    import json
                    import re
                    valor_limpo = re.sub(r'ObjectId\([^)]+\)', 'null', campo_dict)
                    valor_limpo = valor_limpo.replace("'", '"')
                    campo_dict = json.loads(valor_limpo)
                except:
                    return lista
            
            # Se √© dict, extrair booleanos True
            if isinstance(campo_dict, dict):
                for chave, valor in campo_dict.items():
                    if valor is True or (isinstance(valor, str) and valor.lower() == 'true'):
                        lista.append(chave)
            
            return lista
        
        # Extrair endere√ßo completo (address)
        endereco_completo = {
            'postalCode': '',
            'address': '',
            'addressComplement': '',
            'number': '',
            'neighbourhood': '',
            'city': '',
            'state_code': '',
            'state_name': ''
        }
        
        if 'address' in df_labs.columns:
            address_data = row.get('address', '')
            
            # Extrair campos do endere√ßo usando fun√ß√£o auxiliar
            campos_endereco = ['postalCode', 'address', 'addressComplement', 'number', 'neighbourhood', 'city']
            for campo in campos_endereco:
                valor = extrair_de_dict(address_data, campo)
                if valor:
                    endereco_completo[campo] = valor
            
            # Tentar colunas flattenadas
            for campo in campos_endereco:
                coluna_flatten = f'address.{campo}'
                if coluna_flatten in df_labs.columns and not endereco_completo[campo]:
                    valor = str(row.get(coluna_flatten, '')).strip()
                    if valor and valor.lower() != 'nan':
                        endereco_completo[campo] = valor
            
            # Extrair state (objeto aninhado)
            if isinstance(address_data, str):
                try:
                    import json
                    import re
                    valor_limpo = re.sub(r'ObjectId\([^)]+\)', 'null', address_data)
                    valor_limpo = valor_limpo.replace("'", '"')
                    address_dict = json.loads(valor_limpo)
                    if isinstance(address_dict, dict) and 'state' in address_dict:
                        state_data = address_dict.get('state', {})
                        if isinstance(state_data, dict):
                            endereco_completo['state_code'] = str(state_data.get('code', '')).strip()
                            endereco_completo['state_name'] = str(state_data.get('name', '')).strip()
                except:
                    pass
            elif isinstance(address_data, dict) and 'state' in address_data:
                state_data = address_data.get('state', {})
                if isinstance(state_data, dict):
                    endereco_completo['state_code'] = str(state_data.get('code', '')).strip()
                    endereco_completo['state_name'] = str(state_data.get('name', '')).strip()
            
            # Tentar colunas flattenadas para state
            if 'address.state.code' in df_labs.columns and not endereco_completo['state_code']:
                endereco_completo['state_code'] = str(row.get('address.state.code', '')).strip()
            if 'address.state.name' in df_labs.columns and not endereco_completo['state_name']:
                endereco_completo['state_name'] = str(row.get('address.state.name', '')).strip()
        
        # Extrair dados de logistic (days, openingHours, comments)
        logistic_data = {
            'days': [],
            'openingHours': '',
            'comments': ''
        }
        
        # Tentar colunas flattenadas primeiro
        if 'logistic.days' in df_labs.columns:
            days_data = row.get('logistic.days', '')
            if isinstance(days_data, list):
                logistic_data['days'] = [str(item).strip() for item in days_data if item]
            elif isinstance(days_data, str) and days_data:
                try:
                    import json
                    parsed = json.loads(days_data)
                    if isinstance(parsed, list):
                        logistic_data['days'] = [str(item).strip() for item in parsed if item]
                except:
                    logistic_data['days'] = [days_data.strip()] if days_data.strip() else []
        
        if 'logistic.openingHours' in df_labs.columns:
            valor = row.get('logistic.openingHours', '')
            if pd.notna(valor) and str(valor).strip() != '' and str(valor).strip().lower() != 'nan':
                logistic_data['openingHours'] = str(valor).strip()
        
        if 'logistic.comments' in df_labs.columns:
            valor = row.get('logistic.comments', '')
            if pd.notna(valor) and str(valor).strip() != '' and str(valor).strip().lower() != 'nan':
                logistic_data['comments'] = str(valor).strip()
        
        # Fallback: tentar objeto aninhado
        if 'logistic' in df_labs.columns:
            logistic_dict = row.get('logistic', '')
            
            if isinstance(logistic_dict, str) and logistic_dict:
                try:
                    import json
                    import re
                    valor_limpo = re.sub(r'ObjectId\([^)]+\)', 'null', logistic_dict)
                    valor_limpo = valor_limpo.replace("'", '"')
                    logistic_dict = json.loads(valor_limpo)
                except:
                    pass
            
            if isinstance(logistic_dict, dict):
                if not logistic_data['days'] and 'days' in logistic_dict:
                    days_val = logistic_dict.get('days', [])
                    if isinstance(days_val, list):
                        logistic_data['days'] = [str(item).strip() for item in days_val if item]
                
                if not logistic_data['openingHours'] and 'openingHours' in logistic_dict:
                    opening_val = logistic_dict.get('openingHours', '')
                    if opening_val:
                        logistic_data['openingHours'] = str(opening_val).strip()
                
                if not logistic_data['comments'] and 'comments' in logistic_dict:
                    comments_val = logistic_dict.get('comments', '')
                    if comments_val:
                        logistic_data['comments'] = str(comments_val).strip()
        
        # Extrair licensed (booleanos) - tentar todas as possibilidades
        licensed_list = []
        campos_licensed = ['clt', 'cnh', 'cltCnh', 'other', 'online', 'civilService', 'civilServiceAnalysis50', 'otherAnalysis50']
        
        def valor_eh_true(valor):
            """Verifica se um valor deve ser considerado True."""
            if pd.isna(valor) or valor == '':
                return False
            if valor is True:
                return True
            if isinstance(valor, bool):
                return valor
            if isinstance(valor, str):
                return valor.lower() in ['true', '1', 'yes', 't', 'y']
            if isinstance(valor, (int, float)):
                return valor != 0 and valor != 0.0
            return False
        
        # Primeiro: tentar colunas flattenadas com ponto
        for campo in campos_licensed:
            coluna_flatten = f'licensed.{campo}'
            if coluna_flatten in df_labs.columns:
                valor = row.get(coluna_flatten, False)
                if valor_eh_true(valor):
                    licensed_list.append(campo)
        
        # Segundo: tentar sem ponto (caso j√° esteja flattenado no CSV)
        if not licensed_list:
            for campo in campos_licensed:
                if campo in df_labs.columns:
                    valor = row.get(campo, False)
                    if valor_eh_true(valor):
                        licensed_list.append(campo)
        
        # Terceiro: tentar objeto aninhado (dict ou string JSON)
        # Verificar se coluna licensed existe
        if 'licensed' in df_labs.columns and not licensed_list:
            licensed_dict = row.get('licensed', '')
            
            # Se √© string, tentar parsear
            if isinstance(licensed_dict, str) and licensed_dict and licensed_dict.strip():
                try:
                    import json
                    import re
                    import ast
                    # Tentar parsear como JSON
                    valor_limpo = re.sub(r'ObjectId\([^)]+\)', 'null', licensed_dict)
                    valor_limpo = valor_limpo.replace("'", '"')
                    licensed_dict = json.loads(valor_limpo)
                except:
                    try:
                        # Tentar ast.literal_eval (mais seguro que eval)
                        licensed_dict = ast.literal_eval(licensed_dict)
                    except:
                        try:
                            # √öltima tentativa: eval (menos seguro, mas necess√°rio em alguns casos)
                            licensed_dict = eval(licensed_dict)
                        except:
                            pass
            
            # Se √© dict, extrair booleanos True
            if isinstance(licensed_dict, dict):
                for campo in campos_licensed:
                    if campo in licensed_dict:
                        valor = licensed_dict.get(campo, False)
                        if valor_eh_true(valor):
                            if campo not in licensed_list:
                                licensed_list.append(campo)
        
        # Quarto: tentar verificar todas as colunas que contenham "licensed" no nome
        if not licensed_list:
            for col in df_labs.columns:
                if 'licensed' in str(col).lower():
                    # Tentar extrair valor da coluna
                    valor_col = row.get(col, '')
                    # Se a coluna cont√©m um dict/string, tentar parsear
                    if isinstance(valor_col, str) and valor_col and '{' in valor_col:
                        try:
                            import ast
                            valor_col = ast.literal_eval(valor_col)
                        except:
                            pass
                    # Se agora √© dict, processar
                    if isinstance(valor_col, dict):
                        for campo in campos_licensed:
                            if campo in valor_col and valor_eh_true(valor_col.get(campo)):
                                if campo not in licensed_list:
                                    licensed_list.append(campo)
        
        # Extrair allowedMethods (booleanos) - tentar todas as possibilidades
        allowed_methods_list = []
        campos_methods = ['cash', 'credit', 'debit', 'billing_laboratory', 'billing_company', 'billing', 'bank_billet', 'eCredit', 'pix']
        
        # Primeiro: tentar colunas flattenadas com ponto
        for campo in campos_methods:
            coluna_flatten = f'allowedMethods.{campo}'
            if coluna_flatten in df_labs.columns:
                valor = row.get(coluna_flatten, False)
                if valor_eh_true(valor):
                    allowed_methods_list.append(campo)
        
        # Segundo: tentar sem ponto (caso j√° esteja flattenado no CSV)
        if not allowed_methods_list:
            for campo in campos_methods:
                if campo in df_labs.columns:
                    valor = row.get(campo, False)
                    if valor_eh_true(valor):
                        allowed_methods_list.append(campo)
        
        # Terceiro: tentar objeto aninhado (dict ou string JSON)
        # Verificar se coluna allowedMethods existe
        if 'allowedMethods' in df_labs.columns and not allowed_methods_list:
            allowed_methods_dict = row.get('allowedMethods', '')
            
            # Se √© string, tentar parsear
            if isinstance(allowed_methods_dict, str) and allowed_methods_dict and allowed_methods_dict.strip():
                try:
                    import json
                    import re
                    import ast
                    # Tentar parsear como JSON
                    valor_limpo = re.sub(r'ObjectId\([^)]+\)', 'null', allowed_methods_dict)
                    valor_limpo = valor_limpo.replace("'", '"')
                    allowed_methods_dict = json.loads(valor_limpo)
                except:
                    try:
                        # Tentar ast.literal_eval (mais seguro que eval)
                        allowed_methods_dict = ast.literal_eval(allowed_methods_dict)
                    except:
                        try:
                            # √öltima tentativa: eval (menos seguro, mas necess√°rio em alguns casos)
                            allowed_methods_dict = eval(allowed_methods_dict)
                        except:
                            pass
            
            # Se √© dict, extrair booleanos True
            if isinstance(allowed_methods_dict, dict):
                for campo in campos_methods:
                    if campo in allowed_methods_dict:
                        valor = allowed_methods_dict.get(campo, False)
                        if valor_eh_true(valor):
                            if campo not in allowed_methods_list:
                                allowed_methods_list.append(campo)
        
        # Quarto: tentar verificar todas as colunas que contenham "allowedmethods" ou "allowed_methods" no nome
        if not allowed_methods_list:
            for col in df_labs.columns:
                col_lower = str(col).lower()
                if 'allowedmethods' in col_lower or 'allowed_methods' in col_lower:
                    # Tentar extrair valor da coluna
                    valor_col = row.get(col, '')
                    # Se a coluna cont√©m um dict/string, tentar parsear
                    if isinstance(valor_col, str) and valor_col and '{' in valor_col:
                        try:
                            import ast
                            valor_col = ast.literal_eval(valor_col)
                        except:
                            pass
                    # Se agora √© dict, processar
                    if isinstance(valor_col, dict):
                        for campo in campos_methods:
                            if campo in valor_col and valor_eh_true(valor_col.get(campo)):
                                if campo not in allowed_methods_list:
                                    allowed_methods_list.append(campo)
        
        return {
            'contato': contato if contato else '',
            'telefone': telefone if telefone else '',
            'email': email if email else '',
            'endereco': endereco_completo,
            'logistic': logistic_data,
            'licensed': licensed_list,
            'allowedMethods': allowed_methods_list
        }
def _formatar_df_exibicao(df: pd.DataFrame) -> pd.DataFrame:
    """Padroniza exibi√ß√£o: n√∫meros sem NaN/None (0), textos como '‚Äî'."""
    if df is None or df.empty:
        return df
    df_fmt = df.copy()
    for col in df_fmt.columns:
        if pd.api.types.is_numeric_dtype(df_fmt[col]):
            df_fmt[col] = pd.to_numeric(df_fmt[col], errors='coerce').fillna(0)
        else:
            df_fmt[col] = df_fmt[col].astype(object)
            df_fmt[col] = df_fmt[col].where(df_fmt[col].notna(), '‚Äî')
    return df_fmt
class FilterManager:
    """Gerenciador de filtros da interface."""
    def __init__(self):
        self.filtros = {}
    def renderizar_sidebar_filtros(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Renderiza filtros otimizados na sidebar."""
        st.sidebar.markdown('<div class="sidebar-header" style="font-size: 1rem; font-weight: 600; color: var(--primary-color);">üîß Filtros</div>', unsafe_allow_html=True)
        filtros = {}
        # Filtro VIP com op√ß√£o de alternar
        filtros['apenas_vip'] = st.sidebar.toggle(
            "üåü Apenas Clientes VIP",
            value=True,
            help="Ative para mostrar apenas clientes VIP, desative para mostrar todos"
        )
     
        # Separador visual
        st.sidebar.markdown("---")
        # Filtro por representante
        if 'Representante_Nome' in df.columns:
            representantes_lista = (
                df['Representante_Nome']
                .astype(str)
                .str.strip()
                .replace({'nan': '', 'None': ''})
            )
            representantes_opcoes = sorted({r for r in representantes_lista if r})
        else:
            representantes_opcoes = []
        filtros['representantes'] = st.sidebar.multiselect(
            "üë§ Representantes",
            options=representantes_opcoes,
            help="Selecione um ou mais representantes para filtrar os laborat√≥rios exibidos."
        )

        st.sidebar.markdown("---")
        # Filtro por per√≠odo - Anos e Meses (dados mensais)
        st.sidebar.markdown("**üìÖ Per√≠odo de An√°lise (Mensal)**")
        # Verificar anos dispon√≠veis nos dados
        anos_disponiveis = []
        if 'N_Coletas_Jan_24' in df.columns:
            anos_disponiveis.append(2024)
        if 'N_Coletas_Jan_25' in df.columns:
            anos_disponiveis.append(2025)
        if not anos_disponiveis:
            st.sidebar.warning("‚ö†Ô∏è Nenhum dado mensal encontrado")
            anos_disponiveis = [2024, 2025] # fallback
        # Sele√ß√£o de ano
        ano_selecionado = st.sidebar.selectbox(
            "üìä Ano de An√°lise:",
            options=anos_disponiveis,
            index=len(anos_disponiveis)-1, # Padr√£o: √∫ltimo ano dispon√≠vel
            help="Selecione o ano para an√°lise mensal"
        )
        # Mapeamento de meses
        meses_map = {
            'Jan': 'Janeiro', 'Fev': 'Fevereiro', 'Mar': 'Mar√ßo', 'Abr': 'Abril',
            'Mai': 'Maio', 'Jun': 'Junho', 'Jul': 'Julho', 'Ago': 'Agosto',
            'Set': 'Setembro', 'Out': 'Outubro', 'Nov': 'Novembro', 'Dez': 'Dezembro'
        }
        # Meses dispon√≠veis para o ano selecionado
        sufixo_ano = str(ano_selecionado)[-2:] # '24' ou '25'
        meses_disponiveis = []
        for mes_codigo in ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun',
                          'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']:
            coluna_mes = f'N_Coletas_{mes_codigo}_{sufixo_ano}'
            if coluna_mes in df.columns:
                meses_disponiveis.append(mes_codigo)
        if not meses_disponiveis:
            st.sidebar.warning(f"‚ö†Ô∏è Nenhum m√™s encontrado para {ano_selecionado}")
            meses_disponiveis = ['Jan', 'Fev', 'Mar'] # fallback
        # Sele√ß√£o de meses
        meses_opcoes = [f"{mes} - {meses_map.get(mes, mes)}" for mes in meses_disponiveis]
        meses_selecionados_opcoes = st.sidebar.multiselect(
            f"üìÖ Meses de {ano_selecionado}:",
            options=meses_opcoes,
            default=meses_opcoes, # Todos selecionados por padr√£o
            help=f"Selecione os meses de {ano_selecionado} para an√°lise. Deixe todos selecionados para vis√£o completa do ano.",
            key=f"meses_{ano_selecionado}"
        )
        # Converter para c√≥digos de m√™s
        meses_selecionados = []
        for opcao in meses_selecionados_opcoes:
            mes_codigo = opcao.split(' - ')[0]
            if mes_codigo in meses_disponiveis:
                meses_selecionados.append(mes_codigo)
        # Armazenar filtros para uso posterior
        filtros['ano_selecionado'] = ano_selecionado
        filtros['meses_selecionados'] = meses_selecionados
        filtros['sufixo_ano'] = sufixo_ano
        # Mostrar per√≠odo selecionado (texto discreto)
        meses_nomes = [meses_map.get(mes, mes) for mes in meses_selecionados]
        periodo_texto = f"{ano_selecionado}: {', '.join(meses_nomes[:3])}" # Max 3 meses no texto
        if len(meses_selecionados) > 3:
            periodo_texto += f" +{len(meses_selecionados)-3}..."
        st.sidebar.markdown(f"<small>üìä {periodo_texto}</small>", unsafe_allow_html=True)
        self.filtros = filtros
        return filtros
    def aplicar_filtros(self, df: pd.DataFrame, filtros: Dict[str, Any]) -> pd.DataFrame:
        """Aplica filtros otimizados ao DataFrame - Atualizado para coer√™ncia."""
        if df.empty:
            return df
        df_filtrado = df.copy()
        # Filtro VIP (sempre ativo)
        if filtros.get('apenas_vip', False):
            try:
                # Carregar dados VIP
                df_vip = DataManager.carregar_dados_vip()
                if df_vip is not None and not df_vip.empty:
                    # Normalizar CNPJs para match com tratamento de erro
                    df_filtrado['CNPJ_Normalizado'] = df_filtrado['CNPJ_PCL'].apply(
                        lambda x: ''.join(filter(str.isdigit, str(x))) if pd.notna(x) and str(x).strip() != '' else ''
                    )
                    df_vip['CNPJ_Normalizado'] = df_vip['CNPJ'].apply(
                        lambda x: ''.join(filter(str.isdigit, str(x))) if pd.notna(x) and str(x).strip() != '' else ''
                    )
                 
                    # Filtrar apenas registros que est√£o na lista VIP (com valida√ß√£o)
                    if 'CNPJ_Normalizado' in df_filtrado.columns and 'CNPJ_Normalizado' in df_vip.columns:
                        # Remover CNPJs vazios antes do match
                        df_filtrado = df_filtrado[df_filtrado['CNPJ_Normalizado'] != '']
                        df_vip_clean = df_vip[df_vip['CNPJ_Normalizado'] != '']
                     
                        if not df_vip_clean.empty:
                            df_filtrado = df_filtrado[df_filtrado['CNPJ_Normalizado'].isin(df_vip_clean['CNPJ_Normalizado'])]
                        else:
                            # Se n√£o h√° CNPJs v√°lidos na lista VIP, retornar DataFrame vazio
                            return pd.DataFrame()
                    else:
                        # Se as colunas n√£o existem, retornar DataFrame vazio
                        return pd.DataFrame()
                else:
                    # Se n√£o h√° dados VIP, retornar DataFrame vazio
                    return pd.DataFrame()
            except Exception as e:
                # Em caso de erro, retornar DataFrame vazio e log do erro
                st.error(f"Erro ao aplicar filtro VIP: {str(e)}")
                return pd.DataFrame()
        # Filtro por per√≠odo (compatibilidade com filtros antigos)
        if 'Data_Analise' in df_filtrado.columns and filtros.get('data_inicio') and filtros.get('data_fim'):
            try:
                # Garantir que as datas sejam do tipo date
                data_inicio = filtros['data_inicio']
                data_fim = filtros['data_fim']
                # Se for datetime, converter para date
                if hasattr(data_inicio, 'date'):
                    data_inicio = data_inicio.date()
                if hasattr(data_fim, 'date'):
                    data_fim = data_fim.date()
                # Verificar se a coluna Data_Analise √© do tipo datetime
                if df_filtrado['Data_Analise'].dtype == 'object':
                    # Tentar converter para datetime
                    df_filtrado['Data_Analise'] = pd.to_datetime(df_filtrado['Data_Analise'], errors='coerce')
                # Aplicar filtro apenas se a convers√£o foi bem-sucedida
                if df_filtrado['Data_Analise'].dtype.name.startswith('datetime'):
                    df_filtrado = df_filtrado[
                        (df_filtrado['Data_Analise'].dt.date >= data_inicio) &
                        (df_filtrado['Data_Analise'].dt.date <= data_fim)
                    ]
            except Exception as e:
                # Em caso de erro no filtro de data, continuar sem filtrar
                st.warning(f"Aviso: Erro ao aplicar filtro de per√≠odo: {str(e)}")
                pass
        # Filtro por representante
        representantes_sel = filtros.get('representantes', [])
        if representantes_sel and 'Representante_Nome' in df_filtrado.columns:
            df_filtrado = df_filtrado[df_filtrado['Representante_Nome'].isin(representantes_sel)]
        # Para dados mensais, o filtro principal ser√° usado nos c√°lculos dos gr√°ficos
        # Os filtros 'ano_selecionado', 'meses_selecionados' e 'sufixo_ano' s√£o usados
        # diretamente nas fun√ß√µes de c√°lculo dos gr√°ficos
        return df_filtrado
class KPIManager:
    """Gerenciador de c√°lculos de KPIs - Atualizado para coer√™ncia entre telas."""
    @staticmethod
    def calcular_kpis(df: pd.DataFrame) -> KPIMetrics:
        if df.empty:
            return KPIMetrics()
        metrics = KPIMetrics()
        df_recent = df[df['Dias_Sem_Coleta'] <= 90].copy() if 'Dias_Sem_Coleta' in df.columns else df.copy()
        metrics.total_labs = len(df_recent)
        # Distribui√ß√£o por Risco_Diario
        labs_normal = labs_atencao = labs_moderado = labs_alto = labs_critico = 0
        if 'Risco_Diario' in df_recent.columns:
            c = df_recent['Risco_Diario'].value_counts()
            labs_normal = c.get('üü¢ Normal', 0)
            labs_atencao = c.get('üü° Aten√ß√£o', 0)
            labs_moderado = c.get('üü† Moderado', 0)
            labs_alto = c.get('üî¥ Alto', 0)
            labs_critico = c.get('‚ö´ Cr√≠tico', 0)
        metrics.labs_baixo_risco = labs_normal + labs_atencao
        metrics.labs_medio_risco = labs_moderado
        metrics.labs_alto_risco = labs_alto + labs_critico
        metrics.labs_em_risco = labs_moderado + labs_alto + labs_critico
        metrics.labs_critico = labs_critico
        metrics.labs_normal_count = labs_normal
        metrics.labs_atencao_count = labs_atencao
        metrics.labs_moderado_count = labs_moderado
        metrics.labs_alto_count = labs_alto
        metrics.labs_critico_count = labs_critico
        metrics.churn_rate = (metrics.labs_em_risco / metrics.total_labs * 100) if metrics.total_labs else 0
        # Labs abaixo de MM7 por contexto
        def _count_below(column_name: str) -> int:
            if column_name not in df_recent.columns or 'Vol_Hoje' not in df_recent.columns:
                return 0
            serie_ref = pd.to_numeric(df_recent[column_name], errors='coerce')
            vol_hoje = pd.to_numeric(df_recent['Vol_Hoje'], errors='coerce').fillna(0)
            mask_valid = serie_ref.notna() & (serie_ref > 0)
            return int(((vol_hoje < serie_ref) & mask_valid).sum())

        metrics.labs_abaixo_mm7_br = _count_below('MM7_BR')
        metrics.labs_abaixo_mm7_uf = _count_below('MM7_UF')
        denominator = metrics.total_labs or 0
        metrics.labs_abaixo_mm7_br_pct = (
            metrics.labs_abaixo_mm7_br / denominator * 100 if denominator else 0.0
        )
        metrics.labs_abaixo_mm7_uf_pct = (
            metrics.labs_abaixo_mm7_uf / denominator * 100 if denominator else 0.0
        )
        # Total coletas 2025
        meses_2025 = ChartManager._meses_ate_hoje(df_recent, 2025)
        cols = [f'N_Coletas_{m}_25' for m in meses_2025 if f'N_Coletas_{m}_25' in df_recent.columns]
        metrics.total_coletas = int(df_recent[cols].sum().sum()) if cols else 0
        # Volumes di√°rios
        metrics.vol_hoje_total = int(df_recent['Vol_Hoje'].fillna(0).sum()) if 'Vol_Hoje' in df_recent.columns else 0
        metrics.vol_d1_total = int(df_recent['Vol_D1'].fillna(0).sum()) if 'Vol_D1' in df_recent.columns else 0
        # Recupera√ß√£o e zeros consecutivos
        if 'Recuperacao' in df_recent.columns:
            metrics.labs_recuperando = int(df_recent['Recuperacao'].fillna(False).sum())
        if {'Vol_Hoje', 'Vol_D1'}.issubset(df_recent.columns):
            zeros_48h = df_recent[
                df_recent['Vol_Hoje'].fillna(0).eq(0) &
                df_recent['Vol_D1'].fillna(0).eq(0)
            ]
            metrics.labs_sem_coleta_48h = len(zeros_48h)
        # Ativos recentes
        if 'Dias_Sem_Coleta' in df_recent.columns and metrics.total_labs > 0:
            ativos_7d_df = df_recent[df_recent['Dias_Sem_Coleta'] <= 7]
            ativos_30d_df = df_recent[df_recent['Dias_Sem_Coleta'] <= 30]
            metrics.ativos_7d_count = len(ativos_7d_df)
            metrics.ativos_30d_count = len(ativos_30d_df)
            metrics.ativos_7d = metrics.ativos_7d_count / metrics.total_labs * 100
            metrics.ativos_30d = metrics.ativos_30d_count / metrics.total_labs * 100
        return metrics
class ChartManager:
    """Gerenciador de cria√ß√£o de gr√°ficos - Atualizado com corre√ß√µes de bugs e layouts."""
    @staticmethod
    def _meses_ate_hoje(df: pd.DataFrame, ano: int) -> list:
        """Retorna lista de c√≥digos de meses dispon√≠veis at√© o m√™s corrente para o ano informado.
        - Garante ordem cronol√≥gica correta
        - Considera apenas colunas que existem no DataFrame
        - Para anos anteriores ao corrente, considera at√© Dezembro
        """
        meses_ordem = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
        ano_atual = pd.Timestamp.today().year
        limite_mes = pd.Timestamp.today().month if ano == ano_atual else 12
        meses_limite = meses_ordem[:limite_mes]
        sufixo = str(ano)[-2:]
        return [m for m in meses_limite if f'N_Coletas_{m}_{sufixo}' in df.columns]
    
    @staticmethod
    def _obter_ultimo_mes_fechado(df: pd.DataFrame, dia_corte: int = 5) -> dict:
        """
        Retorna informa√ß√µes sobre o √∫ltimo m√™s fechado dispon√≠vel.
        
        Se estamos antes do dia 'dia_corte' do m√™s atual, considera o m√™s anterior.
        Pode retornar dados de 2024 se n√£o houver dados de 2025.
        
        Args:
            df: DataFrame com os dados de coletas
            dia_corte: Dia do m√™s a partir do qual considera o m√™s atual v√°lido (default: 5)
        
        Returns:
            dict com:
                - 'mes': c√≥digo do m√™s (ex: 'Out')
                - 'ano': ano (ex: 2025)
                - 'sufixo': sufixo de 2 d√≠gitos do ano (ex: '25')
                - 'coluna': nome da coluna no DataFrame (ex: 'N_Coletas_Out_25')
                - 'display': string para exibi√ß√£o (ex: 'Out/2025')
                - 'ano_comparacao': ano a ser usado para compara√ß√µes
        """
        meses_ordem = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
        hoje = pd.Timestamp.today()
        dia_atual = hoje.day
        mes_atual = hoje.month
        ano_atual = hoje.year
        
        # Determinar qual m√™s considerar baseado no dia de corte
        if dia_atual < dia_corte:
            # Se estamos antes do dia de corte, usar o m√™s anterior
            if mes_atual == 1:
                # Se estamos em janeiro, voltar para dezembro do ano anterior
                mes_referencia = 12
                ano_referencia = ano_atual - 1
            else:
                mes_referencia = mes_atual - 1
                ano_referencia = ano_atual
        else:
            # Usar o m√™s atual
            mes_referencia = mes_atual
            ano_referencia = ano_atual
        
        # Tentar encontrar coluna come√ßando pelo ano de refer√™ncia
        mes_codigo = meses_ordem[mes_referencia - 1]
        sufixo_ref = str(ano_referencia)[-2:]
        coluna_ref = f'N_Coletas_{mes_codigo}_{sufixo_ref}'
        
        # Verificar se a coluna existe no DataFrame
        if coluna_ref in df.columns:
            return {
                'mes': mes_codigo,
                'ano': ano_referencia,
                'sufixo': sufixo_ref,
                'coluna': coluna_ref,
                'display': f'{mes_codigo}/{ano_referencia}',
                'ano_comparacao': ano_referencia
            }
        
        # Se n√£o encontrou, tentar buscar o √∫ltimo m√™s dispon√≠vel retroativamente
        # Tentar meses anteriores no mesmo ano
        for mes_idx in range(mes_referencia - 1, 0, -1):
            mes_codigo = meses_ordem[mes_idx - 1]
            coluna_teste = f'N_Coletas_{mes_codigo}_{sufixo_ref}'
            if coluna_teste in df.columns:
                return {
                    'mes': mes_codigo,
                    'ano': ano_referencia,
                    'sufixo': sufixo_ref,
                    'coluna': coluna_teste,
                    'display': f'{mes_codigo}/{ano_referencia}',
                    'ano_comparacao': ano_referencia
                }
        
        # Se n√£o encontrou no ano atual/refer√™ncia, tentar ano anterior
        ano_anterior = ano_referencia - 1
        sufixo_anterior = str(ano_anterior)[-2:]
        for mes_idx in range(12, 0, -1):
            mes_codigo = meses_ordem[mes_idx - 1]
            coluna_teste = f'N_Coletas_{mes_codigo}_{sufixo_anterior}'
            if coluna_teste in df.columns:
                return {
                    'mes': mes_codigo,
                    'ano': ano_anterior,
                    'sufixo': sufixo_anterior,
                    'coluna': coluna_teste,
                    'display': f'{mes_codigo}/{ano_anterior}',
                    'ano_comparacao': ano_anterior
                }
        
        # Se n√£o encontrou nenhuma coluna, retornar vazio
        return {
            'mes': None,
            'ano': None,
            'sufixo': None,
            'coluna': None,
            'display': 'N/A',
            'ano_comparacao': ano_atual
        }
    
    @staticmethod
    def criar_grafico_distribuicao_risco(df: pd.DataFrame):
        if df.empty:
            st.info("üìä Nenhum dado dispon√≠vel para o gr√°fico")
            return
        if 'Risco_Diario' not in df.columns:
            st.warning("‚ö†Ô∏è Coluna 'Risco_Diario' n√£o encontrada nos dados.")
            return
        status_counts = df['Risco_Diario'].value_counts()
        cores_map = {
            'üü¢ Normal': '#16A34A',
            'üü° Aten√ß√£o': '#F59E0B',
            'üü† Moderado': '#FB923C',
            'üî¥ Alto': '#DC2626',
            '‚ö´ Cr√≠tico': '#111827'
        }
        fig = px.pie(
            values=status_counts.values,
            names=status_counts.index,
            title="üìä Distribui√ß√£o de Risco Di√°rio<br><sup>Baseado em dias √∫teis e redu√ß√µes vs. MM7_BR/MM7_UF/MM7_CIDADE</sup>",
            color=status_counts.index,
            color_discrete_map=cores_map
        )
        fig.update_traces(
            textposition='inside',
            textinfo='percent+label+value',
            texttemplate='%{label}<br>%{value} labs<br>(%{percent})',
            hovertemplate='<b>%{label}</b><br>%{value} laborat√≥rios<br>%{percent}<extra></extra>'
        )
        fig.update_layout(
            showlegend=True,
            legend=dict(orientation="h", yanchor="bottom", y=-0.2, xanchor="center", x=0.5),
            height=500,
            margin=dict(l=40, r=40, t=40, b=40)
        )
        st.plotly_chart(fig, width='stretch')
    @staticmethod
    def criar_grafico_top_labs(df: pd.DataFrame, top_n: int = 10):
        if df.empty:
            st.info("üìä Nenhum dado dispon√≠vel para o gr√°fico")
            return
        if 'Risco_Diario' not in df.columns:
            st.warning("‚ö†Ô∏è Coluna 'Risco_Diario' n√£o encontrada nos dados.")
            return
        labs_risco = df[df['Risco_Diario'].isin(['üü† Moderado', 'üî¥ Alto', '‚ö´ Cr√≠tico'])].copy()
        if labs_risco.empty:
            st.info("‚úÖ Nenhum laborat√≥rio em risco encontrado!")
            return
        # Ordenar por maior queda vs MM7 e menor volume do dia
        if 'Delta_MM7' in labs_risco.columns:
            labs_risco = labs_risco.sort_values(['Delta_MM7', 'Vol_Hoje'], ascending=[True, True])
        else:
            labs_risco = labs_risco.sort_values('Vol_Hoje', ascending=True)
        cores_map = {'üü† Moderado': '#FB923C', 'üî¥ Alto': '#DC2626', '‚ö´ Cr√≠tico': '#111827'}
        fig = px.bar(
            labs_risco.head(top_n),
            x='Vol_Hoje',
            y='Nome_Fantasia_PCL',
            orientation='h',
            title=f"üö® Top {top_n} Laborat√≥rios em Risco (Di√°rio)<br><sup>Classifica√ß√£o baseada em dias √∫teis</sup>",
            color='Risco_Diario',
            color_discrete_map=cores_map,
            text='Delta_MM7'
        )
        fig.update_traces(texttemplate='%{text:.1f}% vs MM7', textposition='outside')
        fig.update_layout(
            yaxis={'categoryorder': 'total ascending'},
            xaxis_title="Coletas (√öltimo Dia √ötil)",
            yaxis_title="Laborat√≥rio",
            showlegend=True,
            height=500,
            margin=dict(l=40, r=40, t=40, b=100)
        )
        st.plotly_chart(fig, width='stretch')
    @staticmethod
    def criar_grafico_media_diaria(
        df: pd.DataFrame,
        lab_cnpj: Optional[str] = None,
        lab_nome: Optional[str] = None
    ):
        """Cria gr√°fico de m√©dia di√°ria por m√™s usando dados reais de 2025."""
        if df.empty:
            st.info("üìä Nenhum dado dispon√≠vel para o gr√°fico")
            return
        if not lab_cnpj and not lab_nome:
            st.info("üìä Selecione um laborat√≥rio para visualizar a m√©dia di√°ria")
            return

        df_ref = df
        if lab_cnpj and 'CNPJ_Normalizado' not in df_ref.columns and 'CNPJ_PCL' in df_ref.columns:
            df_ref = df_ref.copy()
            df_ref['CNPJ_Normalizado'] = df_ref['CNPJ_PCL'].apply(DataManager.normalizar_cnpj)

        if lab_cnpj and 'CNPJ_Normalizado' in df_ref.columns:
            lab_data = df_ref[df_ref['CNPJ_Normalizado'] == lab_cnpj]
        else:
            lab_data = df_ref[df_ref['Nome_Fantasia_PCL'] == lab_nome]

        if lab_data.empty:
            st.info("üìä Laborat√≥rio n√£o encontrado")
            return

        lab = lab_data.iloc[0]
        nome_exibicao = lab_nome or lab.get('Nome_Fantasia_PCL') or lab_cnpj
        
        # Verificar se temos dados di√°rios reais de 2025
        if 'Dados_Diarios_2025' not in lab or pd.isna(lab['Dados_Diarios_2025']) or lab['Dados_Diarios_2025'] == '{}':
            st.info("üìä Nenhum dado di√°rio dispon√≠vel para 2025. Use o gerador para atualizar os dados.")
            return
        
        import json
        try:
            # Carregar dados di√°rios reais
            dados_diarios = json.loads(lab['Dados_Diarios_2025'])
        except (json.JSONDecodeError, TypeError):
            st.info("üìä Erro ao carregar dados di√°rios. Use o gerador para atualizar os dados.")
            return
        
        if not dados_diarios:
            st.info("üìä Nenhum dado di√°rio dispon√≠vel para 2025.")
            return
        
        # Calcular m√©dia di√°ria real baseada em dias com coleta
        meses_ordem = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
        medias_diarias = []
        meses_com_dados = []
        
        for mes_key, dias_mes in dados_diarios.items():
            # Extrair m√™s do formato "2025-10"
            try:
                ano, mes_num = mes_key.split('-')
                mes_num = int(mes_num)
                if mes_num >= 1 and mes_num <= 12:
                    mes_nome = meses_ordem[mes_num - 1]
                    
                    # Calcular total de coletas e dias com coleta para este m√™s
                    total_coletas = sum(int(coletas) for coletas in dias_mes.values())
                    dias_com_coleta = len(dias_mes)
                    
                    # M√©dia di√°ria = total de coletas / dias com coleta (n√£o dias do m√™s)
                    if dias_com_coleta > 0:
                        media_diaria = total_coletas / dias_com_coleta
                        medias_diarias.append(media_diaria)
                        meses_com_dados.append(mes_nome)
            except (ValueError, IndexError):
                continue
        
        if not medias_diarias:
            st.info("üìä Nenhuma coleta encontrada nos dados di√°rios de 2025.")
            return
        
        # Criar gr√°fico
        fig = px.bar(
            x=meses_com_dados,
            y=medias_diarias,
            title=f"üìä M√©dia Di√°ria Real por M√™s - {nome_exibicao}<br><sup>Baseado em dias com coleta real</sup>",
            color=medias_diarias,
            color_continuous_scale='Greens',
            text=[f"{val:.1f}" for val in medias_diarias]
        )
     
        fig.update_traces(
            texttemplate='%{text} coletas',
            textposition='outside',
            hovertemplate='<b>M√™s:</b> %{x}<br><b>M√©dia Di√°ria:</b> %{y:.1f} coletas<br><sup>Baseado em dias com coleta real</sup><extra></extra>'
        )
     
        fig.update_layout(
            xaxis_title="M√™s",
            yaxis_title="M√©dia Di√°ria (Coletas)",
            showlegend=False,
            height=600,
            margin=dict(l=60, r=60, t=80, b=80),
            autosize=True,
            font=dict(size=14)
        )
     
        st.plotly_chart(fig, width='stretch')
        
        # Explica√ß√£o metodol√≥gica
        with st.expander("‚ÑπÔ∏è Sobre Esta An√°lise", expanded=False):
            st.markdown(f"""
            **Como √© calculada a m√©dia di√°ria real:**
            1. **Base de dados**: Dados reais de coletas de 2025 por dia
            2. **C√°lculo**: Total de coletas do m√™s √∑ dias com coleta (n√£o dias do m√™s)
            3. **Vantagem**: Mostra a produtividade real nos dias de trabalho
            4. **Exemplo**: Se em Outubro houve 8 coletas em 4 dias diferentes, a m√©dia √© 2.0 coletas/dia
            
            **üí° Insight**: Esta an√°lise mostra:
            - Produtividade real nos dias de coleta
            - Padr√µes de intensidade de trabalho
            - Compara√ß√£o mais precisa entre meses
            """)
    @staticmethod
    def criar_grafico_coletas_por_dia(
        df: pd.DataFrame,
        lab_cnpj: Optional[str] = None,
        lab_nome: Optional[str] = None
    ):
        """Cria gr√°fico de coletas por dia do m√™s usando dados reais de 2025."""
        if df.empty:
            st.info("üìä Nenhum dado dispon√≠vel para o gr√°fico")
            return
        if not lab_cnpj and not lab_nome:
            st.info("üìä Selecione um laborat√≥rio para visualizar as coletas por dia")
            return

        df_ref = df
        if lab_cnpj and 'CNPJ_Normalizado' not in df_ref.columns and 'CNPJ_PCL' in df_ref.columns:
            df_ref = df_ref.copy()
            df_ref['CNPJ_Normalizado'] = df_ref['CNPJ_PCL'].apply(DataManager.normalizar_cnpj)

        if lab_cnpj and 'CNPJ_Normalizado' in df_ref.columns:
            lab_data = df_ref[df_ref['CNPJ_Normalizado'] == lab_cnpj]
        else:
            lab_data = df_ref[df_ref['Nome_Fantasia_PCL'] == lab_nome]

        if lab_data.empty:
            st.info("üìä Laborat√≥rio n√£o encontrado")
            return

        lab = lab_data.iloc[0]
        nome_exibicao = lab_nome or lab.get('Nome_Fantasia_PCL') or lab_cnpj

        # Verificar se temos dados di√°rios reais de 2025
        if 'Dados_Diarios_2025' not in lab or pd.isna(lab['Dados_Diarios_2025']) or lab['Dados_Diarios_2025'] == '{}':
            st.info("üìä Nenhum dado di√°rio dispon√≠vel para 2025. Use o gerador para atualizar os dados.")
            return

        import json
        try:
            # Carregar dados di√°rios reais
            dados_diarios = json.loads(lab['Dados_Diarios_2025'])
        except (json.JSONDecodeError, TypeError):
            st.info("üìä Erro ao carregar dados di√°rios. Use o gerador para atualizar os dados.")
            return

        if not dados_diarios:
            st.info("üìä Nenhum dado di√°rio dispon√≠vel para 2025.")
            return

        # Converter dados para DataFrame
        dados_grafico = []
        meses_ordem = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']

        for mes_key, dias_mes in dados_diarios.items():
            # Extrair m√™s do formato "2025-10"
            try:
                ano, mes_num = mes_key.split('-')
                mes_num = int(mes_num)
                if mes_num >= 1 and mes_num <= 12:
                    mes_nome = meses_ordem[mes_num - 1]

                    # Adicionar apenas dias com coletas reais
                    for dia_str, coletas in dias_mes.items():
                        dia = int(dia_str)
                        if coletas > 0:  # S√≥ mostrar dias com coletas
                            dados_grafico.append({
                                'Dia': dia,
                                'M√™s': mes_nome,
                                'Coletas': int(coletas)
                            })
            except (ValueError, IndexError):
                continue

        if not dados_grafico:
            st.info("üìä Nenhuma coleta encontrada nos dados di√°rios de 2025.")
            return

        df_grafico = pd.DataFrame(dados_grafico)

        # Criar gr√°fico de linha interativo
        fig = px.line(
            df_grafico,
            x='Dia',
            y='Coletas',
            color='M√™s',
            title=f"üìÖ Coletas por Dia √ötil do M√™s - {nome_exibicao}",
            markers=True,
            line_shape='linear'
        )

        # Configurar tooltip personalizado com nome correto do m√™s
        fig.update_traces(
            hovertemplate='<b>Dia:</b> %{x}<br><b>M√™s:</b> %{fullData.name}<br><b>Coletas:</b> %{y:.0f}<extra></extra>'
        )

        fig.update_layout(
            xaxis_title="Dia do M√™s (dias √∫teis dispon√≠veis)",
            yaxis_title="N√∫mero de Coletas (dias √∫teis)",
            xaxis=dict(tickmode='linear', tick0=1, dtick=5),
            legend=dict(
                orientation="h",
                yanchor="bottom",
                y=-0.15,
                xanchor="center",
                x=0.5,
                bgcolor="rgba(255,255,255,0.8)",
                bordercolor="rgba(0,0,0,0.2)",
                borderwidth=1
            ),
            height=600,
            margin=dict(l=60, r=60, t=80, b=120),  # Margem inferior maior para legenda
            autosize=True,
            font=dict(size=14),
            # Tornar o gr√°fico mais interativo
            hovermode='x unified',
            # Melhorar a apar√™ncia das linhas
            plot_bgcolor='rgba(0,0,0,0)',
            paper_bgcolor='rgba(0,0,0,0)'
        )

        # Adicionar anota√ß√£o explicativa (dica persistente)
        fig.add_annotation(
            text="üí° Dica: d√™ duplo clique no m√™s na legenda para focar apenas aquela s√©rie. Clique simples mostra/oculta linhas.",
            xref="paper", yref="paper",
            x=0.5, y=-0.25,
            showarrow=False,
            font=dict(size=12, color="gray"),
            xanchor="center"
        )

        st.plotly_chart(fig, width='stretch')
    @staticmethod
    def criar_grafico_media_dia_semana_novo(
        df: pd.DataFrame,
        lab_cnpj: Optional[str] = None,
        lab_nome: Optional[str] = None,
        filtros: dict = None
    ):
        """NOVA VERS√ÉO - Cria gr√°fico de distribui√ß√£o de coletas por dia da semana usando dados reais de 2025."""
        if df.empty:
            st.info("üìä Nenhum dado dispon√≠vel para o gr√°fico")
            return
        if not lab_cnpj and not lab_nome:
            st.info("üìä Selecione um laborat√≥rio para visualizar a distribui√ß√£o semanal")
            return

        df_ref = df
        if lab_cnpj and 'CNPJ_Normalizado' not in df_ref.columns and 'CNPJ_PCL' in df_ref.columns:
            df_ref = df_ref.copy()
            df_ref['CNPJ_Normalizado'] = df_ref['CNPJ_PCL'].apply(DataManager.normalizar_cnpj)

        if lab_cnpj and 'CNPJ_Normalizado' in df_ref.columns:
            lab_data = df_ref[df_ref['CNPJ_Normalizado'] == lab_cnpj]
        else:
            lab_data = df_ref[df_ref['Nome_Fantasia_PCL'] == lab_nome]

        if lab_data.empty:
            st.info("üìä Laborat√≥rio n√£o encontrado")
            return

        lab = lab_data.iloc[0]
        nome_exibicao = lab_nome or lab.get('Nome_Fantasia_PCL') or lab_cnpj
        
        # Verificar se temos dados semanais reais de 2025
        if 'Dados_Semanais_2025' not in lab or pd.isna(lab['Dados_Semanais_2025']) or lab['Dados_Semanais_2025'] == '{}':
            st.info("üìä Nenhum dado semanal dispon√≠vel para 2025. Use o gerador para atualizar os dados.")
            return
        
        import json
        try:
            dados_semanais = json.loads(lab['Dados_Semanais_2025'])
        except (json.JSONDecodeError, TypeError):
            st.info("üìä Erro ao carregar dados semanais. Use o gerador para atualizar os dados.")
            return
        
        if not dados_semanais:
            st.info("üìä Nenhum dado semanal dispon√≠vel para 2025.")
            return
        
        # NOVA IMPLEMENTA√á√ÉO - Criar dados de forma mais simples e direta
        dias_uteis = ['Segunda', 'Ter√ßa', 'Quarta', 'Quinta', 'Sexta']
        cores_dias = {
            'Segunda': '#6BBF47', 'Ter√ßa': '#ff7f0e', 'Quarta': '#52B54B', 'Quinta': '#d62728',
            'Sexta': '#9467bd'
        }
        
        # Criar lista de dados de forma mais direta
        dados_grafico = []
        total_coletas = 0
        
        for dia in dias_uteis:
            coletas = dados_semanais.get(dia, 0)
            total_coletas += coletas
            dados_grafico.append({
                'dia': dia,
                'coletas': coletas,
                'cor': cores_dias[dia]
            })
        
        max_coletas = max((item['coletas'] for item in dados_grafico), default=0)
        y_axis_max = max_coletas * 1.2 if max_coletas > 0 else 10

        # Calcular percentuais
        for item in dados_grafico:
            if total_coletas > 0:
                item['percentual'] = round((item['coletas'] / total_coletas) * 100, 1)
            else:
                item['percentual'] = 0.0
        
        # CRIAR GR√ÅFICO NOVO DO ZERO
        import plotly.graph_objects as go
        
        fig = go.Figure()
        
        # Adicionar barras uma por uma para ter controle total
        for i, row in enumerate(dados_grafico):
            fig.add_trace(go.Bar(
                x=[row['dia']],
                y=[row['coletas']],
                name=row['dia'],
                marker_color=row['cor'],
                text=[f"{row['coletas']} coletas<br>({row['percentual']:.1f}%)"],
                textposition='outside',
                hovertemplate=f"<b>{row['dia']}</b><br>" +
                             f"Coletas: {row['coletas']}<br>" +
                             f"Percentual: {row['percentual']:.1f}% da semana<extra></extra>",
                showlegend=False
            ))
        
        # Configurar layout
        fig.update_layout(
            title=f"üìÖ Distribui√ß√£o Real de Coletas por Dia √ötil da Semana<br><sup>{nome_exibicao} | Total semanal: {total_coletas} coletas √∫teis</sup>",
            xaxis_title="Dia da Semana (dias √∫teis)",
            yaxis_title="Coletas por Dia √ötil",
            height=600,
            margin=dict(l=60, r=60, t=100, b=80),
            font=dict(size=14),
            title_font_size=18,
            yaxis=dict(range=[0, y_axis_max])
        )
        
        # Adicionar linha de m√©dia di√°ria
        if total_coletas > 0:
            media_diaria = total_coletas / len(dias_uteis)
            fig.add_hline(
                y=media_diaria,
                line_dash="dash",
                line_color="red",
                annotation_text=f"M√©dia por dia √∫til: {media_diaria:.1f} coletas",
                annotation_position="top right"
            )
        
        st.plotly_chart(fig, width='stretch')
        
        # M√©tricas
        col1, col2, col3 = st.columns(3)
        with col1:
            dia_max = max(dados_grafico, key=lambda x: x['coletas'])
            st.metric("üìà Dia Mais Forte", dia_max['dia'], f"{dia_max['coletas']:.0f} coletas")
        with col2:
            dia_min = min(dados_grafico, key=lambda x: x['coletas'])
            st.metric("üìâ Dia Mais Fraco", dia_min['dia'], f"{dia_min['coletas']:.0f} coletas")
        with col3:
            max_coletas = max(item['coletas'] for item in dados_grafico)
            min_coletas = min(item['coletas'] for item in dados_grafico)
            variacao = ((max_coletas - min_coletas) / max_coletas * 100) if max_coletas > 0 else 0
            st.metric("üìä Varia√ß√£o (forte vs fraco)", f"{variacao:.1f}%", "forte vs fraco")
        
        # Debug removido ap√≥s valida√ß√£o dos percentuais

    @staticmethod
    def criar_grafico_media_dia_semana(df: pd.DataFrame, lab_selecionado: str = None, filtros: dict = None):
        """Cria gr√°fico de distribui√ß√£o de coletas por dia da semana usando dados reais de 2025."""
        if df.empty:
            st.info("üìä Nenhum dado dispon√≠vel para o gr√°fico")
            return
        if not lab_selecionado:
            st.info("üìä Selecione um laborat√≥rio para visualizar a distribui√ß√£o semanal")
            return
        lab_data = df[df['Nome_Fantasia_PCL'] == lab_selecionado]
        if not lab_data.empty:
            lab = lab_data.iloc[0]
            
            # Verificar se temos dados semanais reais de 2025
            if 'Dados_Semanais_2025' not in lab or pd.isna(lab['Dados_Semanais_2025']) or lab['Dados_Semanais_2025'] == '{}':
                st.info("üìä Nenhum dado semanal dispon√≠vel para 2025. Use o gerador para atualizar os dados.")
                return
            
            import json
            try:
                # Carregar dados semanais reais
                dados_semanais = json.loads(lab['Dados_Semanais_2025'])
            except (json.JSONDecodeError, TypeError):
                st.info("üìä Erro ao carregar dados semanais. Use o gerador para atualizar os dados.")
                return
            
            if not dados_semanais:
                st.info("üìä Nenhum dado semanal dispon√≠vel para 2025.")
                return
            
            # Converter dados para DataFrame
            dias_uteis = ['Segunda', 'Ter√ßa', 'Quarta', 'Quinta', 'Sexta']
            cores_dias = {
                'Segunda': '#6BBF47', # Verde Synvia
                'Ter√ßa': '#ff7f0e', # Laranja
                'Quarta': '#52B54B', # Verde Synvia Escuro
                'Quinta': '#d62728', # Vermelho
                'Sexta': '#9467bd' # Roxo
            }
            
            dados_semana = []
            total_coletas_semana = 0
            
            for dia in dias_uteis:
                coletas_dia = dados_semanais.get(dia, 0)
                total_coletas_semana += coletas_dia
                dados_semana.append({
                    'Dia_Semana': dia,
                    'Coletas_Reais': coletas_dia,
                    'Cor': cores_dias[dia]
                })
            
            df_semana = pd.DataFrame(dados_semana)
            
            # Calcular percentuais corretos baseados nos dados reais
            if total_coletas_semana > 0:
                df_semana['Percentual'] = (df_semana['Coletas_Reais'] / total_coletas_semana * 100).round(1)
            else:
                df_semana['Percentual'] = 0.0
            # Criar t√≠tulo informativo
            periodo_texto = "dados reais de 2025"
            
            # Calcular m√©dia di√°ria correta (soma das coletas semanais / 5 dias √∫teis)
            media_diaria = total_coletas_semana / len(dias_uteis) if total_coletas_semana > 0 else 0
            
            # Gr√°fico de barras
            max_coletas_semana = df_semana['Coletas_Reais'].max() if not df_semana.empty else 0
            y_axis_max = max_coletas_semana * 1.2 if max_coletas_semana > 0 else 10
            fig = px.bar(
                df_semana,
                x='Dia_Semana',
                y='Coletas_Reais',
                title=f"üìÖ Distribui√ß√£o Real de Coletas por Dia √ötil da Semana<br><sup>{lab_selecionado} | Baseado em: {periodo_texto} | Total semanal: {total_coletas_semana:.0f} coletas √∫teis</sup>",
                color='Dia_Semana',
                color_discrete_map=cores_dias,
                text='Coletas_Reais'
            )
            # Usar hovertemplate com c√°lculo direto do percentual
            fig.update_traces(
                texttemplate='%{text:.0f} coletas<br>(%{customdata:.1f}%)',
                textposition='outside',
                customdata=df_semana['Percentual'],
                hovertemplate='<b>%{x}</b><br>Coletas: %{y:.0f}<br>Percentual: %{customdata:.1f}% da semana<extra></extra>'
            )
            fig.update_layout(
                xaxis_title="Dia da Semana (dias √∫teis)",
                yaxis_title="Coletas por Dia √ötil",
                showlegend=False,
                coloraxis_showscale=False,
                height=700,  # Aumentado significativamente para destaque
                margin=dict(l=60, r=60, t=100, b=80),  # Margens aumentadas
                autosize=True,  # Responsivo
                font=dict(size=14),  # Fonte maior para melhor legibilidade
                title_font_size=18,  # T√≠tulo maior
                yaxis=dict(range=[0, y_axis_max])
            )
            # Adicionar linha de refer√™ncia da m√©dia di√°ria
            if media_diaria > 0:
                fig.add_hline(
                    y=media_diaria,
                    line_dash="dash",
                    line_color="red",
                    annotation_text=f"M√©dia por dia √∫til: {media_diaria:.1f} coletas",
                    annotation_position="top right"
                )
            st.plotly_chart(fig, width='stretch')
            # M√©tricas adicionais
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric(
                    "üìà Dia Mais Forte",
                    df_semana.loc[df_semana['Coletas_Reais'].idxmax(), 'Dia_Semana'],
                    f"{df_semana['Coletas_Reais'].max():.0f} coletas"
                )
            with col2:
                st.metric(
                    "üìâ Dia Mais Fraco",
                    df_semana.loc[df_semana['Coletas_Reais'].idxmin(), 'Dia_Semana'],
                    f"{df_semana['Coletas_Reais'].min():.0f} coletas"
                )
            with col3:
                variacao_semanal = (df_semana['Coletas_Reais'].max() - df_semana['Coletas_Reais'].min()) / df_semana['Coletas_Reais'].max() * 100 if df_semana['Coletas_Reais'].max() > 0 else 0
                st.metric(
                    "üìä Varia√ß√£o (forte vs fraco)",
                    f"{variacao_semanal:.1f}%",
                    "forte vs fraco"
                )
            # Explica√ß√£o metodol√≥gica
            with st.expander("‚ÑπÔ∏è Sobre Esta An√°lise", expanded=False):
                st.markdown(f"""
                **Como √© calculada a distribui√ß√£o semanal:**
                1. **Base de dados**: Dados reais de coletas de 2025 ({periodo_texto})
                2. **Distribui√ß√£o real**: Baseada nas datas exatas das coletas (createdAt)
                   - **Total semanal**: {total_coletas_semana:.0f} coletas √∫teis
                   - **Percentuais**: Calculados considerando apenas dias √∫teis
                3. **M√©dia di√°ria √∫til**: {media_diaria:.1f} coletas (total semanal √∑ {len(dias_uteis)})
                **üí° Insight**: Esta an√°lise mostra:
                - Padr√µes reais de coleta do laborat√≥rio
                - Dias com maior/menor movimento baseado em dados hist√≥ricos
                - Oportunidades de otimiza√ß√£o de recursos
                **‚ö†Ô∏è Importante**: Estes s√£o valores estimados baseados em padr√µes hist√≥ricos.
                Dados di√°rios reais forneceriam an√°lise mais precisa.
                """)
    @staticmethod
    def criar_grafico_controle_br_uf_cidade(
        df: pd.DataFrame,
        df_filtrado: Optional[pd.DataFrame] = None,
        lab_cnpj: Optional[str] = None,
        lab_nome: Optional[str] = None,
        usar_mm30: bool = False
    ):
        """
        Cria gr√°fico comparativo de controle BR √ó UF √ó Cidade √ó Lab atual.
        Mostra s√©ries temporais de MM7 ou MM30 (dias √∫teis) para cada contexto.
        
        Args:
            df: DataFrame completo (para calcular contextos BR/UF/Cidade)
            df_filtrado: DataFrame filtrado (para s√©rie atual quando n√£o h√° lab espec√≠fico)
            lab_cnpj: CNPJ do laborat√≥rio espec√≠fico (opcional)
            lab_nome: Nome do laborat√≥rio espec√≠fico (opcional)
            usar_mm30: Se True, usa MM30; se False, usa MM7
        """
        if df.empty:
            st.info("üìä Nenhum dado dispon√≠vel para o gr√°fico de controle")
            return
        
        import json
        from pandas.tseries.offsets import BDay
        
        # Determinar contexto atual (lab espec√≠fico ou conjunto filtrado)
        serie_atual = pd.Series(dtype="float")
        nome_serie_atual = "Conjunto Filtrado"
        lab_data = pd.DataFrame()
        
        if lab_cnpj or lab_nome:
            # Buscar lab espec√≠fico
            df_ref = df.copy()
            if lab_cnpj and 'CNPJ_Normalizado' not in df_ref.columns and 'CNPJ_PCL' in df_ref.columns:
                df_ref['CNPJ_Normalizado'] = df_ref['CNPJ_PCL'].apply(DataManager.normalizar_cnpj)
            
            if lab_cnpj and 'CNPJ_Normalizado' in df_ref.columns:
                lab_data = df_ref[df_ref['CNPJ_Normalizado'] == lab_cnpj]
            else:
                lab_data = df_ref[df_ref['Nome_Fantasia_PCL'] == lab_nome]
            
            if not lab_data.empty:
                lab = lab_data.iloc[0]
                nome_serie_atual = lab.get('Nome_Fantasia_PCL', lab_cnpj or lab_nome)
                if 'Dados_Diarios_2025' in lab and pd.notna(lab['Dados_Diarios_2025']):
                    serie_atual = RiskEngine._serie_diaria_from_json(lab['Dados_Diarios_2025'])
        else:
            # Agregar s√©rie do conjunto filtrado (usar df_filtrado se dispon√≠vel, sen√£o df)
            df_para_serie = df_filtrado if df_filtrado is not None and not df_filtrado.empty else df
            todas_series = []
            for _, row in df_para_serie.iterrows():
                if 'Dados_Diarios_2025' in row and pd.notna(row['Dados_Diarios_2025']):
                    s = RiskEngine._serie_diaria_from_json(row['Dados_Diarios_2025'])
                    if not s.empty:
                        todas_series.append(s)
            
            if todas_series:
                # Agregar todas as s√©ries por data
                todas_datas = set()
                for s in todas_series:
                    todas_datas.update(s.index)
                
                serie_agregada = pd.Series(index=sorted(todas_datas), dtype="float")
                for data in serie_agregada.index:
                    total = sum(s.get(data, 0) for s in todas_series)
                    serie_agregada[data] = total
                
                serie_atual = serie_agregada
        
        # Agregar s√©ries por contexto (BR, UF, Cidade)
        def agregar_por_contexto(df_contexto: pd.DataFrame) -> pd.Series:
            """Agrega coletas di√°rias por contexto."""
            todas_series = []
            for _, row in df_contexto.iterrows():
                if 'Dados_Diarios_2025' in row and pd.notna(row['Dados_Diarios_2025']):
                    s = RiskEngine._serie_diaria_from_json(row['Dados_Diarios_2025'])
                    if not s.empty:
                        todas_series.append(s)
            
            if not todas_series:
                return pd.Series(dtype="float")
            
            # Agregar todas as s√©ries por data
            todas_datas = set()
            for s in todas_series:
                todas_datas.update(s.index)
            
            serie_agregada = pd.Series(index=sorted(todas_datas), dtype="float")
            for data in serie_agregada.index:
                total = sum(s.get(data, 0) for s in todas_series)
                serie_agregada[data] = total
            
            return serie_agregada
        
        # Agregar por BR (todos os labs do DataFrame completo)
        serie_br = agregar_por_contexto(df)
        
        # Agregar por UF (se temos info de UF)
        serie_uf = pd.Series(dtype="float")
        uf_nome = ""
        if lab_cnpj or lab_nome:
            if not lab_data.empty and 'Estado' in lab_data.columns:
                uf_nome = lab_data.iloc[0]['Estado']
                if pd.notna(uf_nome) and uf_nome:
                    df_uf = df[df['Estado'] == uf_nome]
                    serie_uf = agregar_por_contexto(df_uf)
        else:
            # Para conjunto filtrado, usar UF do primeiro lab (se dispon√≠vel)
            if not df.empty and 'Estado' in df.columns:
                uf_nome = df.iloc[0]['Estado']
                if pd.notna(uf_nome) and uf_nome:
                    df_uf = df[df['Estado'] == uf_nome]
                    serie_uf = agregar_por_contexto(df_uf)
        
        # Agregar por Cidade (se temos info de Cidade)
        serie_cidade = pd.Series(dtype="float")
        cidade_nome = ""
        if lab_cnpj or lab_nome:
            if not lab_data.empty and 'Cidade' in lab_data.columns:
                cidade_nome = lab_data.iloc[0]['Cidade']
                if pd.notna(cidade_nome) and cidade_nome:
                    df_cidade = df[df['Cidade'] == cidade_nome]
                    serie_cidade = agregar_por_contexto(df_cidade)
        else:
            # Para conjunto filtrado, usar Cidade do primeiro lab (se dispon√≠vel)
            if not df.empty and 'Cidade' in df.columns:
                cidade_nome = df.iloc[0]['Cidade']
                if pd.notna(cidade_nome) and cidade_nome:
                    df_cidade = df[df['Cidade'] == cidade_nome]
                    serie_cidade = agregar_por_contexto(df_cidade)
        
        # Calcular m√©dias m√≥veis ao longo do tempo (apenas dias √∫teis)
        def calcular_mm_serie(serie: pd.Series, janela: int) -> pd.Series:
            """Calcula m√©dia m√≥vel de janela dias √∫teis ao longo da s√©rie."""
            if serie.empty:
                return pd.Series(dtype="float")
            
            # Garantir que temos apenas dias √∫teis
            serie = serie.sort_index()
            serie_uteis = serie[serie.index.weekday < 5]  # Segunda=0 a Sexta=4
            
            if serie_uteis.empty:
                return pd.Series(dtype="float")
            
            # Calcular MM ao longo do tempo
            mm_serie = serie_uteis.rolling(window=janela, min_periods=1).mean()
            return mm_serie
        
        janela = 30 if usar_mm30 else 7
        mm_label = "MM30" if usar_mm30 else "MM7"
        
        # Calcular MMs para cada contexto
        mm_br = calcular_mm_serie(serie_br, janela)
        mm_uf = calcular_mm_serie(serie_uf, janela)
        mm_cidade = calcular_mm_serie(serie_cidade, janela)
        mm_atual = calcular_mm_serie(serie_atual, janela)
        
        # Preparar dados para o gr√°fico
        dados_grafico = []
        
        # Adicionar s√©rie BR
        for data, valor in mm_br.items():
            dados_grafico.append({
                'Data': data,
                'Valor': valor,
                'Serie': 'üáßüá∑ MM7_BR' if not usar_mm30 else 'üáßüá∑ MM30_BR'
            })
        
        # Adicionar s√©rie UF
        if not mm_uf.empty and uf_nome:
            uf_label = f"üìç MM7_UF ({uf_nome})" if not usar_mm30 else f"üìç MM30_UF ({uf_nome})"
            for data, valor in mm_uf.items():
                dados_grafico.append({
                    'Data': data,
                    'Valor': valor,
                    'Serie': uf_label
                })
        
        # Adicionar s√©rie Cidade
        if not mm_cidade.empty and cidade_nome:
            cidade_label = f"üèôÔ∏è MM7_CIDADE ({cidade_nome})" if not usar_mm30 else f"üèôÔ∏è MM30_CIDADE ({cidade_nome})"
            for data, valor in mm_cidade.items():
                dados_grafico.append({
                    'Data': data,
                    'Valor': valor,
                    'Serie': cidade_label
                })
        
        # Adicionar s√©rie atual
        if not mm_atual.empty:
            atual_label = f"üìä {nome_serie_atual} ({mm_label})"
            for data, valor in mm_atual.items():
                dados_grafico.append({
                    'Data': data,
                    'Valor': valor,
                    'Serie': atual_label
                })
        
        if not dados_grafico:
            st.info("üìä Nenhum dado dispon√≠vel para gerar o gr√°fico de controle")
            return
        
        df_grafico = pd.DataFrame(dados_grafico)
        
        # Criar gr√°fico de linha
        cores_map = {
            'üáßüá∑ MM7_BR': '#DC2626',
            'üáßüá∑ MM30_BR': '#DC2626',
            'üìç MM7_UF': '#3B82F6',
            'üìç MM30_UF': '#3B82F6',
            'üèôÔ∏è MM7_CIDADE': '#10B981',
            'üèôÔ∏è MM30_CIDADE': '#10B981'
        }
        
        # Adicionar cores din√¢micas para s√©rie atual
        for serie_nome in df_grafico['Serie'].unique():
            if serie_nome not in cores_map:
                cores_map[serie_nome] = '#6BBF47'  # Cor padr√£o verde
        
        fig = px.line(
            df_grafico,
            x='Data',
            y='Valor',
            color='Serie',
            title=f"üìä Controle BR √ó UF √ó Cidade √ó {nome_serie_atual}<br><sup>{mm_label} - Apenas dias √∫teis</sup>",
            markers=True,
            line_shape='linear',
            color_discrete_map=cores_map
        )
        
        fig.update_traces(
            hovertemplate='<b>%{fullData.name}</b><br>Data: %{x|%d/%m/%Y}<br>Valor: %{y:.2f}<extra></extra>',
            line=dict(width=2.5)
        )
        
        fig.update_layout(
            xaxis_title="Data (dias √∫teis)",
            yaxis_title=f"M√©dia M√≥vel ({mm_label})",
            legend=dict(
                orientation="h",
                yanchor="bottom",
                y=-0.2,
                xanchor="center",
                x=0.5,
                bgcolor="rgba(255,255,255,0.9)",
                bordercolor="rgba(0,0,0,0.2)",
                borderwidth=1
            ),
            height=600,
            margin=dict(l=60, r=60, t=100, b=120),
            hovermode='x unified',
            plot_bgcolor='rgba(0,0,0,0)',
            paper_bgcolor='rgba(0,0,0,0)',
            font=dict(size=12)
        )
        
        st.plotly_chart(fig, width='stretch')
    
    @staticmethod
    def criar_grafico_evolucao_mensal(
        df: pd.DataFrame,
        lab_cnpj: Optional[str] = None,
        lab_nome: Optional[str] = None,
        chart_key: str = "default"
    ):
        """Cria gr√°fico de evolu√ß√£o mensal - Atualizado com corre√ß√µes de diferen√ßa 2024/2025."""
        if df.empty:
            st.info("üìä Nenhum dado dispon√≠vel para o gr√°fico")
            return
        meses = ChartManager._meses_ate_hoje(df, 2025)
        if not meses:
            st.info("üìä Nenhum m√™s dispon√≠vel at√© a data atual")
            return
        colunas_meses = [f'N_Coletas_{mes}_25' for mes in meses]
        if lab_cnpj or lab_nome:
            # Gr√°fico para laborat√≥rio espec√≠fico
            df_ref = df
            if lab_cnpj and 'CNPJ_Normalizado' not in df_ref.columns and 'CNPJ_PCL' in df_ref.columns:
                df_ref = df_ref.copy()
                df_ref['CNPJ_Normalizado'] = df_ref['CNPJ_PCL'].apply(DataManager.normalizar_cnpj)

            if lab_cnpj and 'CNPJ_Normalizado' in df_ref.columns:
                lab_data = df_ref[df_ref['CNPJ_Normalizado'] == lab_cnpj]
            else:
                lab_data = df_ref[df_ref['Nome_Fantasia_PCL'] == lab_nome]
            if not lab_data.empty:
                lab = lab_data.iloc[0]
                nome_exibicao = lab_nome or lab.get('Nome_Fantasia_PCL') or lab_cnpj
                valores_2025 = [lab.get(col, 0) for col in colunas_meses]
             
                # Dados 2024 (mesmos meses para compara√ß√£o direta)
                colunas_2024 = [f'N_Coletas_{mes}_24' for mes in meses]
                valores_2024 = [lab.get(col, 0) for col in colunas_2024]
             
                # Calcular m√©dias - Corrigido agrupamento temporal
                media_2025 = sum(valores_2025) / len(valores_2025) if valores_2025 else 0
                media_2024 = sum(valores_2024) / len(valores_2024) if valores_2024 else 0
             
                # Criar DataFrame para o gr√°fico
                df_grafico = pd.DataFrame({
                    'M√™s': meses,
                    '2025': valores_2025,
                    '2024': valores_2024,
                    'M√©dia 2025': [media_2025] * len(meses),
                    'M√©dia 2024': [media_2024] * len(meses)
                })
             
                # Criar gr√°fico com m√∫ltiplas linhas
                fig = px.line(
                    df_grafico,
                    x='M√™s',
                    y=['2025', '2024', 'M√©dia 2025', 'M√©dia 2024'],
                    title=f"üìà Evolu√ß√£o Mensal - {nome_exibicao}",
                    markers=True,
                    line_shape='spline'
                )
             
                # Personalizar cores e estilos
                fig.update_traces(
                    mode='lines+markers',
                    hovertemplate='<b>M√™s:</b> %{x}<br><b>Coletas:</b> %{y}<extra></extra>'
                )
             
                # Cores personalizadas
                fig.data[0].line.color = '#6BBF47' # Verde Synvia para 2025
                fig.data[1].line.color = '#ff7f0e' # Laranja para 2024
                fig.data[2].line.color = '#6BBF47' # Verde Synvia para m√©dia 2025
                fig.data[2].line.dash = 'dash'
                fig.data[3].line.color = '#ff7f0e' # Laranja para m√©dia 2024
                fig.data[3].line.dash = 'dash'
                # Ajustar textos de hover para diferenciar coletas x m√©dias
                fig.data[0].hovertemplate = '<b>M√™s:</b> %{x}<br><b>Coletas 2025:</b> %{y:.0f}<extra></extra>'
                fig.data[1].hovertemplate = '<b>M√™s:</b> %{x}<br><b>Coletas 2024:</b> %{y:.0f}<extra></extra>'
                fig.data[2].hovertemplate = '<b>M√™s:</b> %{x}<br><b>M√©dia 2025:</b> %{y:.1f}<extra></extra>'
                fig.data[3].hovertemplate = '<b>M√™s:</b> %{x}<br><b>M√©dia 2024:</b> %{y:.1f}<extra></extra>'
                fig.update_layout(
                    xaxis_title="M√™s",
                    yaxis_title="N√∫mero de Coletas",
                    hovermode='x unified',
                    legend=dict(
                        orientation="h",
                        yanchor="bottom",
                        y=-0.15,
                        xanchor="center",
                        x=0.5
                    ),
                    height=600,  # Aumentado conforme solicitado
                    margin=dict(l=60, r=60, t=60, b=80),  # Margens aumentadas para evitar cortes
                    autosize=True,  # Responsivo
                    showlegend=True
                )
                st.plotly_chart(fig, width='stretch', key=f"evolucao_mensal_lab_{chart_key}")
        else:
            # Gr√°fico agregado
            valores_agregados = [df[col].sum() for col in colunas_meses]
            fig = px.line(
                x=meses,
                y=valores_agregados,
                title="üìà Evolu√ß√£o Mensal Agregada (2025)",
                markers=True,
                line_shape='spline'
            )
            fig.update_traces(
                mode='lines+markers+text',
                text=valores_agregados,
                textposition="top center",
                hovertemplate='<b>M√™s:</b> %{x}<br><b>Total Coletas:</b> %{y}<extra></extra>'
            )
            fig.update_layout(
                xaxis_title="M√™s",
                yaxis_title="Total de Coletas",
                hovermode='x unified',
                height=600,  # Aumentado conforme solicitado
                margin=dict(l=60, r=60, t=60, b=80),  # Margens aumentadas
                autosize=True  # Responsivo
            )
            st.plotly_chart(fig, width='stretch', key=f"evolucao_mensal_agregado_{chart_key}")
class UIManager:
    """Gerenciador da interface do usu√°rio - Atualizado com tabs."""
    @staticmethod
    def renderizar_header():
        """Renderiza o cabe√ßalho principal."""
        st.markdown("""
        <div class="main-header">
            <h1>üìä Syntox Churn</h1>
            <p>Dashboard profissional para an√°lise de reten√ß√£o de laborat√≥rios</p>
        </div>
        """, unsafe_allow_html=True)
    @staticmethod
    def renderizar_kpi_cards(metrics: KPIMetrics):
        """Renderiza cards de KPIs modernos V2 - Otimizado para eliminar redund√¢ncias."""
        # Primeira linha: M√©tricas principais
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            # Card 1: Labs monitorados com breakdown de risco
            risco_breakdown = []
            if metrics.labs_moderado_count > 0:
                risco_breakdown.append(f"üü† {metrics.labs_moderado_count:,}")
            if metrics.labs_alto_count > 0:
                risco_breakdown.append(f"üî¥ {metrics.labs_alto_count:,}")
            if metrics.labs_critico_count > 0:
                risco_breakdown.append(f"‚ö´ {metrics.labs_critico_count:,}")
            
            risco_text = " | ".join(risco_breakdown) if risco_breakdown else "Nenhum em risco"
            recuperacao_text = f"üîÑ Recupera√ß√£o: {metrics.labs_recuperando:,}" if metrics.labs_recuperando > 0 else ""
            delta_text = f"{risco_text}" + (f" | {recuperacao_text}" if recuperacao_text else "")
            
            st.markdown(f"""
            <div class="metric-card" title="Total de laborat√≥rios ativos nos √∫ltimos 90 dias. Breakdown por n√≠vel de risco (üü† Moderado, üî¥ Alto, ‚ö´ Cr√≠tico). Recupera√ß√£o: laborat√≥rios que voltaram a operar acima da MM7 ap√≥s per√≠odo de queda.">
                <div class="metric-value">{metrics.total_labs:,}</div>
                <div class="metric-label">Labs Monitorados (‚â§90 dias)</div>
                <div class="metric-delta" style="font-size:0.85rem;">{delta_text}</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            # Card 2: Coletas do dia
            delta_text = f"D-1: {metrics.vol_d1_total:,} | YTD: {metrics.total_coletas:,}"
            st.markdown(f"""
            <div class="metric-card" title="Total de coletas registradas no √∫ltimo dia √∫til. D-1: volume de coletas do dia √∫til anterior. YTD (Year To Date): soma total de coletas em 2025 at√© o momento.">
                <div class="metric-value">{metrics.vol_hoje_total:,}</div>
                <div class="metric-label">Coletas Hoje</div>
                <div class="metric-delta">{delta_text}</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            # Card 3: Risco cr√≠tico consolidado (sem redund√¢ncia)
            risco_alto_critico = metrics.labs_alto_count + metrics.labs_critico_count
            if risco_alto_critico > 0:
                if metrics.labs_alto_count > 0 and metrics.labs_critico_count > 0:
                    delta_text = f"üî¥ {metrics.labs_alto_count:,} | ‚ö´ {metrics.labs_critico_count:,}"
                elif metrics.labs_critico_count > 0:
                    delta_text = f"‚ö´ {metrics.labs_critico_count:,} cr√≠ticos"
                else:
                    delta_text = f"üî¥ {metrics.labs_alto_count:,} alto"
            else:
                delta_text = "‚úÖ Nenhum"
            
            st.markdown(f"""
            <div class="metric-card" title="Laborat√≥rios em risco alto (üî¥) ou cr√≠tico (‚ö´) pela r√©gua baseada em dias √∫teis e redu√ß√µes vs. MM7_BR/MM7_UF/MM7_CIDADE.">
                <div class="metric-value">{risco_alto_critico:,}</div>
                <div class="metric-label">Risco Alto + Cr√≠tico</div>
                <div class="metric-delta">{delta_text}</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col4:
            # Card 4: Sem coleta 48h com ativos 7D
            delta_class = "positive" if metrics.ativos_7d >= 80 else "negative"
            ativos_label = f"Ativos 7D: {metrics.ativos_7d:.1f}% ({metrics.ativos_7d_count}/{metrics.total_labs})" if metrics.total_labs else "Ativos 7D: --"
            st.markdown(f"""
            <div class="metric-card" title="Laborat√≥rios com dois dias √∫teis consecutivos sem registrar coletas (Vol_Hoje = 0 e Vol_D1 = 0). Ativos 7D: percentual de laborat√≥rios com pelo menos uma coleta nos √∫ltimos 7 dias √∫teis.">
                <div class="metric-value">{metrics.labs_sem_coleta_48h:,}</div>
                <div class="metric-label">Sem Coleta (48h)</div>
                <div class="metric-delta {delta_class}">{ativos_label}</div>
            </div>
            """, unsafe_allow_html=True)
        
        # Segunda linha: Distribui√ß√£o de risco e compara√ß√£o MM7
        col5, col6 = st.columns([1.5, 1])
        
        with col5:
            # Card 5: Distribui√ß√£o completa de risco (consolidado)
            total_risco = metrics.labs_moderado_count + metrics.labs_alto_count + metrics.labs_critico_count
            risco_dist_text = f"üü¢ {metrics.labs_normal_count:,} | üü° {metrics.labs_atencao_count:,} | üü† {metrics.labs_moderado_count:,} | üî¥ {metrics.labs_alto_count:,} | ‚ö´ {metrics.labs_critico_count:,}"
            st.markdown(f"""
            <div class="metric-card" title="Distribui√ß√£o completa de risco di√°rio calculada sobre dias √∫teis e redu√ß√µes m√°ximas vs. MM7 dos contextos BR/UF/Cidade.">
                <div class="metric-value" style="font-size:1.3rem; margin-bottom:0.3rem;">Distribui√ß√£o de Risco</div>
                <div class="metric-delta" style="display:flex; flex-wrap:wrap; gap:0.4rem; font-weight:600; font-size:0.9rem; justify-content:center;">
                    <span>üü¢ {metrics.labs_normal_count:,}</span>
                    <span>üü° {metrics.labs_atencao_count:,}</span>
                    <span>üü† {metrics.labs_moderado_count:,}</span>
                    <span>üî¥ {metrics.labs_alto_count:,}</span>
                    <span>‚ö´ {metrics.labs_critico_count:,}</span>
                </div>
                <div class="metric-label" style="margin-top:0.5rem;">R√©gua de risco (dias √∫teis)</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col6:
            # Card 6: Compara√ß√£o MM7_BR vs MM7_UF (consolidado)
            mm7_br_pct = f"{metrics.labs_abaixo_mm7_br_pct:.1f}%" if metrics.total_labs else "--"
            mm7_uf_pct = f"{metrics.labs_abaixo_mm7_uf_pct:.1f}%" if metrics.total_labs else "--"
            delta_text = f"BR: {mm7_br_pct} | UF: {mm7_uf_pct}"
            
            st.markdown(f"""
            <div class="metric-card" title="Laborat√≥rios abaixo da m√©dia m√≥vel: MM7_BR (nacional) e MM7_UF (estadual), ambas constru√≠das apenas com dias √∫teis.">
                <div class="metric-value" style="font-size:1.3rem; margin-bottom:0.3rem;">Abaixo da MM7</div>
                <div class="metric-delta" style="font-size:0.95rem; font-weight:600; margin:0.5rem 0;">
                    <div>üáßüá∑ BR: {metrics.labs_abaixo_mm7_br:,} ({mm7_br_pct})</div>
                    <div>üìç UF: {metrics.labs_abaixo_mm7_uf:,} ({mm7_uf_pct})</div>
                </div>
                <div class="metric-label">Compara√ß√£o Nacional vs Estadual</div>
            </div>
            """, unsafe_allow_html=True)
class MetricasAvancadas:
    """Classe para m√©tricas avan√ßadas de laborat√≥rios - Atualizado organiza√ß√£o e comparativos."""
 
    @staticmethod
    def calcular_metricas_lab(
        df: pd.DataFrame,
        lab_cnpj: Optional[str] = None,
        lab_nome: Optional[str] = None
    ) -> dict:
        """Calcula m√©tricas avan√ßadas para um laborat√≥rio espec√≠fico - Atualizado score."""

        df_ref = df
        if lab_cnpj and 'CNPJ_Normalizado' not in df_ref.columns and 'CNPJ_PCL' in df_ref.columns:
            df_ref = df_ref.copy()
            df_ref['CNPJ_Normalizado'] = df_ref['CNPJ_PCL'].apply(DataManager.normalizar_cnpj)

        lab_data = pd.DataFrame()
        if lab_cnpj and 'CNPJ_Normalizado' in df_ref.columns:
            lab_data = df_ref[df_ref['CNPJ_Normalizado'] == lab_cnpj]
        if lab_data.empty and lab_nome:
            lab_data = df_ref[df_ref['Nome_Fantasia_PCL'] == lab_nome]

        if lab_data.empty:
            return {}

        lab = lab_data.iloc[0]
     
        # Total de coletas 2025 (at√© o m√™s atual)
        meses_2025 = ChartManager._meses_ate_hoje(df, 2025)
        colunas_2025 = [f'N_Coletas_{mes}_25' for mes in meses_2025]
        total_coletas_2025 = sum(lab.get(col, 0) for col in colunas_2025)
     
        # M√©dia dos √∫ltimos 3 meses (din√¢mico)
        if len(meses_2025) >= 3:
            ultimos_3_meses = meses_2025[-3:]
        else:
            ultimos_3_meses = meses_2025
        colunas_3_meses = [f'N_Coletas_{mes}_25' for mes in ultimos_3_meses]
        media_3_meses = sum(lab.get(col, 0) for col in colunas_3_meses) / len(colunas_3_meses) if colunas_3_meses else 0
     
        # M√©dia di√°ria (√∫ltimos 3 meses)
        dias_3_meses = 90 # Aproximadamente 3 meses
        media_diaria = media_3_meses / 30 if media_3_meses > 0 else 0
     
        # Agudo (7 dias) - coletas nos √∫ltimos 7 dias
        dias_sem_coleta = lab.get('Dias_Sem_Coleta', 0)
        agudo = "Ativo" if dias_sem_coleta <= 7 else "Inativo"
     
        # Cr√¥nico (fechamentos mensais) - baseado na varia√ß√£o
        variacao = lab.get('Variacao_Percentual', 0)
        if variacao > 20:
            cronico = "Crescimento"
        elif variacao < -20:
            cronico = "Decl√≠nio"
        else:
            cronico = "Est√°vel"

        vol_hoje = lab.get('Vol_Hoje', 0)
        vol_hoje = int(vol_hoje) if pd.notna(vol_hoje) else 0
        vol_d1 = lab.get('Vol_D1', 0)
        vol_d1 = int(vol_d1) if pd.notna(vol_d1) else 0
        delta_mm7_val = lab.get('Delta_MM7', None)
        delta_mm7 = round(float(delta_mm7_val), 1) if pd.notna(delta_mm7_val) else None
        delta_d1_val = lab.get('Delta_D1', None)
        delta_d1 = round(float(delta_d1_val), 1) if pd.notna(delta_d1_val) else None
        mm7_val = lab.get('MM7', None)
        mm30_val = lab.get('MM30', None)
        delta_mm30_val = lab.get('Delta_MM30', None)
        mm7 = round(float(mm7_val), 1) if pd.notna(mm7_val) else None
        mm30 = round(float(mm30_val), 1) if pd.notna(mm30_val) else None
        delta_mm30 = round(float(delta_mm30_val), 1) if pd.notna(delta_mm30_val) else None
        risco_diario = lab.get('Risco_Diario', 'N/A')
        if pd.isna(risco_diario):
            risco_diario = 'N/A'
     
        return {
            'total_coletas': int(total_coletas_2025),
            'media_3_meses': round(media_3_meses, 1),
            'media_diaria': round(media_diaria, 1),
            'vol_hoje': vol_hoje,
            'vol_d1': vol_d1,
            'delta_mm7': delta_mm7,
            'delta_d1': delta_d1,
            'mm7': mm7,
            'mm30': mm30,
            'delta_mm30': delta_mm30,
            'agudo': agudo,
            'cronico': cronico,
            'dias_sem_coleta': int(dias_sem_coleta),
            'variacao_percentual': round(variacao, 1),
            'risco_diario': risco_diario
        }
    @staticmethod
    def calcular_metricas_evolucao(
        df: pd.DataFrame,
        lab_cnpj: Optional[str] = None,
        lab_nome: Optional[str] = None
    ) -> dict:
        """Calcula m√©tricas de evolu√ß√£o e comparativos para um laborat√≥rio espec√≠fico - Atualizado organiza√ß√£o e comparativo."""

        df_ref = df
        if lab_cnpj and 'CNPJ_Normalizado' not in df_ref.columns and 'CNPJ_PCL' in df_ref.columns:
            df_ref = df_ref.copy()
            df_ref['CNPJ_Normalizado'] = df_ref['CNPJ_PCL'].apply(DataManager.normalizar_cnpj)

        lab_data = pd.DataFrame()
        if lab_cnpj and 'CNPJ_Normalizado' in df_ref.columns:
            lab_data = df_ref[df_ref['CNPJ_Normalizado'] == lab_cnpj]
        if lab_data.empty and lab_nome:
            lab_data = df_ref[df_ref['Nome_Fantasia_PCL'] == lab_nome]

        if lab_data.empty:
            return {}

        lab = lab_data.iloc[0]
        # Total de coletas 2024 (todos os meses dispon√≠veis)
        meses_2024 = ChartManager._meses_ate_hoje(df, 2024)
        colunas_2024 = [f'N_Coletas_{mes}_24' for mes in meses_2024]
        total_coletas_2024 = sum(lab.get(col, 0) for col in colunas_2024)
        # Total de coletas 2025 (at√© o m√™s atual)
        meses_2025 = ChartManager._meses_ate_hoje(df, 2025)
        colunas_2025 = [f'N_Coletas_{mes}_25' for mes in meses_2025]
        total_coletas_2025 = sum(lab.get(col, 0) for col in colunas_2025)
        # M√©dia de 2024
        media_2024 = total_coletas_2024 / len(colunas_2024) if colunas_2024 else 0
        # M√©dia de 2025
        media_2025 = total_coletas_2025 / len(colunas_2025) if colunas_2025 else 0
        # √öltimo m√™s fechado (usando l√≥gica de dia de corte)
        info_ultimo_mes = ChartManager._obter_ultimo_mes_fechado(df)
        media_ultimo_mes = lab.get(info_ultimo_mes['coluna'], 0) if info_ultimo_mes['coluna'] else 0
        
        # M√°xima hist√≥rica 2024
        max_2024 = max(lab.get(col, 0) for col in colunas_2024) if colunas_2024 else 0
        # M√°xima hist√≥rica 2025
        max_2025 = max(lab.get(col, 0) for col in colunas_2025) if colunas_2025 else 0
        
        # Determinar m√©tricas de compara√ß√£o baseado no ano do √∫ltimo m√™s
        ano_comparacao = info_ultimo_mes['ano_comparacao']
        if ano_comparacao == 2024:
            # Se o √∫ltimo m√™s fechado √© de 2024, comparar com m√©tricas de 2024
            media_comparacao = media_2024
            max_comparacao = max_2024
        else:
            # Se √© 2025, usar m√©tricas de 2025
            media_comparacao = media_2025
            max_comparacao = max_2025
        
        return {
            'total_coletas_2024': int(total_coletas_2024),
            'total_coletas_2025': int(total_coletas_2025),
            'media_2024': round(media_2024, 1),
            'media_2025': round(media_2025, 1),
            'media_ultimo_mes': int(media_ultimo_mes),
            'max_2024': int(max_2024),
            'max_2025': int(max_2025),
            'ultimo_mes_display': info_ultimo_mes['display'],
            'ano_comparacao': ano_comparacao,
            'media_comparacao': round(media_comparacao, 1),
            'max_comparacao': int(max_comparacao)
        }
class AnaliseInteligente:
    """Classe para an√°lises inteligentes e insights autom√°ticos - Atualizado score."""
 
    @staticmethod
    def calcular_insights_automaticos(df: pd.DataFrame) -> pd.DataFrame:
        """Calcula insights autom√°ticos para cada laborat√≥rio."""
        df_insights = df.copy()
     
        # Volume atual (√∫ltimo m√™s fechado com l√≥gica de dia de corte)
        info_ultimo_mes = ChartManager._obter_ultimo_mes_fechado(df_insights)
        if info_ultimo_mes['coluna'] and info_ultimo_mes['coluna'] in df_insights.columns:
            df_insights['Volume_Atual_2025'] = df_insights[info_ultimo_mes['coluna']].fillna(0)
        else:
            df_insights['Volume_Atual_2025'] = 0
     
        # Volume m√°ximo do ano passado
        colunas_2024 = [col for col in df_insights.columns if 'N_Coletas_' in col and '24' in col]
        if colunas_2024:
            df_insights['Volume_Maximo_2024'] = df_insights[colunas_2024].max(axis=1).fillna(0)
        else:
            df_insights['Volume_Maximo_2024'] = 0
     
        # Tend√™ncia de volume (compara√ß√£o atual vs m√°ximo hist√≥rico)
        df_insights['Tendencia_Volume'] = df_insights.apply(
            lambda row: 'Crescimento' if row['Volume_Atual_2025'] > row['Volume_Maximo_2024']
            else 'Decl√≠nio' if row['Volume_Atual_2025'] < row['Volume_Maximo_2024'] * 0.5
            else 'Est√°vel', axis=1
        )
     
        # Insights autom√°ticos
        df_insights['Insights_Automaticos'] = df_insights.apply(
            lambda row: AnaliseInteligente._gerar_insights(row), axis=1
        )
     
        return df_insights
 
    @staticmethod
    def _gerar_insights(row) -> str:
        """Gera insights autom√°ticos baseados nos dados."""
        insights = []
     
        # An√°lise de dias sem coleta
        dias_sem = row.get('Dias_Sem_Coleta', 0)
        if dias_sem > 90:
            insights.append("üö® CR√çTICO: Sem coletas h√° mais de 3 meses")
        elif dias_sem > 60:
            insights.append("‚ö†Ô∏è ALERTA: Sem coletas h√° mais de 2 meses")
        elif dias_sem > 30:
            insights.append("üìâ ATEN√á√ÉO: Sem coletas h√° mais de 1 m√™s")
     
        # An√°lise de volume
        volume_atual = row.get('Volume_Atual_2025', 0)
        volume_max = row.get('Volume_Maximo_2024', 0)
        if volume_max > 0:
            ratio = volume_atual / volume_max
            if ratio > 1.5:
                insights.append("üìà EXCELENTE: Volume 50% acima do hist√≥rico")
            elif ratio > 1.2:
                insights.append("üìä POSITIVO: Volume 20% acima do hist√≥rico")
            elif ratio < 0.3:
                insights.append("üìâ CR√çTICO: Volume 70% abaixo do hist√≥rico")
            elif ratio < 0.6:
                insights.append("‚ö†Ô∏è ALERTA: Volume 40% abaixo do hist√≥rico")
     
        # An√°lise de tend√™ncia
        variacao = row.get('Variacao_Percentual', 0)
        if variacao > 100:
            insights.append("üöÄ CRESCIMENTO: Varia√ß√£o superior a 100%")
        elif variacao > 50:
            insights.append("üìà POSITIVO: Varia√ß√£o superior a 50%")
        elif variacao < -80:
            insights.append("üìâ CR√çTICO: Queda superior a 80%")
        elif variacao < -50:
            insights.append("‚ö†Ô∏è ALERTA: Queda superior a 50%")
     
        return " | ".join(insights) if insights else "‚úÖ Est√°vel"
class ReportManager:
    """Gerenciador de gera√ß√£o de relat√≥rios."""
    @staticmethod
    def gerar_relatorio_automatico(df: pd.DataFrame, metrics: KPIMetrics, tipo: str):
        """Gera relat√≥rio autom√°tico baseado no tipo."""
        if tipo == "semanal":
            ReportManager._gerar_relatorio_semanal(df, metrics)
        elif tipo == "mensal":
            ReportManager._gerar_relatorio_mensal(df, metrics)
    @staticmethod
    def _gerar_relatorio_semanal(df: pd.DataFrame, metrics: KPIMetrics):
        """Gera relat√≥rio semanal."""
        sumario = f"""
        üìä **Relat√≥rio Semanal de Churn - {datetime.now().strftime('%d/%m/%Y')}**
        **KPIs Principais:**
        ‚Ä¢ Total de Coletas: {metrics.total_coletas:,}
        ‚Ä¢ Labs em Risco: {metrics.labs_em_risco:,}
        ‚Ä¢ Ativos (7d): {metrics.ativos_7d:.1f}%
        **Alertas:**
        ‚Ä¢ {metrics.labs_alto_risco:,} laborat√≥rios em alto risco
        ‚Ä¢ {metrics.labs_medio_risco:,} laborat√≥rios em m√©dio risco
        **Recomenda√ß√µes:**
        ‚Ä¢ Focar nos {metrics.labs_alto_risco} labs de alto risco
        ‚Ä¢ Monitorar closely os {metrics.labs_medio_risco} labs de m√©dio risco
        """
        st.success("‚úÖ Relat√≥rio Semanal Gerado!")
        st.code(sumario, language="markdown")
        # Download do relat√≥rio
        st.download_button(
            "üì• Download Relat√≥rio Semanal",
            sumario,
            file_name=f"relatorio_semanal_{datetime.now().strftime('%Y%m%d')}.md",
            mime="text/markdown",
            key="download_relatorio_semanal"
        )
    @staticmethod
    def _gerar_relatorio_mensal(df: pd.DataFrame, metrics: KPIMetrics):
        """Gera relat√≥rio mensal detalhado."""
        # Calcular top varia√ß√µes
        if 'Variacao_Percentual' in df.columns:
            top_quedas = df.nsmallest(10, 'Variacao_Percentual')[['Nome_Fantasia_PCL', 'Variacao_Percentual', 'Estado']].copy()
            top_quedas['Ranking'] = range(1, len(top_quedas) + 1)
            top_quedas = top_quedas[['Ranking', 'Nome_Fantasia_PCL', 'Variacao_Percentual', 'Estado']]
            
            top_recuperacoes = df.nlargest(10, 'Variacao_Percentual')[['Nome_Fantasia_PCL', 'Variacao_Percentual', 'Estado']].copy()
            top_recuperacoes['Ranking'] = range(1, len(top_recuperacoes) + 1)
            top_recuperacoes = top_recuperacoes[['Ranking', 'Nome_Fantasia_PCL', 'Variacao_Percentual', 'Estado']]
        sumario = f"""
        üìä **Relat√≥rio Mensal de Churn - {datetime.now().strftime('%B/%Y').title()}**
        **KPIs Executivos:**
        ‚Ä¢ Total de Laborat√≥rios: {metrics.total_labs:,}
        ‚Ä¢ Taxa de Churn: {metrics.churn_rate:.1f}%
        ‚Ä¢ Net Revenue Retention: {metrics.nrr:.1f}%
        ‚Ä¢ Laborat√≥rios em Risco: {metrics.labs_em_risco:,}
        ‚Ä¢ Ativos (7 dias): {metrics.ativos_7d:.1f}%
        ‚Ä¢ Ativos (30 dias): {metrics.ativos_30d:.1f}%
        **Distribui√ß√£o por Risco:**
        ‚Ä¢ Alto Risco: {metrics.labs_alto_risco:,} ({metrics.labs_alto_risco/metrics.total_labs*100:.1f}%)
        ‚Ä¢ M√©dio Risco: {metrics.labs_medio_risco:,} ({metrics.labs_medio_risco/metrics.total_labs*100:.1f}%)
        ‚Ä¢ Baixo Risco: {metrics.labs_baixo_risco:,} ({metrics.labs_baixo_risco/metrics.total_labs*100:.1f}%)
        ‚Ä¢ Inativos: {metrics.labs_inativos:,} ({metrics.labs_inativos/metrics.total_labs*100:.1f}%)
        **An√°lise de Tend√™ncias:**
        """
        if 'Variacao_Percentual' in df.columns:
            media_variacao = df['Variacao_Percentual'].mean()
            sumario += f"""
        ‚Ä¢ Varia√ß√£o M√©dia: {media_variacao:.1f}%
        ‚Ä¢ Top Recupera√ß√µes: {len(top_recuperacoes)} laborat√≥rios
        ‚Ä¢ Top Quedas: {len(top_quedas)} laborat√≥rios
            """
        st.success("‚úÖ Relat√≥rio Mensal Gerado!")
        with st.expander("üìã Ver Relat√≥rio Completo", expanded=True):
            st.code(sumario, language="markdown")
        # Tabelas detalhadas
        if 'Variacao_Percentual' in df.columns:
            col1, col2 = st.columns(2)
            with col1:
                st.subheader("üìâ Top 10 Quedas")
                st.dataframe(
                    top_quedas,
                    width='stretch',
                    column_config={
                        "Ranking": st.column_config.NumberColumn("üèÜ", width="small", help="Posi√ß√£o no ranking"),
                        "Nome_Fantasia_PCL": st.column_config.TextColumn("Laborat√≥rio", help="Nome do laborat√≥rio"),
                        "Variacao_Percentual": st.column_config.NumberColumn("Varia√ß√£o %", format="%.2f%%", help="Varia√ß√£o percentual"),
                        "Estado": st.column_config.TextColumn("Estado", help="Estado do laborat√≥rio")
                    },
                    hide_index=True
                )
            with col2:
                st.subheader("üìà Top 10 Recupera√ß√µes")
                st.dataframe(
                    top_recuperacoes,
                    width='stretch',
                    column_config={
                        "Ranking": st.column_config.NumberColumn("üèÜ", width="small", help="Posi√ß√£o no ranking"),
                        "Nome_Fantasia_PCL": st.column_config.TextColumn("Laborat√≥rio", help="Nome do laborat√≥rio"),
                        "Variacao_Percentual": st.column_config.NumberColumn("Varia√ß√£o %", format="%.2f%%", help="Varia√ß√£o percentual"),
                        "Estado": st.column_config.TextColumn("Estado", help="Estado do laborat√≥rio")
                    },
                    hide_index=True
                )
        # Download do relat√≥rio
        st.download_button(
            "üì• Download Relat√≥rio Mensal",
            sumario,
            file_name=f"relatorio_mensal_{datetime.now().strftime('%Y%m%d')}.md",
            mime="text/markdown",
            key="download_relatorio_mensal"
        )
def show_toast_once(message: str, key: str):
    """Mostra um toast apenas uma vez por sess√£o."""
    if key not in st.session_state:
        st.toast(message)
        st.session_state[key] = True

def main():
    """Fun√ß√£o principal do dashboard v2.0 - Atualizado com tabs e navega√ß√£o."""
    # ============================================
    # AUTENTICA√á√ÉO MICROSOFT
    # ============================================
    try:
        # Inicializar autenticador Microsoft
        auth = MicrosoftAuth()
        # Verificar autentica√ß√£o
        if not create_login_page(auth):
            # Se n√£o conseguiu fazer login, parar execu√ß√£o
            return
        
        # Verificar e renovar token automaticamente se necess√°rio
        if not AuthManager.check_and_refresh_token(auth):
            # Se falhou ao renovar token, for√ßar novo login
            st.warning("‚ö†Ô∏è Sua sess√£o expirou. Por favor, fa√ßa login novamente.")
            st.rerun()
            return
        
        # Criar cabe√ßalho com informa√ß√µes do usu√°rio
        create_user_header()
    except Exception as e:
        st.error(f"‚ùå Erro no sistema de autentica√ß√£o: {str(e)}")
        st.warning("Verifique as configura√ß√µes de autentica√ß√£o no arquivo secrets.toml")
        return
    # ============================================
    # DASHBOARD PRINCIPAL (APENAS PARA USU√ÅRIOS AUTENTICADOS)
    # ============================================
    # Removido cabe√ßalho principal para layout mais discreto
    # Carregar e preparar dados
    loader_placeholder = st.empty()
    loader_placeholder.markdown(
        """
        <div class="overlay-loader">
            <div class="overlay-loader__content">
                <div class="overlay-loader__spinner"></div>
                <div class="overlay-loader__title">Carregando dados atualizados...</div>
                <div class="overlay-loader__subtitle">Estamos sincronizando as coletas mais recentes. Isso pode levar alguns segundos.</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )
    try:
        df_raw = DataManager.carregar_dados_churn()
        if df_raw is None:
            st.error("‚ùå N√£o foi poss√≠vel carregar os dados. Execute o gerador de dados primeiro.")
            return
        df = DataManager.preparar_dados(df_raw)
        show_toast_once(f"‚úÖ Dados carregados: {len(df):,} laborat√≥rios", "dados_carregados")
    finally:
        loader_placeholder.empty()
    # Indicador de √∫ltima atualiza√ß√£o
    if not df.empty and 'Data_Analise' in df.columns:
        ultima_atualizacao = df['Data_Analise'].max()
        st.markdown(f"**√öltima Atualiza√ß√£o:** {ultima_atualizacao.strftime('%d/%m/%Y %H:%M:%S')}")
    # ========================================
    # NAVEGA√á√ÉO (PRIMEIRO - NO TOPO DA SIDEBAR)
    # ========================================
    # Removido cabe√ßalho "Navega√ß√£o" da sidebar; bot√µes de p√°ginas mantidos abaixo
   
    pages = ["üè† Vis√£o Geral", "üìã An√°lise Detalhada", "üè¢ Ranking Rede", "üîß Manuten√ß√£o VIPs", "üîç An√°lise de Concorrente"]
   
    if "page" not in st.session_state:
        st.session_state.page = pages[0]
   
    for page in pages:
        if st.sidebar.button(page, key=page, width='stretch'):
            st.session_state.page = page
   
    # Separador visual
    st.sidebar.markdown("---")
   
    # Inicializar gerenciadores
    filter_manager = FilterManager()
    # Sidebar com filtros
    filtros = filter_manager.renderizar_sidebar_filtros(df)
    # Aplicar filtros
    df_filtrado = filter_manager.aplicar_filtros(df, filtros)
    # Calcular an√°lises inteligentes
    df_filtrado = AnaliseInteligente.calcular_insights_automaticos(df_filtrado)
    # Calcular KPIs
    metrics = KPIManager.calcular_kpis(df_filtrado)
    # Bot√£o de refresh
    if st.sidebar.button("üîÑ Atualizar Dados", help="Limpar cache e recarregar dados"):
        st.cache_data.clear()
        st.toast("‚úÖ Cache limpo! Os dados ser√£o recarregados automaticamente.")
    # Se√ß√£o de relat√≥rios na sidebar
    st.sidebar.markdown("---")
    st.sidebar.markdown('<div class="sidebar-header" style="font-size: 1rem; font-weight: 600; color: var(--primary-color);">üìÖ Relat√≥rios</div>', unsafe_allow_html=True)
    tipo_relatorio = st.sidebar.selectbox(
        "Tipo de Relat√≥rio",
        ["Semanal", "Mensal"],
        help="Selecione o tipo de relat√≥rio a gerar"
    )
    if st.sidebar.button("üìä Gerar Relat√≥rio", help="Gerar relat√≥rio autom√°tico"):
        ReportManager.gerar_relatorio_automatico(df_filtrado, metrics, tipo_relatorio.lower())
   
    # ========================================
    # RENDERIZA√á√ÉO DA P√ÅGINA SELECIONADA - Atualizado com tabs
    # ========================================
    if st.session_state.page == "üè† Vis√£o Geral":
        st.header("üè† Vis√£o Geral")
        # KPIs principais com cards modernos
        UIManager.renderizar_kpi_cards(metrics)
        # Usar tabs para organiza√ß√£o
        tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs(["üìä Resumo", "üìà Tend√™ncias", "üìä Distribui√ß√£o", "üö® Alto Risco", "üèÜ Top 100 PCLs", "üìä Controle BR/UF/Cidade"])
        with tab1:
            st.subheader("üìä Resumo Geral")
            st.markdown("### üö® Alertas Priorit√°rios")
            if df_filtrado.empty:
                st.info("üìä Nenhum dado dispon√≠vel para avaliar alertas.")
            else:
                if 'Risco_Diario' in df_filtrado.columns:
                    criticos = df_filtrado[df_filtrado['Risco_Diario'] == '‚ö´ Cr√≠tico'].copy()
                    if not criticos.empty:
                        st.error(f"‚ö†Ô∏è {len(criticos)} laborat√≥rio(s) em risco **CR√çTICO** (classifica√ß√£o baseada em dias √∫teis) ‚Äî interven√ß√£o imediata necess√°ria.")
                        colunas_alerta = [
                            'Nome_Fantasia_PCL', 'Estado',
                            'Vol_Hoje',
                            'Vol_D1', 'Delta_D1',
                            'MM7', 'Delta_MM7',
                            'Dias_Sem_Coleta'
                        ]
                        colunas_alerta = [c for c in colunas_alerta if c in criticos.columns]
                        if colunas_alerta:
                            st.dataframe(
                                _formatar_df_exibicao(criticos[colunas_alerta].sort_values('Vol_Hoje', ascending=True).head(10)),
                                width='stretch',
                                column_config={
                                    "Nome_Fantasia_PCL": st.column_config.TextColumn("Laborat√≥rio", help="Nome comercial do laborat√≥rio em risco cr√≠tico"),
                                    "Estado": st.column_config.TextColumn("UF", help="Estado (UF) onde o laborat√≥rio est√° localizado"),
                                    "Vol_Hoje": st.column_config.NumberColumn("Coletas (Hoje)", help="Total de coletas registradas no √∫ltimo dia √∫til"),
                                    "Vol_D1": st.column_config.NumberColumn("Coletas (D-1)", help="Volume de coletas do dia √∫til imediatamente anterior"),
                                    "Delta_D1": st.column_config.NumberColumn("Œî vs D-1", format="%.1f%%", help="Varia√ß√£o percentual: (Vol_Hoje - Vol_D1) / Vol_D1 √ó 100. Indica crescimento ou queda vs. dia √∫til anterior"),
                                    "MM7": st.column_config.NumberColumn("MM7", format="%.3f", help="M√©dia m√≥vel de 7 dias √∫teis - m√©dia aritm√©tica simples dos √∫ltimos 7 dias √∫teis"),
                                    "Delta_MM7": st.column_config.NumberColumn("Œî vs MM7", format="%.1f%%", help="Varia√ß√£o percentual: (Vol_Hoje - MM7) / MM7 √ó 100. Indica performance vs. m√©dia semanal dos √∫ltimos 7 dias √∫teis"),
                                    "Dias_Sem_Coleta": st.column_config.NumberColumn("Dias sem Coleta", help="N√∫mero consecutivo de dias √∫teis sem registrar coletas. Valores altos indicam poss√≠vel inatividade")
                                },
                                hide_index=True
                            )
                    else:
                        st.success("Nenhum laborat√≥rio classificado como ‚ö´ Cr√≠tico hoje.")
                else:
                    st.warning("‚ö†Ô∏è Coluna 'Risco_Diario' ausente ‚Äî imposs√≠vel gerar alertas priorit√°rios.")

                if {'Delta_MM7', 'Risco_Diario'}.issubset(df_filtrado.columns):
                    quedas_relevantes = df_filtrado[
                        (df_filtrado['Delta_MM7'] <= -50) &
                        (df_filtrado['Risco_Diario'].isin(['üü† Moderado', 'üî¥ Alto']))
                    ].copy()
                    if not quedas_relevantes.empty:
                        st.warning(
                            f"üîª {len(quedas_relevantes)} laborat√≥rio(s) com queda ‚â•50% vs MM7 (7 dias √∫teis) e risco elevado ‚Äî priorize contato de recupera√ß√£o."
                        )
                        colunas_queda = [
                            'Nome_Fantasia_PCL', 'Estado',
                            'Vol_Hoje',
                            'Vol_D1', 'Delta_D1',
                            'MM7', 'Delta_MM7',
                            'Risco_Diario', 'Recuperacao'
                        ]
                        colunas_queda = [c for c in colunas_queda if c in quedas_relevantes.columns]
                        if colunas_queda:
                            st.dataframe(
                                _formatar_df_exibicao(quedas_relevantes[colunas_queda].sort_values(['Delta_MM7', 'Vol_Hoje']).head(15)),
                                width='stretch',
                                column_config={
                                    "Nome_Fantasia_PCL": st.column_config.TextColumn("Laborat√≥rio", help="Nome comercial do laborat√≥rio com queda ‚â•50% vs MM7"),
                                    "Estado": st.column_config.TextColumn("UF", help="Estado (UF) onde o laborat√≥rio est√° localizado"),
                                    "Vol_Hoje": st.column_config.NumberColumn("Coletas (Hoje)", help="Total de coletas registradas no √∫ltimo dia √∫til"),
                                    "Vol_D1": st.column_config.NumberColumn("Coletas (D-1)", help="Volume de coletas do dia √∫til imediatamente anterior"),
                                    "Delta_D1": st.column_config.NumberColumn("Œî vs D-1", format="%.1f%%", help="Varia√ß√£o percentual: (Vol_Hoje - Vol_D1) / Vol_D1 √ó 100. Indica crescimento ou queda vs. dia anterior"),
                                    "MM7": st.column_config.NumberColumn("MM7", format="%.3f", help="M√©dia m√≥vel de 7 dias - m√©dia aritm√©tica simples dos √∫ltimos 7 dias (inclui dias sem coleta como zero)"),
                                    "Delta_MM7": st.column_config.NumberColumn("Œî vs MM7", format="%.1f%%", help="Varia√ß√£o percentual: (Vol_Hoje - MM7) / MM7 √ó 100. Valores ‚â§ -50% indicam queda estrutural significativa"),
                                    "Risco_Diario": st.column_config.TextColumn("Risco", help="Classifica√ß√£o de risco: üü¢ Normal, üü° Aten√ß√£o, üü† Moderado, üî¥ Alto, ‚ö´ Cr√≠tico"),
                                    "Recuperacao": st.column_config.CheckboxColumn("Em Recupera√ß√£o", help="Indica que o laborat√≥rio voltou a operar acima da MM7 ap√≥s per√≠odo de queda")
                                },
                                hide_index=True
                            )

                if {'Delta_D1', 'Risco_Diario'}.issubset(df_filtrado.columns):
                    quedas_d1_relevantes = df_filtrado[
                        (df_filtrado['Delta_D1'] <= -40) &
                        (df_filtrado['Risco_Diario'].isin(['üü† Moderado', 'üî¥ Alto']))
                    ].copy()
                    if not quedas_d1_relevantes.empty:
                        st.error(
                            f"üìâ {len(quedas_d1_relevantes)} laborat√≥rio(s) com queda ‚â•40% vs D-1 (dia √∫til anterior) e risco elevado ‚Äî aten√ß√£o imediata necess√°ria."
                        )
                        colunas_queda_d1 = [
                            'Nome_Fantasia_PCL', 'Estado',
                            'Vol_Hoje',
                            'Vol_D1', 'Delta_D1',
                            'MM7', 'Delta_MM7',
                            'Risco_Diario', 'Recuperacao'
                        ]
                        colunas_queda_d1 = [c for c in colunas_queda_d1 if c in quedas_d1_relevantes.columns]
                        if colunas_queda_d1:
                            st.dataframe(
                                _formatar_df_exibicao(quedas_d1_relevantes[colunas_queda_d1].sort_values(['Delta_D1', 'Vol_Hoje']).head(15)),
                                width='stretch',
                                column_config={
                                    "Nome_Fantasia_PCL": st.column_config.TextColumn("Laborat√≥rio", help="Nome comercial do laborat√≥rio com queda ‚â•40% vs D-1"),
                                    "Estado": st.column_config.TextColumn("UF", help="Estado (UF) onde o laborat√≥rio est√° localizado"),
                                    "Vol_Hoje": st.column_config.NumberColumn("Coletas (Hoje)", help="Total de coletas registradas no √∫ltimo dia √∫til"),
                                    "Vol_D1": st.column_config.NumberColumn("Coletas (D-1)", help="Volume de coletas do dia √∫til imediatamente anterior"),
                                    "Delta_D1": st.column_config.NumberColumn("Œî vs D-1", format="%.1f%%", help="Varia√ß√£o percentual: (Vol_Hoje - Vol_D1) / Vol_D1 √ó 100. Valores ‚â§ -40% indicam queda brusca recente"),
                                    "MM7": st.column_config.NumberColumn("MM7", format="%.3f", help="M√©dia m√≥vel de 7 dias - m√©dia aritm√©tica simples dos √∫ltimos 7 dias (inclui dias sem coleta como zero)"),
                                    "Delta_MM7": st.column_config.NumberColumn("Œî vs MM7", format="%.1f%%", help="Varia√ß√£o percentual: (Vol_Hoje - MM7) / MM7 √ó 100. Indica performance vs. m√©dia semanal dos √∫ltimos 7 dias"),
                                    "Risco_Diario": st.column_config.TextColumn("Risco", help="Classifica√ß√£o de risco: üü¢ Normal, üü° Aten√ß√£o, üü† Moderado, üî¥ Alto, ‚ö´ Cr√≠tico"),
                                    "Recuperacao": st.column_config.CheckboxColumn("Em Recupera√ß√£o", help="Indica que o laborat√≥rio voltou a operar acima da MM7 ap√≥s per√≠odo de queda")
                                },
                                hide_index=True
                            )

                if {'Risco_Diario', 'Delta_MM7'}.issubset(df_filtrado.columns):
                    moderados = df_filtrado[
                        (df_filtrado['Risco_Diario'] == 'üü† Moderado') & df_filtrado['Delta_MM7'].notna()
                    ].copy()
                    if not moderados.empty:
                        moderados = moderados.sort_values('Delta_MM7').head(10)
                        st.markdown("#### üü† Risco Moderado ‚Äî Top 10 quedas vs MM7")
                        st.caption("Ordenado por maior queda percentual (ŒîMM7) e limitado aos 10 piores casos. Baseado em dias √∫teis.")
                        colunas_moderado = [
                            'Nome_Fantasia_PCL', 'Estado',
                            'Vol_Hoje',
                            'Vol_D1', 'Delta_D1',
                            'MM7', 'Delta_MM7',
                            'MM30', 'Delta_MM30',
                            'Dias_Sem_Coleta'
                        ]
                        colunas_moderado = [c for c in colunas_moderado if c in moderados.columns]
                        if colunas_moderado:
                            st.dataframe(
                                _formatar_df_exibicao(moderados[colunas_moderado]),
                                width='stretch',
                                column_config={
                                    "Nome_Fantasia_PCL": st.column_config.TextColumn("Laborat√≥rio", help="Nome comercial do laborat√≥rio com risco moderado"),
                                    "Estado": st.column_config.TextColumn("UF", help="Estado (UF) onde o laborat√≥rio est√° localizado"),
                                    "Vol_Hoje": st.column_config.NumberColumn("Coletas (Hoje)", help="Total de coletas registradas no √∫ltimo dia √∫til"),
                                    "Vol_D1": st.column_config.NumberColumn("Coletas (D-1)", help="Volume de coletas do dia √∫til imediatamente anterior"),
                                    "MM7": st.column_config.NumberColumn("MM7", format="%.3f", help="M√©dia m√≥vel de 7 dias √∫teis - m√©dia aritm√©tica simples dos √∫ltimos 7 dias √∫teis"),
                                    "Delta_MM7": st.column_config.NumberColumn("Œî vs MM7", format="%.1f%%", help="Varia√ß√£o percentual: (Vol_Hoje - MM7) / MM7 √ó 100. Indica performance vs. m√©dia semanal dos √∫ltimos 7 dias √∫teis"),
                                    "Delta_D1": st.column_config.NumberColumn("Œî vs D-1", format="%.1f%%", help="Varia√ß√£o percentual: (Vol_Hoje - Vol_D1) / Vol_D1 √ó 100. Indica crescimento ou queda vs. dia √∫til anterior"),
                                    "MM30": st.column_config.NumberColumn("MM30", format="%.3f", help="M√©dia m√≥vel de 30 dias √∫teis - m√©dia aritm√©tica simples dos √∫ltimos 30 dias √∫teis"),
                                    "Delta_MM30": st.column_config.NumberColumn("Œî vs MM30", format="%.1f%%", help="Varia√ß√£o percentual: (Vol_Hoje - MM30) / MM30 √ó 100. Indica performance vs. m√©dia mensal dos √∫ltimos 30 dias √∫teis"),
                                    "Dias_Sem_Coleta": st.column_config.NumberColumn("Dias s/ Coleta", help="N√∫mero consecutivo de dias √∫teis sem registrar coletas. Valores altos indicam poss√≠vel inatividade")
                                },
                                hide_index=True
                            )

                if {'Vol_Hoje', 'Vol_D1'}.issubset(df_filtrado.columns):
                    dois_dias_sem_coleta = df_filtrado[(df_filtrado['Vol_Hoje'] == 0) & (df_filtrado['Vol_D1'] == 0)].copy()
                    if not dois_dias_sem_coleta.empty:
                        st.error(
                            f"üõë {len(dois_dias_sem_coleta)} laborat√≥rio(s) com **dois dias √∫teis consecutivos sem coleta** ‚Äî alinhar com opera√ß√µes/log√≠stica."
                        )
                        colunas_zero = ['Nome_Fantasia_PCL', 'Estado', 'Risco_Diario', 'Vol_D1', 'Dias_Sem_Coleta']
                        colunas_zero = [c for c in colunas_zero if c in dois_dias_sem_coleta.columns]
                        if colunas_zero:
                            st.dataframe(
                                _formatar_df_exibicao(dois_dias_sem_coleta[colunas_zero].head(15)),
                                width='stretch',
                                column_config={
                                    "Nome_Fantasia_PCL": st.column_config.TextColumn("Laborat√≥rio", help="Nome comercial do laborat√≥rio com dois dias consecutivos sem coleta"),
                                    "Estado": st.column_config.TextColumn("UF", help="Estado (UF) onde o laborat√≥rio est√° localizado"),
                                    "Risco_Diario": st.column_config.TextColumn("Risco", help="Classifica√ß√£o de risco: üü¢ Normal, üü° Aten√ß√£o, üü† Moderado, üî¥ Alto, ‚ö´ Cr√≠tico"),
                                    "Vol_D1": st.column_config.NumberColumn("Coletas (D-1)", help="Volume de coletas do dia √∫til imediatamente anterior ao atual (mostra zero para estes casos)"),
                                    "Dias_Sem_Coleta": st.column_config.NumberColumn("Dias sem Coleta", help="N√∫mero consecutivo de dias √∫teis sem registrar coletas. ‚ö†Ô∏è Valores ‚â• 2 indicam necessidade de alinhamento operacional")
                                },
                                hide_index=True
                            )

            st.markdown("---")
            with st.expander("‚ÑπÔ∏è Legenda das m√©tricas di√°rias"):
                st.markdown("""
#### üìä M√©tricas Principais
- **Vol_Hoje**: total de coletas registradas na data de refer√™ncia (√∫ltimo dia √∫til da s√©rie di√°ria).
- **Vol_D1**: volume de coletas do dia √∫til imediatamente anterior ao atual.
- **MM7 / MM30 / MM90**: m√©dias m√≥veis de 7, 30 e 90 dias √∫teis da s√©rie di√°ria, calculadas apenas com dias √∫teis. Exibidas com 3 casas decimais para m√°xima transpar√™ncia nos c√°lculos de varia√ß√£o.
- **Œî vs MM7 / MM30 / MM90**: varia√ß√£o percentual do volume do √∫ltimo dia √∫til em rela√ß√£o √†s respectivas m√©dias m√≥veis (calculada com valores n√£o arredondados).
- **Œî vs D-1**: varia√ß√£o percentual do volume do √∫ltimo dia √∫til comparado ao dia √∫til anterior.
- **DOW_Media**: m√©dia de coletas para o mesmo dia da semana (ex.: todas as segundas) nos √∫ltimos 90 dias √∫teis.

#### üö® Classifica√ß√£o de Risco Di√°rio
A classifica√ß√£o de risco segue uma r√©gua hier√°rquica baseada em m√∫ltiplos crit√©rios, considerando **apenas dias √∫teis**:

**üü¢ Normal**: Volume dentro dos padr√µes esperados (90-120% da MM7 ou 100-120% do D-1).
**üü° Aten√ß√£o**: Volume abaixo do normal mas ainda recuper√°vel (70-90% da MM7 ou 70-100% do D-1).
**üü† Moderado**: Volume significativamente reduzido (50-70% da MM7 ou 60-70% do D-1).
**üî¥ Alto**: Volume cr√≠tico com necessidade de interven√ß√£o (abaixo de 50% da MM7 ou 60% do D-1).
**‚ö´ Cr√≠tico**: Situa√ß√µes extremas (7+ dias √∫teis sem coleta ou 3+ quedas consecutivas de 50%+).

#### ‚ö†Ô∏è Regras de Alerta Espec√≠ficas
- **üîª Queda ‚â•50% vs MM7**: Laborat√≥rios com queda estrutural significativa + risco moderado/alto.
- **üìâ Queda ‚â•40% vs D-1**: Laborat√≥rios com queda brusca recente + risco moderado/alto.

#### üîÑ Outros Indicadores
- **Risco_Diario**: classifica√ß√£o autom√°tica baseada nos limiares acima, calculada sobre dias √∫teis e redu√ß√µes vs. MM7_BR/MM7_UF/MM7_CIDADE.
- **Recuperacao**: indica que o laborat√≥rio voltou a operar acima da MM7 ap√≥s per√≠odo de queda.
- **Sem Coleta (48h)**: quantidade de laborat√≥rios com dois dias √∫teis consecutivos sem registrar coletas (Vol_Hoje = 0 e Vol_D1 = 0).
                """)

            # Adicionar m√©tricas adicionais aqui
        with tab2:
            st.subheader("üìà Tend√™ncias e Varia√ß√µes (Di√°rio - Dias √öteis)")
            if df_filtrado.empty:
                st.info("üìä Nenhum dado dispon√≠vel para esta an√°lise.")
            else:
                col1, col2 = st.columns(2)

                with col1:
                    st.markdown("#### üìâ Maiores Quedas vs MM7 (7 dias √∫teis)")
                    if {'Delta_MM7', 'Vol_Hoje', 'MM7'}.issubset(df_filtrado.columns):
                        quedas_diarias = df_filtrado[df_filtrado['Delta_MM7'].notna()].copy()
                        if not quedas_diarias.empty:
                            quedas_diarias = quedas_diarias.sort_values('Delta_MM7').head(10)
                            colunas_quedas = [
                                'Nome_Fantasia_PCL', 'Estado',
                                'Vol_Hoje',
                                'Vol_D1', 'Delta_D1',
                                'MM7', 'Delta_MM7',
                                'Risco_Diario', 'Dias_Sem_Coleta'
                            ]
                            colunas_quedas = [c for c in colunas_quedas if c in quedas_diarias.columns]
                            st.dataframe(
                                _formatar_df_exibicao(quedas_diarias[colunas_quedas]),
                                width='stretch',
                                column_config={
                                    "Nome_Fantasia_PCL": st.column_config.TextColumn("Laborat√≥rio", help="Nome comercial do laborat√≥rio com maiores quedas vs MM7"),
                                    "Estado": st.column_config.TextColumn("UF", help="Estado (UF) onde o laborat√≥rio est√° localizado"),
                                    "Vol_Hoje": st.column_config.NumberColumn("Coletas (Hoje)", help="Total de coletas registradas no √∫ltimo dia √∫til"),
                                    "Vol_D1": st.column_config.NumberColumn("Coletas (D-1)", help="Volume de coletas do dia √∫til imediatamente anterior"),
                                    "MM7": st.column_config.NumberColumn("MM7", format="%.3f", help="M√©dia m√≥vel de 7 dias √∫teis - m√©dia aritm√©tica simples dos √∫ltimos 7 dias √∫teis"),
                                    "Delta_MM7": st.column_config.NumberColumn("Œî vs MM7", format="%.1f%%", help="Varia√ß√£o percentual: (Vol_Hoje - MM7) / MM7 √ó 100. Ordenado por maior queda (valores mais negativos primeiro). Baseado em dias √∫teis."),
                                    "Delta_D1": st.column_config.NumberColumn("Œî vs D-1", format="%.1f%%", help="Varia√ß√£o percentual: (Vol_Hoje - Vol_D1) / Vol_D1 √ó 100. Indica crescimento ou queda vs. dia √∫til anterior"),
                                    "Risco_Diario": st.column_config.TextColumn("Risco", help="Classifica√ß√£o de risco: üü¢ Normal, üü° Aten√ß√£o, üü† Moderado, üî¥ Alto, ‚ö´ Cr√≠tico"),
                                    "Dias_Sem_Coleta": st.column_config.NumberColumn("Dias s/ Coleta", help="N√∫mero consecutivo de dias √∫teis sem registrar coletas. Valores altos indicam poss√≠vel inatividade")
                                },
                                hide_index=True
                            )
                        if quedas_diarias.empty:
                            st.success("Nenhuma queda relevante detectada hoje.")
                    else:
                        st.warning("‚ö†Ô∏è Colunas necess√°rias para a an√°lise de quedas (Œî vs MM7) n√£o encontradas.")

                with col2:
                    st.markdown("#### üìà Altas vs MM7 (7 dias √∫teis)")
                    if {'Delta_MM7', 'Vol_Hoje', 'MM7'}.issubset(df_filtrado.columns):
                        altas_diarias = df_filtrado[df_filtrado['Delta_MM7'].notna()].copy()
                        altas_diarias = altas_diarias[altas_diarias['Delta_MM7'] > 0]
                        if not altas_diarias.empty:
                            altas_diarias = altas_diarias.sort_values('Delta_MM7', ascending=False).head(10)
                            colunas_altas = [
                                'Nome_Fantasia_PCL', 'Estado',
                                'Vol_Hoje',
                                'Vol_D1', 'Delta_D1',
                                'MM7', 'Delta_MM7',
                                'Risco_Diario', 'Recuperacao'
                            ]
                            colunas_altas = [c for c in colunas_altas if c in altas_diarias.columns]
                            st.dataframe(
                                _formatar_df_exibicao(altas_diarias[colunas_altas]),
                                width='stretch',
                                column_config={
                                    "Nome_Fantasia_PCL": st.column_config.TextColumn("Laborat√≥rio", help="Nome comercial do laborat√≥rio com maiores altas vs MM7"),
                                    "Estado": st.column_config.TextColumn("UF", help="Estado (UF) onde o laborat√≥rio est√° localizado"),
                                    "Vol_Hoje": st.column_config.NumberColumn("Coletas (Hoje)", help="Total de coletas registradas no √∫ltimo dia √∫til"),
                                    "Vol_D1": st.column_config.NumberColumn("Coletas (D-1)", help="Volume de coletas do dia √∫til imediatamente anterior"),
                                    "MM7": st.column_config.NumberColumn("MM7", format="%.3f", help="M√©dia m√≥vel de 7 dias - m√©dia aritm√©tica simples dos √∫ltimos 7 dias (inclui dias sem coleta como zero)"),
                                    "Delta_MM7": st.column_config.NumberColumn("Œî vs MM7", format="%.1f%%", help="Varia√ß√£o percentual: (Vol_Hoje - MM7) / MM7 √ó 100. Ordenado por maior crescimento (valores mais positivos primeiro)"),
                                    "Delta_D1": st.column_config.NumberColumn("Œî vs D-1", format="%.1f%%", help="Varia√ß√£o percentual: (Vol_Hoje - Vol_D1) / Vol_D1 √ó 100. Indica crescimento ou queda vs. dia anterior"),
                                    "Risco_Diario": st.column_config.TextColumn("Risco", help="Classifica√ß√£o de risco: üü¢ Normal, üü° Aten√ß√£o, üü† Moderado, üî¥ Alto, ‚ö´ Cr√≠tico"),
                                    "Recuperacao": st.column_config.CheckboxColumn("Recupera√ß√£o", help="Indica que o laborat√≥rio voltou a operar acima da MM7 ap√≥s per√≠odo de queda")
                                },
                                hide_index=True
                            )
                        else:
                            st.info("Nenhum crescimento significativo vs MM7 identificado hoje.")
                    else:
                        st.warning("‚ö†Ô∏è Colunas necess√°rias para a an√°lise de altas (Œî vs MM7) n√£o encontradas.")

                st.markdown("#### üîÅ Recupera√ß√µes em Andamento")
                if 'Recuperacao' in df_filtrado.columns:
                    recuperacoes = df_filtrado[(df_filtrado['Recuperacao'] == True) & df_filtrado['Delta_MM7'].notna()].copy()
                    if not recuperacoes.empty:
                        # Recalcular Œî vs MM7 para garantir f√≥rmula consistente: (Hoje - MM7) / MM7
                        for col in ['Vol_Hoje', 'MM7']:
                            recuperacoes[col] = pd.to_numeric(recuperacoes[col], errors='coerce')

                        # Fun√ß√£o para calcular percentual igual √† do RiskEngine
                        def pct_safe(a, b):
                            return ((a - b) / b * 100).where(b.notna() & (b != 0), 0.0)

                        recuperacoes['Delta_MM7'] = pct_safe(recuperacoes['Vol_Hoje'], recuperacoes['MM7']).round(1)
                        recuperacoes = recuperacoes.sort_values('Delta_MM7', ascending=False)
                        colunas_recuperacao = [
                            'Nome_Fantasia_PCL', 'Estado',
                            'Vol_Hoje',
                            'Vol_D1', 'Delta_D1',
                            'MM7', 'Delta_MM7',
                            'Risco_Diario', 'Dias_Sem_Coleta'
                        ]
                        colunas_recuperacao = [c for c in colunas_recuperacao if c in recuperacoes.columns]
                        st.dataframe(
                            _formatar_df_exibicao(recuperacoes[colunas_recuperacao].head(10)),
                            width='stretch',
                            column_config={
                                "Nome_Fantasia_PCL": st.column_config.TextColumn("Laborat√≥rio", help="Nome comercial do laborat√≥rio em recupera√ß√£o"),
                                "Estado": st.column_config.TextColumn("UF", help="Estado (UF) onde o laborat√≥rio est√° localizado"),
                                "Vol_Hoje": st.column_config.NumberColumn("Coletas (Hoje)", help="Total de coletas registradas na data de refer√™ncia (dia mais recente)"),
                                "Vol_D1": st.column_config.NumberColumn("Coletas (D-1)", help="Volume de coletas do dia imediatamente anterior ao atual"),
                                "MM7": st.column_config.NumberColumn("MM7", format="%.3f", help="M√©dia m√≥vel de 7 dias - m√©dia aritm√©tica simples dos √∫ltimos 7 dias (inclui dias sem coleta como zero)"),
                                "Delta_MM7": st.column_config.NumberColumn("Œî vs MM7", format="%.1f%%", help="Varia√ß√£o percentual: (Vol_Hoje - MM7) / MM7 √ó 100. Ordenado por maior recupera√ß√£o (valores mais positivos primeiro)"),
                                "Delta_D1": st.column_config.NumberColumn("Œî vs D-1", format="%.1f%%", help="Varia√ß√£o percentual: (Vol_Hoje - Vol_D1) / Vol_D1 √ó 100. Indica crescimento ou queda vs. dia anterior"),
                                "Risco_Diario": st.column_config.TextColumn("Risco", help="Classifica√ß√£o de risco: üü¢ Normal, üü° Aten√ß√£o, üü† Moderado, üî¥ Alto, ‚ö´ Cr√≠tico"),
                                "Dias_Sem_Coleta": st.column_config.NumberColumn("Dias s/ Coleta", help="N√∫mero consecutivo de dias sem registrar coletas. Valores altos indicam poss√≠vel inatividade")
                            },
                            hide_index=True
                        )
                    else:
                        st.info("Nenhuma recupera√ß√£o consistente detectada (labs com Œî vs MM7 positivo e flag de recupera√ß√£o).")
                else:
                    st.warning("‚ö†Ô∏è Coluna 'Recuperacao' n√£o encontrada nos dados.")

            st.markdown("---")
            with st.expander("üìä Como Funcionam os C√°lculos desta Aba"):
                st.markdown("""
#### üî¢ **F√≥rmulas B√°sicas dos Indicadores**

**M√©dias M√≥veis (MM7, MM30, MM90):**
- **MM7**: M√©dia aritm√©tica simples dos √∫ltimos 7 dias √∫teis
- **MM30**: M√©dia aritm√©tica simples dos √∫ltimos 30 dias √∫teis
- **MM90**: M√©dia aritm√©tica simples dos √∫ltimos 90 dias √∫teis
- *Nota: Calculadas apenas com dias √∫teis (exclui finais de semana e feriados)*

**Varia√ß√µes Percentuais (Deltas):**
- **Œî vs MM7** = `(Vol_Hoje - MM7) / MM7 √ó 100` (baseado em dias √∫teis)
- **Œî vs D-1** = `(Vol_Hoje - Vol_D1) / Vol_D1 √ó 100` (dia √∫til anterior)
- **Œî vs MM30** = `(Vol_Hoje - MM30) / MM30 √ó 100` (baseado em dias √∫teis)
- *Nota: Calculados com valores n√£o arredondados para m√°xima precis√£o*

#### üìä **L√≥gica de Cada Tabela**

**1. üìâ Maiores Quedas vs MM7 (7 dias √∫teis)**
- **Filtro**: `Delta_MM7.notna()` (todos com dados dispon√≠veis)
- **Ordena√ß√£o**: Por `Delta_MM7` (maior queda primeiro)
- **Limite**: Top 10 laborat√≥rios
- **Objetivo**: Identificar maiores decl√≠nios estruturais vs. m√©dia semanal de dias √∫teis

**2. üìà Altas vs MM7 (7 dias √∫teis)**
- **Filtro**: `Delta_MM7 > 0` (apenas crescimentos)
- **Ordena√ß√£o**: Por `Delta_MM7` decrescente (maior alta primeiro)
- **Limite**: Top 10 laborat√≥rios
- **Objetivo**: Identificar recupera√ß√µes expressivas vs. m√©dia semanal de dias √∫teis

**3. üîÅ Recupera√ß√µes em Andamento**
- **Filtro**: `Recuperacao == True AND Delta_MM7.notna()`
- **Ordena√ß√£o**: Por `Delta_MM7` decrescente (maior recupera√ß√£o primeiro)
- **Limite**: Top 10 laborat√≥rios
- **Objetivo**: Labs que voltaram acima da MM7 ap√≥s queda

#### üéØ **Exemplo Pr√°tico do C√°lculo**

Dados do laborat√≥rio: `Vol_Hoje = 3`, `MM7 = 0.429`
```
Œî vs MM7 = (3 - 0.429) / 0.429 √ó 100
          = 2.571 / 0.429 √ó 100
          = 6.00 √ó 100
          = 600%
```

**Por que 600%?** Porque o laborat√≥rio coletou ~7 vezes mais que sua m√©dia semanal de dias √∫teis!

#### ‚ö†Ô∏è **Alertas e Regras de Prioriza√ß√£o**

- **üîª Quedas ‚â•50% vs MM7 + Risco Moderado/Alto**: Prioridade m√°xima
- **üìâ Quedas ‚â•40% vs D-1 + Risco Moderado/Alto**: Aten√ß√£o imediata
- **Laborat√≥rios sem coleta por 48h**: Seguir protocolo operacional

#### üîç **Dicas para An√°lise**
- **MM7 pr√≥xima de zero**: Valores percentuais podem parecer "inflados", mas s√£o matematicamente corretos
- **Compare tend√™ncias**: Olhe tanto quedas quanto altas para contexto completo
- **Risco vs Performance**: Laborat√≥rio pode ter alto crescimento mas ainda estar em risco se abaixo dos benchmarks

#### üí° **Como Interpretar Valores Altos de Percentual**
- **0-100%**: Crescimento normal/recupera√ß√£o
- **100-300%**: Crescimento expressivo
- **300%+**: Laborat√≥rio voltou a operar ap√≥s per√≠odo de inatividade
- **Sempre compare com o valor absoluto**: 600% com MM7=0.429 significa apenas ~3 coletas vs. m√©dia de ~0.429

#### üìà **Contexto Executivo**
Para um laborat√≥rio que normalmente coleta 3 vezes por semana (MM7 ‚âà 0.429 em dias √∫teis), registrar 3 coletas no √∫ltimo dia √∫til representa um crescimento de 600% vs. sua m√©dia hist√≥rica de dias √∫teis. Isso indica recupera√ß√£o de opera√ß√£o, n√£o necessariamente "superperformance".
                """)
        with tab3:
            st.subheader("üìä Distribui√ß√£o por Status")
            ChartManager.criar_grafico_distribuicao_risco(df_filtrado)
        with tab4:
            st.subheader("üö® Labs em Risco")
            ChartManager.criar_grafico_top_labs(df_filtrado, top_n=10)
            if 'Risco_Diario' in df_filtrado.columns:
                labs_em_risco = df_filtrado[df_filtrado['Risco_Diario'].isin(['üü† Moderado', 'üî¥ Alto', '‚ö´ Cr√≠tico'])]
            else:
                st.warning("‚ö†Ô∏è Coluna 'Risco_Diario' n√£o encontrada nos dados.")
                labs_em_risco = pd.DataFrame()
            if not labs_em_risco.empty:
                colunas_resumo = ['Nome_Fantasia_PCL', 'Estado', 'Representante_Nome',
                                  'Vol_Hoje', 'Delta_MM7', 'Risco_Diario']
                st.dataframe(
                    _formatar_df_exibicao(labs_em_risco[colunas_resumo]),
                    width='stretch',
                    height=300,
                    column_config={
                        "Nome_Fantasia_PCL": st.column_config.TextColumn("Laborat√≥rio", help="Nome comercial do laborat√≥rio em risco"),
                        "Estado": st.column_config.TextColumn("UF", help="Estado (UF) onde o laborat√≥rio est√° localizado"),
                        "Representante_Nome": st.column_config.TextColumn("Representante", help="Nome do representante comercial respons√°vel pelo laborat√≥rio"),
                        "Vol_Hoje": st.column_config.NumberColumn("Coletas (Hoje)", help="Total de coletas registradas na data de refer√™ncia (dia mais recente)"),
                        "Delta_MM7": st.column_config.NumberColumn("Œî vs MM7", format="%.1f%%", help="Varia√ß√£o percentual: (Vol_Hoje - MM7) / MM7 √ó 100. Indica performance vs. m√©dia semanal dos √∫ltimos 7 dias"),
                        "Risco_Diario": st.column_config.TextColumn("Risco Di√°rio", help="Classifica√ß√£o de risco: üü¢ Normal, üü° Aten√ß√£o, üü† Moderado, üî¥ Alto, ‚ö´ Cr√≠tico")
                    },
                    hide_index=True
                )
            else:
                st.success("‚úÖ Nenhum laborat√≥rio em risco encontrado!")
        with tab5:
            st.subheader("üèÜ Top 100 PCLs - Maiores Coletas")
            
            # Calcular total de coletas para cada laborat√≥rio
            if not df_filtrado.empty:
                # Calcular total de coletas 2025
                meses_2025 = ChartManager._meses_ate_hoje(df_filtrado, 2025)
                colunas_2025 = [f'N_Coletas_{mes}_25' for mes in meses_2025]
                colunas_existentes = [col for col in colunas_2025 if col in df_filtrado.columns]
                
                if colunas_existentes:
                    df_filtrado['Total_Coletas_2025_Calculado'] = df_filtrado[colunas_existentes].sum(axis=1)
                else:
                    df_filtrado['Total_Coletas_2025_Calculado'] = 0
                
                # Criar ranking dos top 100
                top_100 = df_filtrado.nlargest(100, 'Total_Coletas_2025_Calculado')
                
                # Preparar dados para exibi√ß√£o
                ranking_data = []
                for idx, (_, row) in enumerate(top_100.iterrows(), 1):
                    ranking_data.append({
                        'Ranking': idx,
                        'CNPJ': row.get('CNPJ_PCL', 'N/A'),
                        'Laborat√≥rio': row.get('Nome_Fantasia_PCL', 'N/A'),
                        'Coletas': int(row.get('Total_Coletas_2025_Calculado', 0)),
                        'Representante': row.get('Representante_Nome', 'N/A'),
                        'Estado': row.get('Estado', 'N/A'),
                        'Cidade': row.get('Cidade', 'N/A')
                    })
                
                df_ranking = pd.DataFrame(ranking_data)
                
                # Filtros de busca
                col1, col2 = st.columns([2, 1])
                with col1:
                    busca = st.text_input("üîç Fa√ßa sua Pesquisa", placeholder="Digite CNPJ, nome do laborat√≥rio ou representante...")
                with col2:
                    estado_filtro = st.selectbox("üìç Estado", ["Todos"] + sorted(df_ranking['Estado'].unique().tolist()))
                
                # Aplicar filtros
                df_filtrado_ranking = df_ranking.copy()
                
                if busca:
                    mask = (
                        df_filtrado_ranking['CNPJ'].str.contains(busca, case=False, na=False) |
                        df_filtrado_ranking['Laborat√≥rio'].str.contains(busca, case=False, na=False) |
                        df_filtrado_ranking['Representante'].str.contains(busca, case=False, na=False)
                    )
                    df_filtrado_ranking = df_filtrado_ranking[mask]
                
                if estado_filtro != "Todos":
                    df_filtrado_ranking = df_filtrado_ranking[df_filtrado_ranking['Estado'] == estado_filtro]
                
                # Exibir tabela
                if not df_filtrado_ranking.empty:
                    # Estilizar a tabela
                    st.markdown("""
                    <style>
                    .ranking-table {
                        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                        border-collapse: collapse;
                        width: 100%;
                        margin-top: 1rem;
                    }
                    .ranking-table th {
                        background: linear-gradient(135deg, #6BBF47 0%, #52B54B 100%);
                        color: white;
                        padding: 12px 8px;
                        text-align: left;
                        font-weight: 600;
                        font-size: 0.9rem;
                    }
                    .ranking-table td {
                        padding: 10px 8px;
                        border-bottom: 1px solid #e9ecef;
                        font-size: 0.85rem;
                    }
                    .ranking-table tr:nth-child(even) {
                        background-color: #f8f9fa;
                    }
                    .ranking-table tr:hover {
                        background-color: #e8f5e9;
                        transition: background-color 0.2s;
                    }
                    .ranking-number {
                        font-weight: bold;
                        color: #6BBF47;
                        text-align: center;
                    }
                    .coletas-number {
                        font-weight: bold;
                        color: #28a745;
                        text-align: right;
                    }
                    </style>
                    """, unsafe_allow_html=True)
                    
                    # Mostrar estat√≠sticas
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.metric("üèÜ Total de Labs", f"{len(df_filtrado_ranking):,}")
                    with col2:
                        total_coletas = df_filtrado_ranking['Coletas'].sum()
                        st.metric("üìä Total de Coletas", f"{total_coletas:,}")
                    with col3:
                        media_coletas = df_filtrado_ranking['Coletas'].mean()
                        st.metric("üìà M√©dia por Lab", f"{media_coletas:.0f}")
                    with col4:
                        top_coletas = df_filtrado_ranking['Coletas'].max() if not df_filtrado_ranking.empty else 0
                        st.metric("ü•á Maior Volume", f"{top_coletas:,}")
                    
                    # Exibir tabela com formata√ß√£o
                    st.dataframe(
                        df_filtrado_ranking[['Ranking', 'CNPJ', 'Laborat√≥rio', 'Coletas', 'Representante', 'Estado', 'Cidade']],
                        width='stretch',
                        height=600,
                        column_config={
                            "Ranking": st.column_config.NumberColumn(
                                "Ranking",
                                help="Posi√ß√£o do laborat√≥rio no ranking geral por volume de coletas em 2025. Ranking 1 = maior volume",
                                format="%d",
                                width="small"
                            ),
                            "CNPJ": st.column_config.TextColumn(
                                "CNPJ",
                                help="CNPJ (Cadastro Nacional de Pessoa Jur√≠dica) do laborat√≥rio. Identificador √∫nico",
                                width="medium"
                            ),
                            "Laborat√≥rio": st.column_config.TextColumn(
                                "Laborat√≥rio",
                                help="Nome comercial/fantasia do laborat√≥rio. Top 100 laborat√≥rios por volume total de coletas em 2025",
                                width="large"
                            ),
                            "Coletas": st.column_config.NumberColumn(
                                "Coletas",
                                help="Soma total de coletas em 2025 at√© o momento (todos os meses dispon√≠veis at√© hoje). Ordena√ß√£o por este valor (maior para menor)",
                                format="%d",
                                width="small"
                            ),
                            "Representante": st.column_config.TextColumn(
                                "Representante",
                                help="Nome do representante comercial respons√°vel pelo laborat√≥rio",
                                width="medium"
                            ),
                            "Estado": st.column_config.TextColumn(
                                "Estado",
                                help="Estado (UF) onde o laborat√≥rio est√° localizado. Permite filtrar e agrupar por regi√£o geogr√°fica",
                                width="small"
                            ),
                            "Cidade": st.column_config.TextColumn(
                                "Cidade",
                                help="Cidade onde o laborat√≥rio est√° localizado. Permite an√°lise mais granular por localiza√ß√£o",
                                width="medium"
                            )
                        },
                        hide_index=True
                    )
                    
                    # Bot√µes de download
                    col_download1, col_download2 = st.columns(2)
                    
                    with col_download1:
                        csv_data = df_filtrado_ranking.to_csv(index=False, encoding='utf-8')
                        st.download_button(
                            "üì• Download CSV",
                            csv_data,
                            file_name=f"ranking_top_100_pcls_{datetime.now().strftime('%Y%m%d')}.csv",
                            mime="text/csv",
                            width='stretch'
                        )
                    
                    with col_download2:
                        # Preparar dados para Excel
                        excel_buffer = BytesIO()
                        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                            # Adicionar metadados na primeira aba
                            metadata_df = pd.DataFrame({
                                'M√©trica': ['Total de Laborat√≥rios', 'Total de Coletas', 'M√©dia por Laborat√≥rio', 'Maior Volume', 'Data de Gera√ß√£o'],
                                'Valor': [
                                    f"{len(df_filtrado_ranking):,}",
                                    f"{df_filtrado_ranking['Coletas'].sum():,}",
                                    f"{df_filtrado_ranking['Coletas'].mean():.0f}",
                                    f"{df_filtrado_ranking['Coletas'].max():,}",
                                    datetime.now().strftime('%d/%m/%Y %H:%M:%S')
                                ]
                            })
                            metadata_df.to_excel(writer, sheet_name='Resumo', index=False)
                            
                            # Adicionar ranking na segunda aba
                            df_filtrado_ranking.to_excel(writer, sheet_name='Ranking Top 100', index=False)
                            
                            # Formata√ß√£o da planilha
                            workbook = writer.book
                            
                            # Formatar aba de resumo
                            summary_sheet = writer.sheets['Resumo']
                            summary_sheet.column_dimensions['A'].width = 25
                            summary_sheet.column_dimensions['B'].width = 20
                            
                            # Formatar aba de ranking
                            ranking_sheet = writer.sheets['Ranking Top 100']
                            ranking_sheet.column_dimensions['A'].width = 8   # Ranking
                            ranking_sheet.column_dimensions['B'].width = 18  # CNPJ
                            ranking_sheet.column_dimensions['C'].width = 40  # Laborat√≥rio
                            ranking_sheet.column_dimensions['D'].width = 12  # Coletas
                            ranking_sheet.column_dimensions['E'].width = 25  # Representante
                            ranking_sheet.column_dimensions['F'].width = 8   # Estado
                            ranking_sheet.column_dimensions['G'].width = 20  # Cidade
                            
                            # Aplicar formata√ß√£o condicional para destacar top 10
                            from openpyxl.styles import PatternFill, Font
                            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                            bold_font = Font(bold=True)
                            
                            for row in range(2, min(12, len(df_filtrado_ranking) + 2)):  # Top 10
                                for col in range(1, 8):
                                    cell = ranking_sheet.cell(row=row, column=col)
                                    cell.fill = yellow_fill
                                    cell.font = bold_font
                        
                        excel_data = excel_buffer.getvalue()
                        st.download_button(
                            "üìä Download Excel",
                            excel_data,
                            file_name=f"ranking_top_100_pcls_{datetime.now().strftime('%Y%m%d')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            width='stretch'
                        )
                else:
                    st.info("üîç Nenhum resultado encontrado para os filtros aplicados.")
            else:
                st.warning("‚ö†Ô∏è Nenhum dado dispon√≠vel para gerar o ranking.")
        
        with tab6:
            st.subheader("üìä Controle BR √ó UF √ó Cidade")
            st.markdown("""
            **Compara√ß√£o de m√©dias m√≥veis (MM7/MM30) entre contextos geogr√°ficos e o laborat√≥rio/conjunto selecionado.**
            
            Este gr√°fico permite visualizar como o desempenho do laborat√≥rio ou conjunto filtrado se compara com:
            - üáßüá∑ **MM7_BR / MM30_BR**: M√©dia m√≥vel nacional (todos os laborat√≥rios)
            - üìç **MM7_UF / MM30_UF**: M√©dia m√≥vel do estado (UF) do laborat√≥rio
            - üèôÔ∏è **MM7_CIDADE / MM30_CIDADE**: M√©dia m√≥vel da cidade do laborat√≥rio
            - üìä **S√©rie atual**: Laborat√≥rio espec√≠fico ou conjunto filtrado
            
            **Nota**: Todas as s√©ries s√£o calculadas apenas com dias √∫teis (exclui finais de semana e feriados).
            """)
            
            if df_filtrado.empty:
                st.info("üìä Nenhum dado dispon√≠vel para o gr√°fico de controle")
            else:
                # Toggle para MM7/MM30
                col_toggle1, col_toggle2 = st.columns([1, 4])
                with col_toggle1:
                    usar_mm30 = st.toggle("Usar MM30", value=False, help="Alternar entre MM7 (7 dias √∫teis) e MM30 (30 dias √∫teis)")
                
                # Op√ß√£o para selecionar laborat√≥rio espec√≠fico ou usar conjunto filtrado
                st.markdown("#### Sele√ß√£o de Contexto")
                modo_visualizacao = st.radio(
                    "Escolha o contexto para compara√ß√£o:",
                    ["Conjunto Filtrado", "Laborat√≥rio Espec√≠fico"],
                    horizontal=True,
                    help="Conjunto Filtrado: agrega todos os laborat√≥rios que passaram pelos filtros aplicados. Laborat√≥rio Espec√≠fico: seleciona um laborat√≥rio individual para an√°lise."
                )
                
                lab_cnpj_selecionado = None
                lab_nome_selecionado = None
                
                if modo_visualizacao == "Laborat√≥rio Espec√≠fico":
                    # Buscar laborat√≥rios dispon√≠veis
                    labs_disponiveis = df_filtrado[['CNPJ_PCL', 'Nome_Fantasia_PCL', 'Estado', 'Cidade']].copy()
                    labs_disponiveis = labs_disponiveis.dropna(subset=['Nome_Fantasia_PCL'])
                    labs_disponiveis['Display'] = labs_disponiveis.apply(
                        lambda x: f"{x['Nome_Fantasia_PCL']} ({x.get('Estado', 'N/A')})",
                        axis=1
                    )
                    
                    if not labs_disponiveis.empty:
                        lab_selecionado = st.selectbox(
                            "Selecione o laborat√≥rio:",
                            options=labs_disponiveis['Display'].tolist(),
                            help="Selecione um laborat√≥rio espec√≠fico para comparar com os contextos BR/UF/Cidade"
                        )
                        
                        lab_info = labs_disponiveis[labs_disponiveis['Display'] == lab_selecionado].iloc[0]
                        lab_nome_selecionado = lab_info['Nome_Fantasia_PCL']
                        lab_cnpj_selecionado = lab_info.get('CNPJ_PCL')
                    else:
                        st.warning("‚ö†Ô∏è Nenhum laborat√≥rio dispon√≠vel nos dados filtrados")
                        modo_visualizacao = "Conjunto Filtrado"
                
                # Renderizar gr√°fico
                if modo_visualizacao == "Conjunto Filtrado" or (modo_visualizacao == "Laborat√≥rio Espec√≠fico" and lab_nome_selecionado):
                    # Usar df completo para calcular contextos BR/UF/Cidade
                    ChartManager.criar_grafico_controle_br_uf_cidade(
                        df=df,  # DataFrame completo para contextos
                        df_filtrado=df_filtrado,  # DataFrame filtrado para s√©rie atual
                        lab_cnpj=lab_cnpj_selecionado,
                        lab_nome=lab_nome_selecionado,
                        usar_mm30=usar_mm30
                    )
                    
                    st.markdown("---")
                    with st.expander("‚ÑπÔ∏è Como interpretar este gr√°fico"):
                        st.markdown("""
                        #### üìä **Interpreta√ß√£o do Gr√°fico**
                        
                        **Posicionamento relativo:**
                        - Se a s√©rie atual est√° **acima** das linhas de contexto (BR/UF/Cidade), o desempenho est√° melhor que a m√©dia do contexto
                        - Se est√° **abaixo**, h√° oportunidade de melhoria comparado ao contexto
                        
                        **Tend√™ncias:**
                        - **Linhas ascendentes**: Crescimento consistente
                        - **Linhas descendentes**: Decl√≠nio que requer aten√ß√£o
                        - **Linhas est√°veis**: Manuten√ß√£o de padr√£o
                        
                        **Compara√ß√µes √∫teis:**
                        - Compare primeiro com **MM7_UF** (contexto estadual mais pr√≥ximo)
                        - Use **MM7_BR** para vis√£o macro nacional
                        - **MM7_CIDADE** mostra contexto local mais espec√≠fico
                        
                        **MM7 vs MM30:**
                        - **MM7**: Mais sens√≠vel a varia√ß√µes recentes (√∫ltimos 7 dias √∫teis)
                        - **MM30**: Vis√£o mais suavizada e de m√©dio prazo (√∫ltimos 30 dias √∫teis)
                        
                        **Dias √∫teis:**
                        - Todas as s√©ries consideram apenas dias √∫teis (segunda a sexta)
                        - Feriados e finais de semana s√£o exclu√≠dos automaticamente
                        """)
                else:
                    st.info("üìä Selecione um laborat√≥rio para visualizar o gr√°fico de controle")
    
    elif st.session_state.page == "üìã An√°lise Detalhada":
        st.header("üìã An√°lise Detalhada")
        # Filtros avan√ßados com design moderno
        st.markdown("""
        <div style="background: linear-gradient(135deg, #6BBF47 0%, #52B54B 100%);
                    color: white; padding: 1.5rem; border-radius: 10px;
                    margin-bottom: 1rem; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
            <h3 style="margin: 0; font-size: 1.3rem;">üîç Busca Inteligente de Laborat√≥rios</h3>
            <p style="margin: 0.5rem 0 0 0; opacity: 0.9;">
                Busque por CNPJ (com ou sem formata√ß√£o) ou nome do laborat√≥rio
            </p>
        </div>
        """, unsafe_allow_html=True)
        # Sele√ß√£o de laborat√≥rio espec√≠fico
        if not df_filtrado.empty:
            if 'CNPJ_Normalizado' not in df_filtrado.columns:
                df_filtrado['CNPJ_Normalizado'] = df_filtrado['CNPJ_PCL'].apply(DataManager.normalizar_cnpj)
            df_filtrado['CNPJ_Normalizado'] = df_filtrado['CNPJ_Normalizado'].fillna('')

            labs_catalogo = df_filtrado[
                ['CNPJ_PCL', 'CNPJ_Normalizado', 'Nome_Fantasia_PCL', 'Razao_Social_PCL', 'Cidade', 'Estado']
            ].copy()
            labs_catalogo = labs_catalogo[labs_catalogo['CNPJ_Normalizado'] != ""]
            labs_catalogo['CNPJ_Normalizado'] = labs_catalogo['CNPJ_Normalizado'].astype(str)
            labs_catalogo = labs_catalogo.drop_duplicates('CNPJ_Normalizado')

            def formatar_cnpj_display(cnpj_val):
                digitos = ''.join(filter(str.isdigit, str(cnpj_val))) if pd.notna(cnpj_val) else ''
                if len(digitos) == 14:
                    return f"{digitos[:2]}.{digitos[2:5]}.{digitos[5:8]}/{digitos[8:12]}-{digitos[12:]}"
                return digitos or "N/A"

            def montar_rotulo(row):
                nome = row.get('Nome_Fantasia_PCL') or row.get('Razao_Social_PCL') or "Laborat√≥rio sem nome"
                cidade = row.get('Cidade') or ''
                estado = row.get('Estado') or ''
                if cidade and estado:
                    local = f"{cidade}/{estado}"
                elif cidade:
                    local = cidade
                elif estado:
                    local = estado
                else:
                    local = "Localidade n√£o informada"
                cnpj_fmt = formatar_cnpj_display(row.get('CNPJ_PCL') or row.get('CNPJ_Normalizado'))
                return f"{nome} - {local} (CNPJ: {cnpj_fmt})"

            lab_display_map = {str(row['CNPJ_Normalizado']): montar_rotulo(row) for _, row in labs_catalogo.iterrows()}
            lab_nome_map = {
                str(row['CNPJ_Normalizado']): row.get('Nome_Fantasia_PCL') or row.get('Razao_Social_PCL') or str(row['CNPJ_Normalizado'])
                for _, row in labs_catalogo.iterrows()
            }
            lista_cnpjs_ordenada = sorted(lab_display_map.keys(), key=lambda cnpj: lab_display_map[cnpj].lower())
            lista_cnpjs_validos = set(lista_cnpjs_ordenada)

            LAB_STATE_KEY = 'lab_cnpj_selecionado'
            lab_cnpj_estado = st.session_state.get(LAB_STATE_KEY, "") or ""
            if lab_cnpj_estado and lab_cnpj_estado not in lista_cnpjs_validos:
                lab_cnpj_estado = ""
                st.session_state[LAB_STATE_KEY] = ""

            opcoes_select = [""] + lista_cnpjs_ordenada
            index_padrao = opcoes_select.index(lab_cnpj_estado) if lab_cnpj_estado in lista_cnpjs_validos else 0

            # Layout melhorado com 3 colunas - ajustado para melhor alinhamento
            col1, col2, col3 = st.columns([4, 1.5, 2.5])
            with col1:
                # Campo de busca aprimorado
                busca_lab = st.text_input(
                    "üîé Buscar",
                    placeholder="CNPJ (com/sem formata√ß√£o) ou Nome do laborat√≥rio",
                    help="Digite CNPJ (com ou sem pontos/tra√ßos) ou nome do laborat√≥rio/raz√£o social",
                    key="busca_avancada"
                )
            with col2:
                # Bot√£o de busca r√°pida
                buscar_btn = st.button("üîé Buscar", type="primary", width='stretch')
            with col3:
                # Sele√ß√£o por dropdown como alternativa
                lab_selecionado = st.selectbox(
                    "üìã Lista R√°pida:",
                    options=opcoes_select,
                    index=index_padrao,
                    format_func=lambda cnpj: "Selecione um laborat√≥rio" if cnpj == "" else lab_display_map.get(cnpj, cnpj),
                    help="Ou selecione um laborat√≥rio da lista completa"
                )
                lab_selecionado = lab_selecionado or ""
                if lab_selecionado != st.session_state.get(LAB_STATE_KEY, ""):
                    st.session_state[LAB_STATE_KEY] = lab_selecionado
            lab_cnpj_estado = st.session_state.get(LAB_STATE_KEY, "") or ""
            # Informa√ß√µes de ajuda - Atualizado espa√ßamento dica busca
            with st.expander("üí° Dicas de Busca", expanded=False):
                st.markdown("""
                **üî¢ Para CNPJ:**
                - Apenas n√∫meros: `51865434001248`
                - Com formata√ß√£o: `51.865.434/0012-48`
                **üè• Para Nome:**
                - Nome fantasia ou raz√£o social
                - Busca parcial e sem distin√ß√£o de mai√∫sculas/min√∫sculas
                **üìä Resultados:**
                - 1 resultado: Selecionado automaticamente
                - M√∫ltiplos: Lista para escolher o correto
                """)
            # Estado da busca
            lab_final = None
            lab_final_cnpj = lab_cnpj_estado if lab_cnpj_estado in lista_cnpjs_validos else ""
            # Verificar se h√° busca ativa ou laborat√≥rio selecionado
            busca_ativa = buscar_btn or (busca_lab and len(busca_lab.strip()) > 2)
            tem_selecao = bool(lab_cnpj_estado)
            if busca_ativa or tem_selecao:
                # L√≥gica de busca aprimorada
                if busca_ativa and busca_lab:
                    busca_normalizada = busca_lab.strip()
                    # Verificar se √© CNPJ (com ou sem formata√ß√£o)
                    cnpj_limpo = ''.join(filter(str.isdigit, busca_normalizada))
                    if len(cnpj_limpo) >= 1:
                        if len(cnpj_limpo) >= 14:
                            lab_encontrado = df_filtrado[df_filtrado['CNPJ_Normalizado'] == cnpj_limpo]
                        else:
                            lab_encontrado = df_filtrado[df_filtrado['CNPJ_Normalizado'].str.startswith(cnpj_limpo)]
                    else:
                        # Buscar por nome (case insensitive e parcial) - apenas nome fantasia e raz√£o social
                        lab_encontrado = df_filtrado[
                            df_filtrado['Nome_Fantasia_PCL'].str.contains(busca_normalizada, case=False, na=False) |
                            df_filtrado['Razao_Social_PCL'].str.contains(busca_normalizada, case=False, na=False)
                        ]
                    lab_encontrado = lab_encontrado[lab_encontrado['CNPJ_Normalizado'] != ""].drop_duplicates('CNPJ_Normalizado')
                    if not lab_encontrado.empty:
                        if len(lab_encontrado) == 1:
                            lab_info_unico = lab_encontrado.iloc[0]
                            lab_final = lab_info_unico.get('Nome_Fantasia_PCL') or lab_info_unico.get('Razao_Social_PCL')
                            lab_final_cnpj = str(lab_info_unico.get('CNPJ_Normalizado', ''))
                            st.toast(
                                f"‚úÖ Laborat√≥rio encontrado: {lab_final} (CNPJ: {formatar_cnpj_display(lab_final_cnpj)})"
                            )
                            st.session_state[LAB_STATE_KEY] = lab_final_cnpj
                        else:
                            # M√∫ltiplos resultados - mostrar op√ß√µes
                            st.info(f"üîç Encontrados {len(lab_encontrado)} laborat√≥rios. Selecione um:")
                            opcoes_df = lab_encontrado.head(10)
                            opcoes_cnpjs = [""] + opcoes_df['CNPJ_Normalizado'].astype(str).tolist()
                            if 'multiplo_resultados' in st.session_state:
                                valor_multi = st.session_state['multiplo_resultados']
                                if valor_multi not in opcoes_cnpjs:
                                    st.session_state['multiplo_resultados'] = ""
                            lab_selecionado_multiplo = st.selectbox(
                                "Selecione o laborat√≥rio correto:",
                                options=opcoes_cnpjs,
                                format_func=lambda cnpj: "Selecione" if cnpj == "" else lab_display_map.get(cnpj, cnpj),
                                key="multiplo_resultados"
                            )
                            if lab_selecionado_multiplo:
                                lab_final_cnpj = str(lab_selecionado_multiplo)
                                lab_final = lab_nome_map.get(lab_final_cnpj, lab_final_cnpj)
                                st.session_state[LAB_STATE_KEY] = lab_final_cnpj
                    else:
                        # N√£o encontrou - verificar se existe na base completa e qual filtro est√° impedindo
                        cnpj_limpo = ''.join(filter(str.isdigit, busca_normalizada))
                        lab_na_base_completa = None
                        
                        if len(cnpj_limpo) >= 1:
                            if len(cnpj_limpo) >= 14:
                                lab_na_base_completa = df[df['CNPJ_Normalizado'] == cnpj_limpo]
                            else:
                                lab_na_base_completa = df[df['CNPJ_Normalizado'].str.startswith(cnpj_limpo)]
                        else:
                            # Buscar por nome na base completa
                            lab_na_base_completa = df[
                                df['Nome_Fantasia_PCL'].str.contains(busca_normalizada, case=False, na=False) |
                                df['Razao_Social_PCL'].str.contains(busca_normalizada, case=False, na=False)
                            ]
                        
                        lab_na_base_completa = lab_na_base_completa[lab_na_base_completa['CNPJ_Normalizado'] != ""].drop_duplicates('CNPJ_Normalizado')
                        
                        if not lab_na_base_completa.empty:
                            # Encontrou na base completa mas n√£o nos filtros atuais
                            st.warning("‚ö†Ô∏è Nenhum laborat√≥rio encontrado com os filtros atuais")
                            
                            # Identificar quais filtros est√£o ativos
                            filtros_ativos = []
                            if filtros.get('apenas_vip', False):
                                filtros_ativos.append("**Apenas VIPs**")
                            if filtros.get('representantes'):
                                filtros_ativos.append(f"**Representante(s)**: {', '.join(filtros['representantes'])}")
                            if filtros.get('ufs'):
                                filtros_ativos.append(f"**UF(s)**: {', '.join(filtros['ufs'])}")
                            
                            if filtros_ativos:
                                st.info(f"üí° Encontramos **{len(lab_na_base_completa)} laborat√≥rio(s)** na base completa, mas est√°(√£o) sendo filtrado(s) por:\n\n" + 
                                       "\n".join([f"- {f}" for f in filtros_ativos]))
                                
                                st.caption("üí° **Dica**: Desative os filtros na barra lateral para visualizar este laborat√≥rio.")
                            else:
                                st.info(f"üí° Encontramos **{len(lab_na_base_completa)} laborat√≥rio(s)** na base completa, mas est√°(√£o) fora do per√≠odo selecionado ou de outros filtros aplicados.")
                        else:
                            # N√£o encontrou nem na base completa
                            st.warning("‚ö†Ô∏è Nenhum laborat√≥rio encontrado com os crit√©rios informados")
                            st.caption("Este laborat√≥rio n√£o est√° na nossa base de dados.")
                elif tem_selecao:
                    # Laborat√≥rio selecionado diretamente da lista
                    lab_final_cnpj = st.session_state.get(LAB_STATE_KEY, "")
                    lab_final = lab_nome_map.get(lab_final_cnpj, lab_final_cnpj)
                # Renderizar dados do laborat√≥rio encontrado/selecionado
                if lab_final_cnpj:
                    st.markdown("---") # Separador antes dos dados
                    # Verificar se √© VIP
                    df_vip = DataManager.carregar_dados_vip()
                    if lab_final_cnpj:
                        lab_data = df_filtrado[df_filtrado['CNPJ_Normalizado'] == lab_final_cnpj]
                    else:
                        lab_data = df_filtrado[df_filtrado['Nome_Fantasia_PCL'] == lab_final]
                    info_vip = None
                    if not lab_data.empty and df_vip is not None:
                        cnpj_lab = lab_data.iloc[0].get('CNPJ_PCL', '')
                        info_vip = VIPManager.buscar_info_vip(cnpj_lab, df_vip)
                    # Container principal para informa√ß√µes do laborat√≥rio
                    st.markdown(f"""
                        <div style="background: linear-gradient(135deg, #6BBF47 0%, #52B54B 100%);
                                    color: white; padding: 2rem; border-radius: 15px;
                                    margin-bottom: 2rem; box-shadow: 0 8px 25px rgba(0,0,0,0.15);">
                            <div style="display: flex; align-items: center;">
                                <div style="font-size: 2rem; margin-right: 1rem;">üè•</div>
                                <div>
                                    <h2 style="margin: 0; font-size: 1.8rem; font-weight: 600;">{lab_final}</h2>
                                </div>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                    # Armazenar informa√ß√µes da rede para filtro autom√°tico na tabela
                    if info_vip and 'rede' in info_vip:
                        st.session_state['rede_lab_pesquisado'] = info_vip['rede']
                    else:
                        st.session_state['rede_lab_pesquisado'] = None
                    # Ficha T√©cnica Comercial
                    st.markdown("""
                        <div style="background: white; border-radius: 8px; padding: 1.5rem; margin-bottom: 2rem;
                                    border: 1px solid #e9ecef; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                            <h3 style="margin: 0 0 1rem 0; color: #2c3e50; font-weight: 600; border-bottom: 2px solid #6BBF47; padding-bottom: 0.5rem;">
                                üìã Ficha T√©cnica Comercial
                            </h3>
                        """, unsafe_allow_html=True)
                    # Informa√ß√µes de contato e localiza√ß√£o
                    if lab_final_cnpj:
                        lab_data = df_filtrado[df_filtrado['CNPJ_Normalizado'] == lab_final_cnpj]
                    else:
                        lab_data = df_filtrado[df_filtrado['Nome_Fantasia_PCL'] == lab_final]
                    if not lab_data.empty:
                            lab_info = lab_data.iloc[0]
                         
                            # CNPJ formatado
                            cnpj_raw = str(lab_info.get('CNPJ_PCL', ''))
                            cnpj_formatado = f"{cnpj_raw[:2]}.{cnpj_raw[2:5]}.{cnpj_raw[5:8]}/{cnpj_raw[8:12]}-{cnpj_raw[12:14]}" if len(cnpj_raw) == 14 else cnpj_raw
                         
                            # Carregar dados de laboratories.csv
                            df_labs = DataManager.carregar_laboratories()
                            info_lab = None
                            if df_labs is not None and not df_labs.empty:
                                info_lab = VIPManager.buscar_info_laboratory(cnpj_raw, df_labs)
                            
                            # Prioridade: 1) laboratories.csv, 2) matriz VIP (legado), 3) dados do laborat√≥rio
                            contato = ''
                            telefone = ''
                            email = ''
                            
                            if info_lab:
                                contato = info_lab.get('contato', '')
                                telefone = info_lab.get('telefone', '')
                                email = info_lab.get('email', '')
                            
                            # Fallback para dados VIP (legado) se n√£o encontrado no laboratories.csv
                            if not contato and info_vip:
                                contato = info_vip.get('contato', '')
                            if not telefone and info_vip:
                                telefone = info_vip.get('telefone', '')
                            if not email and info_vip:
                                email = info_vip.get('email', '')
                            
                            # √öltimo fallback: dados do lab_info
                            # Nota: lab_info pode n√£o ter campo 'Contato' separado, ent√£o apenas telefone/email
                            if not telefone:
                                telefone = lab_info.get('Telefone', 'N/A')
                            if not email:
                                email = lab_info.get('Email', 'N/A')
                            # Se contato ainda estiver vazio, deixar como 'N/A' (n√£o h√° fallback no lab_info)
                            
                            representante = lab_info.get('Representante_Nome', 'N/A')
                            
                            # Limpar dados vazios
                            telefone = telefone if telefone and telefone != 'N/A' and telefone != '' else 'N/A'
                            email = email if email and email != 'N/A' and email != '' else 'N/A'
                            contato = contato if contato and contato != '' else 'N/A'
                            representante = representante if representante and representante != 'N/A' else 'N/A'
                            
                            # Extrair novos dados do info_lab
                            endereco_completo = info_lab.get('endereco', {}) if info_lab else {}
                            logistic_data = info_lab.get('logistic', {}) if info_lab else {}
                            licensed_list = info_lab.get('licensed', []) if info_lab else []
                            allowed_methods_list = info_lab.get('allowedMethods', []) if info_lab else []
                            
                            # Mapear dias da semana para portugu√™s
                            dias_semana_map = {
                                'mon': 'Segunda', 'tue': 'Ter√ßa', 'wed': 'Quarta', 
                                'thu': 'Quinta', 'fri': 'Sexta', 'sat': 'S√°bado', 'sun': 'Domingo'
                            }
                            
                            # Formatar dias de funcionamento
                            dias_funcionamento = []
                            if logistic_data.get('days'):
                                for dia in logistic_data.get('days', []):
                                    dias_funcionamento.append(dias_semana_map.get(dia.lower(), dia.capitalize()))
                            dias_funcionamento_str = ', '.join(dias_funcionamento) if dias_funcionamento else 'N/A'
                            horario_funcionamento = logistic_data.get('openingHours', '') if logistic_data.get('openingHours') else 'N/A'
                            
                            # Formatar endere√ßo completo
                            endereco_linha1 = ''
                            endereco_linha2 = ''
                            cep_formatado = 'N/A'
                            if endereco_completo:
                                endereco_parts = []
                                if endereco_completo.get('address'):
                                    endereco_parts.append(endereco_completo.get('address', ''))
                                if endereco_completo.get('number'):
                                    endereco_parts.append(f"n¬∫ {endereco_completo.get('number', '')}")
                                if endereco_completo.get('addressComplement'):
                                    endereco_parts.append(endereco_completo.get('addressComplement', ''))
                                endereco_linha1 = ', '.join(endereco_parts) if endereco_parts else 'N/A'
                                
                                endereco_parts2 = []
                                if endereco_completo.get('neighbourhood'):
                                    endereco_parts2.append(endereco_completo.get('neighbourhood', ''))
                                if endereco_completo.get('city'):
                                    endereco_parts2.append(endereco_completo.get('city', ''))
                                if endereco_completo.get('state_code'):
                                    endereco_parts2.append(endereco_completo.get('state_code', ''))
                                endereco_linha2 = ' - '.join(endereco_parts2) if endereco_parts2 else 'N/A'
                                
                                # Formatar CEP
                                if endereco_completo.get('postalCode'):
                                    cep_raw = str(endereco_completo.get('postalCode', '')).strip()
                                    if len(cep_raw) == 8:
                                        cep_formatado = f"{cep_raw[:5]}-{cep_raw[5:]}"
                                    else:
                                        cep_formatado = cep_raw
                            
                            # Fallback para dados do lab_info se endere√ßo completo n√£o dispon√≠vel
                            if not endereco_linha1 or endereco_linha1 == 'N/A':
                                endereco_linha1 = 'N/A'
                            if not endereco_linha2 or endereco_linha2 == 'N/A':
                                endereco_linha2 = f"{lab_info.get('Cidade', 'N/A')} - {lab_info.get('Estado', 'N/A')}"
                            
                            # Formatar licen√ßas
                            licencas_map = {
                                'clt': 'CLT', 'cnh': 'CNH', 'cltCnh': 'CLT/CNH',
                                'other': 'Outros', 'online': 'Online',
                                'civilService': 'Concurso P√∫blico', 
                                'civilServiceAnalysis50': 'Concurso P√∫blico (50)',
                                'otherAnalysis50': 'Outros (50)'
                            }
                            licencas_formatadas = [licencas_map.get(l, l) for l in licensed_list] if licensed_list else []
                            licencas_str = ', '.join(licencas_formatadas) if licencas_formatadas else 'N/A'
                            
                            # Formatar m√©todos de pagamento
                            metodos_map = {
                                'cash': 'Dinheiro', 'credit': 'Cr√©dito', 'debit': 'D√©bito',
                                'billing_laboratory': 'Faturamento Lab', 'billing_company': 'Faturamento Empresa',
                                'billing': 'Faturamento', 'bank_billet': 'Boleto',
                                'eCredit': 'e-Cr√©dito', 'pix': 'PIX'
                            }
                            metodos_formatados = [metodos_map.get(m, m) for m in allowed_methods_list] if allowed_methods_list else []
                            metodos_str = ', '.join(metodos_formatados) if metodos_formatados else 'N/A'
                         
                            st.markdown(f"""
                            <div style="background: #f8f9fa; border-radius: 6px; padding: 1rem; margin-bottom: 1rem; border-left: 4px solid #6c757d;">
                                <div style="font-size: 0.9rem; color: #666; margin-bottom: 0.5rem; font-weight: 600;">INFORMA√á√ïES DE CONTATO</div>
                                <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 1rem;">
                                    <div>
                                        <div style="font-size: 0.8rem; color: #666; margin-bottom: 0.3rem;">CNPJ</div>
                                        <div style="font-size: 1rem; font-weight: bold; color: #495057;">{cnpj_formatado}</div>
                                    </div>
                                    <div>
                                        <div style="font-size: 0.8rem; color: #666; margin-bottom: 0.3rem;">CEP</div>
                                        <div style="font-size: 1rem; font-weight: bold; color: #495057;">{cep_formatado}</div>
                                    </div>
                                    <div style="grid-column: 1 / -1;">
                                        <div style="font-size: 0.8rem; color: #666; margin-bottom: 0.3rem;">Endere√ßo</div>
                                        <div style="font-size: 1rem; font-weight: bold; color: #495057;">{endereco_linha1}</div>
                                        <div style="font-size: 0.9rem; color: #6c757d; margin-top: 0.2rem;">{endereco_linha2}</div>
                                    </div>
                                    <div>
                                        <div style="font-size: 0.8rem; color: #666; margin-bottom: 0.3rem;">Localiza√ß√£o</div>
                                        <div style="font-size: 1rem; font-weight: bold; color: #495057;">{lab_info.get('Cidade', 'N/A')} - {lab_info.get('Estado', 'N/A')}</div>
                                    </div>
                                    <div>
                                        <div style="font-size: 0.8rem; color: #666; margin-bottom: 0.3rem;">Contato</div>
                                        <div style="font-size: 1rem; font-weight: bold; color: #495057;">{contato}</div>
                                    </div>
                                    <div>
                                        <div style="font-size: 0.8rem; color: #666; margin-bottom: 0.3rem;">Telefone</div>
                                        <div style="font-size: 1rem; font-weight: bold; color: #495057;">{telefone}</div>
                                    </div>
                                    <div>
                                        <div style="font-size: 0.8rem; color: #666; margin-bottom: 0.3rem;">Email</div>
                                        <div style="font-size: 1rem; font-weight: bold; color: #495057;">{email}</div>
                                    </div>
                                    <div>
                                        <div style="font-size: 0.8rem; color: #666; margin-bottom: 0.3rem;">Representante</div>
                                        <div style="font-size: 1rem; font-weight: bold; color: #495057;">{representante}</div>
                                    </div>
                                    <div>
                                        <div style="font-size: 0.8rem; color: #666; margin-bottom: 0.3rem;">Dias de Funcionamento</div>
                                        <div style="font-size: 1rem; font-weight: bold; color: #495057;">{dias_funcionamento_str}</div>
                                    </div>
                                    <div>
                                        <div style="font-size: 0.8rem; color: #666; margin-bottom: 0.3rem;">Hor√°rio</div>
                                        <div style="font-size: 1rem; font-weight: bold; color: #495057;">{horario_funcionamento}</div>
                                    </div>
                                    <div>
                                        <div style="font-size: 0.8rem; color: #666; margin-bottom: 0.3rem;">Licen√ßas</div>
                                        <div style="font-size: 1rem; font-weight: bold; color: #495057;">{licencas_str}</div>
                                    </div>
                                    <div>
                                        <div style="font-size: 0.8rem; color: #666; margin-bottom: 0.3rem;">M√©todos de Pagamento</div>
                                        <div style="font-size: 1rem; font-weight: bold; color: #495057;">{metodos_str}</div>
                                    </div>
                                </div>
                            </div>
                            """, unsafe_allow_html=True)

                            # Bloco de pre√ßos praticados
                            def _formatar_preco_valor(valor):
                                try:
                                    if pd.notna(valor):
                                        return f"R$ {float(valor):.2f}".replace('.', ',')
                                except Exception:
                                    pass
                                return "N/A"

                            price_labels = {
                                'CLT': 'CLT',
                                'CNH': 'CNH',
                                'Civil_Service': 'Concurso P√∫blico',
                                'Civil_Service50': 'Concurso P√∫blico (50)',
                                'CLT_CNH': 'CLT / CNH',
                                'Outros': 'Outros',
                                'Outros50': 'Outros (50)'
                            }

                            price_cards = []
                            possui_preco = False
                            for key, cfg in PRICE_CATEGORIES.items():
                                prefix = cfg['prefix']
                                label = price_labels.get(prefix, prefix.replace('_', ' '))
                                total = lab_info.get(f'Preco_{prefix}_Total', np.nan)
                                coleta = lab_info.get(f'Preco_{prefix}_Coleta', np.nan)
                                exame = lab_info.get(f'Preco_{prefix}_Exame', np.nan)

                                possui_valores = any(pd.notna(v) for v in [total, coleta, exame])
                                if not possui_valores:
                                    continue

                                possui_preco = True

                                price_cards.append(
                                    f"<div style=\"background: white; border-radius: 8px; padding: 1rem; "
                                    f"box-shadow: 0 2px 6px rgba(0,0,0,0.08);\">"
                                    f"<div style=\"font-size: 0.85rem; color: #6c757d; text-transform: uppercase; "
                                    f"letter-spacing: 0.5px; margin-bottom: 0.6rem; font-weight: 700;\">"
                                    f"{label}"
                                    f"</div>"
                                    f"<div style=\"display: flex; justify-content: space-between; font-size: 0.85rem; "
                                    f"color: #6c757d; margin-bottom: 0.4rem;\">"
                                    f"<span>Coleta</span>"
                                    f"<strong style=\"color: #495057;\">{_formatar_preco_valor(coleta)}</strong>"
                                    f"</div>"
                                    f"<div style=\"display: flex; justify-content: space-between; font-size: 0.85rem; "
                                    f"color: #6c757d; margin-bottom: 0.4rem;\">"
                                    f"<span>Exame</span>"
                                    f"<strong style=\"color: #495057;\">{_formatar_preco_valor(exame)}</strong>"
                                    f"</div>"
                                    f"<div style=\"display: flex; justify-content: space-between; font-size: 0.85rem; "
                                    f"color: #6c757d;\">"
                                    f"<span>Total</span>"
                                    f"<strong style=\"color: #495057;\">{_formatar_preco_valor(total)}</strong>"
                                    f"</div>"
                                    f"</div>"
                                )

                            if possui_preco or pd.notna(lab_info.get('Voucher_Commission', np.nan)):
                                voucher_valor = lab_info.get('Voucher_Commission', np.nan)
                                voucher_fmt = f"{float(voucher_valor):.0f}%" if pd.notna(voucher_valor) else "N/A"
                                data_preco = lab_info.get('Data_Preco_Atualizacao')
                                if isinstance(data_preco, pd.Timestamp):
                                    data_preco_fmt = data_preco.tz_localize(None).strftime("%d/%m/%Y %H:%M")
                                else:
                                    data_preco_fmt = "N/A"

                                if price_cards:
                                    cards_html = "".join(price_cards)
                                else:
                                    cards_html = (
                                        "<div style=\"background: white; border-radius: 8px; padding: 1rem; "
                                        "color: #6c757d; text-align: center;\">Nenhum pre√ßo cadastrado.</div>"
                                    )

                                st.markdown(f"""
                                    <div style="background: #f8f9fa; border-radius: 6px; padding: 1rem; margin-bottom: 1rem; border-left: 4px solid #0d6efd;">
                                        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 1rem;">
                                            <div style="font-size: 0.9rem; color: #0d6efd; font-weight: 700; text-transform: uppercase;">Tabela de Pre√ßos</div>
                                            <div style="font-size: 0.8rem; color: #6c757d;">
                                                Atualizado em <strong>{data_preco_fmt}</strong> ‚Ä¢ Voucher: <strong>{voucher_fmt}</strong>
                                            </div>
                                        </div>
                                        <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 1rem;">
                                            {cards_html}
                                        </div>
                                    </div>
                                """, unsafe_allow_html=True)

                    # Informa√ß√µes VIP se dispon√≠vel
                    if info_vip:
                        st.markdown(f"""
                            <div style="background: #f8f9fa; border-radius: 6px; padding: 1rem; margin-bottom: 1rem; border-left: 4px solid #6BBF47;">
                                <div style="display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 1rem; text-align: center;">
                                    <div>
                                        <div style="font-size: 0.8rem; color: #666; margin-bottom: 0.3rem;">RANKING GERAL</div>
                                        <div style="font-size: 1.2rem; font-weight: bold; color: #FFD700;">{info_vip.get('ranking', 'N/A')}</div>
                                    </div>
                                    <div>
                                        <div style="font-size: 0.8rem; color: #666; margin-bottom: 0.3rem;">RANKING REDE</div>
                                        <div style="font-size: 1.2rem; font-weight: bold; color: #FFA500;">{info_vip.get('ranking_rede', 'N/A')}</div>
                                    </div>
                                    <div>
                                        <div style="font-size: 0.8rem; color: #666; margin-bottom: 0.3rem;">REDE</div>
                                        <div style="font-size: 1.1rem; font-weight: bold; color: #6BBF47;">{info_vip.get('rede', 'N/A')}</div>
                                    </div>
                                </div>
                            </div>
                            """, unsafe_allow_html=True)
                    # M√©tricas comerciais essenciais
                    metricas = MetricasAvancadas.calcular_metricas_lab(
                        df_filtrado,
                        lab_cnpj=lab_final_cnpj,
                        lab_nome=lab_final
                    )
                    if metricas:
                        # Dados de Performance
                        st.markdown(f"""
                            <div style="background: #f8f9fa; border-radius: 6px; padding: 1rem; margin-bottom: 1rem; border-left: 4px solid #28a745;">
                                <div style="font-size: 0.9rem; color: #666; margin-bottom: 0.5rem; font-weight: 600;">PERFORMANCE 2025</div>
                                <div style="display: grid; grid-template-columns: repeat(4, 1fr); gap: 1rem; text-align: center;">
                                    <div>
                                        <div style="font-size: 0.8rem; color: #666;">Total Coletas</div>
                                        <div style="font-size: 1.3rem; font-weight: bold; color: #28a745;">{metricas['total_coletas']:,}</div>
                                    </div>
                                    <div>
                                        <div style="font-size: 0.8rem; color: #666;">M√©dia 3 Meses</div>
                                        <div style="font-size: 1.3rem; font-weight: bold; color: #28a745;">{metricas['media_3_meses']:.1f}</div>
                                    </div>
                                    <div>
                                        <div style="font-size: 0.8rem; color: #666;">M√©dia Di√°ria</div>
                                        <div style="font-size: 1.3rem; font-weight: bold; color: #28a745;">{metricas['media_diaria']:.1f}</div>
                                    </div>
                                    <div>
                                        <div style="font-size: 0.8rem; color: #666;">Coletas (Hoje)</div>
                                        <div style="font-size: 1.3rem; font-weight: bold; color: #28a745;">{metricas['vol_hoje']:,}</div>
                                    </div>
                                </div>
                            </div>
                            """, unsafe_allow_html=True)
                        # Status e Risco
                        status_color = "#28a745" if metricas['agudo'] == "Ativo" else "#dc3545"
                        risco_color = "#28a745" if metricas['dias_sem_coleta'] <= 7 else "#ffc107" if metricas['dias_sem_coleta'] <= 30 else "#dc3545"
                        risco_diario = metricas.get('risco_diario', 'N/A')
                        cores_risco = {
                            'üü¢ Normal': '#16A34A',
                            'üü° Aten√ß√£o': '#F59E0B',
                            'üü† Moderado': '#FB923C',
                            'üî¥ Alto': '#DC2626',
                            '‚ö´ Cr√≠tico': '#111827'
                        }
                        risco_diario_color = cores_risco.get(risco_diario, "#6c757d")
                        delta_mm7 = metricas.get('delta_mm7')
                        if isinstance(delta_mm7, (int, float)):
                            delta_mm7_color = "#28a745" if delta_mm7 >= 0 else "#dc3545"
                            delta_mm7_display = f"{delta_mm7:.1f}%"
                        else:
                            delta_mm7_color = "#6c757d"
                            delta_mm7_display = "--"
                        mm7_val = metricas.get('mm7')
                        if isinstance(mm7_val, (int, float)):
                            mm7_display = f"{mm7_val:.1f}"
                            mm7_color = "#2563eb"
                        else:
                            mm7_display = "--"
                            mm7_color = "#6c757d"
                        mm30_val = metricas.get('mm30')
                        if isinstance(mm30_val, (int, float)):
                            mm30_display = f"{mm30_val:.1f}"
                            mm30_color = "#2563eb"
                        else:
                            mm30_display = "--"
                            mm30_color = "#6c757d"
                        delta_mm30_val = metricas.get('delta_mm30')
                        if isinstance(delta_mm30_val, (int, float)):
                            delta_mm30_color = "#28a745" if delta_mm30_val >= 0 else "#dc3545"
                            delta_mm30_display = f"{delta_mm30_val:.1f}%"
                        else:
                            delta_mm30_color = "#6c757d"
                            delta_mm30_display = "--"
                     
                        st.markdown(f"""
                            <div style="background: #f8f9fa; border-radius: 6px; padding: 1rem; margin-bottom: 1rem; border-left: 4px solid {risco_color};">
                                <div style="font-size: 0.9rem; color: #666; margin-bottom: 0.5rem; font-weight: 600;">STATUS & RISCO</div>
                                <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(140px, 1fr)); gap: 1rem; text-align: center;">
                                    <div>
                                        <div style="font-size: 0.8rem; color: #666;">Status Atual</div>
                                        <div style="font-size: 1.1rem; font-weight: bold; color: {status_color};">{metricas['agudo']}</div>
                                    </div>
                                    <div>
                                        <div style="font-size: 0.8rem; color: #666;">Dias sem Coleta</div>
                                        <div style="font-size: 1.1rem; font-weight: bold; color: {risco_color};">{metricas['dias_sem_coleta']}</div>
                                    </div>
                                    <div>
                                        <div style="font-size: 0.8rem; color: #666;">Risco Di√°rio</div>
                                        <div style="font-size: 1.1rem; font-weight: bold; color: {risco_diario_color};">{risco_diario}</div>
                                    </div>
                                    <div>
                                        <div style="font-size: 0.8rem; color: #666;">Œî vs MM7</div>
                                        <div style="font-size: 1.1rem; font-weight: bold; color: {delta_mm7_color};">{delta_mm7_display}</div>
                                    </div>
                                    <div>
                                        <div style="font-size: 0.8rem; color: #666;">MM7 (M√©dia 7d)</div>
                                        <div style="font-size: 1.1rem; font-weight: bold; color: {mm7_color};">{mm7_display}</div>
                                    </div>
                                    <div>
                                        <div style="font-size: 0.8rem; color: #666;">MM30 (M√©dia 30d)</div>
                                        <div style="font-size: 1.1rem; font-weight: bold; color: {mm30_color};">{mm30_display}</div>
                                    </div>
                                    <div>
                                        <div style="font-size: 0.8rem; color: #666;">Œî vs MM30</div>
                                        <div style="font-size: 1.1rem; font-weight: bold; color: {delta_mm30_color};">{delta_mm30_display}</div>
                                    </div>
                                </div>
                            </div>
                            """, unsafe_allow_html=True)
                        # Hist√≥rico de Performance - Reorganizado conforme solicita√ß√£o
                        # Calcular m√°xima de coletas hist√≥rica (respeitando meses dispon√≠veis)
                        metricas_evolucao = MetricasAvancadas.calcular_metricas_evolucao(
                            df_filtrado,
                            lab_cnpj=lab_final_cnpj,
                            lab_nome=lab_final
                        )
                        st.markdown(f"""
                            <div style="background: #f8f9fa; border-radius: 6px; padding: 1rem; margin-bottom: 1rem; border-left: 4px solid #17a2b8;">
                                <div style="font-size: 0.9rem; color: #666; margin-bottom: 0.5rem; font-weight: 600;">HIST√ìRICO DE PERFORMANCE</div>
                                <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 1rem; text-align: center;">
                                    <div>
                                        <div style="font-size: 0.8rem; color: #666;">M√©dia 2024</div>
                                        <div style="font-size: 1.3rem; font-weight: bold; color: #17a2b8;">{metricas_evolucao['media_2024']:.1f}</div>
                                    </div>
                                    <div>
                                        <div style="font-size: 0.8rem; color: #666;">M√©dia 2025</div>
                                        <div style="font-size: 1.3rem; font-weight: bold; color: #17a2b8;">{metricas_evolucao['media_2025']:.1f}</div>
                                    </div>
                                    <div>
                                        <div style="font-size: 0.8rem; color: #666;">M√°xima 2024</div>
                                        <div style="font-size: 1.3rem; font-weight: bold; color: #17a2b8;">{metricas_evolucao['max_2024']:,}</div>
                                    </div>
                                    <div>
                                        <div style="font-size: 0.8rem; color: #666;">M√°xima 2025</div>
                                        <div style="font-size: 1.3rem; font-weight: bold; color: #17a2b8;">{metricas_evolucao['max_2025']:,}</div>
                                    </div>
                                </div>
                            </div>
                            """, unsafe_allow_html=True)
                        # Adiciona tamb√©m os cards de Totais e Comparativos ao bloco de Hist√≥rico
                        st.markdown(f"""
                        <div style="background: #f8f9fa; border-radius: 6px; padding: 1rem; margin-bottom: 1rem; border-left: 4px solid #28a745;">
                            <div style="font-size: 0.9rem; color: #666; margin-bottom: 0.5rem; font-weight: 600;">TOTAIS DE COLETAS</div>
                            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 1rem; text-align: center;">
                                <div>
                                    <div style="font-size: 0.8rem; color: #666;">Total 2024</div>
                                    <div style="font-size: 1.4rem; font-weight: bold; color: #28a745;">{metricas_evolucao['total_coletas_2024']:,}</div>
                                </div>
                                <div>
                                    <div style="font-size: 0.8rem; color: #666;">Total 2025</div>
                                    <div style="font-size: 1.4rem; font-weight: bold; color: #6BBF47;">{metricas_evolucao['total_coletas_2025']:,}</div>
                                </div>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)

                        variacao_ultimo_vs_media = ((metricas_evolucao['media_ultimo_mes'] - metricas_evolucao['media_comparacao']) / metricas_evolucao['media_comparacao'] * 100) if metricas_evolucao['media_comparacao'] > 0 else 0
                        percentual_maxima = (metricas_evolucao['media_ultimo_mes'] / metricas_evolucao['max_comparacao'] * 100) if metricas_evolucao['max_comparacao'] > 0 else 0
                        cor_variacao = "#28a745" if variacao_ultimo_vs_media >= 0 else "#dc3545"
                        cor_percentual = "#28a745" if percentual_maxima >= 80 else "#ffc107" if percentual_maxima >= 50 else "#dc3545"
                        ano_comp = metricas_evolucao['ano_comparacao']
                        st.markdown(f"""
                        <div style="background: #f8f9fa; border-radius: 6px; padding: 1rem; margin-bottom: 1rem; border-left: 4px solid #6f42c1;">
                            <div style="font-size: 0.9rem; color: #666; margin-bottom: 0.5rem; font-weight: 600;">COMPARATIVOS</div>
                            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 1rem; text-align: center;">
                                <div>
                                    <div style="font-size: 0.8rem; color: #666;">√öltimo M√™s ({metricas_evolucao['ultimo_mes_display']}) vs M√©dia {ano_comp}</div>
                                    <div style="font-size: 1.2rem; font-weight: bold; color: {cor_variacao};">
                                        {'+' if variacao_ultimo_vs_media >= 0 else ''}{variacao_ultimo_vs_media:.1f}%
                                    </div>
                                    <div style="font-size: 0.7rem; color: #666;">{metricas_evolucao['media_ultimo_mes']:,} vs {metricas_evolucao['media_comparacao']:.1f}</div>
                                </div>
                                <div>
                                    <div style="font-size: 0.8rem; color: #666;">√öltimo M√™s ({metricas_evolucao['ultimo_mes_display']}) vs M√°xima {ano_comp}</div>
                                    <div style="font-size: 1.2rem; font-weight: bold; color: {cor_percentual};">
                                        {percentual_maxima:.1f}%
                                    </div>
                                    <div style="font-size: 0.7rem; color: #666;">{metricas_evolucao['media_ultimo_mes']:,} vs {metricas_evolucao['max_comparacao']:,}</div>
                                </div>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                        # Gr√°fico de Evolu√ß√£o Mensal abaixo dos comparativos
                        st.markdown("---")
                        st.subheader("üìà Evolu√ß√£o Mensal")
                        ChartManager.criar_grafico_evolucao_mensal(
                            df_filtrado,
                            lab_cnpj=lab_final_cnpj,
                            lab_nome=lab_final,
                            chart_key="historico"
                        )
                    st.markdown("</div>", unsafe_allow_html=True)
                    
                    # Se√ß√£o de Gr√°ficos com Abas - Refatorado conforme solicita√ß√£o
                    st.markdown("""
                        <div style="background: white; border-radius: 8px; padding: 1.5rem; margin-bottom: 2rem;
                                    border: 1px solid #e9ecef; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                            <h3 style="margin: 0 0 1rem 0; color: #2c3e50; font-weight: 600; border-bottom: 2px solid #6BBF47; padding-bottom: 0.5rem;">
                                üìä An√°lise Visual Detalhada
                            </h3>
                        """, unsafe_allow_html=True)
                    
                    # Criar abas para organizar os gr√°ficos (Resumo Executivo removido)
                    tab_distribuicao, tab_media_diaria, tab_coletas_dia = st.tabs([
                        "üìä Distribui√ß√£o por Dia √ötil", "üìÖ M√©dia Di√°ria", "üìà Coletas por Dia √ötil"
                    ])
                    
                    with tab_distribuicao:
                        st.subheader("üìä Distribui√ß√£o de Coletas por Dia √ötil da Semana")
                        # Gr√°fico com destaque maior conforme solicitado
                        ChartManager.criar_grafico_media_dia_semana_novo(
                            df_filtrado,
                            lab_cnpj=lab_final_cnpj,
                            lab_nome=lab_final,
                            filtros=filtros
                        )
                    
                    with tab_media_diaria:
                        st.subheader("üìä M√©dia Di√°ria por M√™s")
                        ChartManager.criar_grafico_media_diaria(
                            df_filtrado,
                            lab_cnpj=lab_final_cnpj,
                            lab_nome=lab_final
                        )

                    with tab_coletas_dia:
                        st.subheader("üìà Coletas por Dia √ötil do M√™s")
                        ChartManager.criar_grafico_coletas_por_dia(
                            df_filtrado,
                            lab_cnpj=lab_final_cnpj,
                            lab_nome=lab_final
                        )

        # Conte√∫do √∫nico da an√°lise detalhada
        # Carregar dados VIP para an√°lise de rede
        df_vip_tabela = DataManager.carregar_dados_vip()
        # Adicionar informa√ß√µes de rede se dispon√≠vel
        df_tabela = df_filtrado.copy()
        mostrar_rede = False
        if df_vip_tabela is not None and not df_vip_tabela.empty:
            # Merge dos dados com informa√ß√µes VIP
            df_tabela['CNPJ_Normalizado'] = df_tabela['CNPJ_PCL'].apply(
                lambda x: ''.join(filter(str.isdigit, str(x))) if pd.notna(x) else ''
            )
            df_vip_tabela['CNPJ_Normalizado'] = df_vip_tabela['CNPJ'].apply(
                lambda x: ''.join(filter(str.isdigit, str(x))) if pd.notna(x) else ''
            )
            # Verificar quais colunas VIP est√£o dispon√≠veis
            colunas_vip_disponiveis = ['CNPJ_Normalizado']
            colunas_vip_opcionais = ['Rede', 'Ranking', 'Ranking Rede']
            for col in colunas_vip_opcionais:
                if col in df_vip_tabela.columns:
                    colunas_vip_disponiveis.append(col)
            # Fazer merge apenas com colunas dispon√≠veis
            if len(colunas_vip_disponiveis) > 1: # Mais que apenas CNPJ_Normalizado
                df_tabela = df_tabela.merge(
                    df_vip_tabela[colunas_vip_disponiveis],
                    on='CNPJ_Normalizado',
                    how='left'
                )
                mostrar_rede = 'Rede' in colunas_vip_disponiveis
            else:
                # Se n√£o h√° colunas VIP dispon√≠veis, n√£o fazer merge
                mostrar_rede = False
        # Filtro por rede (simplificado)
        if mostrar_rede and 'Rede' in df_tabela.columns:
            redes_disponiveis = ["Todas"] + sorted(df_tabela['Rede'].dropna().unique().tolist())
            # Usar rede do laborat√≥rio pesquisado como padr√£o, se dispon√≠vel
            rede_padrao = st.session_state.get('rede_lab_pesquisado', "Todas")
            if rede_padrao not in redes_disponiveis:
                rede_padrao = "Todas"
            # Aplicar filtro autom√°tico se h√° rede selecionada
            if rede_padrao != "Todas":
                rede_filtro = rede_padrao
                # Mostrar indicador de filtro autom√°tico
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, #e8f5e9, #f1f8f1); border-radius: 6px; padding: 0.8rem; margin-bottom: 1rem;">
                    <span style="color: #52B54B; font-size: 0.9rem;">üéØ <strong>Filtro autom√°tico ativo:</strong> mostrando apenas laborat√≥rios da rede <strong>"{rede_padrao}"</strong></span>
                </div>
                """, unsafe_allow_html=True)

                # Bot√£o para limpar filtro autom√°tico
                if st.button("üîÑ Mostrar Todas as Redes", key="limpar_filtro_auto", help="Mostrar laborat√≥rios de todas as redes"):
                    st.session_state['rede_lab_pesquisado'] = None
                    st.toast("‚úÖ Filtro de rede limpo! Todas as redes ser√£o exibidas.")
            else:
                # Sele√ß√£o manual de rede
                rede_filtro = st.selectbox(
                    "üè¢ Filtrar por Rede:",
                    options=redes_disponiveis,
                    index=0, # Sempre "Todas" por padr√£o
                    help="Selecione uma rede para filtrar",
                    key="filtro_rede_tabela"
                )
        else:
            rede_filtro = "Todas"
        # Aplicar filtros
        df_tabela_filtrada = df_tabela.copy()
        # Filtro por rede
        if rede_filtro != "Todas" and mostrar_rede:
            df_tabela_filtrada = df_tabela_filtrada[df_tabela_filtrada['Rede'] == rede_filtro]
        # Mostrar informa√ß√µes da rede se filtrada
        if rede_filtro != "Todas" and mostrar_rede and not df_tabela_filtrada.empty:
            # Estat√≠sticas da rede
            stats_rede = {
                'total_labs': len(df_tabela_filtrada),
                'volume_total': df_tabela_filtrada['Volume_Total_2025'].sum() if 'Volume_Total_2025' in df_tabela_filtrada.columns else 0,
                'media_volume': df_tabela_filtrada['Volume_Total_2025'].mean() if 'Volume_Total_2025' in df_tabela_filtrada.columns else 0,
                'labs_risco_alto': (
                    df_tabela_filtrada['Risco_Diario'].isin(['üî¥ Alto', '‚ö´ Cr√≠tico']).sum()
                    if 'Risco_Diario' in df_tabela_filtrada.columns else 0
                ),
                'labs_ativos': len(df_tabela_filtrada[df_tabela_filtrada['Dias_Sem_Coleta'] <= 30]) if 'Dias_Sem_Coleta' in df_tabela_filtrada.columns else 0
            }
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, #e8f5e9, #f1f8f1); border-radius: 8px; padding: 1rem; margin-bottom: 1rem;">
                <h4 style="margin: 0 0 0.5rem 0; color: #52B54B;">üìä Estat√≠sticas da Rede: {rede_filtro}</h4>
                <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 1rem;">
                    <div style="text-align: center;">
                        <div style="font-size: 1.5rem; font-weight: bold; color: #6BBF47;">{stats_rede['total_labs']}</div>
                        <div style="font-size: 0.8rem; color: #666;">Laborat√≥rios</div>
                    </div>
                    <div style="text-align: center;">
                        <div style="font-size: 1.5rem; font-weight: bold; color: #6BBF47;">{stats_rede['volume_total']:,.0f}</div>
                        <div style="font-size: 0.8rem; color: #666;">Volume Total</div>
                    </div>
                    <div style="text-align: center;">
                        <div style="font-size: 1.5rem; font-weight: bold; color: #6BBF47;">{stats_rede['media_volume']:.0f}</div>
                        <div style="font-size: 0.8rem; color: #666;">M√©dia por Lab</div>
                    </div>
                    <div style="text-align: center;">
                        <div style="font-size: 1.5rem; font-weight: bold; color: #f44336;">{stats_rede['labs_risco_alto']}</div>
                        <div style="font-size: 0.8rem; color: #666;">Alto Risco</div>
                    </div>
                    <div style="text-align: center;">
                        <div style="font-size: 1.5rem; font-weight: bold; color: #4caf50;">{stats_rede['labs_ativos']}</div>
                        <div style="font-size: 0.8rem; color: #666;">Ativos (30d)</div>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)

        # Configurar colunas da tabela
        colunas_principais = [
            'CNPJ_PCL', 'Nome_Fantasia_PCL', 'Estado', 'Cidade', 'Representante_Nome',
            'Risco_Diario', 'Dias_Sem_Coleta',
            'Volume_Atual_2025', 'Volume_Maximo_2024', 'Tendencia_Volume',
            # Ordem l√≥gica: (Hoje), (D-1, ŒîD-1), (MM7, ŒîMM7), (MM30, ŒîMM30)
            'Vol_Hoje',
            'Vol_D1', 'Delta_D1',
            'MM7', 'Delta_MM7',
            'MM30', 'Delta_MM30',
            # M√©dias m√≥veis adicionais no final
            'MM90', 'Delta_MM90'
        ]

        # Adicionar colunas de coletas mensais (2024 e 2025)
        meses_nomes = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
        # Mapeamento dos c√≥digos dos meses para nomes completos em portugu√™s
        meses_nomes_completos = {
            "Jan": "Janeiro", "Fev": "Fevereiro", "Mar": "Mar√ßo", "Abr": "Abril",
            "Mai": "Maio", "Jun": "Junho", "Jul": "Julho", "Ago": "Agosto",
            "Set": "Setembro", "Out": "Outubro", "Nov": "Novembro", "Dez": "Dezembro"
        }
        mes_limite_2025 = min(datetime.now().month, 12)
        
        # Colunas de 2024 (todos os meses)
        cols_2024 = [f'N_Coletas_{m}_24' for m in meses_nomes]
        # Colunas de 2025 (at√© o m√™s atual)
        cols_2025 = [f'N_Coletas_{m}_25' for m in meses_nomes[:mes_limite_2025]]

        colunas_principais.extend(cols_2024 + cols_2025)

        # Adicionar colunas de rede se dispon√≠vel
        if mostrar_rede:
            colunas_principais.extend(['Rede', 'Ranking', 'Ranking Rede'])
        colunas_existentes = [col for col in colunas_principais if col in df_tabela_filtrada.columns]
        if not df_tabela_filtrada.empty and colunas_existentes:
            df_exibicao = df_tabela_filtrada[colunas_existentes].copy()
            # Formata√ß√£o de colunas
            if 'Variacao_Percentual' in df_exibicao.columns:
                df_exibicao['Variacao_Percentual'] = df_exibicao['Variacao_Percentual'].round(2)
            if 'Volume_Atual_2025' in df_exibicao.columns:
                df_exibicao['Volume_Atual_2025'] = df_exibicao['Volume_Atual_2025'].astype(int)
            if 'Volume_Maximo_2024' in df_exibicao.columns:
                df_exibicao['Volume_Maximo_2024'] = df_exibicao['Volume_Maximo_2024'].astype(int)
            # Criar configura√ß√£o de colunas de forma mais expl√≠cita
            column_config = {
                "CNPJ_PCL": st.column_config.TextColumn(
                    "üìÑ CNPJ",
                    help="CNPJ (Cadastro Nacional de Pessoa Jur√≠dica) do laborat√≥rio. Identificador √∫nico para busca e identifica√ß√£o"
                ),
                "Nome_Fantasia_PCL": st.column_config.TextColumn(
                    "üè• Nome Fantasia",
                    help="Nome comercial/fantasia do laborat√≥rio. Use para busca r√°pida por nome"
                ),
                "Estado": st.column_config.TextColumn(
                    "üó∫Ô∏è Estado",
                    help="Estado (UF) onde o laborat√≥rio est√° localizado. Permite filtrar e agrupar por regi√£o geogr√°fica"
                ),
                "Cidade": st.column_config.TextColumn(
                    "üèôÔ∏è Cidade",
                    help="Cidade onde o laborat√≥rio est√° localizado. Permite an√°lise mais granular por localiza√ß√£o"
                ),
                "Representante_Nome": st.column_config.TextColumn(
                    "üë§ Representante",
                    help="Nome do representante comercial respons√°vel pelo laborat√≥rio. √ötil para contato direto e gest√£o de relacionamento"
                ),
                "Risco_Diario": st.column_config.TextColumn(
                    "Risco Di√°rio",
                    help="Classifica√ß√£o de risco di√°ria: üü¢ Normal, üü° Aten√ß√£o, üü† Moderado, üî¥ Alto, ‚ö´ Cr√≠tico. Baseado em volume atual vs. m√©dias m√≥veis e padr√µes hist√≥ricos"
                ),
                "Dias_Sem_Coleta": st.column_config.NumberColumn(
                    "Dias Sem Coleta",
                    help="N√∫mero consecutivo de dias sem registrar coletas. Valores altos indicam poss√≠vel inatividade do laborat√≥rio"
                ),
                # Removido da tabela principal: Varia√ß√£o %
                "Volume_Atual_2025": st.column_config.NumberColumn(
                    "Volume Atual 2025",
                    help="Soma total de coletas em 2025 at√© o momento (todos os meses dispon√≠veis at√© hoje)"
                ),
                "Volume_Maximo_2024": st.column_config.NumberColumn(
                    "Maior M√™s 2024",
                    help="Maior volume mensal de coletas em 2024 (melhor m√™s individual do ano)"
                ),
                "Tendencia_Volume": st.column_config.TextColumn(
                    "Tend√™ncia",
                    help="Tend√™ncia de volume calculada comparando Volume Atual 2025 vs. Maior M√™s 2024: Crescimento (>100%), Decl√≠nio (<50%), Est√°vel (50-100%)"
                )
            }

            column_config.update({
                "Vol_Hoje": st.column_config.NumberColumn(
                    "Coletas (Hoje)",
                    help="Total de coletas registradas na data de refer√™ncia (dia mais recente da s√©rie di√°ria)"
                ),
                "Vol_D1": st.column_config.NumberColumn(
                    "Coletas (D-1)",
                    help="Volume de coletas do dia imediatamente anterior ao atual"
                ),
                "MM7": st.column_config.NumberColumn(
                    "MM7",
                    format="%.3f",
                    help="M√©dia m√≥vel de 7 dias - m√©dia aritm√©tica simples dos √∫ltimos 7 dias (inclui dias sem coleta como zero)"
                ),
                "MM30": st.column_config.NumberColumn(
                    "MM30",
                    format="%.3f",
                    help="M√©dia m√≥vel de 30 dias - m√©dia aritm√©tica simples dos √∫ltimos 30 dias (inclui dias sem coleta como zero)"
                ),
                "MM90": st.column_config.NumberColumn(
                    "MM90",
                    format="%.3f",
                    help="M√©dia m√≥vel de 90 dias - m√©dia aritm√©tica simples dos √∫ltimos 90 dias (inclui dias sem coleta como zero)"
                ),
                "Delta_D1": st.column_config.NumberColumn(
                    "Œî vs D-1",
                    format="%.1f%%",
                    help="Varia√ß√£o percentual: (Vol_Hoje - Vol_D1) / Vol_D1 √ó 100. Indica crescimento ou queda vs. dia anterior"
                ),
                "Delta_MM7": st.column_config.NumberColumn(
                    "Œî vs MM7",
                    format="%.1f%%",
                    help="Varia√ß√£o percentual: (Vol_Hoje - MM7) / MM7 √ó 100. Indica performance vs. m√©dia semanal dos √∫ltimos 7 dias"
                ),
                "Delta_MM30": st.column_config.NumberColumn(
                    "Œî vs MM30",
                    format="%.1f%%",
                    help="Varia√ß√£o percentual: (Vol_Hoje - MM30) / MM30 √ó 100. Indica performance vs. m√©dia mensal dos √∫ltimos 30 dias"
                ),
                "Delta_MM90": st.column_config.NumberColumn(
                    "Œî vs MM90",
                    format="%.1f%%",
                    help="Varia√ß√£o percentual: (Vol_Hoje - MM90) / MM90 √ó 100. Indica performance vs. m√©dia trimestral dos √∫ltimos 90 dias"
                )
            })
            
            # Adicionar configura√ß√µes para colunas mensais de 2024
            for col in cols_2024:
                if col in df_exibicao.columns:
                    mes_codigo = col.split('_')[2]  # Corrigido: pegar o terceiro elemento (√≠ndice 2)
                    mes_nome = meses_nomes_completos.get(mes_codigo, mes_codigo)
                    # Usar configura√ß√£o mais simples
                    column_config[col] = st.column_config.NumberColumn(
                        f"{mes_nome}/24",
                        help=f"Total de coletas realizadas em {mes_nome} de 2024. Permite an√°lise de sazonalidade e compara√ß√£o ano a ano"
                    )
            
            # Adicionar configura√ß√µes para colunas mensais de 2025
            for col in cols_2025:
                if col in df_exibicao.columns:
                    mes_codigo = col.split('_')[2]  # Corrigido: pegar o terceiro elemento (√≠ndice 2)
                    mes_nome = meses_nomes_completos.get(mes_codigo, mes_codigo)
                    # Usar configura√ß√£o mais simples
                    column_config[col] = st.column_config.NumberColumn(
                        f"{mes_nome}/25",
                        help=f"Total de coletas realizadas em {mes_nome} de 2025. Permite acompanhamento do desempenho mensal atual"
                    )
            
            # Adicionar colunas de rede se dispon√≠vel
            if 'Rede' in df_exibicao.columns:
                column_config["Rede"] = st.column_config.TextColumn(
                    "üè¢ Rede",
                    help="Nome da rede √† qual o laborat√≥rio pertence. Permite agrupar e comparar laborat√≥rios da mesma rede"
                )
            if 'Ranking' in df_exibicao.columns:
                column_config["Ranking"] = st.column_config.TextColumn(
                    "üèÜ Ranking",
                    help="Posi√ß√£o do laborat√≥rio no ranking geral por volume de coletas. Ranking 1 = maior volume"
                )
            if 'Ranking_Rede' in df_exibicao.columns:
                column_config["Ranking_Rede"] = st.column_config.TextColumn(
                    "üèÖ Ranking Rede",
                    help="Posi√ß√£o do laborat√≥rio no ranking dentro de sua pr√≥pria rede. Permite identificar l√≠deres regionais por rede"
                )
            
            # Renomear as colunas diretamente no dataframe para exibir nomes completos dos meses
            df_exibicao_renamed = _formatar_df_exibicao(df_exibicao)
            rename_dict = {}
            
            # Renomear colunas principais para nomes mais leg√≠veis
            rename_dict.update({
                "CNPJ_PCL": "CNPJ",
                "Nome_Fantasia_PCL": "Nome Fantasia",
                "Representante_Nome": "Representante",
                "Risco_Diario": "Risco Di√°rio",
                "Dias_Sem_Coleta": "Dias Sem Coleta",
                "Volume_Atual_2025": "Volume Atual 2025",
                "Volume_Maximo_2024": "Maior M√™s 2024",
                "Tendencia_Volume": "Tend√™ncia",
                "Vol_Hoje": "Coletas (Hoje)",
                "Vol_D1": "D-1",
                "MM7": "MM7",
                "MM30": "MM30",
                "MM90": "MM90",
                "Delta_D1": "Œî vs D-1",
                "Delta_MM7": "Œî vs MM7",
                "Delta_MM30": "Œî vs MM30",
                "Delta_MM90": "Œî vs MM90",
                "Ranking_Rede": "Ranking Rede"
            })
            
            # Renomear colunas de 2024
            for col in cols_2024:
                if col in df_exibicao_renamed.columns:
                    mes_codigo = col.split('_')[2]  # Corrigido: pegar o terceiro elemento (√≠ndice 2)
                    mes_nome = meses_nomes_completos.get(mes_codigo, mes_codigo)
                    rename_dict[col] = f"{mes_nome}/24"
            
            # Renomear colunas de 2025
            for col in cols_2025:
                if col in df_exibicao_renamed.columns:
                    mes_codigo = col.split('_')[2]  # Corrigido: pegar o terceiro elemento (√≠ndice 2)
                    mes_nome = meses_nomes_completos.get(mes_codigo, mes_codigo)
                    rename_dict[col] = f"{mes_nome}/25"
            
            df_exibicao_renamed = df_exibicao_renamed.rename(columns=rename_dict)
            
            # Atualizar column_config com os nomes renomeados
            column_config_renamed = {}
            for col_original, config in column_config.items():
                if col_original in rename_dict:
                    col_nomeada = rename_dict[col_original]
                    column_config_renamed[col_nomeada] = config
                elif col_original in df_exibicao_renamed.columns:
                    column_config_renamed[col_original] = config
            
            # Mostrar tabela com contador
            st.markdown(f"**Mostrando {len(df_exibicao_renamed)} laborat√≥rios**")
            st.dataframe(
                df_exibicao_renamed,
                width='stretch',
                height=500,
                hide_index=True,
                column_config=column_config_renamed
            )
            
            # Bot√µes de download
            col_download1, col_download2 = st.columns(2)
            with col_download1:
                csv_data = df_exibicao.to_csv(index=False, encoding='utf-8')
                st.download_button(
                    label="üì• Download CSV",
                    data=csv_data,
                    file_name=f"dados_laboratorios_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv",
                    key="download_csv_tabela"
                )
            with col_download2:
                excel_buffer = BytesIO()
                df_exibicao.to_excel(excel_buffer, index=False, engine='openpyxl')
                excel_data = excel_buffer.getvalue()
                st.download_button(
                    label="üì• Download Excel",
                    data=excel_data,
                    file_name=f"dados_laboratorios_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_excel_tabela"
                )
        else:
            st.info("üìã Nenhum laborat√≥rio encontrado com os filtros aplicados.")
        
        # ========================================
        # DADOS DO CONCORRENTE GRALAB (FINAL DA P√ÅGINA)
        # ========================================
        if lab_final_cnpj:  # S√≥ mostrar se houver laborat√≥rio selecionado
            st.markdown("---")
            st.markdown("""
            <div style="background: linear-gradient(135deg, #ffd700 0%, #ffed4e 100%);
                        color: #333; padding: 2rem; border-radius: 12px;
                        margin: 2rem 0; box-shadow: 0 6px 12px rgba(0,0,0,0.15);">
                <h2 style="margin: 0; font-size: 1.8rem; color: #b8860b; font-weight: 700;">
                    üèÜ Dados no Concorrente Gralab (CunhaLab)
                </h2>
                <p style="margin: 0.5rem 0 0 0; font-size: 1.1rem; color: #666;">
                    Compare os dados deste laborat√≥rio com a base do concorrente
                </p>
            </div>
            """, unsafe_allow_html=True)
            
            try:
                dados_gralab = DataManager.carregar_dados_gralab()
                
                if dados_gralab and 'Dados Completos' in dados_gralab:
                    df_gralab = dados_gralab['Dados Completos']
                    
                    # Buscar laborat√≥rio pelo CNPJ normalizado
                    lab_gralab = df_gralab[df_gralab['CNPJ_Normalizado'] == lab_final_cnpj]
                    
                    if not lab_gralab.empty:
                        lab_g = lab_gralab.iloc[0]
                        
                        # Verificar se est√° na aba EntradaSaida
                        status_movimentacao = ""
                        if 'EntradaSaida' in dados_gralab:
                            df_entrada_saida = dados_gralab['EntradaSaida']
                            lab_entrada = df_entrada_saida[df_entrada_saida['CNPJ_Normalizado'] == lab_final_cnpj]
                            if not lab_entrada.empty:
                                tipo_mov = lab_entrada.iloc[0].get('Tipo Movimenta√ß√£o', '')
                                status_lab = lab_entrada.iloc[0].get('Status', '')
                                if tipo_mov or status_lab:
                                    status_movimentacao = f"<div style='margin-top: 1rem; padding: 1rem; background: #fff3cd; border-radius: 8px; border-left: 4px solid #ffc107;'>"
                                    status_movimentacao += f"<strong style='font-size: 1.1rem;'>Movimenta√ß√£o:</strong> {tipo_mov} | <strong>Status:</strong> {status_lab}</div>"
                        
                        # Extrair pre√ßos
                        preco_cnh = lab_g.get('Pre√ßo CNH', 'N/A')
                        preco_concurso = lab_g.get('Pre√ßo Concurso', 'N/A')
                        preco_clt = lab_g.get('Pre√ßo CLT', 'N/A')
                        
                        # Formatar pre√ßos
                        def formatar_preco(preco):
                            try:
                                if pd.notna(preco) and preco != '' and preco != 'N/A':
                                    return f"R$ {float(preco):.2f}"
                                return "N/A"
                            except:
                                return "N/A"
                        
                        preco_cnh_fmt = formatar_preco(preco_cnh)
                        preco_concurso_fmt = formatar_preco(preco_concurso)
                        preco_clt_fmt = formatar_preco(preco_clt)
                        
                        st.markdown(f"""
                        <div style="background: linear-gradient(135deg, #fff8dc 0%, #fffacd 100%); 
                                    border-radius: 12px; padding: 2rem; margin: 1rem 0 2rem 0;
                                    border: 3px solid #ffd700; box-shadow: 0 4px 12px rgba(255,215,0,0.4);">
                            <h3 style="margin: 0 0 1.5rem 0; color: #b8860b; font-weight: 700; font-size: 1.5rem;">
                                ‚úÖ Laborat√≥rio Encontrado na Base do Gralab (CunhaLab)
                            </h3>
                            <div style="background: white; border-radius: 10px; padding: 1.5rem; margin-bottom: 1rem; box-shadow: 0 2px 6px rgba(0,0,0,0.1);">
                                <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 1.5rem;">
                                    <div>
                                        <div style="font-size: 0.9rem; color: #666; margin-bottom: 0.5rem; font-weight: 600;">NOME</div>
                                        <div style="font-size: 1.2rem; font-weight: 700; color: #2c3e50;">{lab_g.get('Nome', 'N/A')}</div>
                                    </div>
                                    <div>
                                        <div style="font-size: 0.9rem; color: #666; margin-bottom: 0.5rem; font-weight: 600;">CIDADE / UF</div>
                                        <div style="font-size: 1.2rem; font-weight: 700; color: #2c3e50;">{lab_g.get('Cidade', 'N/A')} / {lab_g.get('UF', 'N/A')}</div>
                                    </div>
                                    <div>
                                        <div style="font-size: 0.9rem; color: #666; margin-bottom: 0.5rem; font-weight: 600;">TELEFONE</div>
                                        <div style="font-size: 1.1rem; font-weight: 600; color: #2c3e50;">{lab_g.get('Telefone', 'N/A')}</div>
                                    </div>
                                    <div>
                                        <div style="font-size: 0.9rem; color: #666; margin-bottom: 0.5rem; font-weight: 600;">ENDERE√áO</div>
                                        <div style="font-size: 1rem; color: #2c3e50;">{lab_g.get('Endereco', 'N/A')[:60]}...</div>
                                    </div>
                                </div>
                            </div>
                            <div style="background: linear-gradient(135deg, #e3f2fd 0%, #bbdefb 100%); 
                                        border-radius: 10px; padding: 1.5rem; margin-top: 1rem; 
                                        border-left: 5px solid #2196f3; box-shadow: 0 2px 6px rgba(33,150,243,0.3);">
                                <div style="font-size: 1.1rem; color: #1565c0; margin-bottom: 1rem; font-weight: 700;">üí∞ PRE√áOS PRATICADOS PELO CONCORRENTE</div>
                                <div style="display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 1.5rem; text-align: center;">
                                    <div style="background: white; border-radius: 8px; padding: 1rem; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                                        <div style="font-size: 0.9rem; color: #666; margin-bottom: 0.5rem; font-weight: 600;">üé´ CNH</div>
                                        <div style="font-size: 1.5rem; font-weight: bold; color: #2196f3;">{preco_cnh_fmt}</div>
                                    </div>
                                    <div style="background: white; border-radius: 8px; padding: 1rem; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                                        <div style="font-size: 0.9rem; color: #666; margin-bottom: 0.5rem; font-weight: 600;">üìù Concurso</div>
                                        <div style="font-size: 1.5rem; font-weight: bold; color: #2196f3;">{preco_concurso_fmt}</div>
                                    </div>
                                    <div style="background: white; border-radius: 8px; padding: 1rem; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                                        <div style="font-size: 0.9rem; color: #666; margin-bottom: 0.5rem; font-weight: 600;">üëî CLT</div>
                                        <div style="font-size: 1.5rem; font-weight: bold; color: #2196f3;">{preco_clt_fmt}</div>
                                    </div>
                                </div>
                            </div>
                            {status_movimentacao}
                        </div>
                        """, unsafe_allow_html=True)
                    else:
                        st.info("‚ÑπÔ∏è Este laborat√≥rio n√£o est√° cadastrado na base do Gralab (CunhaLab)")
                else:
                    st.warning("‚ö†Ô∏è N√£o foi poss√≠vel carregar dados do Gralab (CunhaLab)")
            except Exception as e:
                st.error(f"‚ùå Erro ao carregar dados do Gralab (CunhaLab): {e}")

        # Fechar container principal
        st.markdown("</div>", unsafe_allow_html=True)
    elif st.session_state.page == "üè¢ Ranking Rede":
        st.header("üè¢ Ranking por Rede")
        # Carregar dados VIP para an√°lise de rede
        df_vip = DataManager.carregar_dados_vip()
        if df_vip is not None and not df_vip.empty:
            # Merge dos dados principais com dados VIP
            df_com_rede = df_filtrado.copy()
            # Adicionar coluna CNPJ normalizado para match
            df_com_rede['CNPJ_Normalizado'] = df_com_rede['CNPJ_PCL'].apply(
                lambda x: ''.join(filter(str.isdigit, str(x))) if pd.notna(x) else ''
            )
            df_vip['CNPJ_Normalizado'] = df_vip['CNPJ'].apply(
                lambda x: ''.join(filter(str.isdigit, str(x))) if pd.notna(x) else ''
            )
            # Merge dos dados
            df_com_rede = df_com_rede.merge(
                df_vip[['CNPJ_Normalizado', 'Rede', 'Ranking', 'Ranking Rede']],
                on='CNPJ_Normalizado',
                how='left'
            )
            # Filtros espec√≠ficos para ranking de rede
            st.markdown("""
            <div style="background: linear-gradient(135deg, #6BBF47 0%, #52B54B 100%);
                        color: white; padding: 1rem; border-radius: 8px;
                        margin: 1rem 0; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                <h4 style="margin: 0;">üîç Filtros Gerais de Redes</h4>
            </div>
            """, unsafe_allow_html=True)
            col1, col2, col3 = st.columns(3)
            with col1:
                redes_disponiveis = sorted(df_com_rede['Rede'].dropna().unique())
                rede_selecionada = st.multiselect(
                    "üè¢ Redes:",
                    options=redes_disponiveis,
                    default=redes_disponiveis if len(redes_disponiveis) <= 5 else [],
                    help="Selecione as redes para an√°lise"
                )
            with col2:
                rankings_rede = sorted(df_com_rede['Ranking Rede'].dropna().unique())
                ranking_rede_selecionado = st.multiselect(
                    "üèÖ Ranking Rede:",
                    options=rankings_rede,
                    default=rankings_rede if len(rankings_rede) <= 5 else [],
                    help="Selecione os rankings de rede"
                )
            with col3:
                # Categorias de redes (ouro, prata, bronze, diamante)
                categorias_rede = []
                if 'Ranking Rede' in df_com_rede.columns:
                    df_cats = df_com_rede.copy()
                    df_cats['Categoria_Rede'] = df_cats['Ranking Rede'].apply(
                        lambda x: 'Diamante' if str(x).upper() in ['DIAMANTE', 'DIAMOND'] else
                                 'Ouro' if str(x).upper() in ['OURO', 'GOLD', 'ORO'] else
                                 'Prata' if str(x).upper() in ['PRATA', 'SILVER', 'PLATA'] else
                                 'Bronze' if str(x).upper() in ['BRONZE', 'BRONCE'] else
                                 'Outros'
                    )
                    categorias_rede = sorted(df_cats['Categoria_Rede'].unique())
                categoria_selecionada = st.multiselect(
                    "üèÜ Categoria Rede:",
                    options=categorias_rede,
                    default=categorias_rede if len(categorias_rede) <= 4 else [],
                    help="Filtrar por categoria da rede (Diamante, Ouro, Prata, Bronze)"
                )
            # Quarta coluna para tipo de an√°lise
            col4 = st.columns(1)[0]
            with col4:
                tipo_analise = st.selectbox(
                    "üìä Tipo de An√°lise:",
                    options=["Vis√£o Geral", "Por Volume", "Por Performance", "Por Risco", "üîÑ Compara√ß√£o de Redes"],
                    help="Escolha o tipo de an√°lise a ser realizada"
                )
            # Aplicar filtros
            df_rede_filtrado = df_com_rede.copy()
            # Nota explicativa sobre filtros
            st.info("üí° **Dica:** Use os filtros acima para an√°lise geral. Para explora√ß√£o detalhada de uma rede espec√≠fica, role para baixo at√© a se√ß√£o 'Explorador Detalhado por Rede'.")
            if rede_selecionada:
                df_rede_filtrado = df_rede_filtrado[df_rede_filtrado['Rede'].isin(rede_selecionada)]
            if ranking_rede_selecionado:
                df_rede_filtrado = df_rede_filtrado[df_rede_filtrado['Ranking Rede'].isin(ranking_rede_selecionado)]
            # Aplicar filtro de categoria de rede
            if categoria_selecionada:
                df_cats_filtro = df_rede_filtrado.copy()
                df_cats_filtro['Categoria_Rede'] = df_cats_filtro['Ranking Rede'].apply(
                    lambda x: 'Diamante' if str(x).upper() in ['DIAMANTE', 'DIAMOND'] else
                             'Ouro' if str(x).upper() in ['OURO', 'GOLD', 'ORO'] else
                             'Prata' if str(x).upper() in ['PRATA', 'SILVER', 'PLATA'] else
                             'Bronze' if str(x).upper() in ['BRONZE', 'BRONCE'] else
                             'Outros'
                )
                df_rede_filtrado = df_cats_filtro[df_cats_filtro['Categoria_Rede'].isin(categoria_selecionada)]
            # ========================================
            # C√ÅLCULO GLOBAL DE ESTAT√çSTICAS DE REDES
            # ========================================
            # Calcular rede_stats para uso em todas as an√°lises
            rede_stats = pd.DataFrame() # Inicializar vazio por seguran√ßa
            if not df_rede_filtrado.empty and 'Rede' in df_rede_filtrado.columns:
                # Remover duplicatas baseado no CNPJ antes da contagem
                df_sem_duplicatas_rede = df_rede_filtrado.drop_duplicates(subset=['CNPJ_PCL'], keep='first')
                # Estat√≠sticas expandidas por rede
                rede_stats = df_sem_duplicatas_rede.groupby('Rede').agg(
                    Qtd_Labs=('Nome_Fantasia_PCL', 'count'),
                    Volume_Total=('Volume_Total_2025', 'sum'),
                    Volume_Medio=('Volume_Total_2025', 'mean'),
                    Volume_Std=('Volume_Total_2025', 'std'),
                    Estado_Principal=('Estado', lambda x: x.mode().iloc[0] if not x.mode().empty else 'N/A'),
                    Cidades_Unicas=('Cidade', 'nunique'),
                    Labs_Churn=('Risco_Diario', lambda x: x.isin(['üü† Moderado', 'üî¥ Alto', '‚ö´ Cr√≠tico']).sum())
                ).reset_index()
                # Adicionar mais m√©tricas calculadas
                rede_stats['Taxa_Churn'] = (rede_stats['Labs_Churn'] / rede_stats['Qtd_Labs'] * 100).round(1)
                rede_stats['Volume_por_Lab'] = (rede_stats['Volume_Total'] / rede_stats['Qtd_Labs']).round(0)
                # Adicionar categoria da rede se dispon√≠vel
                if 'Ranking Rede' in df_sem_duplicatas_rede.columns:
                    rede_ranking = df_sem_duplicatas_rede.groupby('Rede')['Ranking Rede'].first().reset_index()
                    rede_stats = rede_stats.merge(rede_ranking, on='Rede', how='left')
                    # Adicionar categoria
                    rede_stats['Categoria_Rede'] = rede_stats['Ranking Rede'].apply(
                        lambda x: 'Diamante' if str(x).upper() in ['DIAMANTE', 'DIAMOND'] else
                                 'Ouro' if str(x).upper() in ['OURO', 'GOLD', 'ORO'] else
                                 'Prata' if str(x).upper() in ['PRATA', 'SILVER', 'PLATA'] else
                                 'Bronze' if str(x).upper() in ['BRONZE', 'BRONCE'] else
                                 'Outros'
                    )
                else:
                    rede_stats['Ranking Rede'] = 'N/A'
                    rede_stats['Categoria_Rede'] = 'N/A'
            if not df_rede_filtrado.empty:
                # An√°lise baseada no tipo selecionado
                if tipo_analise == "Vis√£o Geral":
                    # Cards de m√©tricas gerais
                    col1, col2, col3, col4 = st.columns(4)
                    total_redes = len(rede_stats) if not rede_stats.empty else 0
                    total_labs_rede = rede_stats['Qtd_Labs'].sum() if not rede_stats.empty else 0
                    volume_total_rede = rede_stats['Volume_Total'].sum() if not rede_stats.empty else 0
                    with col1:
                        st.metric("üè¢ Total de Redes", total_redes)
                    with col2:
                        st.metric("üè• Labs nas Redes", f"{total_labs_rede:,}")
                    with col3:
                        st.metric("üì¶ Volume Total", f"{volume_total_rede:,}")
                    with col4:
                        media_por_rede = volume_total_rede / total_redes if total_redes > 0 else 0
                        st.metric("üìä M√©dia por Rede", f"{media_por_rede:,.0f}")
                    # ========================================
                    # CARDS DE LOCALIDADE E VOLUMES
                    # ========================================
                    st.markdown("""
                    <div style="background: linear-gradient(135deg, #6BBF47 0%, #8FD968 100%);
                                color: white; padding: 1rem; border-radius: 8px;
                                margin: 1rem 0; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                        <h4 style="margin: 0;">üìç Distribui√ß√£o por Localidade</h4>
                    </div>
                    """, unsafe_allow_html=True)
                    # Cards de localidade
                    col1, col2, col3, col4, col5, col6 = st.columns(6)
                    # Calcular m√©tricas por estado
                    df_sem_duplicatas_local = df_rede_filtrado.drop_duplicates(subset=['CNPJ_PCL'], keep='first')
                    # N√∫mero total de laborat√≥rios
                    total_labs = len(df_sem_duplicatas_local)
                    # Por estado
                    estados_stats = df_sem_duplicatas_local.groupby('Estado').agg({
                        'Nome_Fantasia_PCL': 'count',
                        'Volume_Total_2025': ['sum', 'mean']
                    }).round(2)
                    # Achatar colunas multi-√≠ndice
                    estados_stats.columns = ['Qtd_Labs', 'Volume_Total', 'Volume_Medio']
                    estados_stats = estados_stats.reset_index()
                    # Top 5 estados por quantidade
                    top_estados = estados_stats.nlargest(5, 'Qtd_Labs')
                    # Por cidade
                    cidades_stats = df_sem_duplicatas_local.groupby('Cidade').agg({
                        'Nome_Fantasia_PCL': 'count',
                        'Volume_Total_2025': ['sum', 'mean']
                    }).round(2)
                    cidades_stats.columns = ['Qtd_Labs', 'Volume_Total', 'Volume_Medio']
                    cidades_stats = cidades_stats.reset_index()
                    # Top 5 cidades por quantidade
                    top_cidades = cidades_stats.nlargest(5, 'Qtd_Labs')
                    with col1:
                        st.metric("üè• Total Labs", f"{total_labs:,}")
                    with col2:
                        total_estados = df_sem_duplicatas_local['Estado'].nunique()
                        st.metric("üó∫Ô∏è Estados", f"{total_estados}")
                    with col3:
                        total_cidades = df_sem_duplicatas_local['Cidade'].nunique()
                        st.metric("üèôÔ∏è Cidades", f"{total_cidades}")
                    with col4:
                        volume_total_3m = df_sem_duplicatas_local['Volume_Total_2025'].sum()
                        st.metric("üì¶ Vol. Total 2025", f"{volume_total_3m:,.0f}")
                    with col5:
                        volume_medio_3m = df_sem_duplicatas_local['Volume_Total_2025'].mean()
                        st.metric("üìä Vol. M√©dio 2025", f"{volume_medio_3m:,.0f}")
                    with col6:
                        volume_medio_por_lab = volume_total_3m / total_labs if total_labs > 0 else 0
                        st.metric("üìà Vol/Lab", f"{volume_medio_por_lab:,.0f}")
                    # Tabelas detalhadas por localidade
                    col1, col2 = st.columns(2)
                    with col1:
                        st.subheader("üìç Top Estados")
                        # Adicionar ranking para top_estados
                        top_estados_display = top_estados.copy()
                        top_estados_display['Ranking'] = range(1, len(top_estados_display) + 1)
                        top_estados_display = top_estados_display[['Ranking', 'Estado', 'Qtd_Labs', 'Volume_Total', 'Volume_Medio']]
                        st.dataframe(
                            top_estados_display,
                            width='stretch',
                            column_config={
                                "Ranking": st.column_config.NumberColumn("üèÜ", width="small", help="Posi√ß√£o no ranking"),
                                "Estado": st.column_config.TextColumn("üèõÔ∏è Estado"),
                                "Qtd_Labs": st.column_config.NumberColumn("üè• Labs"),
                                "Volume_Total": st.column_config.NumberColumn("üì¶ Vol. Total", format="%.0f"),
                                "Volume_Medio": st.column_config.NumberColumn("üìä Vol. M√©dio", format="%.0f")
                            },
                            hide_index=True
                        )
                    with col2:
                        st.subheader("üèôÔ∏è Top Cidades")
                        # Adicionar ranking para top_cidades
                        top_cidades_display = top_cidades.copy()
                        top_cidades_display['Ranking'] = range(1, len(top_cidades_display) + 1)
                        top_cidades_display = top_cidades_display[['Ranking', 'Cidade', 'Qtd_Labs', 'Volume_Total', 'Volume_Medio']]
                        st.dataframe(
                            top_cidades_display,
                            width='stretch',
                            column_config={
                                "Ranking": st.column_config.NumberColumn("üèÜ", width="small", help="Posi√ß√£o no ranking"),
                                "Cidade": st.column_config.TextColumn("üèôÔ∏è Cidade"),
                                "Qtd_Labs": st.column_config.NumberColumn("üè• Labs"),
                                "Volume_Total": st.column_config.NumberColumn("üì¶ Vol. Total", format="%.0f"),
                                "Volume_Medio": st.column_config.NumberColumn("üìä Vol. M√©dio", format="%.0f")
                            },
                            hide_index=True
                        )
                elif tipo_analise == "Por Volume":
                    st.subheader("üì¶ An√°lise por Volume de Coletas")
                    # Ranking de redes por volume - remover duplicatas antes da contagem
                    df_sem_duplicatas_volume = df_rede_filtrado.drop_duplicates(subset=['CNPJ_PCL'], keep='first')
                    volume_por_rede = df_sem_duplicatas_volume.groupby('Rede')['Volume_Total_2025'].agg(['sum', 'mean', 'count']).reset_index()
                    volume_por_rede.columns = ['Rede', 'Volume_Total', 'Volume_Medio', 'Qtd_Labs']
                    volume_por_rede = volume_por_rede.sort_values('Volume_Total', ascending=False)
                    # Gr√°fico de ranking
                    fig_ranking = px.bar(
                        volume_por_rede.head(10),
                        x='Rede',
                        y='Volume_Total',
                        title="üèÜ Top 10 Redes por Volume Total",
                        color='Volume_Medio',
                        color_continuous_scale='Viridis',
                        text='Volume_Total'
                    )
                    fig_ranking.update_traces(texttemplate='%{text:.0f}', textposition='outside')
                    max_volume = volume_por_rede.head(10)['Volume_Total'].max()
                    y_axis_max = max_volume * 1.2 if max_volume > 0 else 10
                    fig_ranking.update_layout(
                        xaxis_tickangle=-45,
                        height=500,
                        margin=dict(l=60, r=60, t=80, b=80),
                        yaxis=dict(range=[0, y_axis_max])
                    )
                    st.plotly_chart(fig_ranking, width='stretch')
                    # Tabela detalhada
                    # Adicionar ranking para volume_por_rede
                    volume_por_rede_display = volume_por_rede.round(2).copy()
                    volume_por_rede_display['Ranking'] = range(1, len(volume_por_rede_display) + 1)
                    volume_por_rede_display = volume_por_rede_display[['Ranking', 'Rede', 'Volume_Total', 'Volume_Medio', 'Qtd_Labs']]
                    st.dataframe(
                        volume_por_rede_display,
                        width='stretch',
                        column_config={
                            "Ranking": st.column_config.NumberColumn("üèÜ", width="small", help="Posi√ß√£o no ranking"),
                            "Rede": st.column_config.TextColumn("üè¢ Rede"),
                            "Volume_Total": st.column_config.NumberColumn("üì¶ Volume Total", format="%.0f"),
                            "Volume_Medio": st.column_config.NumberColumn("üìä Volume M√©dio", format="%.1f"),
                            "Qtd_Labs": st.column_config.NumberColumn("üè• Qtd Labs")
                        },
                        hide_index=True
                    )
                elif tipo_analise == "Por Performance":
                    st.subheader("üìà An√°lise de Performance por Rede")
                    # Performance por rede (baseado em crescimento/variacao) - remover duplicatas
                    if 'Variacao_Percentual' in df_rede_filtrado.columns:
                        df_sem_duplicatas_perf = df_rede_filtrado.drop_duplicates(subset=['CNPJ_PCL'], keep='first')
                        perf_rede = df_sem_duplicatas_perf.groupby('Rede').agg({
                            'Variacao_Percentual': ['mean', 'count'],
                            'Volume_Total_2025': 'sum'
                        }).reset_index()
                        perf_rede.columns = ['Rede', 'Variacao_Media', 'Qtd_Labs', 'Volume_Total']
                        perf_rede = perf_rede.sort_values('Variacao_Media', ascending=False)
                        col1, col2 = st.columns(2)
                        with col1:
                            # Performance por varia√ß√£o
                            fig_perf = px.bar(
                                perf_rede.head(10),
                                x='Rede',
                                y='Variacao_Media',
                                title="üìà Top 10 Redes por Performance (Varia√ß√£o %)",
                                color='Variacao_Media',
                                color_continuous_scale='RdYlGn',
                                text='Variacao_Media'
                            )
                            fig_perf.update_traces(texttemplate='%{text:.1f}%', textposition='outside')
                            max_var = perf_rede.head(10)['Variacao_Media'].max()
                            y_axis_max = max_var * 1.15 if max_var > 0 else 1
                            fig_perf.update_layout(
                                xaxis_tickangle=-45,
                                height=500,
                                margin=dict(l=60, r=60, t=80, b=80),
                                yaxis=dict(range=[0, y_axis_max])
                            )
                            st.plotly_chart(fig_perf, width='stretch')
                        with col2:
                            # Scatter plot: Volume vs Performance
                            fig_scatter = px.scatter(
                                perf_rede,
                                x='Volume_Total',
                                y='Variacao_Media',
                                size='Qtd_Labs',
                                color='Rede',
                                title="üìä Volume vs Performance por Rede",
                                labels={'Volume_Total': 'Volume Total', 'Variacao_Media': 'Varia√ß√£o M√©dia %'}
                            )
                            fig_scatter.update_layout(height=500, margin=dict(l=40, r=40, t=40, b=40))
                            st.plotly_chart(fig_scatter, width='stretch')
                        # Tabela de performance
                        st.dataframe(
                            perf_rede.round(2),
                            width='stretch',
                            column_config={
                                "Rede": st.column_config.TextColumn("üè¢ Rede"),
                                "Variacao_Media": st.column_config.NumberColumn("üìà Varia√ß√£o M√©dia %", format="%.2f%%"),
                                "Qtd_Labs": st.column_config.NumberColumn("üè• Qtd Labs"),
                                "Volume_Total": st.column_config.NumberColumn("üì¶ Volume Total", format="%.0f")
                            },
                            hide_index=True
                        )
                elif tipo_analise == "Por Risco":
                    st.subheader("‚ö†Ô∏è An√°lise de Risco por Rede")
                    if 'Risco_Diario' not in df_rede_filtrado.columns:
                        st.warning("‚ö†Ô∏è Coluna 'Risco_Diario' n√£o encontrada nos dados.")
                    else:
                        df_risco = df_rede_filtrado.drop_duplicates(subset=['CNPJ_PCL'], keep='first')
                        labs_risco = df_risco[df_risco['Risco_Diario'].isin(['üü† Moderado', 'üî¥ Alto', '‚ö´ Cr√≠tico'])]
                        cores_map = {
                            'üü¢ Normal': '#16A34A',
                            'üü° Aten√ß√£o': '#F59E0B',
                            'üü† Moderado': '#FB923C',
                            'üî¥ Alto': '#DC2626',
                            '‚ö´ Cr√≠tico': '#111827'
                        }
                        if labs_risco.empty:
                            st.success("‚úÖ Nenhuma rede com laborat√≥rios em risco elevado.")
                        else:
                            resumo_rede = labs_risco.groupby('Rede').agg(
                                Labs_Risco=('CNPJ_PCL', 'count'),
                                Vol_Hoje_Medio=('Vol_Hoje', 'mean'),
                                Delta_MM7_Medio=('Delta_MM7', 'mean'),
                            Recuperando=('Recuperacao', lambda x: x.fillna(False).astype(bool).sum())
                            ).reset_index()
                            resumo_rede['Delta_MM7_Medio'] = resumo_rede['Delta_MM7_Medio'].round(1)
                            resumo_rede['Vol_Hoje_Medio'] = resumo_rede['Vol_Hoje_Medio'].round(1)
                            resumo_rede = resumo_rede.sort_values(['Labs_Risco', 'Delta_MM7_Medio'], ascending=[False, True])
                            col1, col2 = st.columns(2)
                            with col1:
                                fig_top = px.bar(
                                    resumo_rede.head(10),
                                    x='Labs_Risco',
                                    y='Rede',
                                    orientation='h',
                                    title="üö® Redes com Mais Labs em Risco",
                                    color='Delta_MM7_Medio',
                                    color_continuous_scale='Reds',
                                    text='Labs_Risco'
                                )
                                fig_top.update_traces(texttemplate='%{text}', textposition='outside')
                                fig_top.update_layout(xaxis_title="Laborat√≥rios em risco", yaxis_title="Rede",
                                                      height=500, margin=dict(l=40, r=40, t=40, b=40))
                                st.plotly_chart(fig_top, width='stretch')
                            with col2:
                                resumo_rede_delta = resumo_rede.sort_values('Delta_MM7_Medio')
                                fig_delta = px.bar(
                                    resumo_rede_delta.head(10),
                                    x='Delta_MM7_Medio',
                                    y='Rede',
                                    orientation='h',
                                    title="üìâ Redes com Maior Queda vs MM7",
                                    color='Labs_Risco',
                                    color_continuous_scale='Reds',
                                    text='Delta_MM7_Medio'
                                )
                                fig_delta.update_traces(texttemplate='%{text:.1f}%', textposition='outside')
                                fig_delta.update_layout(xaxis_title="Œî vs MM7 (%)", yaxis_title="Rede",
                                                        height=500, margin=dict(l=40, r=40, t=40, b=40))
                                st.plotly_chart(fig_delta, width='stretch')
                        st.dataframe(
                            resumo_rede,
                            width='stretch',
                            column_config={
                                "Rede": st.column_config.TextColumn("üè¢ Rede"),
                                "Labs_Risco": st.column_config.NumberColumn("üö® Labs em Risco"),
                                "Vol_Hoje_Medio": st.column_config.NumberColumn("üì¶ Vol. M√©dio (Hoje)", format="%.1f"),
                                "Delta_MM7_Medio": st.column_config.NumberColumn("Œî M√©dio vs MM7", format="%.1f%%"),
                                "Recuperando": st.column_config.NumberColumn("üîÅ Em Recupera√ß√£o")
                            },
                            hide_index=True
                        )
                        risco_status = df_risco.groupby(['Rede', 'Risco_Diario']).size().reset_index(name='Qtd')
                        fig_status = px.bar(
                            risco_status,
                            x='Rede',
                            y='Qtd',
                            color='Risco_Diario',
                            title="üìä Distribui√ß√£o de Risco Di√°rio por Rede",
                            color_discrete_map=cores_map,
                            barmode='stack'
                        )
                        fig_status.update_layout(xaxis_tickangle=-45, height=500, margin=dict(l=40, r=40, t=40, b=40))
                        st.plotly_chart(fig_status, width='stretch')
                        # Destaques de risco cr√≠tico
                        redes_criticas = labs_risco[labs_risco['Risco_Diario'] == '‚ö´ Cr√≠tico']['Rede'].value_counts()
                        if not redes_criticas.empty:
                            st.error("üö® Redes com laborat√≥rios em risco cr√≠tico detectadas!")
                            for rede, qtd in redes_criticas.items():
                                st.write(f"‚Ä¢ **{rede}**: {qtd} laborat√≥rio(s) cr√≠tico(s)")
                elif tipo_analise == "üîÑ Compara√ß√£o de Redes":
                    st.subheader("üîÑ Compara√ß√£o Direta de Redes")
                    # Seletor de redes para compara√ß√£o (m√°ximo 5 para legibilidade)
                    redes_para_comparar = st.multiselect(
                        "üè¢ Selecione at√© 5 redes para comparar:",
                        options=sorted(rede_stats['Rede'].unique()),
                        default=sorted(rede_stats['Rede'].unique())[:3] if len(rede_stats) >= 3 else sorted(rede_stats['Rede'].unique()),
                        max_selections=5,
                        help="Escolha as redes que deseja comparar diretamente"
                    )
                    if redes_para_comparar:
                        # Filtrar dados apenas das redes selecionadas
                        redes_comparacao = rede_stats[rede_stats['Rede'].isin(redes_para_comparar)].copy()
                        if not redes_comparacao.empty:
                            # ========================================
                            # DASHBOARD DE COMPARA√á√ÉO
                            # ========================================
                            # Cards de compara√ß√£o r√°pida
                            st.markdown("### üìä Compara√ß√£o R√°pida")
                            col1, col2, col3, col4 = st.columns(4)
                            with col1:
                                maior_qtd = redes_comparacao.loc[redes_comparacao['Qtd_Labs'].idxmax()]
                                st.metric(
                                    "üè• Maior Qtd Labs",
                                    f"{int(maior_qtd['Qtd_Labs'])}",
                                    f"{maior_qtd['Rede'][:15]}..."
                                )
                            with col2:
                                maior_volume = redes_comparacao.loc[redes_comparacao['Volume_Total'].idxmax()]
                                st.metric(
                                    "üì¶ Maior Volume",
                                    f"{maior_volume['Volume_Total']:,.0f}",
                                    f"{maior_volume['Rede'][:15]}..."
                                )
                            with col3:
                                menor_churn = redes_comparacao.loc[redes_comparacao['Taxa_Churn'].idxmin()]
                                st.metric(
                                    "‚úÖ Menor Churn",
                                    f"{menor_churn['Taxa_Churn']:.1f}%",
                                    f"{menor_churn['Rede'][:15]}..."
                                )
                            with col4:
                                maior_risco = redes_comparacao.loc[redes_comparacao['Labs_Churn'].idxmax()]
                                st.metric(
                                    "‚ö†Ô∏è Mais Labs em Risco",
                                    f"{int(maior_risco['Labs_Churn'])}",
                                    f"{maior_risco['Rede'][:15]}..."
                                )
                            # ========================================
                            # GR√ÅFICOS COMPARATIVOS
                            # ========================================
                            st.markdown("### üìà Compara√ß√µes Visuais")
                            # Gr√°fico de barras comparativo - m√∫ltiplas m√©tricas
                            col1, col2 = st.columns(2)
                            with col1:
                                # Compara√ß√£o por quantidade de laborat√≥rios e volume
                                fig_comp1 = go.Figure()
                                for _, rede in redes_comparacao.iterrows():
                                    fig_comp1.add_trace(go.Bar(
                                        name=f"{rede['Rede'][:12]}...",
                                        x=['Labs', 'Volume (k)'],
                                        y=[rede['Qtd_Labs'], rede['Volume_Total']/1000],
                                        text=[f"{int(rede['Qtd_Labs'])}", f"{rede['Volume_Total']/1000:.0f}k"],
                                        textposition='auto',
                                    ))
                                fig_comp1.update_layout(
                                    title="üè• Labs vs üì¶ Volume por Rede",
                                    barmode='group',
                                    height=400
                                )
                                st.plotly_chart(fig_comp1, width='stretch')
                            with col2:
                                # Compara√ß√£o de performance (volume m√©dio e taxa churn)
                                fig_comp2 = go.Figure()
                                for _, rede in redes_comparacao.iterrows():
                                    fig_comp2.add_trace(go.Scatter(
                                        name=f"{rede['Rede'][:12]}...",
                                        x=[rede['Volume_Medio']],
                                        y=[rede['Taxa_Churn']],
                                        mode='markers+text',
                                        text=f"{rede['Rede'][:8]}...",
                                        textposition="top center",
                                        marker=dict(size=15)
                                    ))
                                fig_comp2.update_layout(
                                    title="üí∞ Volume M√©dio vs üìâ Taxa Churn",
                                    xaxis_title="Volume M√©dio por Lab",
                                    yaxis_title="Taxa Churn (%)",
                                    height=400
                                )
                                st.plotly_chart(fig_comp2, width='stretch')
                            # ========================================
                            # TABELA COMPARATIVA DETALHADA
                            # ========================================
                            st.markdown("### üìã Compara√ß√£o Detalhada")
                            # Reordenar colunas para melhor visualiza√ß√£o
                            cols_comparacao = [
                                'Rede', 'Categoria_Rede', 'Qtd_Labs', 'Labs_Churn', 'Taxa_Churn',
                                'Volume_Total', 'Volume_Medio', 'Volume_por_Lab'
                            ]
                            # Adicionar indicadores visuais de risco
                            redes_comparacao_display = redes_comparacao[cols_comparacao].copy()
                            # Fun√ß√£o para adicionar indicadores de risco
                            def adicionar_indicador_risco(row):
                                indicadores = []
                                # Indicador de alto churn
                                if row['Taxa_Churn'] > 30:
                                    indicadores.append("üî¥")
                                elif row['Taxa_Churn'] > 15:
                                    indicadores.append("üü†")
                                else:
                                    indicadores.append("üü¢")
                                # Indicador de concentra√ß√£o de labs em risco
                                proporcao_risco = (row['Labs_Churn'] / row['Qtd_Labs']) if row['Qtd_Labs'] else 0
                                if proporcao_risco >= 0.5:
                                    indicadores.append("‚ö†Ô∏è")
                                elif proporcao_risco >= 0.3:
                                    indicadores.append("‚ö°")
                                # Indicador de baixa efici√™ncia (volume por lab)
                                media_geral = redes_comparacao['Volume_por_Lab'].mean()
                                if row['Volume_por_Lab'] < media_geral * 0.7:
                                    indicadores.append("üìâ")
                                return ' '.join(indicadores) if indicadores else "‚úÖ"
                            redes_comparacao_display['üö® Indicadores'] = redes_comparacao_display.apply(adicionar_indicador_risco, axis=1)
                            # Reordenar para colocar indicadores primeiro
                            cols_final = ['üö® Indicadores'] + cols_comparacao
                            redes_comparacao_display = redes_comparacao_display[cols_final]
                            st.dataframe(
                                redes_comparacao_display.round(2),
                                width='stretch',
                                column_config={
                                    "üö® Indicadores": st.column_config.TextColumn("üö® Alertas", width="small"),
                                    "Rede": st.column_config.TextColumn("üè¢ Rede", width="medium"),
                                    "Categoria_Rede": st.column_config.TextColumn("üèÜ Categoria", width="small"),
                                    "Qtd_Labs": st.column_config.NumberColumn("üè• Labs", format="%d"),
                                    "Labs_Churn": st.column_config.NumberColumn("‚ùå Churn", format="%d"),
                                    "Taxa_Churn": st.column_config.NumberColumn("üìâ % Churn", format="%.1f%%"),
                                    "Volume_Total": st.column_config.NumberColumn("üì¶ Vol. Total", format="%.0f"),
                                    "Volume_Medio": st.column_config.NumberColumn("üìä Vol. M√©dio", format="%.0f"),
                                    "Volume_por_Lab": st.column_config.NumberColumn("üí∞ Vol/Lab", format="%.0f")
                                },
                                hide_index=True
                            )
                            # ========================================
                            # RANKING COMPARATIVO
                            # ========================================
                            st.markdown("### üèÜ Rankings Comparativos")
                            col1, col2, col3 = st.columns(3)
                            with col1:
                                st.subheader("ü•á Por Volume Total")
                                ranking_volume = redes_comparacao.sort_values('Volume_Total', ascending=False)[['Rede', 'Volume_Total']]
                                for idx, row in ranking_volume.iterrows():
                                    medal = "ü•á" if idx == 0 else "ü•à" if idx == 1 else "ü•â" if idx == 2 else "üìä"
                                    st.write(f"{medal} {row['Rede'][:20]}...: {row['Volume_Total']:,.0f}")
                            with col2:
                                st.subheader("ü•á Por Efici√™ncia")
                                ranking_eficiencia = redes_comparacao.sort_values('Volume_por_Lab', ascending=False)[['Rede', 'Volume_por_Lab']]
                                for idx, row in ranking_eficiencia.iterrows():
                                    medal = "ü•á" if idx == 0 else "ü•à" if idx == 1 else "ü•â" if idx == 2 else "üìä"
                                    st.write(f"{medal} {row['Rede'][:20]}...: {row['Volume_por_Lab']:,.0f}")
                            with col3:
                                st.subheader("ü•á Por Menor Risco")
                                ranking_risco = redes_comparacao.sort_values('Taxa_Churn', ascending=True)[['Rede', 'Taxa_Churn']]
                                for idx, row in ranking_risco.iterrows():
                                    medal = "ü•á" if idx == 0 else "ü•à" if idx == 1 else "ü•â" if idx == 2 else "üìä"
                                    st.write(f"{medal} {row['Rede'][:20]}...: {row['Taxa_Churn']:.1f}%")
                        else:
                            st.warning("‚ö†Ô∏è Nenhuma rede encontrada com os crit√©rios selecionados.")
                    else:
                        st.info("‚ÑπÔ∏è Selecione pelo menos uma rede para iniciar a compara√ß√£o.")
            else:
                st.warning("‚ö†Ô∏è Nenhum dado encontrado com os filtros aplicados.")
        else:
            st.warning("‚ö†Ô∏è Dados VIP n√£o dispon√≠veis. Verifique se o arquivo Excel foi carregado corretamente.")
    elif st.session_state.page == "üîß Manuten√ß√£o VIPs":
        st.header("üîß Manuten√ß√£o de Dados VIP")
        st.markdown("""
        <div style="background: linear-gradient(135deg, #6BBF47 0%, #52B54B 100%);
                    color: white; padding: 1rem; border-radius: 8px; margin-bottom: 2rem;">
            <h3 style="margin: 0; color: white;">Gerenciamento de Laborat√≥rios VIP</h3>
            <p style="margin: 0.5rem 0 0 0; opacity: 0.9;">Adicione, edite e gerencie laborat√≥rios VIP com valida√ß√£o completa e hist√≥rico de altera√ß√µes.</p>
        </div>
        """, unsafe_allow_html=True)
     
        # Importar m√≥dulos necess√°rios
        try:
            from vip_history_manager import VIPHistoryManager
            from vip_integration import VIPIntegration
            import json
            import shutil
        except ImportError as e:
            st.error(f"Erro ao importar m√≥dulos VIP: {e}")
            st.stop()
     
        # Inicializar gerenciadores
        history_manager = VIPHistoryManager(OUTPUT_DIR)
        vip_integration = VIPIntegration(OUTPUT_DIR)
     
        # Sub-abas para diferentes funcionalidades
        sub_tab1, sub_tab2, sub_tab3, sub_tab4 = st.tabs([
            "üìã Visualizar VIPs",
            "‚ûï Adicionar VIP",
            "‚úèÔ∏è Editar VIP",
            "üìä Hist√≥rico"
        ])
     
        with sub_tab1:
            st.subheader("üìã Lista de Laborat√≥rios VIP")
         
            # Carregar dados VIP
            df_vip = DataManager.carregar_dados_vip()
         
            if df_vip is not None and not df_vip.empty:
                # Filtros
                col1, col2, col3 = st.columns(3)
             
                with col1:
                    ranking_filtro = st.selectbox(
                        "üèÜ Ranking:",
                        options=["Todos"] + sorted(df_vip['Ranking'].dropna().unique().tolist()),
                        help="Filtrar por ranking individual"
                    )
             
                with col2:
                    ranking_rede_filtro = st.selectbox(
                        "üèÖ Ranking Rede:",
                        options=["Todos"] + sorted(df_vip['Ranking Rede'].dropna().unique().tolist()),
                        help="Filtrar por ranking de rede"
                    )
             
                with col3:
                    rede_filtro = st.selectbox(
                        "üè¢ Rede:",
                        options=["Todas"] + sorted(df_vip['Rede'].dropna().unique().tolist()),
                        help="Filtrar por rede"
                    )
             
                # Aplicar filtros
                df_filtrado = df_vip.copy()
             
                if ranking_filtro != "Todos":
                    df_filtrado = df_filtrado[df_filtrado['Ranking'] == ranking_filtro]
             
                if ranking_rede_filtro != "Todos":
                    df_filtrado = df_filtrado[df_filtrado['Ranking Rede'] == ranking_rede_filtro]
             
                if rede_filtro != "Todas":
                    df_filtrado = df_filtrado[df_filtrado['Rede'] == rede_filtro]
             
                # Estat√≠sticas
                col1, col2, col3, col4 = st.columns(4)
             
                with col1:
                    st.metric("üìä Total VIPs", len(df_filtrado))
             
                with col2:
                    st.metric("üèÜ Rankings", len(df_filtrado['Ranking'].unique()))
             
                with col3:
                    st.metric("üè¢ Redes", len(df_filtrado['Rede'].unique()))
             
                with col4:
                    st.metric("üèÖ Rankings Rede", len(df_filtrado['Ranking Rede'].unique()))
             
                # Tabela de dados
                st.subheader("üìã Dados VIP Filtrados")
             
                # Configurar colunas para exibi√ß√£o
                colunas_exibir = ['CNPJ', 'RAZ√ÉO SOCIAL', 'NOME FANTASIA', 'Cidade ', 'UF',
                                'Ranking', 'Ranking Rede', 'Rede', 'STATUS']
             
                colunas_existentes = [col for col in colunas_exibir if col in df_filtrado.columns]
             
                if colunas_existentes:
                    st.dataframe(
                        df_filtrado[colunas_existentes],
                        width='stretch',
                        height=400,
                        column_config={
                            "CNPJ": st.column_config.TextColumn("üìÑ CNPJ", help="CNPJ do laborat√≥rio"),
                            "RAZ√ÉO SOCIAL": st.column_config.TextColumn("üè¢ Raz√£o Social"),
                            "NOME FANTASIA": st.column_config.TextColumn("üè• Nome Fantasia"),
                            "Cidade ": st.column_config.TextColumn("üèôÔ∏è Cidade"),
                            "UF": st.column_config.TextColumn("üó∫Ô∏è Estado"),
                            "Ranking": st.column_config.TextColumn("üèÜ Ranking"),
                            "Ranking Rede": st.column_config.TextColumn("üèÖ Ranking Rede"),
                            "Rede": st.column_config.TextColumn("üè¢ Rede"),
                            "STATUS": st.column_config.TextColumn("üìä Status")
                        },
                        hide_index=True
                    )
                else:
                    st.warning("Nenhuma coluna v√°lida encontrada para exibi√ß√£o")
            else:
                st.warning("‚ö†Ô∏è Nenhum dado VIP encontrado. Execute primeiro o script de normaliza√ß√£o.")
     
        with sub_tab2:
            st.subheader("‚ûï Adicionar Novo Laborat√≥rio VIP")
         
            # Formul√°rio para adicionar adicionar VIP
            with st.form("form_adicionar_vip"):
                col1, col2 = st.columns(2)
             
                with col1:
                    cnpj_novo = st.text_input(
                        "üìÑ CNPJ:",
                        placeholder="00.000.000/0000-00",
                        help="CNPJ do laborat√≥rio (ser√° validado automaticamente)"
                    )
                 
                    razao_social = st.text_input(
                        "üè¢ Raz√£o Social:",
                        placeholder="Nome da empresa"
                    )
                 
                    nome_fantasia = st.text_input(
                        "üè• Nome Fantasia:",
                        placeholder="Nome comercial"
                    )
                 
                    cidade = st.text_input(
                        "üèôÔ∏è Cidade:",
                        placeholder="Nome da cidade"
                    )
             
                with col2:
                    uf = st.selectbox(
                        "üó∫Ô∏è Estado:",
                        options=[""] + ESTADOS_BRASIL,
                        help="Selecione o estado"
                    )
             
                    ranking = st.selectbox(
                        "üèÜ Ranking:",
                        options=list(CATEGORIAS_RANKING.keys()),
                        help="Ranking individual do laborat√≥rio"
                    )
                 
                    ranking_rede = st.selectbox(
                        "üèÖ Ranking Rede:",
                        options=list(CATEGORIAS_RANKING_REDE.keys()),
                        help="Ranking da rede"
                    )
                 
                    rede = st.text_input(
                        "üè¢ Rede:",
                        placeholder="Nome da rede"
                    )
             
                contato = st.text_input(
                    "üë§ Contato:",
                    placeholder="Nome do contato"
                )
             
                telefone = st.text_input(
                    "üìû Telefone/WhatsApp:",
                    placeholder="(00) 00000-0000"
                )
             
                observacoes = st.text_area(
                    "üìù Observa√ß√µes:",
                    placeholder="Observa√ß√µes adicionais (opcional)"
                )
             
                submitted = st.form_submit_button("‚ûï Adicionar VIP", type="primary")
             
                if submitted:
                    # Valida√ß√µes
                    erros = []
                 
                    # Validar CNPJ
                    if not cnpj_novo:
                        erros.append("CNPJ √© obrigat√≥rio")
                    else:
                        valido, mensagem = vip_integration.validar_cnpj(cnpj_novo)
                        if not valido:
                            erros.append(f"CNPJ inv√°lido: {mensagem}")
                        elif vip_integration.verificar_cnpj_vip_existe(cnpj_novo):
                            erros.append("CNPJ j√° existe na lista VIP")
                 
                    # Validar campos obrigat√≥rios
                    if not razao_social:
                        erros.append("Raz√£o Social √© obrigat√≥rio")
                 
                    if not nome_fantasia:
                        erros.append("Nome Fantasia √© obrigat√≥rio")
                 
                    if not uf:
                        erros.append("Estado √© obrigat√≥rio")
                 
                    if not rede:
                        erros.append("Rede √© obrigat√≥ria")
                 
                    if erros:
                        for erro in erros:
                            st.error(f"‚ùå {erro}")
                    else:
                        # Auto-completar dados se CNPJ existe nos laborat√≥rios
                        dados_lab = vip_integration.buscar_laboratorio_por_cnpj(cnpj_novo)
                        if dados_lab:
                            if not razao_social:
                                razao_social = dados_lab.get('razao_social', '')
                            if not nome_fantasia:
                                nome_fantasia = dados_lab.get('nome_fantasia', '')
                            if not cidade:
                                cidade = dados_lab.get('cidade', '')
                            if not uf:
                                uf = dados_lab.get('estado', '')
                     
                        # Criar backup antes de adicionar
                        if VIP_AUTO_BACKUP:
                            try:
                                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                                backup_path = os.path.join(VIP_BACKUP_DIR, f"vip_backup_{timestamp}.csv")
                                os.makedirs(VIP_BACKUP_DIR, exist_ok=True)
                             
                                if os.path.exists(os.path.join(OUTPUT_DIR, VIP_CSV_FILE)):
                                    shutil.copy2(os.path.join(OUTPUT_DIR, VIP_CSV_FILE), backup_path)
                                    st.toast(f"‚úÖ Backup criado: {backup_path}")
                            except Exception as e:
                                st.warning(f"‚ö†Ô∏è Erro ao criar backup: {e}")
                     
                        # Adicionar novo VIP
                        try:
                            # Carregar dados existentes
                            df_vip_atual = DataManager.carregar_dados_vip()
                            if df_vip_atual is None:
                                df_vip_atual = pd.DataFrame()
                         
                            # Criar novo registro
                            novo_registro = {
                                'CNPJ': cnpj_novo,
                                'RAZ√ÉO SOCIAL': razao_social,
                                'NOME FANTASIA': nome_fantasia,
                                'Cidade ': cidade,
                                'UF': uf,
                                'Contato PCL': contato,
                                'Whatsapp/telefone': telefone,
                                'REP': '', # Ser√° preenchido automaticamente se CNPJ existir
                                'CS': '', # Ser√° preenchido automaticamente se CNPJ existir
                                'STATUS': 'ATIVO',
                                'Ranking': ranking,
                                'Ranking Rede': ranking_rede,
                                'Rede': rede
                            }
                         
                            # Adicionar ao DataFrame
                            df_novo = pd.DataFrame([novo_registro])
                            df_vip_atualizado = pd.concat([df_vip_atual, df_novo], ignore_index=True)
                         
                            # Salvar CSV atualizado
                            caminho_csv = os.path.join(OUTPUT_DIR, VIP_CSV_FILE)
                            df_vip_atualizado.to_csv(caminho_csv, index=False, encoding='utf-8-sig')
                         
                            # Registrar no hist√≥rico
                            history_manager.registrar_insercao(
                                cnpj=cnpj_novo,
                                dados_novos=novo_registro,
                                usuario="streamlit_user",
                                observacoes=observacoes
                            )
                         
                            # Limpar cache
                            DataManager.carregar_dados_vip.clear()
                         
                            st.toast(f"‚úÖ Laborat√≥rio VIP adicionado com sucesso!")
                            st.success(f"üìÑ CNPJ: {cnpj_novo}")
                            st.success(f"üè• Nome: {nome_fantasia}")
                         
                            # Mostrar sugest√µes de laborat√≥rios similares
                            sugestoes = vip_integration.obter_sugestoes_laboratorios(limite=5)
                            if sugestoes:
                                st.info("üí° Outros laborat√≥rios que ainda n√£o s√£o VIP:")
                                for sug in sugestoes[:3]:
                                    st.write(f"‚Ä¢ {sug['nome_fantasia']} ({sug['cnpj']}) - {sug['estado']}")
                         
                        except Exception as e:
                            st.error(f"‚ùå Erro ao adicionar VIP: {e}")
     
        with sub_tab3:
            st.subheader("‚úèÔ∏è Editar Laborat√≥rio VIP")
         
            # Carregar dados VIP
            df_vip = DataManager.carregar_dados_vip()
         
            if df_vip is not None and not df_vip.empty:
                # Selecionar VIP para editar
                col1, col2 = st.columns([2, 1])
             
                with col1:
                    # Busca por CNPJ ou nome
                    busca = st.text_input(
                        "üîç Buscar VIP:",
                        placeholder="Digite CNPJ ou nome do laborat√≥rio"
                    )
             
                with col2:
                    if busca:
                        # Filtrar resultados
                        mask = (
                            df_vip['CNPJ'].str.contains(busca, case=False, na=False) |
                            df_vip['NOME FANTASIA'].str.contains(busca, case=False, na=False) |
                            df_vip['RAZ√ÉO SOCIAL'].str.contains(busca, case=False, na=False)
                        )
                        df_filtrado = df_vip[mask]
                    else:
                        df_filtrado = df_vip
             
                if not df_filtrado.empty:
                    # Selecionar VIP
                    vip_selecionado = st.selectbox(
                        "üìã Selecionar VIP para editar:",
                        options=df_filtrado.index,
                        format_func=lambda x: f"{df_filtrado.loc[x, 'NOME FANTASIA']} - {df_filtrado.loc[x, 'CNPJ']}",
                        help="Selecione o laborat√≥rio VIP para editar"
                    )
                 
                    if vip_selecionado is not None:
                        vip_data = df_filtrado.loc[vip_selecionado]
                     
                        st.markdown("---")
                        st.subheader(f"‚úèÔ∏è Editando: {vip_data['NOME FANTASIA']}")
                     
                        # Formul√°rio de edi√ß√£o
                        with st.form("form_editar_vip"):
                            col1, col2 = st.columns(2)
                         
                            with col1:
                                cnpj_edit = st.text_input(
                                    "üìÑ CNPJ:",
                                    value=vip_data['CNPJ'],
                                    disabled=True, # CNPJ n√£o pode ser alterado
                                    help="CNPJ n√£o pode ser alterado"
                                )
                             
                                razao_social_edit = st.text_input(
                                    "üè¢ Raz√£o Social:",
                                    value=vip_data.get('RAZ√ÉO SOCIAL', '')
                                )
                             
                                nome_fantasia_edit = st.text_input(
                                    "üè• Nome Fantasia:",
                                    value=vip_data.get('NOME FANTASIA', '')
                                )
                             
                                cidade_edit = st.text_input(
                                    "üèôÔ∏è Cidade:",
                                    value=vip_data.get('Cidade ', '')
                                )
                         
                            with col2:
                                uf_edit = st.selectbox(
                                    "üó∫Ô∏è Estado:",
                                    options=ESTADOS_BRASIL,
                                    index=ESTADOS_BRASIL.index(vip_data.get('UF', '')) if vip_data.get('UF', '') in ESTADOS_BRASIL else 0
                                )
                             
                                ranking_edit = st.selectbox(
                                    "üèÜ Ranking:",
                                    options=list(CATEGORIAS_RANKING.keys()),
                                    index=list(CATEGORIAS_RANKING.keys()).index(vip_data.get('Ranking', 'BRONZE')) if vip_data.get('Ranking', '') in CATEGORIAS_RANKING else 0
                                )
                             
                                ranking_rede_edit = st.selectbox(
                                    "üèÖ Ranking Rede:",
                                    options=list(CATEGORIAS_RANKING_REDE.keys()),
                                    index=list(CATEGORIAS_RANKING_REDE.keys()).index(vip_data.get('Ranking Rede', 'BRONZE')) if vip_data.get('Ranking Rede', '') in CATEGORIAS_RANKING_REDE else 0
                                )
                             
                                rede_edit = st.text_input(
                                    "üè¢ Rede:",
                                    value=vip_data.get('Rede', '')
                                )
                         
                            contato_edit = st.text_input(
                                "üë§ Contato:",
                                value=vip_data.get('Contato PCL', '')
                            )
                         
                            telefone_edit = st.text_input(
                                "üìû Telefone/WhatsApp:",
                                value=vip_data.get('Whatsapp/telefone', '')
                            )
                         
                            status_edit = st.selectbox(
                                "üìä Status:",
                                options=['ATIVO', 'INATIVO', 'DELETADO'],
                                index=['ATIVO', 'INATIVO', 'DELETADO'].index(vip_data.get('STATUS', 'ATIVO'))
                            )
                         
                            observacoes_edit = st.text_area(
                                "üìù Observa√ß√µes da Edi√ß√£o:",
                                placeholder="Descreva as altera√ß√µes realizadas"
                            )
                         
                            submitted_edit = st.form_submit_button("üíæ Salvar Altera√ß√µes", type="primary")
                         
                            if submitted_edit:
                                # Verificar se houve altera√ß√µes
                                alteracoes = []
                             
                                if razao_social_edit != vip_data.get('RAZ√ÉO SOCIAL', ''):
                                    alteracoes.append(('RAZ√ÉO SOCIAL', vip_data.get('RAZ√ÉO SOCIAL', ''), razao_social_edit))
                             
                                if nome_fantasia_edit != vip_data.get('NOME FANTASIA', ''):
                                    alteracoes.append(('NOME FANTASIA', vip_data.get('NOME FANTASIA', ''), nome_fantasia_edit))
                             
                                if ranking_edit != vip_data.get('Ranking', ''):
                                    alteracoes.append(('Ranking', vip_data.get('Ranking', ''), ranking_edit))
                             
                                if ranking_rede_edit != vip_data.get('Ranking Rede', ''):
                                    alteracoes.append(('Ranking Rede', vip_data.get('Ranking Rede', ''), ranking_rede_edit))
                             
                                if rede_edit != vip_data.get('Rede', ''):
                                    alteracoes.append(('Rede', vip_data.get('Rede', ''), rede_edit))
                             
                                if status_edit != vip_data.get('STATUS', ''):
                                    alteracoes.append(('STATUS', vip_data.get('STATUS', ''), status_edit))
                             
                                if alteracoes:
                                    # Criar backup antes de editar
                                    if VIP_AUTO_BACKUP:
                                        try:
                                            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                                            backup_path = os.path.join(VIP_BACKUP_DIR, f"vip_backup_{timestamp}.csv")
                                            os.makedirs(VIP_BACKUP_DIR, exist_ok=True)
                                         
                                            if os.path.exists(os.path.join(OUTPUT_DIR, VIP_CSV_FILE)):
                                                shutil.copy2(os.path.join(OUTPUT_DIR, VIP_CSV_FILE), backup_path)
                                                st.toast(f"‚úÖ Backup criado: {backup_path}")
                                        except Exception as e:
                                            st.warning(f"‚ö†Ô∏è Erro ao criar backup: {e}")
                                 
                                    # Atualizar dados
                                    try:
                                        # Atualizar DataFrame
                                        df_vip_atualizado = df_vip.copy()
                                        df_vip_atualizado.loc[vip_selecionado, 'RAZ√ÉO SOCIAL'] = razao_social_edit
                                        df_vip_atualizado.loc[vip_selecionado, 'NOME FANTASIA'] = nome_fantasia_edit
                                        df_vip_atualizado.loc[vip_selecionado, 'Cidade '] = cidade_edit
                                        df_vip_atualizado.loc[vip_selecionado, 'UF'] = uf_edit
                                        df_vip_atualizado.loc[vip_selecionado, 'Ranking'] = ranking_edit
                                        df_vip_atualizado.loc[vip_selecionado, 'Ranking Rede'] = ranking_rede_edit
                                        df_vip_atualizado.loc[vip_selecionado, 'Rede'] = rede_edit
                                        df_vip_atualizado.loc[vip_selecionado, 'Contato PCL'] = contato_edit
                                        df_vip_atualizado.loc[vip_selecionado, 'Whatsapp/telefone'] = telefone_edit
                                        df_vip_atualizado.loc[vip_selecionado, 'STATUS'] = status_edit
                                     
                                        # Salvar CSV atualizado
                                        caminho_csv = os.path.join(OUTPUT_DIR, VIP_CSV_FILE)
                                        df_vip_atualizado.to_csv(caminho_csv, index=False, encoding='utf-8-sig')
                                     
                                        # Registrar altera√ß√µes no hist√≥rico
                                        for campo, valor_anterior, valor_novo in alteracoes:
                                            history_manager.registrar_edicao(
                                                cnpj=vip_data['CNPJ'],
                                                campo_alterado=campo,
                                                valor_anterior=valor_anterior,
                                                valor_novo=valor_novo,
                                                dados_antes=vip_data.to_dict(),
                                                dados_depois=df_vip_atualizado.loc[vip_selecionado].to_dict(),
                                                usuario="streamlit_user",
                                                observacoes=observacoes_edit
                                            )
                                     
                                        # Limpar cache
                                        DataManager.carregar_dados_vip.clear()
                                     
                                        st.toast(f"‚úÖ Laborat√≥rio VIP atualizado com sucesso!")
                                        st.success(f"üìù {len(alteracoes)} campo(s) alterado(s)")
                                     
                                        # Mostrar resumo das altera√ß√µes
                                        for campo, valor_anterior, valor_novo in alteracoes:
                                            st.info(f"üîÑ {campo}: '{valor_anterior}' ‚Üí '{valor_novo}'")
                                     
                                    except Exception as e:
                                        st.error(f"‚ùå Erro ao atualizar VIP: {e}")
                                else:
                                    st.info("‚ÑπÔ∏è Nenhuma altera√ß√£o detectada")
            else:
                st.warning("‚ö†Ô∏è Nenhum dado VIP encontrado. Execute primeiro o script de normaliza√ß√£o.")
     
        with sub_tab4:
            st.subheader("üìä Hist√≥rico de Altera√ß√µes")
         
            # Estat√≠sticas do hist√≥rico
            stats = history_manager.obter_estatisticas()
         
            if stats.get('total_alteracoes', 0) > 0:
                col1, col2, col3, col4 = st.columns(4)
             
                with col1:
                    st.metric("üìä Total Altera√ß√µes", stats['total_alteracoes'])
             
                with col2:
                    st.metric("‚ûï Inser√ß√µes", stats['por_tipo'].get('insercao', 0))
             
                with col3:
                    st.metric("‚úèÔ∏è Edi√ß√µes", stats['por_tipo'].get('edicao', 0))
             
                with col4:
                    st.metric("üóëÔ∏è Exclus√µes", stats['por_tipo'].get('exclusao', 0))
             
                # Filtros para hist√≥rico
                col1, col2, col3 = st.columns(3)
             
                with col1:
                    tipo_filtro = st.selectbox(
                        "üîç Tipo de Altera√ß√£o:",
                        options=["Todos"] + list(stats['por_tipo'].keys()),
                        help="Filtrar por tipo de altera√ß√£o"
                    )
             
                with col2:
                    cnpj_filtro = st.text_input(
                        "üìÑ CNPJ:",
                        placeholder="Digite CNPJ para filtrar",
                        help="Filtrar por CNPJ espec√≠fico"
                    )
             
                with col3:
                    dias_filtro = st.selectbox(
                        "üìÖ Per√≠odo:",
                        options=["Todos", "√öltimos 7 dias", "√öltimos 30 dias", "√öltimos 90 dias"],
                        help="Filtrar por per√≠odo"
                    )
             
                # Obter hist√≥rico filtrado
                if cnpj_filtro:
                    historico_filtrado = history_manager.buscar_historico_cnpj(cnpj_filtro)
                else:
                    historico_filtrado = history_manager.historico
             
                # Filtrar por tipo
                if tipo_filtro != "Todos":
                    historico_filtrado = [alt for alt in historico_filtrado if alt['tipo'] == tipo_filtro]
             
                # Filtrar por per√≠odo
                if dias_filtro != "Todos":
                    dias = {"√öltimos 7 dias": 7, "√öltimos 30 dias": 30, "√öltimos 90 dias": 90}[dias_filtro]
                    data_limite = datetime.now() - timedelta(days=dias)
                    historico_filtrado = [alt for alt in historico_filtrado
                                        if datetime.fromisoformat(alt['timestamp']) >= data_limite]
             
                # Mostrar hist√≥rico
                if historico_filtrado:
                    st.subheader(f"üìã Hist√≥rico Filtrado ({len(historico_filtrado)} registros)")
                 
                    # Ordenar por timestamp (mais recente primeiro)
                    historico_filtrado.sort(key=lambda x: x['timestamp'], reverse=True)
                 
                    for i, alt in enumerate(historico_filtrado[:20]): # Mostrar apenas os 20 mais recentes
                        with st.expander(f"{alt['tipo'].title()} - {alt['cnpj']} - {alt['timestamp'][:19]}"):
                            col1, col2 = st.columns(2)
                         
                            with col1:
                                st.write(f"**Tipo:** {alt['tipo'].title()}")
                                st.write(f"**CNPJ:** {alt['cnpj']}")
                                st.write(f"**Data/Hora:** {alt['timestamp'][:19]}")
                                st.write(f"**Usu√°rio:** {alt.get('usuario', 'N/A')}")
                         
                            with col2:
                                if alt['tipo'] == 'edicao':
                                    st.write(f"**Campo:** {alt.get('campo_alterado', 'N/A')}")
                                    st.write(f"**De:** {alt.get('valor_anterior', 'N/A')}")
                                    st.write(f"**Para:** {alt.get('valor_novo', 'N/A')}")
                             
                                if alt.get('observacoes'):
                                    st.write(f"**Observa√ß√µes:** {alt['observacoes']}")
                 
                    # Bot√£o para exportar hist√≥rico
                    if st.button("üì• Exportar Hist√≥rico CSV"):
                        try:
                            caminho_export = history_manager.exportar_historico_csv()
                            if caminho_export:
                                st.toast(f"‚úÖ Hist√≥rico exportado: {caminho_export}")
                        except Exception as e:
                            st.error(f"‚ùå Erro ao exportar hist√≥rico: {e}")
                else:
                    st.info("‚ÑπÔ∏è Nenhum registro encontrado com os filtros aplicados")
            else:
                st.info("‚ÑπÔ∏è Nenhuma altera√ß√£o registrada ainda")
    
    # ========================================
    # AN√ÅLISE DE CONCORRENTE (GRALAB)
    # ========================================
    elif st.session_state.page == "üîç An√°lise de Concorrente":
        st.header("üîç An√°lise de Concorrente - Gralab (CunhaLab)")
        
        st.markdown("""
        <div style="background: linear-gradient(135deg, #ffd700 0%, #ffed4e 100%);
                    color: #333; padding: 1.5rem; border-radius: 10px;
                    margin-bottom: 2rem; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
            <h3 style="margin: 0; font-size: 1.4rem; color: #b8860b;">üìä An√°lise Comparativa de Mercado</h3>
            <p style="margin: 0.5rem 0 0 0; font-size: 1rem; color: #666;">
                Compare nossa base de laborat√≥rios com o concorrente Gralab (CunhaLab) para identificar oportunidades e amea√ßas.
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        # Carregar dados
        with st.spinner("Carregando dados do Gralab (CunhaLab)..."):
            dados_gralab = DataManager.carregar_dados_gralab()
        
        if not dados_gralab or 'Dados Completos' not in dados_gralab:
            st.error("‚ùå N√£o foi poss√≠vel carregar os dados do Gralab. Verifique a conex√£o com o SharePoint.")
        else:
            df_gralab = dados_gralab['Dados Completos']
            
            # Normalizar CNPJs da nossa base (usar df completo, n√£o df_filtrado)
            # Isso garante que todos os nossos clientes sejam considerados na compara√ß√£o
            if 'CNPJ_Normalizado' not in df.columns:
                df['CNPJ_Normalizado'] = df['CNPJ_PCL'].apply(DataManager.normalizar_cnpj)
            
            # Obter conjuntos de CNPJs (usar df completo para ter todos os clientes)
            cnpjs_nossos = set(df['CNPJ_Normalizado'].dropna().unique())
            cnpjs_gralab = set(df_gralab['CNPJ_Normalizado'].dropna().unique())
            
            # Calcular intersec√ß√µes
            cnpjs_comuns = cnpjs_nossos & cnpjs_gralab
            cnpjs_so_nossos = cnpjs_nossos - cnpjs_gralab
            cnpjs_so_gralab = cnpjs_gralab - cnpjs_nossos
            
            # ========================================
            # KPIs COMPARATIVOS
            # ========================================
            st.subheader("üìä Vis√£o Geral Comparativa")
            
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                pct_comuns = (len(cnpjs_comuns) / len(cnpjs_nossos) * 100) if len(cnpjs_nossos) > 0 else 0
                st.metric(
                    label="ü§ù Labs em Comum",
                    value=f"{len(cnpjs_comuns)}",
                    delta=f"{pct_comuns:.1f}% da nossa base"
                )
            
            with col2:
                pct_exclusivos_nossos = (len(cnpjs_so_nossos) / len(cnpjs_nossos) * 100) if len(cnpjs_nossos) > 0 else 0
                st.metric(
                    label="üîµ Exclusivos Nossos",
                    value=f"{len(cnpjs_so_nossos)}",
                    delta=f"{pct_exclusivos_nossos:.1f}% da nossa base"
                )
            
            with col3:
                pct_exclusivos_gralab = (len(cnpjs_so_gralab) / len(cnpjs_gralab) * 100) if len(cnpjs_gralab) > 0 else 0
                st.metric(
                    label="üü† Exclusivos Gralab",
                    value=f"{len(cnpjs_so_gralab)}",
                    delta=f"{pct_exclusivos_gralab:.1f}% do Gralab"
                )
            
            with col4:
                st.metric(
                    label="üìä Total Gralab",
                    value=f"{len(cnpjs_gralab)}",
                    delta=f"vs {len(cnpjs_nossos)} nossos"
                )
            
            # ========================================
            # GR√ÅFICOS COMPARATIVOS
            # ========================================
            st.markdown("---")
            st.subheader("üìà An√°lise Visual")
            
            col_g1, col_g2 = st.columns(2)
            
            with col_g1:
                # Gr√°fico de Pizza - Distribui√ß√£o
                import plotly.graph_objects as go
                
                fig_pizza = go.Figure(data=[go.Pie(
                    labels=['Em Comum', 'S√≥ Nossos', 'S√≥ Gralab'],
                    values=[len(cnpjs_comuns), len(cnpjs_so_nossos), len(cnpjs_so_gralab)],
                    marker=dict(colors=['#6BBF47', '#3B82F6', '#FB923C']),
                    hole=0.4,
                    textinfo='label+percent+value',
                    textposition='outside'
                )])
                
                fig_pizza.update_layout(
                    title="Distribui√ß√£o de Laborat√≥rios",
                    height=400,
                    showlegend=True
                )
                
                st.plotly_chart(fig_pizza, width='stretch')
            
            with col_g2:
                # Gr√°fico de Barras - Top UFs em Comum
                if len(cnpjs_comuns) > 0:
                    # Usar df completo para an√°lise geogr√°fica completa
                    df_comuns = df[df['CNPJ_Normalizado'].isin(cnpjs_comuns)]
                    top_ufs = df_comuns['Estado'].value_counts().head(10)
                    
                    import plotly.express as px
                    
                    fig_ufs = px.bar(
                        x=top_ufs.index,
                        y=top_ufs.values,
                        labels={'x': 'UF', 'y': 'Quantidade'},
                        title="Top 10 UFs com Labs em Comum",
                        color=top_ufs.values,
                        color_continuous_scale='Greens'
                    )
                    
                    fig_ufs.update_layout(
                        height=400,
                        showlegend=False,
                        xaxis_title="Estado",
                        yaxis_title="Quantidade de Laborat√≥rios"
                    )
                    
                    st.plotly_chart(fig_ufs, width='stretch')
                else:
                    st.info("Nenhum laborat√≥rio em comum para an√°lise geogr√°fica")
            
            # ========================================
            # TABELAS DETALHADAS COM TABS
            # ========================================
            st.markdown("---")
            st.subheader("üìã An√°lise Detalhada")
            
            tab1, tab2, tab3, tab4, tab5 = st.tabs([
                "ü§ù Labs em Comum",
                "üîµ Exclusivos Nossos",
                "üü† Exclusivos Gralab",
                "üîÑ Movimenta√ß√µes",
                "üí∞ An√°lise de Pre√ßos"
            ])
            
            with tab1:
                st.markdown("### ü§ù Laborat√≥rios em Ambas as Bases")
                
                if len(cnpjs_comuns) > 0:
                    # Criar DataFrame combinado (usar df completo para pegar todos os labs)
                    df_comuns_nossos = df[df['CNPJ_Normalizado'].isin(cnpjs_comuns)][
                        ['CNPJ_PCL', 'CNPJ_Normalizado', 'Nome_Fantasia_PCL', 'Cidade', 'Estado']
                    ].drop_duplicates('CNPJ_Normalizado')
                    
                    # Selecionar colunas dispon√≠veis do Gralab
                    colunas_gralab_desejadas = ['CNPJ_Normalizado', 'Nome', 'Cidade', 'UF', 'Pre√ßo CNH', 'Pre√ßo Concurso', 'Pre√ßo CLT']
                    colunas_gralab_disponiveis = [col for col in colunas_gralab_desejadas if col in df_gralab.columns]
                    
                    df_comuns_gralab = df_gralab[df_gralab['CNPJ_Normalizado'].isin(cnpjs_comuns)][
                        colunas_gralab_disponiveis
                    ]
                    
                    # Merge
                    df_comparacao = pd.merge(
                        df_comuns_nossos,
                        df_comuns_gralab,
                        on='CNPJ_Normalizado',
                        how='inner',
                        suffixes=('_Nosso', '_Gralab')
                    )
                    
                    # Filtros
                    col_f1, col_f2 = st.columns(2)
                    with col_f1:
                        ufs_disponiveis = ['Todos'] + sorted(df_comparacao['Estado'].dropna().unique().tolist())
                        uf_filtro = st.selectbox("Filtrar por UF:", ufs_disponiveis, key="uf_comuns")
                    
                    with col_f2:
                        if uf_filtro != 'Todos':
                            df_temp = df_comparacao[df_comparacao['Estado'] == uf_filtro]
                            cidades_disponiveis = ['Todas'] + sorted(df_temp['Cidade_Nosso'].dropna().unique().tolist())
                        else:
                            cidades_disponiveis = ['Todas'] + sorted(df_comparacao['Cidade_Nosso'].dropna().unique().tolist())
                        cidade_filtro = st.selectbox("Filtrar por Cidade:", cidades_disponiveis, key="cidade_comuns")
                    
                    # Aplicar filtros
                    df_exibir = df_comparacao.copy()
                    if uf_filtro != 'Todos':
                        df_exibir = df_exibir[df_exibir['Estado'] == uf_filtro]
                    if cidade_filtro != 'Todas':
                        df_exibir = df_exibir[df_exibir['Cidade_Nosso'] == cidade_filtro]
                    
                    # Selecionar colunas dispon√≠veis para exibi√ß√£o
                    colunas_exibir = ['CNPJ_PCL', 'Nome_Fantasia_PCL']
                    if 'Nome' in df_exibir.columns:
                        colunas_exibir.append('Nome')
                    colunas_exibir.extend(['Cidade_Nosso', 'Estado'])
                    
                    # Adicionar colunas de pre√ßo se dispon√≠veis
                    for col_preco in ['Pre√ßo CNH', 'Pre√ßo Concurso', 'Pre√ßo CLT']:
                        if col_preco in df_exibir.columns:
                            colunas_exibir.append(col_preco)
                    
                    # Renomear colunas para exibi√ß√£o
                    df_exibir_final = df_exibir[colunas_exibir].copy()
                    
                    rename_map = {
                        'CNPJ_PCL': 'CNPJ',
                        'Nome_Fantasia_PCL': 'Nome (Nossa Base)',
                        'Nome': 'Nome (Gralab/CunhaLab)',
                        'Cidade_Nosso': 'Cidade',
                        'Estado': 'UF',
                        'Pre√ßo CNH': 'Pre√ßo CNH (Gralab/CunhaLab)',
                        'Pre√ßo Concurso': 'Pre√ßo Concurso (Gralab/CunhaLab)',
                        'Pre√ßo CLT': 'Pre√ßo CLT (Gralab/CunhaLab)'
                    }
                    
                    df_exibir_final = df_exibir_final.rename(columns={k: v for k, v in rename_map.items() if k in df_exibir_final.columns})
                    
                    st.dataframe(df_exibir_final, width='stretch', height=400, hide_index=True)
                    
                    # Bot√µes de download
                    col_d1, col_d2 = st.columns(2)
                    
                    with col_d1:
                        csv = df_exibir_final.to_csv(index=False, encoding='utf-8-sig')
                        st.download_button(
                            label="üì• Download CSV",
                            data=csv,
                            file_name=f"labs_em_comum_gralab_{datetime.now().strftime('%Y%m%d')}.csv",
                            mime="text/csv",
                            key="download_comuns_csv"
                        )
                    
                    with col_d2:
                        excel_buffer = BytesIO()
                        df_exibir_final.to_excel(excel_buffer, index=False, engine='openpyxl')
                        excel_data = excel_buffer.getvalue()
                        st.download_button(
                            label="üìä Download Excel",
                            data=excel_data,
                            file_name=f"labs_em_comum_gralab_{datetime.now().strftime('%Y%m%d')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_comuns_excel"
                        )
                else:
                    st.info("Nenhum laborat√≥rio em comum encontrado")
            
            with tab2:
                st.markdown("### üîµ Laborat√≥rios Exclusivos da Nossa Base")
                st.caption("Laborat√≥rios que temos mas o Gralab (CunhaLab) n√£o tem - potencial para prote√ß√£o")
                
                if len(cnpjs_so_nossos) > 0:
                    # Usar df completo para pegar todos os labs exclusivos nossos
                    df_exclusivos_nossos = df[df['CNPJ_Normalizado'].isin(cnpjs_so_nossos)][
                        ['CNPJ_PCL', 'Nome_Fantasia_PCL', 'Cidade', 'Estado', 'Vol_Hoje', 'Risco_Diario']
                    ].drop_duplicates('CNPJ_PCL')
                    
                    df_exclusivos_nossos = df_exclusivos_nossos.rename(columns={
                        'CNPJ_PCL': 'CNPJ',
                        'Nome_Fantasia_PCL': 'Nome',
                        'Estado': 'UF',
                        'Vol_Hoje': 'Volume Hoje',
                        'Risco_Diario': 'Risco'
                    })
                    
                    st.dataframe(df_exclusivos_nossos, width='stretch', height=400, hide_index=True)
                    
                    # Bot√µes de download
                    col_d1, col_d2 = st.columns(2)
                    
                    with col_d1:
                        csv = df_exclusivos_nossos.to_csv(index=False, encoding='utf-8-sig')
                        st.download_button(
                            label="üì• Download CSV",
                            data=csv,
                            file_name=f"labs_exclusivos_nossos_{datetime.now().strftime('%Y%m%d')}.csv",
                            mime="text/csv",
                            key="download_exclusivos_nossos_csv"
                        )
                    
                    with col_d2:
                        excel_buffer = BytesIO()
                        df_exclusivos_nossos.to_excel(excel_buffer, index=False, engine='openpyxl')
                        excel_data = excel_buffer.getvalue()
                        st.download_button(
                            label="üìä Download Excel",
                            data=excel_data,
                            file_name=f"labs_exclusivos_nossos_{datetime.now().strftime('%Y%m%d')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_exclusivos_nossos_excel"
                        )
                else:
                    st.info("Nenhum laborat√≥rio exclusivo encontrado")
            
            with tab3:
                st.markdown("### üü† Laborat√≥rios Exclusivos do Gralab (CunhaLab)")
                st.caption("Laborat√≥rios que o Gralab (CunhaLab) tem mas n√£o temos - oportunidade de prospec√ß√£o")
                
                if len(cnpjs_so_gralab) > 0:
                    # Filtrar labs exclusivos do Gralab
                    df_exclusivos_gralab = df_gralab[df_gralab['CNPJ_Normalizado'].isin(cnpjs_so_gralab)].copy()
                    
                    # Selecionar colunas dispon√≠veis - sempre incluir CNPJ_Normalizado
                    colunas_disponiveis = []
                    
                    # Verificar se tem coluna CNPJ ou usar CNPJ_Normalizado
                    if 'CNPJ' in df_exclusivos_gralab.columns:
                        colunas_disponiveis.append('CNPJ')
                    elif 'CNPJ_Normalizado' in df_exclusivos_gralab.columns:
                        colunas_disponiveis.append('CNPJ_Normalizado')
                    
                    # Adicionar outras colunas desejadas
                    colunas_desejadas = ['Nome', 'Cidade', 'UF', 'Telefone', 'Pre√ßo CNH', 'Pre√ßo Concurso', 'Pre√ßo CLT']
                    
                    for col in colunas_desejadas:
                        if col in df_exclusivos_gralab.columns:
                            colunas_disponiveis.append(col)
                    
                    if colunas_disponiveis:
                        df_exclusivos_gralab_filtrado = df_exclusivos_gralab[colunas_disponiveis].copy()
                        
                        # Renomear CNPJ_Normalizado para CNPJ se necess√°rio
                        if 'CNPJ_Normalizado' in df_exclusivos_gralab_filtrado.columns and 'CNPJ' not in df_exclusivos_gralab_filtrado.columns:
                            df_exclusivos_gralab_filtrado = df_exclusivos_gralab_filtrado.rename(columns={'CNPJ_Normalizado': 'CNPJ'})
                        
                        # Usar CNPJ para drop_duplicates
                        if 'CNPJ' in df_exclusivos_gralab_filtrado.columns:
                            df_exclusivos_gralab_filtrado = df_exclusivos_gralab_filtrado.drop_duplicates('CNPJ')
                        
                        st.dataframe(df_exclusivos_gralab_filtrado, width='stretch', height=400, hide_index=True)
                        
                        # Bot√µes de download
                        col_d1, col_d2 = st.columns(2)
                        
                        with col_d1:
                            csv = df_exclusivos_gralab_filtrado.to_csv(index=False, encoding='utf-8-sig')
                            st.download_button(
                                label="üì• Download CSV",
                                data=csv,
                                file_name=f"labs_exclusivos_gralab_{datetime.now().strftime('%Y%m%d')}.csv",
                                mime="text/csv",
                                key="download_exclusivos_gralab_csv"
                            )
                        
                        with col_d2:
                            excel_buffer = BytesIO()
                            df_exclusivos_gralab_filtrado.to_excel(excel_buffer, index=False, engine='openpyxl')
                            excel_data = excel_buffer.getvalue()
                            st.download_button(
                                label="üìä Download Excel",
                                data=excel_data,
                                file_name=f"labs_exclusivos_gralab_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="download_exclusivos_gralab_excel"
                            )
                    else:
                        st.warning("‚ö†Ô∏è Colunas esperadas n√£o encontradas no arquivo do Gralab")
                else:
                    st.info("Nenhum laborat√≥rio exclusivo do Gralab (CunhaLab) encontrado")
            
            with tab4:
                st.markdown("### üîÑ Movimenta√ß√µes do Gralab (CunhaLab)")
                st.caption("Credenciamentos e descredenciamentos registrados")
                
                if 'EntradaSaida' in dados_gralab:
                    df_entrada_saida = dados_gralab['EntradaSaida'].copy()
                    
                    if not df_entrada_saida.empty:
                        # Formatar datas para padr√£o brasileiro
                        for col_data in ['Data Entrada', 'Data Sa√≠da', '√öltima Verifica√ß√£o']:
                            if col_data in df_entrada_saida.columns:
                                df_entrada_saida[col_data] = pd.to_datetime(df_entrada_saida[col_data], errors='coerce').dt.strftime('%d/%m/%Y')
                                df_entrada_saida[col_data] = df_entrada_saida[col_data].replace('NaT', '')
                        
                        # Filtros
                        col_mov1, col_mov2 = st.columns(2)
                        
                        with col_mov1:
                            tipo_mov_filtro = st.multiselect(
                                "Tipo de Movimenta√ß√£o:",
                                options=['Todos', 'Credenciamento', 'Descredenciamento'],
                                default=['Todos'],
                                key="tipo_mov_gralab"
                            )
                        
                        with col_mov2:
                            if 'UF' in df_entrada_saida.columns:
                                uf_mov_filtro = st.multiselect(
                                    "UF:",
                                    options=['Todos'] + sorted(df_entrada_saida['UF'].dropna().unique().tolist()),
                                    default=['Todos'],
                                    key="uf_mov_gralab"
                                )
                        
                        # Aplicar filtros
                        df_mov_filtrado = df_entrada_saida.copy()
                        
                        if tipo_mov_filtro and 'Todos' not in tipo_mov_filtro:
                            df_mov_filtrado = df_mov_filtrado[df_mov_filtrado['Tipo Movimenta√ß√£o'].isin(tipo_mov_filtro)]
                        
                        if 'UF' in df_entrada_saida.columns and uf_mov_filtro and 'Todos' not in uf_mov_filtro:
                            df_mov_filtrado = df_mov_filtrado[df_mov_filtrado['UF'].isin(uf_mov_filtro)]
                        
                        # Selecionar colunas para exibi√ß√£o
                        colunas_exibir_mov = []
                        
                        # Sempre tentar incluir CNPJ
                        if 'CNPJ' in df_mov_filtrado.columns:
                            colunas_exibir_mov.append('CNPJ')
                        elif 'CNPJ_Normalizado' in df_mov_filtrado.columns:
                            colunas_exibir_mov.append('CNPJ_Normalizado')
                        
                        # Outras colunas importantes
                        colunas_desejadas_mov = [
                            'Nome', 'Cidade', 'UF', 'Data Entrada', 'Data Sa√≠da', 
                            'Tipo Movimenta√ß√£o', '√öltima Verifica√ß√£o',
                            'Pre√ßo CNH', 'Pre√ßo Concurso', 'Pre√ßo CLT'
                        ]
                        
                        for col in colunas_desejadas_mov:
                            if col in df_mov_filtrado.columns:
                                colunas_exibir_mov.append(col)
                        
                        if colunas_exibir_mov:
                            df_mov_exibir = df_mov_filtrado[colunas_exibir_mov].copy()
                            
                            # Renomear CNPJ_Normalizado se necess√°rio
                            if 'CNPJ_Normalizado' in df_mov_exibir.columns and 'CNPJ' not in df_mov_exibir.columns:
                                df_mov_exibir = df_mov_exibir.rename(columns={'CNPJ_Normalizado': 'CNPJ'})
                            
                            # Adicionar coluna indicando se √© nosso cliente
                            def verificar_nosso_cliente(cnpj):
                                if pd.isna(cnpj) or cnpj == '':
                                    return 'N/A'
                                # Normalizar CNPJ para compara√ß√£o
                                cnpj_normalizado = DataManager.normalizar_cnpj(str(cnpj))
                                if cnpj_normalizado in cnpjs_nossos:
                                    return '‚úÖ Sim'
                                return '‚ùå N√£o'
                            
                            if 'CNPJ' in df_mov_exibir.columns:
                                df_mov_exibir['Nosso Cliente'] = df_mov_exibir['CNPJ'].apply(verificar_nosso_cliente)
                            
                            # M√©tricas resumidas
                            col_m1, col_m2, col_m3 = st.columns(3)
                            
                            total_movimentacoes = len(df_mov_exibir)
                            credenciamentos = len(df_mov_exibir[df_mov_exibir['Tipo Movimenta√ß√£o'] == 'Credenciamento']) if 'Tipo Movimenta√ß√£o' in df_mov_exibir.columns else 0
                            descredenciamentos = len(df_mov_exibir[df_mov_exibir['Tipo Movimenta√ß√£o'] == 'Descredenciamento']) if 'Tipo Movimenta√ß√£o' in df_mov_exibir.columns else 0
                            
                            with col_m1:
                                st.metric("Total Movimenta√ß√µes", total_movimentacoes)
                            with col_m2:
                                st.metric("‚úÖ Credenciamentos", credenciamentos)
                            with col_m3:
                                st.metric("‚ùå Descredenciamentos", descredenciamentos)
                            
                            st.markdown("---")
                            
                            # Tabela
                            st.dataframe(df_mov_exibir, width='stretch', height=400, hide_index=True)
                            
                            # Bot√µes de download
                            col_d1, col_d2 = st.columns(2)
                            
                            with col_d1:
                                csv = df_mov_exibir.to_csv(index=False, encoding='utf-8-sig')
                                st.download_button(
                                    label="üì• Download CSV",
                                    data=csv,
                                    file_name=f"movimentacoes_gralab_{datetime.now().strftime('%Y%m%d')}.csv",
                                    mime="text/csv",
                                    key="download_movimentacoes_gralab_csv"
                                )
                            
                            with col_d2:
                                excel_buffer = BytesIO()
                                df_mov_exibir.to_excel(excel_buffer, index=False, engine='openpyxl')
                                excel_data = excel_buffer.getvalue()
                                st.download_button(
                                    label="üìä Download Excel",
                                    data=excel_data,
                                    file_name=f"movimentacoes_gralab_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key="download_movimentacoes_gralab_excel"
                                )
                        else:
                            st.warning("‚ö†Ô∏è Nenhuma coluna dispon√≠vel para exibi√ß√£o")
                    else:
                        st.info("Nenhuma movimenta√ß√£o registrada")
                else:
                    st.info("Aba 'EntradaSaida' n√£o encontrada no arquivo do Gralab (CunhaLab)")
            
            with tab5:
                st.markdown("### üí∞ An√°lise de Pre√ßos (Labs em Comum)")
                
                if len(cnpjs_comuns) > 0:
                    df_precos = df_gralab[df_gralab['CNPJ_Normalizado'].isin(cnpjs_comuns)].copy()
                    
                    # Converter pre√ßos para num√©rico
                    for col in ['Pre√ßo CNH', 'Pre√ßo Concurso', 'Pre√ßo CLT']:
                        if col in df_precos.columns:
                            df_precos[col] = pd.to_numeric(df_precos[col], errors='coerce')
                    
                    # Estat√≠sticas
                    col_s1, col_s2, col_s3 = st.columns(3)
                    
                    with col_s1:
                        st.markdown("#### üé´ CNH")
                        if 'Pre√ßo CNH' in df_precos.columns:
                            precos_cnh = df_precos['Pre√ßo CNH'].dropna()
                            if len(precos_cnh) > 0:
                                st.metric("M√©dia", f"R$ {precos_cnh.mean():.2f}")
                                st.metric("Mediana", f"R$ {precos_cnh.median():.2f}")
                                st.metric("M√≠n / M√°x", f"R$ {precos_cnh.min():.2f} / R$ {precos_cnh.max():.2f}")
                            else:
                                st.info("Sem dados")
                    
                    with col_s2:
                        st.markdown("#### üìù Concurso")
                        if 'Pre√ßo Concurso' in df_precos.columns:
                            precos_concurso = df_precos['Pre√ßo Concurso'].dropna()
                            if len(precos_concurso) > 0:
                                st.metric("M√©dia", f"R$ {precos_concurso.mean():.2f}")
                                st.metric("Mediana", f"R$ {precos_concurso.median():.2f}")
                                st.metric("M√≠n / M√°x", f"R$ {precos_concurso.min():.2f} / R$ {precos_concurso.max():.2f}")
                            else:
                                st.info("Sem dados")
                    
                    with col_s3:
                        st.markdown("#### üëî CLT")
                        if 'Pre√ßo CLT' in df_precos.columns:
                            precos_clt = df_precos['Pre√ßo CLT'].dropna()
                            if len(precos_clt) > 0:
                                st.metric("M√©dia", f"R$ {precos_clt.mean():.2f}")
                                st.metric("Mediana", f"R$ {precos_clt.median():.2f}")
                                st.metric("M√≠n / M√°x", f"R$ {precos_clt.min():.2f} / R$ {precos_clt.max():.2f}")
                            else:
                                st.info("Sem dados")
                    
                    # Boxplot de distribui√ß√£o
                    st.markdown("---")
                    st.markdown("#### üìä Distribui√ß√£o de Pre√ßos")
                    
                    # Preparar dados para boxplot
                    dados_boxplot = []
                    for col, nome in [('Pre√ßo CNH', 'CNH'), ('Pre√ßo Concurso', 'Concurso'), ('Pre√ßo CLT', 'CLT')]:
                        if col in df_precos.columns:
                            valores = df_precos[col].dropna()
                            for valor in valores:
                                dados_boxplot.append({'Tipo': nome, 'Pre√ßo': valor})
                    
                    if dados_boxplot:
                        df_boxplot = pd.DataFrame(dados_boxplot)
                        
                        import plotly.express as px
                        fig_box = px.box(
                            df_boxplot,
                            x='Tipo',
                            y='Pre√ßo',
                            color='Tipo',
                            title="Distribui√ß√£o de Pre√ßos por Tipo de Exame",
                            labels={'Pre√ßo': 'Pre√ßo (R$)', 'Tipo': 'Tipo de Exame'}
                        )
                        
                        fig_box.update_layout(height=400, showlegend=False)
                        st.plotly_chart(fig_box, width='stretch')
                    else:
                        st.info("Sem dados de pre√ßos dispon√≠veis para an√°lise")
                else:
                    st.info("Nenhum laborat√≥rio em comum para an√°lise de pre√ßos")
    
    st.markdown("""
    <div class="footer">
        <p>üìä <strong>Syntox Churn</strong> - Dashboard profissional de an√°lise de reten√ß√£o de laborat√≥rios</p>
        <p>Desenvolvido com ‚ù§Ô∏è para otimizar a gest√£o de relacionamento com PCLs</p>
    </div>
    """, unsafe_allow_html=True)
if __name__ == "__main__":
    main()